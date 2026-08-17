## -*- coding: utf-8 -*-
import os
import glob
import copy
import json
import argparse
import itertools
import re
from collections import Counter
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import math
import cv2
import numpy as np
import pandas as pd
import trimesh
from tqdm import tqdm
from scipy.spatial import cKDTree
from scipy.optimize import minimize
from scipy.spatial.transform import Rotation as SciPyRotation

import nvdiffrast.torch as dr
import torch
from pytorch_lightning import seed_everything

from sam2.sam2.build_sam import build_sam2
from sam2.sam2.sam2_image_predictor import SAM2ImagePredictor

from estimater import MultiView, ScorePredictor, PoseRefinePredictor


SCRIPT_BUILD_ID = "2026-07-27-simple-first-frame-axis-v12"


# -----------------------------------------------------------------------------
# Geometry and visualization utilities
# -----------------------------------------------------------------------------

def as_mesh(mesh_or_scene) -> trimesh.Trimesh:
    """Convert a Trimesh or Scene into one Trimesh."""
    if isinstance(mesh_or_scene, trimesh.Trimesh):
        return mesh_or_scene
    if isinstance(mesh_or_scene, trimesh.Scene):
        geoms = [g for g in mesh_or_scene.geometry.values() if isinstance(g, trimesh.Trimesh) and len(g.faces) > 0]
        if not geoms:
            raise ValueError("The loaded mesh scene contains no valid TriangleMesh geometry.")
        return trimesh.util.concatenate(geoms)
    raise TypeError(f"Unsupported mesh type: {type(mesh_or_scene)}")


def mesh_diameter(mesh: trimesh.Trimesh, exact_limit: int = 5000) -> float:
    """Return the maximum pairwise model-point distance in model units.

    An AABB diagonal is only an upper bound and changes every
    diameter-normalized BOP threshold. The fallback below evaluates convex-hull
    vertices. If the hull is unusually dense, it uses a deterministic subset;
    publication runs should then pass --model_diameter_m or --models_info.
    """
    try:
        vertices = np.asarray(mesh.convex_hull.vertices, dtype=np.float64)
    except Exception:
        vertices = np.asarray(mesh.vertices, dtype=np.float64)
    if len(vertices) < 2:
        return 0.0

    if len(vertices) > exact_limit:
        rng = np.random.default_rng(0)
        sampled_idx = rng.choice(len(vertices), size=exact_limit, replace=False)
        extrema_idx = np.unique(np.concatenate([
            np.argmin(vertices, axis=0),
            np.argmax(vertices, axis=0),
        ]))
        vertices = vertices[np.unique(np.concatenate([sampled_idx, extrema_idx]))]
        print(
            f"[Warning] Convex hull has more than {exact_limit} vertices; "
            "the fallback diameter is approximate. Use --model_diameter_m "
            "or --models_info for an official diameter."
        )

    max_d2 = 0.0
    block = 256
    for start in range(0, len(vertices), block):
        a = vertices[start:start + block]
        d2 = np.sum((a[:, None, :] - vertices[None, :, :]) ** 2, axis=2)
        max_d2 = max(max_d2, float(np.max(d2)))
    return float(np.sqrt(max_d2))


def ensure_multiview_mesh_visuals(
    mesh: trimesh.Trimesh,
    fallback_rgb: Tuple[int, int, int] = (128, 128, 128),
) -> Tuple[trimesh.Trimesh, str]:
    """Guarantee visuals accepted by FoundationPose ``make_mesh_tensors``.

    MultiView follows FoundationPose and assumes every ``TextureVisuals`` instance
    has both a readable material image and one UV coordinate per vertex. Some
    geometry-only OBJ files are nevertheless loaded by trimesh as textured
    meshes whose ``material.image`` is ``None``. That reaches
    ``None.convert('RGB')`` and crashes before registration starts.

    A missing or malformed texture is converted to deterministic neutral-gray
    per-vertex colors. FoundationPose then takes its supported vertex-color
    path and no texture image or UV coordinates are required. Valid texture
    data is preserved unchanged.
    """
    if not isinstance(mesh, trimesh.Trimesh):
        raise TypeError(f"Expected trimesh.Trimesh, got {type(mesh)}")
    rgb = np.asarray(fallback_rgb, dtype=np.int64).reshape(-1)
    if rgb.shape != (3,) or np.any(rgb < 0) or np.any(rgb > 255):
        raise ValueError(
            f"fallback_rgb must contain three values in [0,255], got {fallback_rgb}"
        )

    visual = mesh.visual
    if isinstance(visual, trimesh.visual.texture.TextureVisuals):
        material = getattr(visual, "material", None)
        image = getattr(material, "image", None)
        uv = getattr(visual, "uv", None)
        uv_array = None if uv is None else np.asarray(uv)
        texture_ok = (
            image is not None
            and callable(getattr(image, "convert", None))
            and uv_array is not None
            and uv_array.shape == (len(mesh.vertices), 2)
            and np.all(np.isfinite(uv_array))
        )
        if texture_ok:
            return mesh, "texture"

        rgba = np.tile(
            np.array([rgb[0], rgb[1], rgb[2], 255], dtype=np.uint8),
            (len(mesh.vertices), 1),
        )
        mesh.visual = trimesh.visual.ColorVisuals(
            mesh=mesh,
            vertex_colors=rgba,
        )
        reason = []
        if image is None:
            reason.append("material.image=None")
        elif not callable(getattr(image, "convert", None)):
            reason.append("material.image is not PIL-compatible")
        if uv_array is None:
            reason.append("UV missing")
        elif uv_array.shape != (len(mesh.vertices), 2):
            reason.append(
                f"UV shape={uv_array.shape}, expected=({len(mesh.vertices)},2)"
            )
        elif not np.all(np.isfinite(uv_array)):
            reason.append("UV contains NaN/Inf")
        return mesh, "fallback_vertex_color:" + ",".join(reason)

    colors = getattr(visual, "vertex_colors", None)
    color_array = None if colors is None else np.asarray(colors)
    if (
        color_array is None
        or color_array.ndim != 2
        or color_array.shape[0] != len(mesh.vertices)
        or color_array.shape[1] < 3
    ):
        rgba = np.tile(
            np.array([rgb[0], rgb[1], rgb[2], 255], dtype=np.uint8),
            (len(mesh.vertices), 1),
        )
        mesh.visual = trimesh.visual.ColorVisuals(
            mesh=mesh,
            vertex_colors=rgba,
        )
        return mesh, "fallback_vertex_color:missing_or_invalid_vertex_colors"
    return mesh, "vertex_color"


def load_bbox_centered_mesh(mesh_path: str, mesh_scale: float) -> Tuple[trimesh.Trimesh, np.ndarray, float]:
    """Load the explicitly specified OBJ, scale it to metres, and bbox-centre it.

    This is a fixed real-dataset convention, not an automatic mesh-frame guess.
    The returned centre offset is reported only for diagnostics.
    """
    if not os.path.isfile(mesh_path):
        raise FileNotFoundError(f"Mesh file does not exist: {mesh_path}")
    mesh = as_mesh(trimesh.load(mesh_path, process=False)).copy()
    vertices = np.asarray(mesh.vertices, dtype=np.float64) * float(mesh_scale)
    if vertices.ndim != 2 or vertices.shape[1] != 3 or len(vertices) == 0:
        raise ValueError(f"Invalid mesh vertices from {mesh_path}: {vertices.shape}")
    center = 0.5 * (vertices.min(axis=0) + vertices.max(axis=0))
    mesh.vertices = vertices - center.reshape(1, 3)
    mesh, visual_source = ensure_multiview_mesh_visuals(mesh)
    if visual_source.startswith("fallback_vertex_color:"):
        print(
            "[Mesh visual] No usable texture was found; converted the mesh to "
            f"neutral-gray vertex colors for MultiView ({visual_source})."
        )
    diameter = mesh_diameter(mesh)
    return mesh, center.astype(np.float64), diameter


_MODEL_UNIT_TO_METRES = {
    "m": 1.0,
    "meter": 1.0,
    "metre": 1.0,
    "meters": 1.0,
    "metres": 1.0,
    "cm": 1e-2,
    "centimeter": 1e-2,
    "centimetre": 1e-2,
    "mm": 1e-3,
    "millimeter": 1e-3,
    "millimetre": 1e-3,
    "um": 1e-6,
    "micrometer": 1e-6,
    "micrometre": 1e-6,
}


def model_unit_scale_to_metres(unit: str) -> float:
    """Convert the object_pose_setup model-unit label to metres."""
    normalized = str(unit).strip().lower().replace(" ", "")
    if normalized not in _MODEL_UNIT_TO_METRES:
        raise ValueError(
            "Unsupported object_model_unit "
            f"{unit!r}. Expected one of m, cm, mm or um."
        )
    return float(_MODEL_UNIT_TO_METRES[normalized])


def _metadata_matrix(
    value,
    label: str,
    allow_rotation_3x3: bool = False,
) -> np.ndarray:
    """Parse one rigid transform stored in JSON metadata."""
    matrix = np.asarray(value, dtype=np.float64)
    if allow_rotation_3x3 and matrix.shape == (3, 3):
        transform = np.eye(4, dtype=np.float64)
        transform[:3, :3] = matrix
        matrix = transform
    if matrix.shape != (4, 4):
        raise ValueError(f"{label} must be 4x4, got {matrix.shape}")
    return validate_rigid_transform(matrix, label)


def _resolve_metadata_path(
    raw_path: Optional[str],
    dataset_root: str,
    seq_dir: str,
) -> Optional[str]:
    if raw_path is None or not str(raw_path).strip():
        return None
    value = os.path.expanduser(str(raw_path).strip())
    if os.path.isabs(value):
        return os.path.abspath(value)
    candidates = [
        os.path.join(dataset_root, value),
        os.path.join(seq_dir, value),
        os.path.join(os.path.dirname(seq_dir), value),
    ]
    existing = next((path for path in candidates if os.path.exists(path)), None)
    return os.path.abspath(existing or candidates[0])


def _paths_refer_to_same_file(path_a: Optional[str], path_b: Optional[str]) -> bool:
    if not path_a or not path_b:
        return False
    a = os.path.abspath(os.path.expanduser(path_a))
    b = os.path.abspath(os.path.expanduser(path_b))
    try:
        if os.path.exists(a) and os.path.exists(b):
            return bool(os.path.samefile(a, b))
    except OSError:
        pass
    return os.path.normcase(os.path.realpath(a)) == os.path.normcase(
        os.path.realpath(b)
    )


def _model_bbox_scale_score(
    vertices_mesh_m: np.ndarray,
    T_object_mesh: np.ndarray,
    bbox_min_object_m: np.ndarray,
    bbox_max_object_m: np.ndarray,
) -> Tuple[float, float, np.ndarray]:
    """Compare transformed model size with the metadata object bbox.

    Sorted extents are used for the scale test so a valid axis permutation does
    not look like a unit error. The returned ratio is based on AABB diagonals.
    """
    vertices_object = transform_points(vertices_mesh_m, T_object_mesh)
    observed_extent = np.ptp(vertices_object, axis=0)
    expected_extent = (
        np.asarray(bbox_max_object_m, dtype=np.float64)
        - np.asarray(bbox_min_object_m, dtype=np.float64)
    )
    if np.any(expected_extent <= 1e-9):
        raise ValueError(
            "object_bbox_min_m/max_m define a degenerate metadata bbox: "
            f"extent={expected_extent.tolist()}"
        )
    observed_sorted = np.sort(np.maximum(observed_extent, 1e-12))
    expected_sorted = np.sort(np.maximum(expected_extent, 1e-12))
    log_error = float(
        np.mean(np.abs(np.log(observed_sorted / expected_sorted)))
    )
    observed_diag = float(np.linalg.norm(observed_extent))
    expected_diag = float(np.linalg.norm(expected_extent))
    diagonal_ratio = observed_diag / max(expected_diag, 1e-12)
    return log_error, diagonal_ratio, observed_extent


def _axis_alignment_rotation_from_metadata(
    value,
) -> Tuple[Optional[np.ndarray], str]:
    """Parse object_pose_setup.object_model_axis_alignment.

    Capture metadata normally stores labels such as ``x_pos_90``. Matrix
    values and signed-axis maps are accepted as well so the fixed model
    convention is read from metadata instead of inferred from a prediction.
    """
    if value is None:
        return None, "not_provided"
    if isinstance(value, (list, tuple, np.ndarray)):
        rotation = _metadata_matrix(
            value,
            "object_pose_setup.object_model_axis_alignment",
            allow_rotation_3x3=True,
        )[:3, :3]
        return rotation, "metadata_axis_alignment_matrix"

    raw = str(value).strip()
    compact = raw.lower().replace(" ", "")
    normalized = compact.replace("-", "_")
    if normalized in {
        "", "none", "identity", "same", "already_aligned",
        "dataset_gt", "gt", "+x,+y,+z", "x,y,z",
    }:
        return np.eye(3, dtype=np.float64), (
            f"metadata_axis_alignment_label:{raw or 'identity'}"
        )
    if "," in compact or ";" in compact:
        rotation, description = parse_axis_map(compact)
        return rotation, f"metadata_axis_map:{description}"

    match = re.fullmatch(
        r"([xyz])_?(pos|positive|plus|p|neg|negative|minus|n)_?"
        r"(90|180|270)(?:_?deg)?",
        normalized,
    )
    if match is not None:
        axis_name, sign_token, angle_token = match.groups()
        sign = (
            1.0
            if sign_token in {"pos", "positive", "plus", "p"}
            else -1.0
        )
    else:
        compact_match = re.fullmatch(
            r"([xyz])([+-])(90|180|270)(?:deg)?",
            compact.replace("_", ""),
        )
        if compact_match is None:
            raise ValueError(
                "Unsupported object_pose_setup.object_model_axis_alignment "
                f"{value!r}. Expected identity, a signed-axis map, or a label "
                "such as x_pos_90 / x_neg_90."
            )
        axis_name, sign_token, angle_token = compact_match.groups()
        sign = 1.0 if sign_token == "+" else -1.0

    axis = _AXIS_UNIT_VECTORS[axis_name]
    angle_rad = sign * math.radians(float(angle_token))
    rotation = SciPyRotation.from_rotvec(axis * angle_rad).as_matrix()
    return (
        project_rotation_to_so3(rotation),
        f"metadata_axis_alignment_label:{raw}",
    )


def _object_setup_rotation(
    setup: Dict[str, object],
) -> Tuple[np.ndarray, str]:
    label_rotation, label_source = (
        _axis_alignment_rotation_from_metadata(
            setup.get("object_model_axis_alignment")
        )
    )
    matrix_rotation: Optional[np.ndarray] = None
    matrix_source = "not_provided"
    rotation_value = setup.get("T_gt_model_rotation")
    if rotation_value is not None:
        matrix_rotation = _metadata_matrix(
            rotation_value,
            "object_pose_setup.T_gt_model_rotation",
            allow_rotation_3x3=True,
        )[:3, :3]
        matrix_source = "object_pose_setup.T_gt_model_rotation"
    elif setup.get("T_gt_model_m") is not None:
        matrix_rotation = _metadata_matrix(
            setup["T_gt_model_m"],
            "object_pose_setup.T_gt_model_m",
        )[:3, :3]
        matrix_source = "rotation_of_object_pose_setup.T_gt_model_m"

    if label_rotation is not None and matrix_rotation is not None:
        disagreement_deg = rotation_distance_deg_from_matrices(
            label_rotation, matrix_rotation
        )
        if disagreement_deg > 1e-3:
            raise ValueError(
                "object_pose_setup axis metadata is internally inconsistent: "
                "object_model_axis_alignment="
                f"{setup.get('object_model_axis_alignment')!r} and "
                f"{matrix_source} differ by {disagreement_deg:.6f} degree."
            )
        return (
            project_rotation_to_so3(matrix_rotation),
            f"{label_source}+validated_against_{matrix_source}",
        )
    if label_rotation is not None:
        return project_rotation_to_so3(label_rotation), label_source
    if matrix_rotation is not None:
        return project_rotation_to_so3(matrix_rotation), matrix_source
    raise KeyError(
        "object_pose_setup must contain object_model_axis_alignment, "
        "T_gt_model_rotation or T_gt_model_m"
    )


def load_mesh_from_object_pose_setup(
    args,
    setup: Dict[str, object],
    seq_dir: str,
) -> Dict[str, object]:
    """Load one immutable estimator mesh/model-to-GT convention.

    The returned ``T_object_mesh`` always satisfies

        p_GT_object = T_object_mesh @ p_estimator_mesh.

    Model-source rules:

    * ``dataset_gt``: ``dataset_model_path`` is already in metres and in the GT
      object frame, so no extra transform or bbox-centering is applied.
    * ``raw_bbox_centered``: the source OBJ is scaled and bbox-centred here;
      only the metadata axis rotation is applied. In particular the translation
      in ``T_gt_model_m`` is intentionally not applied a second time.
    * ``raw_uncentered``: the source vertices are not re-centred and the full
      metadata ``T_gt_model_m`` is used.

    An explicit --T_object_mesh/--axis_map remains available as a legacy
    override, but it never changes from a prediction or from a camera view.
    """
    if not isinstance(setup, dict) or not setup:
        raise ValueError("object_pose_setup is missing or empty")

    required_bbox = ("object_bbox_min_m", "object_bbox_max_m")
    missing_bbox = [key for key in required_bbox if key not in setup]
    if missing_bbox:
        raise KeyError(
            "object_pose_setup is missing required GT bbox fields: "
            + ", ".join(missing_bbox)
        )
    bbox_min = np.asarray(
        setup["object_bbox_min_m"], dtype=np.float64
    ).reshape(3)
    bbox_max = np.asarray(
        setup["object_bbox_max_m"], dtype=np.float64
    ).reshape(3)
    if not (
        np.all(np.isfinite(bbox_min))
        and np.all(np.isfinite(bbox_max))
        and np.all(bbox_max > bbox_min)
    ):
        raise ValueError(
            "Invalid object_pose_setup bbox: "
            f"min={bbox_min.tolist()}, max={bbox_max.tolist()}"
        )

    mesh_path = os.path.abspath(os.path.expanduser(args.obj_mesh))
    if not os.path.isfile(mesh_path):
        raise FileNotFoundError(f"Mesh file does not exist: {mesh_path}")
    dataset_model_path = _resolve_metadata_path(
        setup.get("dataset_model_path"),
        dataset_root=args.dataset_root,
        seq_dir=seq_dir,
    )
    source_model_path = _resolve_metadata_path(
        setup.get("object_model_path"),
        dataset_root=args.dataset_root,
        seq_dir=seq_dir,
    )

    requested_mode = str(args.model_frame_mode).strip().lower()
    if requested_mode == "auto":
        if _paths_refer_to_same_file(mesh_path, dataset_model_path):
            model_frame_mode = "dataset_gt"
        else:
            origin_mode = str(
                setup.get("object_model_origin_mode", "bbox_center")
            ).strip().lower()
            if origin_mode in {
                "bbox_center", "bbox-centre", "bboxcenter",
                "center", "centre", "centered", "centred",
            }:
                model_frame_mode = "raw_bbox_centered"
            else:
                model_frame_mode = "raw_uncentered"
    else:
        model_frame_mode = requested_mode

    raw_mesh = as_mesh(trimesh.load(mesh_path, process=False)).copy()
    raw_vertices = np.asarray(raw_mesh.vertices, dtype=np.float64)
    if (
        raw_vertices.ndim != 2
        or raw_vertices.shape[1] != 3
        or len(raw_vertices) == 0
        or not np.all(np.isfinite(raw_vertices))
    ):
        raise ValueError(
            f"Invalid mesh vertices from {mesh_path}: {raw_vertices.shape}"
        )

    metadata_unit = str(setup.get("object_model_unit", "")).strip()
    metadata_scale = (
        1.0
        if model_frame_mode == "dataset_gt"
        else model_unit_scale_to_metres(metadata_unit)
    )
    if args.mesh_scale is not None:
        candidate_scales = [float(args.mesh_scale)]
        scale_source = "cli_override_checked_against_metadata_bbox"
    else:
        # Test the metadata-derived scale against common metric unit encodings.
        # The bbox score chooses between metre/mm/cm without applying one global
        # scale to all six objects.
        candidate_scales = []
        for candidate in (metadata_scale, 1.0, 1e-3, 1e-2, 1e-6):
            if not any(
                abs(candidate - existing)
                <= max(1e-15, 1e-12 * abs(existing))
                for existing in candidate_scales
            ):
                candidate_scales.append(float(candidate))
        scale_source = (
            "object_pose_setup.object_model_unit+metadata_bbox_auto_check"
        )

    if (
        model_frame_mode == "raw_uncentered"
        and "T_gt_model_m" not in setup
    ):
        raise KeyError(
            "raw_uncentered model mode requires "
            "object_pose_setup.T_gt_model_m; an uncentred raw OBJ cannot be "
            "placed in the GT object frame from an axis label alone."
        )
    full_T_gt_model = _metadata_matrix(
        setup.get("T_gt_model_m", np.eye(4)),
        "object_pose_setup.T_gt_model_m",
    )
    object_setup_rotation, object_setup_rotation_source = (
        _object_setup_rotation(setup)
    )
    rotation_T = np.eye(4, dtype=np.float64)
    rotation_T[:3, :3] = object_setup_rotation

    explicit_transform_requested = bool(
        args.T_object_mesh is not None
        or args.axis_map not in (None, "", "none")
    )
    explicit_transform = None
    explicit_transform_source = None
    if explicit_transform_requested:
        explicit_transform, explicit_transform_source = load_T_object_mesh(args)

    candidates: List[Dict[str, object]] = []
    for scale in candidate_scales:
        vertices_scaled = raw_vertices * float(scale)
        center_removed = np.zeros(3, dtype=np.float64)
        if model_frame_mode == "raw_bbox_centered":
            center_removed = 0.5 * (
                vertices_scaled.min(axis=0) + vertices_scaled.max(axis=0)
            )
            vertices_mesh = vertices_scaled - center_removed[None, :]
            metadata_T_object_mesh = rotation_T
            transform_source = (
                "metadata_rotation_only_after_script_bbox_center"
            )
        elif model_frame_mode == "raw_uncentered":
            vertices_mesh = vertices_scaled
            metadata_T_object_mesh = full_T_gt_model
            transform_source = "metadata_full_T_gt_model_m"
        elif model_frame_mode == "dataset_gt":
            vertices_mesh = vertices_scaled
            metadata_T_object_mesh = np.eye(4, dtype=np.float64)
            transform_source = "dataset_model_already_in_gt_frame"
        else:
            raise ValueError(
                "--model_frame_mode must be auto, dataset_gt, "
                "raw_bbox_centered or raw_uncentered"
            )

        T_object_mesh = (
            explicit_transform
            if explicit_transform is not None
            else metadata_T_object_mesh
        )
        score, ratio, observed_extent = _model_bbox_scale_score(
            vertices_mesh,
            T_object_mesh,
            bbox_min,
            bbox_max,
        )
        candidates.append({
            "scale": float(scale),
            "vertices_mesh": vertices_mesh,
            "center_removed": center_removed,
            "T_object_mesh": np.asarray(
                T_object_mesh, dtype=np.float64
            ).copy(),
            "transform_source": (
                f"explicit:{explicit_transform_source}"
                if explicit_transform is not None
                else transform_source
            ),
            "bbox_score": float(score),
            "bbox_diagonal_ratio": float(ratio),
            "observed_bbox_extent_m": observed_extent,
        })

    selected = min(
        candidates,
        key=lambda item: (
            float(item["bbox_score"]),
            0 if np.isclose(item["scale"], metadata_scale) else 1,
        ),
    )
    resolved_scale = float(selected["scale"])
    bbox_ratio = float(selected["bbox_diagonal_ratio"])
    tolerance = float(args.model_bbox_scale_tolerance)
    if bbox_ratio < 1.0 / tolerance or bbox_ratio > tolerance:
        trials = ", ".join(
            f"scale={item['scale']:g}:ratio="
            f"{item['bbox_diagonal_ratio']:.6g}"
            for item in candidates
        )
        raise ValueError(
            "Loaded proxy-model size is incompatible with "
            "object_pose_setup.object_bbox_min_m/max_m even after unit "
            f"checking (selected ratio={bbox_ratio:.6g}, allowed "
            f"[{1.0 / tolerance:.6g}, {tolerance:.6g}]). Trials: {trials}. "
            "Check --obj_mesh, metadata object_model_unit and "
            "--model_frame_mode."
        )
    if bbox_ratio < 0.67 or bbox_ratio > 1.5:
        print(
            "[Warning] Proxy-model and metadata bbox sizes differ: "
            f"diagonal ratio={bbox_ratio:.4f}. Scale is still accepted because "
            "it is within --model_bbox_scale_tolerance; verify that this is "
            "the intended proxy model."
        )

    mesh = raw_mesh.copy()
    mesh.vertices = np.asarray(
        selected["vertices_mesh"], dtype=np.float64
    )
    mesh, visual_source = ensure_multiview_mesh_visuals(mesh)
    if visual_source.startswith("fallback_vertex_color:"):
        print(
            "[Mesh visual] No usable texture was found; converted the mesh to "
            f"neutral-gray vertex colors for MultiView ({visual_source})."
        )
    diameter = mesh_diameter(mesh)
    T_object_mesh = validate_rigid_transform(
        selected["T_object_mesh"],
        "fixed metadata-derived T_object_mesh",
    )
    return {
        "mesh": mesh,
        "mesh_visual_source": visual_source,
        "mesh_path": mesh_path,
        "dataset_model_path": dataset_model_path,
        "source_model_path": source_model_path,
        "model_frame_mode": model_frame_mode,
        "object_model_unit": metadata_unit,
        "mesh_scale": resolved_scale,
        "mesh_scale_source": scale_source,
        "mesh_center_offset_after_scale_m": np.asarray(
            selected["center_removed"], dtype=np.float64
        ),
        "T_object_mesh": T_object_mesh,
        "T_object_mesh_source": str(selected["transform_source"]),
        "object_model_axis_alignment": copy.deepcopy(
            setup.get("object_model_axis_alignment")
        ),
        "object_model_axis_alignment_source": (
            object_setup_rotation_source
        ),
        "object_model_origin_mode": str(
            setup.get("object_model_origin_mode", "")
        ),
        "computed_diameter_m": float(diameter),
        "bbox_min_object_m": bbox_min,
        "bbox_max_object_m": bbox_max,
        "bbox_extent_object_m": bbox_max - bbox_min,
        "bbox_diagonal_ratio_model_to_metadata": bbox_ratio,
        "model_bbox_extent_object_m": np.asarray(
            selected["observed_bbox_extent_m"], dtype=np.float64
        ),
        "scale_candidates_audit": [
            {
                "scale": float(item["scale"]),
                "bbox_score": float(item["bbox_score"]),
                "bbox_diagonal_ratio": float(
                    item["bbox_diagonal_ratio"]
                ),
                "observed_bbox_extent_m": np.asarray(
                    item["observed_bbox_extent_m"],
                    dtype=np.float64,
                ).tolist(),
            }
            for item in candidates
        ],
        "object_pose_setup": copy.deepcopy(setup),
    }


def mask_bbox_xyxy(mask: np.ndarray) -> Optional[np.ndarray]:
    mask = np.asarray(mask) > 0
    ys, xs = np.where(mask)
    if xs.size == 0:
        return None
    return np.array([xs.min(), ys.min(), xs.max(), ys.max()], dtype=np.float64)


def projected_bbox_xyxy(points_obj: np.ndarray, pose: np.ndarray, K: np.ndarray) -> Optional[np.ndarray]:
    pts_cam = transform_points(points_obj, pose)
    uv = project_points(pts_cam, K)
    valid = np.isfinite(uv).all(axis=1)
    if valid.sum() < 2:
        return None
    uv = uv[valid]
    return np.array([uv[:, 0].min(), uv[:, 1].min(), uv[:, 0].max(), uv[:, 1].max()], dtype=np.float64)


def bbox_iou_xyxy(a: Optional[np.ndarray], b: Optional[np.ndarray]) -> float:
    if a is None or b is None:
        return 0.0
    x0 = max(float(a[0]), float(b[0])); y0 = max(float(a[1]), float(b[1]))
    x1 = min(float(a[2]), float(b[2])); y1 = min(float(a[3]), float(b[3]))
    iw = max(0.0, x1 - x0 + 1.0); ih = max(0.0, y1 - y0 + 1.0)
    inter = iw * ih
    area_a = max(0.0, float(a[2] - a[0] + 1.0)) * max(0.0, float(a[3] - a[1] + 1.0))
    area_b = max(0.0, float(b[2] - b[0] + 1.0)) * max(0.0, float(b[3] - b[1] + 1.0))
    union = area_a + area_b - inter
    return float(inter / union) if union > 1e-9 else 0.0


def transform_points(points_xyz: np.ndarray, pose_4x4: np.ndarray) -> np.ndarray:
    """Transform Nx3 object-frame points by a 4x4 object-to-camera pose."""
    return points_xyz @ pose_4x4[:3, :3].T + pose_4x4[:3, 3][None, :]


def project_points(points_cam: np.ndarray, K: np.ndarray) -> np.ndarray:
    """Project camera-frame points to pixel coordinates."""
    z = points_cam[:, 2:3]
    valid = z[:, 0] > 1e-6
    uv = np.full((points_cam.shape[0], 2), np.nan, dtype=np.float32)
    if valid.any():
        p = points_cam[valid] @ K.T
        uv[valid, 0] = p[:, 0] / p[:, 2]
        uv[valid, 1] = p[:, 1] / p[:, 2]
    return uv


def rotation_error_deg(pred_pose: np.ndarray, gt_pose: np.ndarray) -> float:
    """Geodesic SO(3) rotation error in degrees."""
    r_delta = pred_pose[:3, :3] @ gt_pose[:3, :3].T
    cos_theta = (np.trace(r_delta) - 1.0) / 2.0
    cos_theta = np.clip(cos_theta, -1.0, 1.0)
    return float(np.degrees(np.arccos(cos_theta)))


def translation_error_cm(pred_pose: np.ndarray, gt_pose: np.ndarray) -> float:
    """Euclidean translation error in centimeters. Poses are in meters."""
    return float(np.linalg.norm(pred_pose[:3, 3] - gt_pose[:3, 3]) * 100.0)


def add_error_m(vertices_xyz: np.ndarray, pred_pose: np.ndarray, gt_pose: np.ndarray) -> float:
    """ADD error in meters using sampled model vertices."""
    pred_pts = transform_points(vertices_xyz, pred_pose)
    gt_pts = transform_points(vertices_xyz, gt_pose)
    return float(np.linalg.norm(pred_pts - gt_pts, axis=1).mean())


def add_s_error_m(vertices_xyz: np.ndarray, pred_pose: np.ndarray, gt_pose: np.ndarray) -> float:
    """ADD-S error in meters using nearest-neighbor matching in 3D."""
    pred_pts = transform_points(vertices_xyz, pred_pose)
    gt_pts = transform_points(vertices_xyz, gt_pose)
    # Match the ADI/ADD-S direction used by the BOP toolkit: each GT model
    # point is paired with its nearest estimated point.
    tree = cKDTree(pred_pts)
    dists, _ = tree.query(gt_pts, k=1, workers=-1)
    return float(np.mean(dists))


def sample_vertices(vertices: np.ndarray, max_points: int, seed: int = 0) -> np.ndarray:
    vertices = np.asarray(vertices, dtype=np.float32)
    if max_points <= 0:
        return vertices
    if vertices.shape[0] <= max_points:
        return vertices
    rng = np.random.default_rng(seed)
    idx = rng.choice(vertices.shape[0], size=max_points, replace=False)
    return vertices[idx]


def parse_float_list(value: str, label: str) -> np.ndarray:
    """Parse a comma-separated list into a finite float64 vector."""
    try:
        result = np.asarray(
            [float(x.strip()) for x in str(value).split(",") if x.strip()],
            dtype=np.float64,
        )
    except Exception as exc:
        raise ValueError(f"Cannot parse {label}: {value!r}") from exc
    if result.size == 0 or not np.all(np.isfinite(result)):
        raise ValueError(f"{label} must contain finite numbers: {value!r}")
    if np.any(result <= 0) or np.any(np.diff(result) <= 0):
        raise ValueError(f"{label} must be positive and strictly increasing: {value!r}")
    return result


def to_json_safe(value):
    """Recursively convert common runtime objects to JSON-native values."""
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, dict):
        return {str(key): to_json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [to_json_safe(item) for item in value]
    if isinstance(value, os.PathLike):
        return os.fspath(value)
    if isinstance(value, argparse.Namespace):
        return to_json_safe(vars(value))
    return value


def argparse_config_dict(args: argparse.Namespace) -> dict:
    """Return serializable CLI configuration without private runtime state."""
    return {
        key: to_json_safe(value)
        for key, value in vars(args).items()
        if not key.startswith("_")
    }


def threshold_recall_contribution(error: float, thresholds: np.ndarray) -> float:
    """Per-instance contribution whose dataset mean equals BOP average recall."""
    if not np.isfinite(error):
        return 0.0
    return float(np.mean(float(error) < np.asarray(thresholds, dtype=np.float64)))


def vsd_ar_contribution(vsd_errors: np.ndarray, correctness_thresholds: np.ndarray) -> float:
    """Per-instance VSD AR contribution averaged over taus and thresholds."""
    errors = np.asarray(vsd_errors, dtype=np.float64).reshape(-1)
    thresholds = np.asarray(correctness_thresholds, dtype=np.float64).reshape(-1)
    if errors.size == 0 or not np.all(np.isfinite(errors)):
        return 0.0
    return float(np.mean(errors[:, None] < thresholds[None, :]))


def mean_projection_error_px(
    vertices_xyz: np.ndarray,
    pred_pose: np.ndarray,
    gt_pose: np.ndarray,
    K: np.ndarray,
) -> float:
    pred_uv = project_points(transform_points(vertices_xyz, pred_pose), K)
    gt_uv = project_points(transform_points(vertices_xyz, gt_pose), K)
    valid = np.isfinite(pred_uv).all(axis=1) & np.isfinite(gt_uv).all(axis=1)
    if not valid.any():
        return float("nan")
    return float(np.linalg.norm(pred_uv[valid] - gt_uv[valid], axis=1).mean())


def symmetry_aware_rotation_error_deg(
    pred_pose: np.ndarray,
    gt_pose: np.ndarray,
    symmetries: List[Dict[str, np.ndarray]],
) -> float:
    """Minimum geodesic rotation error over declared object symmetries."""
    errors = []
    for sym in symmetries:
        gt_sym = gt_pose.copy()
        gt_sym[:3, :3] = gt_pose[:3, :3] @ np.asarray(sym["R"], dtype=np.float64)
        errors.append(rotation_error_deg(pred_pose, gt_sym))
    return float(min(errors)) if errors else rotation_error_deg(pred_pose, gt_pose)


def relative_pose_errors(
    prev_pred: Optional[np.ndarray],
    pred: np.ndarray,
    prev_gt: Optional[np.ndarray],
    gt: np.ndarray,
) -> Tuple[float, float]:
    """Return frame-to-frame relative translation [mm] and rotation [deg]."""
    if prev_pred is None or prev_gt is None:
        return float("nan"), float("nan")
    pred_rel = np.linalg.inv(prev_pred) @ pred
    gt_rel = np.linalg.inv(prev_gt) @ gt
    rel_error = np.linalg.inv(gt_rel) @ pred_rel
    t_mm = float(np.linalg.norm(rel_error[:3, 3]) * 1000.0)
    r_deg = rotation_error_deg(rel_error, np.eye(4, dtype=np.float64))
    return t_mm, r_deg


def make_bbox_corners(vertices_xyz: np.ndarray) -> np.ndarray:
    """Axis-aligned 3D bounding-box corners in object coordinates."""
    vertices = np.asarray(vertices_xyz, dtype=np.float64)
    return make_bbox_corners_from_bounds(
        vertices.min(axis=0), vertices.max(axis=0)
    )


def make_bbox_corners_from_bounds(
    lower_xyz: np.ndarray,
    upper_xyz: np.ndarray,
) -> np.ndarray:
    """Return 3D corners from metadata min/max in the same object frame."""
    vmin = np.asarray(lower_xyz, dtype=np.float64).reshape(3)
    vmax = np.asarray(upper_xyz, dtype=np.float64).reshape(3)
    if not (
        np.all(np.isfinite(vmin))
        and np.all(np.isfinite(vmax))
        and np.all(vmax > vmin)
    ):
        raise ValueError(
            f"Invalid bbox bounds: min={vmin.tolist()}, max={vmax.tolist()}"
        )
    x0, y0, z0 = vmin
    x1, y1, z1 = vmax
    return np.array([
        [x0, y0, z0], [x1, y0, z0], [x1, y1, z0], [x0, y1, z0],
        [x0, y0, z1], [x1, y0, z1], [x1, y1, z1], [x0, y1, z1],
    ], dtype=np.float32)


BBOX_EDGES = [
    (0, 1), (1, 2), (2, 3), (3, 0),
    (4, 5), (5, 6), (6, 7), (7, 4),
    (0, 4), (1, 5), (2, 6), (3, 7),
]


def draw_line_if_valid(img, p0, p1, color, thickness=2):
    if np.any(np.isnan(p0)) or np.any(np.isnan(p1)):
        return
    p0 = tuple(np.round(p0).astype(int))
    p1 = tuple(np.round(p1).astype(int))
    h, w = img.shape[:2]
    if (-w <= p0[0] <= 2 * w and -h <= p0[1] <= 2 * h and
        -w <= p1[0] <= 2 * w and -h <= p1[1] <= 2 * h):
        cv2.line(img, p0, p1, color, thickness, cv2.LINE_AA)


def draw_projected_bbox(img, bbox_corners_obj, pose, K, color, thickness=2):
    corners_cam = transform_points(bbox_corners_obj, pose)
    corners_uv = project_points(corners_cam, K)
    for i0, i1 in BBOX_EDGES:
        draw_line_if_valid(img, corners_uv[i0], corners_uv[i1], color, thickness)



def draw_pose_rays_monocolor(img, pose, K, axis_length, color, thickness=2):
    """Draw three object-frame direction rays using one color for one pose.

    GT is drawn in green and prediction in red, as requested. This avoids confusion
    with the usual XYZ-axis colors and makes GT/pred identity visually explicit.
    """
    pts_obj = np.array([
        [0.0, 0.0, 0.0],
        [axis_length, 0.0, 0.0],
        [0.0, axis_length, 0.0],
        [0.0, 0.0, axis_length],
    ], dtype=np.float32)
    uv = project_points(transform_points(pts_obj, pose), K)
    origin = uv[0]
    for j in [1, 2, 3]:
        draw_line_if_valid(img, origin, uv[j], color, thickness)
    if not np.any(np.isnan(origin)):
        cv2.circle(img, tuple(np.round(origin).astype(int)), 4, color, -1, cv2.LINE_AA)


def draw_origin_gap(img, gt_pose, pred_pose, K):
    origins = np.array([[0.0, 0.0, 0.0]], dtype=np.float32)
    gt_uv = project_points(transform_points(origins, gt_pose), K)[0]
    pred_uv = project_points(transform_points(origins, pred_pose), K)[0]
    draw_line_if_valid(img, gt_uv, pred_uv, (255, 255, 255), 2)


def draw_text_panel(img, lines, origin=(12, 24), line_height=24):
    if not lines:
        return
    x, y = origin
    max_width = max(cv2.getTextSize(line, cv2.FONT_HERSHEY_SIMPLEX, 0.58, 1)[0][0] for line in lines)
    panel_h = line_height * len(lines) + 14
    overlay = img.copy()
    cv2.rectangle(overlay, (x - 8, y - 20), (x + max_width + 18, y - 20 + panel_h), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.45, img, 0.55, 0, img)
    for idx, line in enumerate(lines):
        cv2.putText(img, line, (x, y + idx * line_height), cv2.FONT_HERSHEY_SIMPLEX,
                    0.58, (255, 255, 255), 1, cv2.LINE_AA)


def draw_mask_contour(img, mask):
    contours, _ = cv2.findContours((mask.astype(np.uint8) * 255), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(img, contours, -1, (0, 255, 255), 1, cv2.LINE_AA)


SINGLE_VIEW_METRIC_COLUMN_MAP = [
    ("frame_idx", "frame_idx"),
    ("ADD_mm", "ADD_mm"),
    ("ADDS_mm", "ADD-S_mm"),
    ("VSD_error_mean", "VSD"),
    ("MSSD_mm", "MSSD_mm"),
    ("MSPD_px", "MSPD_px"),
    ("BOP_AR_contribution", "BOP_AR"),
    ("Translation_error_mm", "Translation_error_mm"),
    ("Rotation_error_deg", "Rotation_error_deg"),
]

FUSION_METRIC_COLUMN_MAP = [
    *SINGLE_VIEW_METRIC_COLUMN_MAP,
    ("Right_view_contribution", "Right_view_contribution"),
    ("Left_view_contribution", "Left_view_contribution"),
]


def compact_metrics_dataframe(
    records: List[Dict],
    include_contributions: Optional[bool] = None,
) -> pd.DataFrame:
    """Return exactly the requested per-frame fields.

    ``frame_idx`` is retained as the only identifier because otherwise a bad
    frame cannot be located. Right/left contribution columns exist only in the
    fusion sheet.
    """
    source = pd.DataFrame(records)
    if include_contributions is None:
        views = set(source.get("camera_view", pd.Series(dtype=str)).astype(str))
        include_contributions = views == {"fusion"}
    column_map = (
        FUSION_METRIC_COLUMN_MAP
        if include_contributions
        else SINGLE_VIEW_METRIC_COLUMN_MAP
    )
    compact = pd.DataFrame(index=source.index)
    for source_name, output_name in column_map:
        if source_name in source.columns:
            compact[output_name] = source[source_name]
        else:
            compact[output_name] = np.nan
    return compact


def write_metrics_excel(
    records: List[Dict],
    xlsx_path: str,
    args,
    metadata: Optional[Dict] = None,
) -> None:
    """Write one compact single-stream worksheet without auxiliary sheets."""
    if not records:
        return
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise ImportError(
            "Writing .xlsx requires openpyxl. Install it in the MultiView environment "
            "or use the simultaneously written CSV file."
        ) from exc

    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    per_frame = compact_metrics_dataframe(records)
    view_names = set(
        pd.DataFrame(records).get(
            "camera_view", pd.Series(dtype=str)
        ).astype(str)
    )
    sheet_name = next(iter(view_names)) if len(view_names) == 1 else "metrics"
    sheet_name = sheet_name if sheet_name in {"right", "left", "fusion"} else "metrics"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        per_frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(xlsx_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 34
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column in enumerate(ws.iter_cols(), start=1):
            values = [str(cell.value) if cell.value is not None else "" for cell in column[:200]]
            width = min(42, max(10, max((len(v) for v in values), default=8) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            for cell in column[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws = wb[sheet_name]
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    for name in (
        "BOP_AR",
        "Right_view_contribution",
        "Left_view_contribution",
    ):
        if name in header_to_col and ws.max_row >= 2:
            col = get_column_letter(header_to_col[name])
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="F8696B",
                    mid_type="num", mid_value=0.5, mid_color="FFEB84",
                    end_type="num", end_value=1, end_color="63BE7B",
                ),
            )
            for cell in ws[col][1:]:
                cell.number_format = "0.000"
    wb.save(xlsx_path)


def write_three_stream_metrics_excel(
    right_records: List[Dict],
    left_records: List[Dict],
    fusion_records: List[Dict],
    xlsx_path: str,
) -> None:
    """Write exactly three sheets: right, left and fusion."""
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise ImportError(
            "Writing .xlsx requires openpyxl in the MultiView environment."
        ) from exc

    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    sheets = {
        "right": compact_metrics_dataframe(
            right_records, include_contributions=False
        ),
        "left": compact_metrics_dataframe(
            left_records, include_contributions=False
        ),
        "fusion": compact_metrics_dataframe(
            fusion_records, include_contributions=True
        ),
    }
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name in ("right", "left", "fusion"):
            sheets[sheet_name].to_excel(
                writer, sheet_name=sheet_name, index=False
            )

    wb = load_workbook(xlsx_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 34
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for col_idx, column in enumerate(ws.iter_cols(), start=1):
            values = [
                str(cell.value) if cell.value is not None else ""
                for cell in column[:200]
            ]
            width = min(
                28,
                max(11, max((len(value) for value in values), default=8) + 2),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            for cell in column[1:]:
                cell.alignment = Alignment(vertical="top")
                if col_idx == 1:
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.000"

        header_to_col = {cell.value: cell.column for cell in ws[1]}
        color_columns = ["BOP_AR"]
        if ws.title == "fusion":
            color_columns.extend([
                "Right_view_contribution",
                "Left_view_contribution",
            ])
        for name in color_columns:
            if name not in header_to_col or ws.max_row < 2:
                continue
            col = get_column_letter(header_to_col[name])
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="F8696B",
                    mid_type="num", mid_value=0.5, mid_color="FFEB84",
                    end_type="num", end_value=1, end_color="63BE7B",
                ),
            )
    wb.save(xlsx_path)


def save_metrics_bundle(
    records: List[Dict],
    output_dir: str,
    basename: str,
    args,
    metadata: Optional[Dict] = None,
) -> Dict[str, str]:
    if not records:
        return {}
    os.makedirs(output_dir, exist_ok=True)
    df = compact_metrics_dataframe(records)
    csv_path = os.path.join(output_dir, f"{basename}.csv")
    xlsx_path = os.path.join(output_dir, f"{basename}.xlsx")
    df.to_csv(csv_path, index=False)
    write_metrics_excel(records, xlsx_path, args=args, metadata=metadata)
    return {
        "csv": csv_path,
        "xlsx": xlsx_path,
    }


def save_fusion_metrics_workbook(
    right_records: List[Dict],
    left_records: List[Dict],
    fusion_records: List[Dict],
    output_dir: str,
) -> str:
    """Save the single requested three-sheet fusion workbook."""
    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = os.path.join(
        output_dir, "pose_metrics_fusion_per_frame.xlsx"
    )
    write_three_stream_metrics_excel(
        right_records=right_records,
        left_records=left_records,
        fusion_records=fusion_records,
        xlsx_path=xlsx_path,
    )
    return xlsx_path


# -----------------------------------------------------------------------------
# Dataset reader
# -----------------------------------------------------------------------------

def _name_match_key(value: str) -> str:
    """Normalize spaces/underscores/hyphens for conservative name matching."""
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def infer_object_name_from_dataset(dataset_root: str, view_arg: str) -> str:
    """Infer one object prefix from ``evaluation/*_left`` and ``*_right``.

    This supports the intended one-object-per-command workflow. Inference is
    accepted only when it is unambiguous (or matches the dataset directory name);
    it never silently selects one object from a multi-object evaluation folder.
    """
    dataset_root = os.path.abspath(dataset_root)
    evaluation_dir = os.path.join(dataset_root, "evaluation")
    if not os.path.isdir(evaluation_dir):
        raise FileNotFoundError(
            f"Missing evaluation directory under --dataset_root: {evaluation_dir}"
        )
    entries = sorted(
        name for name in os.listdir(evaluation_dir)
        if os.path.isdir(os.path.join(evaluation_dir, name))
    )
    requested = (
        {"left", "right"} if str(view_arg).lower().strip() == "both"
        else {str(view_arg).lower().strip()}
    )
    prefixes_by_view = {"left": set(), "right": set()}
    for name in entries:
        for view in ("left", "right"):
            suffix = f"_{view}"
            if name.lower().endswith(suffix) and len(name) > len(suffix):
                prefixes_by_view[view].add(name[:-len(suffix)])

    candidate_sets = [prefixes_by_view[view] for view in requested]
    candidates = (
        set.intersection(*candidate_sets) if candidate_sets else set()
    )
    if len(candidates) == 1:
        return next(iter(candidates))

    dataset_name = os.path.basename(os.path.normpath(dataset_root))
    dataset_key = _name_match_key(dataset_name)
    basename_matches = sorted(
        name for name in candidates if _name_match_key(name) == dataset_key
    )
    if len(basename_matches) == 1:
        return basename_matches[0]

    # Also support a compact single-object layout:
    #   evaluation/left/... and evaluation/right/...
    if all(view in entries for view in requested):
        return dataset_name

    available = {
        view: sorted(prefixes_by_view[view]) for view in ("left", "right")
    }
    if not candidates:
        raise ValueError(
            "Cannot infer one object name for the requested view(s) from "
            f"{evaluation_dir}. Found prefixes: {available}. Pass "
            "--object_name explicitly, or use evaluation/<name>_left and "
            "evaluation/<name>_right."
        )
    raise ValueError(
        f"More than one object prefix exists in {evaluation_dir}: "
        f"{sorted(candidates)}. This command evaluates one object at a time; "
        "pass --object_name explicitly."
    )


class RealSceneMultiViewReader:
    """Strict reader for the real captured RGB-D dataset."""

    def __init__(self, dataset_root: str, view: str, args, start_frame: int = 0, stride: int = 1):
        self.dataset_root = os.path.abspath(dataset_root)
        self.args = args
        self.view = self._normalize_view(view, args)
        self.seq_name = f"{args.object_name}_{self.view}"
        evaluation_dir = os.path.join(self.dataset_root, "evaluation")
        named_dir = os.path.join(evaluation_dir, self.seq_name)
        compact_dir = os.path.join(evaluation_dir, self.view)
        if os.path.isdir(named_dir):
            self.seq_dir = named_dir
        elif os.path.isdir(compact_dir):
            self.seq_dir = compact_dir
        else:
            self.seq_dir = named_dir
        if not os.path.isdir(self.seq_dir):
            raise FileNotFoundError(
                "Real-scene sequence directory does not exist. Tried: "
                f"{named_dir} and {compact_dir}"
            )
        self.K = self._load_K()
        self.rgb_files = self._load_rgb_file_list(start_frame=start_frame, stride=stride)
        self.frame_ids = [self._frame_id_from_path(p) for p in self.rgb_files]
        self._frame_meta_cache: Dict[int, Dict[str, object]] = {}
        self.meta_path_map = self._load_meta_path_map()
        self._validate_frame_metadata()
        self.gt_pose_path_map = self._load_gt_pose_path_map()
        self._validate_gt_trajectory_files()

    @staticmethod
    def _normalize_view(view: str, args) -> str:
        view = view.lower().strip()
        object_name = args.object_name
        if view in ["l", "left", f"{object_name}_left"]:
            return "left"
        if view in ["r", "right", f"{object_name}_right"]:
            return "right"
        raise ValueError(f"Unsupported view: {view}. Use left, right, or both.")

    @staticmethod
    def _frame_id_from_path(path: str) -> int:
        stem = os.path.splitext(os.path.basename(path))[0]
        try:
            return int(stem)
        except ValueError:
            digits = "".join(ch for ch in stem if ch.isdigit())
            if not digits:
                raise ValueError(f"Cannot parse frame id from file name: {path}")
            return int(digits)

    def _load_rgb_file_list(self, start_frame: int, stride: int) -> List[str]:
        rgb_dir = os.path.join(self.seq_dir, "rgb")
        if not os.path.isdir(rgb_dir):
            raise FileNotFoundError(f"Missing RGB directory: {rgb_dir}")
        files: List[str] = []
        for suffix in ("*.png", "*.jpg", "*.jpeg"):
            files.extend(glob.glob(os.path.join(rgb_dir, suffix)))
        files = sorted(files, key=self._frame_id_from_path)
        files = [p for p in files if self._frame_id_from_path(p) >= start_frame]
        if stride > 1:
            files = files[::stride]
        if not files:
            raise FileNotFoundError(f"No RGB images found in {rgb_dir}")
        return files

    def _load_K(self) -> np.ndarray:
        path = os.path.join(self.seq_dir, "K.txt")
        if not os.path.isfile(path):
            raise FileNotFoundError(f"Missing camera intrinsics: {path}")
        K = np.loadtxt(path).astype(np.float32)
        if K.shape != (3, 3) or not np.all(np.isfinite(K)):
            raise ValueError(f"Invalid K matrix in {path}: shape={K.shape}")
        return K

    def _load_meta_path_map(self) -> Dict[int, str]:
        """Map frame ids to the capture metadata containing T_cam_base."""
        meta_dir = os.path.join(self.seq_dir, "meta")
        path_map: Dict[int, str] = {}
        if os.path.isdir(meta_dir):
            for path in sorted(glob.glob(os.path.join(meta_dir, "*.json"))):
                frame_id = self._frame_id_from_path(path)
                if frame_id in path_map:
                    raise ValueError(
                        f"Duplicate metadata files resolve to frame {frame_id}: "
                        f"{path_map[frame_id]} and {path}"
                    )
                path_map[frame_id] = path
        missing = [
            frame_id for frame_id in self.frame_ids
            if frame_id not in path_map
        ]
        if missing and not self.args.allow_legacy_camera_extrinsics:
            raise FileNotFoundError(
                "Per-frame metadata is required for correct camera/GT "
                f"conversion. Missing meta/<frame>.json for {missing[:20]} in "
                f"{meta_dir}. Use --allow_legacy_camera_extrinsics only for an "
                "older dataset that truly has no per-frame T_cam_base."
            )
        if missing:
            print(
                "[Warning] Missing per-frame metadata for "
                f"{len(missing)} frame(s) in {self.seq_name}; the explicitly "
                "enabled legacy fixed camera calibration will be used only "
                "for those frames."
            )
        return path_map

    def get_meta_path(self, frame_id: int) -> Optional[str]:
        return self.meta_path_map.get(int(frame_id))

    def get_frame_meta(self, frame_id: int) -> Dict[str, object]:
        frame_id = int(frame_id)
        if frame_id in self._frame_meta_cache:
            return copy.deepcopy(self._frame_meta_cache[frame_id])
        path = self.get_meta_path(frame_id)
        if path is None:
            return {}
        with open(path, "r", encoding="utf-8") as file:
            payload = json.load(file)
        if not isinstance(payload, dict):
            raise TypeError(
                f"Frame metadata must be a JSON object: {path}"
            )
        self._frame_meta_cache[frame_id] = payload
        return copy.deepcopy(payload)

    def get_T_cam_base(
        self,
        frame_id: int,
    ) -> Tuple[np.ndarray, str]:
        """Return this frame's authoritative base-to-camera transform."""
        meta = self.get_frame_meta(frame_id)
        if "T_cam_base" in meta:
            path = self.get_meta_path(frame_id)
            return (
                _metadata_matrix(
                    meta["T_cam_base"],
                    f"T_cam_base in {path}",
                ).copy(),
                str(path),
            )
        if not self.args.allow_legacy_camera_extrinsics:
            path = self.get_meta_path(frame_id)
            raise KeyError(
                f"Metadata {path} does not contain T_cam_base. GT must use "
                "T_cam_object_gt = T_cam_base_meta @ T_base_object_gt."
            )
        T_base_cam, legacy_path = resolve_T_base_cam(
            self.args, self.view
        )
        return np.linalg.inv(T_base_cam), (
            f"legacy_inverse_of:{legacy_path}"
        )

    def get_object_pose_setup(
        self,
        frame_id: int,
    ) -> Dict[str, object]:
        meta = self.get_frame_meta(frame_id)
        setup = meta.get("object_pose_setup")
        if isinstance(setup, dict) and setup:
            return copy.deepcopy(setup)
        if self.args.allow_legacy_object_setup:
            return {}
        raise KeyError(
            f"Metadata {self.get_meta_path(frame_id)} does not contain a "
            "non-empty object_pose_setup. Model unit, axis, origin, "
            "T_gt_model_m and GT bbox must come from capture metadata."
        )

    def get_reference_object_pose_setup(self) -> Dict[str, object]:
        return self.get_object_pose_setup(self.frame_ids[0])

    @staticmethod
    def _setup_matrix_or_vector(
        setup: Dict[str, object],
        key: str,
    ) -> Optional[np.ndarray]:
        if key not in setup:
            return None
        return np.asarray(setup[key], dtype=np.float64)

    def _validate_frame_metadata(self) -> None:
        """Validate camera transforms and fixed object setup for all frames."""
        if not self.frame_ids:
            return
        reference_setup: Optional[Dict[str, object]] = None
        reference_frame: Optional[int] = None
        string_keys = (
            "object_model_unit",
            "object_model_axis_alignment",
            "object_model_origin_mode",
        )
        numeric_keys = (
            "T_gt_model_rotation",
            "T_gt_model_m",
            "object_bbox_min_m",
            "object_bbox_max_m",
            "object_bbox_size_m",
        )
        for frame_id in self.frame_ids:
            meta = self.get_frame_meta(frame_id)
            if "T_cam_base" in meta:
                _metadata_matrix(
                    meta["T_cam_base"],
                    f"T_cam_base frame {frame_id:04d} ({self.view})",
                )
            elif not self.args.allow_legacy_camera_extrinsics:
                raise KeyError(
                    f"meta/{frame_id:04d}.json has no T_cam_base"
                )

            setup = meta.get("object_pose_setup")
            if not isinstance(setup, dict) or not setup:
                if self.args.allow_legacy_object_setup:
                    continue
                raise KeyError(
                    f"meta/{frame_id:04d}.json has no object_pose_setup"
                )
            if reference_setup is None:
                reference_setup = setup
                reference_frame = frame_id
                continue
            for key in string_keys:
                if str(setup.get(key, "")) != str(
                    reference_setup.get(key, "")
                ):
                    raise ValueError(
                        f"object_pose_setup.{key} changes between frames "
                        f"{reference_frame:04d} and {frame_id:04d}: "
                        f"{reference_setup.get(key)!r} vs {setup.get(key)!r}"
                    )
            for key in numeric_keys:
                reference_value = self._setup_matrix_or_vector(
                    reference_setup, key
                )
                current_value = self._setup_matrix_or_vector(setup, key)
                if reference_value is None and current_value is None:
                    continue
                if (
                    reference_value is None
                    or current_value is None
                    or reference_value.shape != current_value.shape
                    or not np.allclose(
                        reference_value,
                        current_value,
                        atol=1e-9,
                        rtol=1e-7,
                    )
                ):
                    raise ValueError(
                        f"object_pose_setup.{key} changes between frames "
                        f"{reference_frame:04d} and {frame_id:04d}; one fixed "
                        "GT object/model coordinate system is required."
                    )

    def camera_extrinsic_summary(self) -> Dict[str, object]:
        """Report metadata T_cam_base variation without assuming it is fixed."""
        transforms = []
        sources = []
        for frame_id in self.frame_ids:
            transform, source = self.get_T_cam_base(frame_id)
            transforms.append(transform)
            sources.append(source)
        reference = transforms[0]
        translation_deltas_mm = [
            float(
                np.linalg.norm(
                    transform[:3, 3] - reference[:3, 3]
                ) * 1000.0
            )
            for transform in transforms
        ]
        rotation_deltas_deg = [
            rotation_distance_deg_from_matrices(
                transform[:3, :3], reference[:3, :3]
            )
            for transform in transforms
        ]
        return {
            "source": "per_frame_meta/T_cam_base",
            "first_source": sources[0],
            "frame_count": int(len(transforms)),
            "first_T_cam_base": reference.tolist(),
            "max_translation_change_mm": float(
                max(translation_deltas_mm, default=0.0)
            ),
            "max_rotation_change_deg": float(
                max(rotation_deltas_deg, default=0.0)
            ),
        }

    @staticmethod
    def _validate_pose(pose: np.ndarray, path: str) -> np.ndarray:
        pose = np.asarray(pose, dtype=np.float64)
        if pose.shape != (4, 4):
            raise ValueError(f"GT pose must be 4x4: {path}, got {pose.shape}")
        if not np.all(np.isfinite(pose)):
            raise ValueError(f"GT pose contains NaN/Inf: {path}")
        if not np.allclose(pose[3], [0, 0, 0, 1], atol=1e-5):
            raise ValueError(f"Invalid homogeneous last row in {path}: {pose[3].tolist()}")
        return pose

    def _load_gt_pose_path_map(self) -> Dict[int, str]:
        """Map every RGB frame ID to its own GT pose file.

        The real-scene gt_pose matrices are used verbatim. Their translation is
        already expressed in metres in the robot-base coordinate system. No
        scaling, fallback to the first frame, interpolation, or pose caching is
        applied.
        """
        gt_dir = os.path.join(self.seq_dir, "gt_pose")
        if not os.path.isdir(gt_dir):
            raise FileNotFoundError(f"Missing GT directory: {gt_dir}")

        path_map: Dict[int, str] = {}
        for path in sorted(glob.glob(os.path.join(gt_dir, "*.txt"))):
            frame_id = self._frame_id_from_path(path)
            if frame_id in path_map:
                raise ValueError(
                    f"Duplicate GT pose files resolve to frame {frame_id}: "
                    f"{path_map[frame_id]} and {path}"
                )
            # Validate now, but load again for the requested frame so the visualised
            # GT is always the exact matrix stored in that frame's TXT file.
            self._validate_pose(np.loadtxt(path), path)
            path_map[frame_id] = path

        if not path_map:
            raise FileNotFoundError(f"No 4x4 GT pose TXT files found in {gt_dir}")

        missing = [fid for fid in self.frame_ids if fid not in path_map]
        if missing:
            raise FileNotFoundError(
                f"Missing per-frame GT poses for RGB frames: {missing[:20]}"
            )
        return path_map

    def _validate_gt_trajectory_files(self) -> None:
        """Print a diagnostic warning when all per-frame GT translations are identical."""
        translations = []
        for frame_id in self.frame_ids:
            path = self.gt_pose_path_map[frame_id]
            pose = self._validate_pose(np.loadtxt(path), path)
            translations.append(pose[:3, 3])
        translations = np.asarray(translations, dtype=np.float64)
        span = np.ptp(translations, axis=0) if len(translations) else np.zeros(3)
        self.gt_translation_span_m = span
        if len(translations) > 1 and float(np.linalg.norm(span)) < 1e-6:
            print(
                "[Warning] All gt_pose translations are numerically identical. "
                "The program is reading a different TXT for every frame, so verify "
                "that the dataset writer saved changing robot poses."
            )

    def __len__(self) -> int:
        return len(self.rgb_files)

    def get_frame_id(self, local_index: int) -> int:
        return self.frame_ids[local_index]

    def get_rgb(self, frame_id: int) -> np.ndarray:
        matches = [p for p in self.rgb_files if self._frame_id_from_path(p) == frame_id]
        if not matches:
            raise FileNotFoundError(f"Cannot find RGB for frame {frame_id:04d}")
        bgr = cv2.imread(matches[0], cv2.IMREAD_COLOR)
        if bgr is None:
            raise RuntimeError(f"Failed to read RGB image: {matches[0]}")
        return cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)

    def get_depth(self, frame_id: int) -> np.ndarray:
        npy_path = os.path.join(self.seq_dir, "depth_npy", f"{frame_id:04d}.npy")
        if os.path.isfile(npy_path):
            depth = np.load(npy_path).astype(np.float32)
            depth *= float(self.args.depth_npy_scale)
            return depth
        png_path = os.path.join(self.seq_dir, "depth", f"{frame_id:04d}.png")
        if os.path.isfile(png_path):
            depth_png = cv2.imread(png_path, cv2.IMREAD_UNCHANGED)
            if depth_png is None:
                raise RuntimeError(f"Failed to read depth image: {png_path}")
            return depth_png.astype(np.float32) * float(self.args.depth_png_scale)
        raise FileNotFoundError(
            f"Cannot find depth for frame {frame_id:04d}: expected {npy_path} or {png_path}"
        )

    def get_mask(self, frame_id: int, fallback_to_first: bool = False) -> np.ndarray:
        mask_dir = os.path.join(self.seq_dir, "mask")
        candidates = [
            os.path.join(mask_dir, f"{frame_id:04d}.png"),
            os.path.join(mask_dir, f"{frame_id}.png"),
        ]
        if fallback_to_first:
            candidates.extend([
                os.path.join(mask_dir, f"{self.frame_ids[0]:04d}.png"),
                os.path.join(mask_dir, "0000.png"),
            ])
        path = next((p for p in candidates if os.path.isfile(p)), None)
        if path is None:
            raise FileNotFoundError(f"Missing mask for frame {frame_id:04d} in {mask_dir}")
        mask_img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        if mask_img is None:
            raise RuntimeError(f"Failed to read mask: {path}")
        return mask_img > 0

    def get_gt_pose_path(self, frame_id: int) -> str:
        if frame_id not in self.gt_pose_path_map:
            raise KeyError(f"No GT pose file for frame {frame_id:04d}")
        return self.gt_pose_path_map[frame_id]

    def get_gt_pose_base(self, frame_id: int) -> np.ndarray:
        """Load this frame's exact T_base_object matrix directly from disk.

        Translation is used verbatim in metres. This function never returns the
        first-frame pose as a fallback and never applies a unit scale.
        """
        path = self.get_gt_pose_path(frame_id)
        return self._validate_pose(np.loadtxt(path), path).copy()

    def get_raw_gt_translation(self, frame_id: int) -> np.ndarray:
        return self.get_gt_pose_base(frame_id)[:3, 3].copy()


# -----------------------------------------------------------------------------
# SAM2 re-mask utilities
# -----------------------------------------------------------------------------

class SAM2BoxSegmenter:
    """Persistent SAM2 predictor for periodic re-registration.

    Compared with the previous simple box-only version, this one supports:
      1) multimask output;
      2) optional positive/negative point prompts;
      3) returning candidate masks with SAM2 scores.

    The extra prompts are important for the right view: a loose box may contain both
    the pipe and the robot gripper. Positive points projected from the current pose
    prior make SAM2 more likely to select the thin pipe instead of the larger robot.
    """

    def __init__(self, checkpoint: str, model_cfg: str, device: str = "auto"):
        if device == "auto":
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
        else:
            self.device = device
        self.predictor = SAM2ImagePredictor(build_sam2(model_cfg, checkpoint))

    def predict_prompt(
        self,
        rgb: np.ndarray,
        box_xyxy: np.ndarray,
        point_coords: Optional[np.ndarray] = None,
        point_labels: Optional[np.ndarray] = None,
        multimask_output: bool = True,
    ) -> Tuple[List[np.ndarray], List[float]]:
        box_xyxy = np.asarray(box_xyxy, dtype=np.float32).reshape(1, 4)
        if point_coords is not None:
            point_coords = np.asarray(point_coords, dtype=np.float32).reshape(-1, 2)
            point_labels = np.asarray(point_labels, dtype=np.int32).reshape(-1)
        if self.device == "cuda" and torch.cuda.is_available():
            with torch.inference_mode(), torch.autocast("cuda", dtype=torch.bfloat16):
                self.predictor.set_image(rgb)
                masks, scores, _ = self.predictor.predict(
                    point_coords=point_coords,
                    point_labels=point_labels,
                    box=box_xyxy,
                    multimask_output=multimask_output,
                )
        else:
            with torch.inference_mode():
                self.predictor.set_image(rgb)
                masks, scores, _ = self.predictor.predict(
                    point_coords=point_coords,
                    point_labels=point_labels,
                    box=box_xyxy,
                    multimask_output=multimask_output,
                )
        masks = [m.astype(np.bool_) for m in masks]
        scores = [float(s) for s in np.asarray(scores).reshape(-1)]
        return masks, scores

    def predict_box(self, rgb: np.ndarray, box_xyxy: np.ndarray) -> np.ndarray:
        masks, _ = self.predict_prompt(rgb, box_xyxy, None, None, multimask_output=False)
        return masks[0]


def pad_and_clip_box_xyxy(box: Optional[np.ndarray], image_shape: Tuple[int, int, int], pad_rel: float) -> Optional[np.ndarray]:
    if box is None:
        return None
    h, w = image_shape[:2]
    x1, y1, x2, y2 = [float(v) for v in box]
    bw = max(1.0, x2 - x1)
    bh = max(1.0, y2 - y1)
    x1 -= pad_rel * bw
    x2 += pad_rel * bw
    y1 -= pad_rel * bh
    y2 += pad_rel * bh
    x1 = np.clip(x1, 0, w - 1)
    x2 = np.clip(x2, 0, w - 1)
    y1 = np.clip(y1, 0, h - 1)
    y2 = np.clip(y2, 0, h - 1)
    if x2 <= x1 or y2 <= y1:
        return None
    return np.array([x1, y1, x2, y2], dtype=np.float32)


def is_valid_sam_box(box: Optional[np.ndarray], min_size: float) -> bool:
    if box is None:
        return False
    x1, y1, x2, y2 = [float(v) for v in box]
    return (x2 - x1) >= min_size and (y2 - y1) >= min_size




def union_boxes_xyxy(boxes: List[Optional[np.ndarray]]) -> Optional[np.ndarray]:
    valid = [b for b in boxes if b is not None and np.all(np.isfinite(b))]
    if not valid:
        return None
    arr = np.stack(valid, axis=0).astype(np.float64)
    return np.array([arr[:, 0].min(), arr[:, 1].min(), arr[:, 2].max(), arr[:, 3].max()], dtype=np.float64)


def box_center_xy(box: Optional[np.ndarray]) -> Optional[np.ndarray]:
    if box is None:
        return None
    return np.array([(float(box[0]) + float(box[2])) * 0.5,
                     (float(box[1]) + float(box[3])) * 0.5], dtype=np.float64)


def box_area_xyxy(box: Optional[np.ndarray]) -> float:
    if box is None:
        return 0.0
    return max(0.0, float(box[2] - box[0] + 1.0)) * max(0.0, float(box[3] - box[1] + 1.0))


def predict_next_pose_constant_velocity(prev2: Optional[np.ndarray], prev1: Optional[np.ndarray]) -> Optional[np.ndarray]:
    """Constant-velocity prediction in SE(3) using the last two estimated poses.

    This is deliberately simple and only used to enlarge the SAM search region when
    the pipe moves quickly. It is not used for metrics.
    """
    if prev2 is None or prev1 is None:
        return prev1
    if not (np.all(np.isfinite(prev2)) and np.all(np.isfinite(prev1))):
        return prev1
    try:
        delta = prev1 @ np.linalg.inv(prev2)
        pred = delta @ prev1
        if np.all(np.isfinite(pred)):
            return pred
    except Exception:
        pass
    return prev1


def projected_positive_negative_points(
    pose: Optional[np.ndarray],
    vertices_obj: np.ndarray,
    bbox_corners_obj: np.ndarray,
    K: np.ndarray,
    image_shape: Tuple[int, int, int],
    n_pos: int = 8,
) -> Tuple[Optional[np.ndarray], Optional[np.ndarray]]:
    """Generate SAM2 point prompts from the pose prior.

    Positive points are sampled from projected mesh vertices. Negative points are
    placed around the projected full-object box. They help SAM2 avoid selecting the
    robot when the prompt box contains both pipe and gripper.
    """
    if pose is None or vertices_obj is None or len(vertices_obj) == 0:
        return None, None
    h, w = image_shape[:2]
    pts_cam = transform_points(vertices_obj, pose)
    uv = project_points(pts_cam, K)
    valid = np.isfinite(uv).all(axis=1)
    if valid.sum() < 4:
        return None, None
    uv = uv[valid]
    inside = (uv[:, 0] >= 0) & (uv[:, 0] < w) & (uv[:, 1] >= 0) & (uv[:, 1] < h)
    uv = uv[inside]
    if len(uv) < 2:
        return None, None
    # Uniformly sample along the visible projected model points.
    idx = np.linspace(0, len(uv) - 1, num=min(n_pos, len(uv)), dtype=int)
    pos = uv[idx].astype(np.float32)

    box = projected_bbox_xyxy(bbox_corners_obj, pose, K)
    box = pad_and_clip_box_xyxy(box, image_shape, pad_rel=0.05)
    neg = []
    if box is not None:
        x1, y1, x2, y2 = [float(v) for v in box]
        cx, cy = (x1 + x2) * 0.5, (y1 + y2) * 0.5
        bw, bh = max(1.0, x2 - x1), max(1.0, y2 - y1)
        raw_neg = [
            [cx - 0.70 * bw, cy - 0.70 * bh],
            [cx + 0.70 * bw, cy - 0.70 * bh],
            [cx - 0.70 * bw, cy + 0.70 * bh],
            [cx + 0.70 * bw, cy + 0.70 * bh],
        ]
        for x, y in raw_neg:
            if 0 <= x < w and 0 <= y < h:
                neg.append([x, y])
    if neg:
        pts = np.vstack([pos, np.asarray(neg, dtype=np.float32)]).astype(np.float32)
        labels = np.concatenate([np.ones(len(pos), dtype=np.int32), np.zeros(len(neg), dtype=np.int32)])
    else:
        pts = pos.astype(np.float32)
        labels = np.ones(len(pos), dtype=np.int32)
    return pts, labels


def keep_largest_connected_component(mask: np.ndarray) -> np.ndarray:
    mask_u8 = (mask.astype(np.uint8) > 0).astype(np.uint8)
    num, labels, stats, _ = cv2.connectedComponentsWithStats(mask_u8, connectivity=8)
    if num <= 1:
        return mask.astype(np.bool_)
    areas = stats[1:, cv2.CC_STAT_AREA]
    idx = 1 + int(np.argmax(areas))
    return (labels == idx)


def mask_quality_stats(mask: np.ndarray, expected_box: Optional[np.ndarray]) -> Dict[str, float]:
    m = keep_largest_connected_component(mask)
    area = int(m.sum())
    box = mask_bbox_xyxy(m)
    barea = box_area_xyxy(box)
    fill = float(area / max(barea, 1.0))
    iou = bbox_iou_xyxy(box, expected_box)
    cdist_norm = 9999.0
    if box is not None and expected_box is not None:
        c0 = box_center_xy(box)
        c1 = box_center_xy(expected_box)
        diag = math.sqrt(max(1.0, box_area_xyxy(expected_box)))
        cdist_norm = float(np.linalg.norm(c0 - c1) / max(diag, 1.0))
    return {
        "area": float(area),
        "box_area": float(barea),
        "fill_ratio": fill,
        "box_iou": float(iou),
        "center_dist_norm": float(cdist_norm),
    }


def validate_sam_mask(mask: np.ndarray, expected_box: Optional[np.ndarray], args) -> Tuple[bool, str, Dict[str, float]]:
    """Reject masks that are likely robot/background or severe partial masks.

    This gate addresses two observed failure modes:
      - right view SAM2 selects the robot/gripper rather than the pipe;
      - under occlusion SAM2 returns only a small visible fragment, causing a bad
        re-register pose and an apparently shrunken red box.
    """
    if mask is None:
        return False, "mask_none", {}
    mask = keep_largest_connected_component(mask)
    stats = mask_quality_stats(mask, expected_box)
    if stats["area"] < args.min_mask_pixels:
        return False, f"too_few_pixels:{stats['area']:.0f}", stats
    if expected_box is not None:
        exp_area = box_area_xyxy(expected_box)
        if exp_area > 1.0:
            ratio = stats["box_area"] / exp_area
            stats["box_area_ratio_to_expected"] = float(ratio)
            if ratio < args.mask_min_box_area_ratio:
                return False, f"partial_mask_box_ratio:{ratio:.3f}", stats
            if ratio > args.mask_max_box_area_ratio:
                return False, f"too_large_mask_box_ratio:{ratio:.3f}", stats
            if stats["center_dist_norm"] > args.mask_max_center_dist_norm:
                return False, f"center_too_far:{stats['center_dist_norm']:.3f}", stats
    if stats["fill_ratio"] > args.mask_max_fill_ratio:
        return False, f"too_filled_like_robot:{stats['fill_ratio']:.3f}", stats
    if stats["fill_ratio"] < args.mask_min_fill_ratio:
        return False, f"too_sparse:{stats['fill_ratio']:.3f}", stats
    return True, "ok", stats


def select_sam_box(args, reader, frame_id: int, rgb: np.ndarray, bbox_corners: np.ndarray,
                   gt_pose_q: np.ndarray, pred_pose_q_prev: Optional[np.ndarray],
                   pred_pose_q_prev2: Optional[np.ndarray] = None) -> Tuple[Optional[np.ndarray], str]:
    """Choose a SAM2 box prompt.

    For fast motion, the default `auto` mode uses the union of the last-pose box and
    a constant-velocity motion-predicted box. For diagnostics, `gt` can be used, but
    it should not be reported as a fair tracking result.
    """
    pred_box = None
    motion_box = None
    union_pred_motion_box = None

    if pred_pose_q_prev is not None:
        pred_box = projected_bbox_xyxy(bbox_corners, pred_pose_q_prev, reader.K)
        pred_box = pad_and_clip_box_xyxy(pred_box, rgb.shape, args.sam_box_pad)

    motion_pose = predict_next_pose_constant_velocity(pred_pose_q_prev2, pred_pose_q_prev)
    if motion_pose is not None:
        motion_box = projected_bbox_xyxy(bbox_corners, motion_pose, reader.K)
        motion_box = pad_and_clip_box_xyxy(motion_box, rgb.shape, args.sam_motion_box_pad)

    union_pred_motion_box = union_boxes_xyxy([pred_box, motion_box])
    union_pred_motion_box = pad_and_clip_box_xyxy(union_pred_motion_box, rgb.shape, args.sam_union_box_pad)

    saved_box = None
    try:
        saved_mask = reader.get_mask(frame_id, fallback_to_first=False)
        saved_box = mask_bbox_xyxy(saved_mask)
        saved_box = pad_and_clip_box_xyxy(saved_box, rgb.shape, args.sam_box_pad)
    except Exception:
        saved_box = None

    cmap = {
        "pred_box": pred_box,
        "motion_box": motion_box,
        "pred_motion_union_box": union_pred_motion_box,
        "saved_mask_box": saved_box,
    }

    if args.sam_box_source == "pred":
        order = ["pred_motion_union_box", "motion_box", "pred_box"]
    elif args.sam_box_source == "motion":
        order = ["motion_box", "pred_motion_union_box", "pred_box"]
    elif args.sam_box_source == "saved_mask":
        order = ["saved_mask_box"]
    elif args.sam_box_source == "auto":
        order = ["pred_motion_union_box", "motion_box", "pred_box", "saved_mask_box"]
    else:
        raise ValueError(f"Unsupported --sam_box_source: {args.sam_box_source}")

    for name in order:
        box = cmap.get(name)
        if is_valid_sam_box(box, args.sam_min_box_size):
            return box, name

    if args.sam_fallback_full_image:
        h, w = rgb.shape[:2]
        return np.array([0, 0, w - 1, h - 1], dtype=np.float32), "full_image_box"
    return None, "no_valid_box"



def parse_xyxy_box_string(box_str: str) -> Optional[np.ndarray]:
    """Parse an xyxy box string such as '303,279,415,411'."""
    if box_str is None:
        return None
    s = str(box_str).strip()
    if s == "" or s.lower() in ["none", "null", "disable", "disabled"]:
        return None
    parts = [p.strip() for p in s.replace(";", ",").split(",") if p.strip() != ""]
    if len(parts) != 4:
        raise ValueError(f"Invalid xyxy box string: {box_str}. Expected x1,y1,x2,y2")
    return np.array([float(p) for p in parts], dtype=np.float32)


def normalize_manual_box_xyxy(box: Optional[np.ndarray], image_shape: Tuple[int, int, int],
                              min_side: float = 120.0, pad_rel: float = 0.0) -> Optional[Tuple[np.ndarray, bool, str]]:
    """Return a valid clipped xyxy box.

    The user-provided right-view box may have y1 == y2. A zero-height or zero-width
    box cannot be used by SAM2, so this function expands degenerate boxes around
    their center to at least `min_side` pixels. This keeps the user's manually
    measured center while making the prompt usable.
    """
    if box is None:
        return None
    h, w = image_shape[:2]
    x1, y1, x2, y2 = [float(v) for v in box]
    changed = False
    notes = []

    # Sort endpoints in case they are provided in reverse order.
    if x2 < x1:
        x1, x2 = x2, x1
        changed = True
        notes.append("swap_x")
    if y2 < y1:
        y1, y2 = y2, y1
        changed = True
        notes.append("swap_y")

    bw = x2 - x1
    bh = y2 - y1
    cx = 0.5 * (x1 + x2)
    cy = 0.5 * (y1 + y2)

    if bw < min_side:
        x1 = cx - 0.5 * min_side
        x2 = cx + 0.5 * min_side
        changed = True
        notes.append(f"expand_w_to_{min_side:g}")
    if bh < min_side:
        y1 = cy - 0.5 * min_side
        y2 = cy + 0.5 * min_side
        changed = True
        notes.append(f"expand_h_to_{min_side:g}")

    box2 = np.array([x1, y1, x2, y2], dtype=np.float32)
    box2 = pad_and_clip_box_xyxy(box2, image_shape, pad_rel=pad_rel)
    if box2 is None:
        return None
    return box2.astype(np.float32), changed, ";".join(notes) if notes else "as_given"


def manual_first_box_for_reader(args, reader: RealSceneMultiViewReader, rgb: np.ndarray) -> Tuple[Optional[np.ndarray], str, Dict]:
    """Get the user-provided first-frame box for the current view."""
    if not args.use_manual_first_bbox:
        return None, "manual_first_bbox_disabled", {}
    box_str = args.left_first_bbox if reader.view == "left" else args.right_first_bbox
    raw_box = parse_xyxy_box_string(box_str)
    if raw_box is None:
        return None, f"manual_first_bbox_missing_{reader.view}", {}
    normalized = normalize_manual_box_xyxy(
        raw_box, rgb.shape,
        min_side=float(args.manual_bbox_min_side),
        pad_rel=float(args.manual_bbox_pad),
    )
    if normalized is None:
        return None, f"manual_first_bbox_invalid_{reader.view}", {"raw_box": raw_box.tolist()}
    box, changed, note = normalized
    info = {
        "view": reader.view,
        "raw_box_xyxy": [float(x) for x in raw_box],
        "used_box_xyxy": [float(x) for x in box],
        "changed": bool(changed),
        "note": note,
        "min_side": float(args.manual_bbox_min_side),
        "pad_rel": float(args.manual_bbox_pad),
    }
    return box, f"manual_first_bbox_{reader.view}_{note}", info


def make_first_mask_from_manual_bbox(args, reader: RealSceneMultiViewReader, frame_id: int, rgb: np.ndarray,
                                     sam2_segmenter: Optional[SAM2BoxSegmenter],
                                     view_root: str) -> Tuple[Optional[np.ndarray], str]:
    """Use the measured first-frame bbox as the SAM2 prompt for initial register()."""
    box, source, info = manual_first_box_for_reader(args, reader, rgb)
    if box is None:
        return None, source
    if sam2_segmenter is None:
        return None, f"{source}_no_sam2"

    masks, scores = sam2_segmenter.predict_prompt(
        rgb,
        box,
        point_coords=None,
        point_labels=None,
        multimask_output=args.sam_multimask,
    )

    best = None
    best_tuple = None
    expected_box = box
    for mi, (cand_mask, sam_score) in enumerate(zip(masks, scores)):
        cand_mask = keep_largest_connected_component(cand_mask)
        ok, reason, stats = validate_sam_mask(cand_mask, expected_box, args)
        score = float(sam_score)
        score += 2.0 if ok else -1.0
        score += 2.0 * float(stats.get("box_iou", 0.0))
        score -= 0.6 * float(stats.get("center_dist_norm", 0.0))
        score -= 0.4 * max(0.0, float(stats.get("fill_ratio", 0.0)) - args.pipe_preferred_max_fill_ratio)
        if best is None or score > best:
            best = score
            best_tuple = (cand_mask, ok, reason, stats, sam_score, mi)

    if best_tuple is None:
        return None, f"{source}_sam2_no_candidates"

    mask, ok, reason, stats, sam_score, mask_index = best_tuple
    if args.reject_bad_sam_masks and not ok:
        # For the initial frame, do not silently continue. A wrong first mask makes
        # all following tracking meaningless.
        if args.save_sam_masks:
            rej_dir = os.path.join(view_root, "sam2_first_mask_rejected")
            os.makedirs(rej_dir, exist_ok=True)
            cv2.imwrite(os.path.join(rej_dir, f"{frame_id:04d}_{reason}.png"), mask.astype(np.uint8) * 255)
            with open(os.path.join(rej_dir, f"{frame_id:04d}_{reason}.json"), "w") as f:
                json.dump({**info, "reason": reason, "stats": stats, "sam_score": float(sam_score)}, f, indent=2)
        return None, f"{source}_sam2_rejected_{reason}"

    mask = mask.astype(np.bool_)
    if args.save_sam_masks:
        first_dir = os.path.join(view_root, "sam2_first_mask")
        os.makedirs(first_dir, exist_ok=True)
        cv2.imwrite(os.path.join(first_dir, f"{frame_id:04d}.png"), mask.astype(np.uint8) * 255)
        with open(os.path.join(first_dir, f"{frame_id:04d}.json"), "w") as f:
            json.dump({
                **info,
                "frame_id": int(frame_id),
                "mask_pixels": int(mask.sum()),
                "mask_index": int(mask_index),
                "sam_score": float(sam_score),
                "validation_ok": bool(ok),
                "validation_reason": reason,
                "validation_stats": stats,
            }, f, indent=2)

    if args.write_first_mask_to_dataset:
        mask_dir = os.path.join(reader.seq_dir, "mask")
        os.makedirs(mask_dir, exist_ok=True)
        cv2.imwrite(os.path.join(mask_dir, f"{frame_id:04d}.png"), mask.astype(np.uint8) * 255)

    return mask, f"manual_bbox_sam2:{source}:{reason}"

def make_register_mask(args, reader, frame_id: int, rgb: np.ndarray, bbox_corners: np.ndarray,
                       gt_pose_q: np.ndarray, pred_pose_q_prev: Optional[np.ndarray],
                       sam2_segmenter: Optional[SAM2BoxSegmenter], view_root: str,
                       pred_pose_q_prev2: Optional[np.ndarray] = None,
                       vertices_for_prompt: Optional[np.ndarray] = None,
                       depth: Optional[np.ndarray] = None) -> Tuple[Optional[np.ndarray], str]:
    """Return the mask used by register().

    First-frame behavior is unchanged: use the verified saved mask. Reinit frames use
    SAM2 with a motion-aware box and mask quality validation. Invalid masks are
    rejected; in that case the caller should skip re-registering and keep track_one.
    """
    # The first registration frame is decisive. By default, ignore potentially bad
    # saved masks and regenerate the first mask from the user-measured bbox.
    is_first_registration_frame = (frame_id == reader.frame_ids[0])
    if is_first_registration_frame and args.use_manual_first_bbox:
        manual_mask, manual_source = make_first_mask_from_manual_bbox(
            args=args,
            reader=reader,
            frame_id=frame_id,
            rgb=rgb,
            sam2_segmenter=sam2_segmenter,
            view_root=view_root,
        )
        if manual_mask is not None and manual_mask.sum() >= args.min_mask_pixels:
            return manual_mask.astype(np.bool_), manual_source
        if args.abort_on_bad_first_mask:
            return None, manual_source
        # Only fall back to the saved mask if explicitly allowed by disabling abort.

    try:
        saved_mask = reader.get_mask(frame_id, fallback_to_first=False)
        if saved_mask.sum() >= args.min_mask_pixels:
            if (
                frame_id == reader.frame_ids[0]
                or args.prefer_saved_mask
                or args.tracking_profile == "slender_tool"
            ):
                return keep_largest_connected_component(saved_mask).astype(np.bool_), f"saved_mask_{frame_id:04d}"
    except Exception:
        saved_mask = None

    if sam2_segmenter is None:
        if args.fallback_first_mask_for_reinit:
            try:
                m0 = reader.get_mask(frame_id, fallback_to_first=True)
                if m0.sum() >= args.min_mask_pixels:
                    return keep_largest_connected_component(m0).astype(np.bool_), "fallback_first_mask"
            except Exception:
                pass
        return None, "no_sam2_no_mask"

    box, box_source = select_sam_box(
        args, reader, frame_id, rgb, bbox_corners, gt_pose_q,
        pred_pose_q_prev, pred_pose_q_prev2=pred_pose_q_prev2,
    )
    if box is None:
        return None, box_source

    # Expected box for validating the returned mask. Prefer motion prediction, then
    # previous pose. In diagnostic gt mode, use gt box as expected.
    motion_pose = predict_next_pose_constant_velocity(pred_pose_q_prev2, pred_pose_q_prev)
    expected_pose = motion_pose if motion_pose is not None else pred_pose_q_prev
    expected_box = projected_bbox_xyxy(bbox_corners, expected_pose, reader.K) if expected_pose is not None else box
    expected_box = pad_and_clip_box_xyxy(expected_box, rgb.shape, args.expected_box_pad)

    point_coords, point_labels = (None, None)
    if args.sam_use_pose_points and expected_pose is not None and vertices_for_prompt is not None:
        point_coords, point_labels = projected_positive_negative_points(
            expected_pose, vertices_for_prompt, bbox_corners, reader.K, rgb.shape,
            n_pos=args.sam_num_positive_points,
        )

    masks, scores = sam2_segmenter.predict_prompt(
        rgb,
        box,
        point_coords=point_coords,
        point_labels=point_labels,
        multimask_output=args.sam_multimask,
    )

    best = None
    best_tuple = None
    for mi, (cand_mask, sam_score) in enumerate(zip(masks, scores)):
        cand_mask = keep_largest_connected_component(cand_mask)
        ok, reason, stats = validate_sam_mask(cand_mask, expected_box, args)
        # Higher is better. Favor valid masks, SAM confidence, expected-box overlap,
        # and reasonable center. Penalize high fill ratio because robot parts often
        # fill the box more densely than the thin pipe.
        score = float(sam_score)
        score += 2.0 if ok else -2.0
        score += 1.5 * float(stats.get("box_iou", 0.0))
        score -= 0.8 * float(stats.get("center_dist_norm", 0.0))
        score -= 0.5 * max(0.0, float(stats.get("fill_ratio", 0.0)) - args.pipe_preferred_max_fill_ratio)
        if best is None or score > best:
            best = score
            best_tuple = (cand_mask, ok, reason, stats, sam_score, mi)

    if best_tuple is None:
        return None, f"sam2_no_candidates_{box_source}"

    mask, ok, reason, stats, sam_score, mask_index = best_tuple
    if not ok and args.reject_bad_sam_masks:
        if args.save_sam_masks:
            sam_dir = os.path.join(view_root, "sam2_masks_rejected")
            os.makedirs(sam_dir, exist_ok=True)
            cv2.imwrite(os.path.join(sam_dir, f"{frame_id:04d}_{reason}.png"), mask.astype(np.uint8) * 255)
        return None, f"sam2_rejected_{box_source}_{reason}"

    if args.save_sam_masks:
        sam_dir = os.path.join(view_root, "sam2_masks")
        os.makedirs(sam_dir, exist_ok=True)
        cv2.imwrite(os.path.join(sam_dir, f"{frame_id:04d}.png"), mask.astype(np.uint8) * 255)
        with open(os.path.join(sam_dir, f"{frame_id:04d}.json"), "w") as f:
            json.dump({
                "frame_id": int(frame_id),
                "box_source": box_source,
                "box_xyxy": [float(x) for x in box],
                "expected_box_xyxy": None if expected_box is None else [float(x) for x in expected_box],
                "mask_pixels": int(mask.sum()),
                "mask_index": int(mask_index),
                "sam_score": float(sam_score),
                "validation_ok": bool(ok),
                "validation_reason": reason,
                "validation_stats": stats,
                "use_pose_points": bool(point_coords is not None),
            }, f, indent=2)

    return mask.astype(np.bool_), f"sam2_{box_source}_{reason}"

# -----------------------------------------------------------------------------
# MultiView tracking and evaluation
# -----------------------------------------------------------------------------

def init_multiview(mesh: trimesh.Trimesh, save_root: str, debug: int = 0) -> MultiView:
    glctx = dr.RasterizeCudaContext()
    dummy_mesh = copy.deepcopy(trimesh.primitives.Box(extents=np.ones((3)), transform=np.eye(4)))
    dummy_mesh = trimesh.Trimesh(vertices=dummy_mesh.vertices.copy(), faces=dummy_mesh.faces.copy())
    est = MultiView(
        mesh=dummy_mesh,
        scorer=ScorePredictor(),
        refiner=PoseRefinePredictor(),
        debug_dir=save_root,
        debug=debug,
        glctx=glctx,
    )
    est.reset_object(mesh=mesh, symmetry_tfs=None)
    return est


def validate_rigid_transform(T: np.ndarray, label: str, atol: float = 1e-4) -> np.ndarray:
    """Validate a 4x4 rigid transform without changing its convention."""
    T = np.asarray(T, dtype=np.float64)
    if T.shape != (4, 4):
        raise ValueError(f"{label} must be 4x4, got {T.shape}")
    if not np.all(np.isfinite(T)):
        raise ValueError(f"{label} contains NaN/Inf")
    if not np.allclose(T[3], [0.0, 0.0, 0.0, 1.0], atol=atol):
        raise ValueError(f"{label} has invalid homogeneous last row: {T[3].tolist()}")
    R = T[:3, :3]
    orth_err = float(np.linalg.norm(R.T @ R - np.eye(3), ord="fro"))
    det = float(np.linalg.det(R))
    if orth_err > 5e-3 or abs(det - 1.0) > 5e-3:
        raise ValueError(
            f"{label} rotation is not a valid SO(3) matrix: "
            f"orthogonality_error={orth_err:.6g}, det={det:.6g}"
        )
    return T


def load_transform_txt(path: str, label: str, translation_scale: float = 1.0) -> np.ndarray:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Missing {label}: {path}")
    T = np.loadtxt(path).astype(np.float64)
    T = validate_rigid_transform(T, f"{label} ({path})")
    T = T.copy()
    T[:3, 3] *= float(translation_scale)
    return T


_AXIS_UNIT_VECTORS = {
    "x": np.array([1.0, 0.0, 0.0], dtype=np.float64),
    "y": np.array([0.0, 1.0, 0.0], dtype=np.float64),
    "z": np.array([0.0, 0.0, 1.0], dtype=np.float64),
}


def parse_axis_map(axis_map: str) -> Tuple[np.ndarray, str]:
    """Parse a fixed mesh-axis to robot-object-axis mapping.

    The three comma-separated tokens describe the directions of mesh +X, +Y and
    +Z, expressed in the robot/GT object frame. For example:

        --axis_map "+y,-x,+z"

    means mesh +X -> object +Y, mesh +Y -> object -X and mesh +Z -> object +Z.
    Only proper right-handed signed permutations are accepted.
    """
    raw = str(axis_map).strip().lower().replace(" ", "")
    if raw in {"", "none", "identity", "same", "+x,+y,+z", "x,y,z"}:
        return np.eye(3, dtype=np.float64), "+x,+y,+z"
    tokens = raw.replace(";", ",").split(",")
    if len(tokens) != 3:
        raise ValueError(
            "--axis_map must contain three signed axes for mesh X,Y,Z, e.g. "
            "'+y,-x,+z'."
        )
    columns = []
    used_axes = []
    normalized_tokens = []
    for token in tokens:
        if not token:
            raise ValueError(f"Invalid empty token in --axis_map {axis_map!r}")
        sign = -1.0 if token.startswith("-") else 1.0
        axis = token[1:] if token[0] in "+-" else token
        if axis not in _AXIS_UNIT_VECTORS:
            raise ValueError(
                f"Invalid axis token {token!r} in --axis_map. Use signed x, y or z."
            )
        used_axes.append(axis)
        columns.append(sign * _AXIS_UNIT_VECTORS[axis])
        normalized_tokens.append(("+" if sign > 0 else "-") + axis)
    if len(set(used_axes)) != 3:
        raise ValueError(
            f"--axis_map must use x, y and z exactly once, got {axis_map!r}."
        )
    R_object_mesh = np.column_stack(columns)
    determinant = float(np.linalg.det(R_object_mesh))
    if determinant < 0.999:
        raise ValueError(
            f"--axis_map {axis_map!r} is left-handed (det={determinant:.0f}). "
            "Flip one sign so the mapping is a proper rotation."
        )
    return R_object_mesh, ",".join(normalized_tokens)


def axis_map_from_rotation(R_object_mesh: np.ndarray, atol: float = 1e-6) -> str:
    """Return the canonical axis-map string for a signed-permutation rotation."""
    R = np.asarray(R_object_mesh, dtype=np.float64)
    if R.shape != (3, 3):
        raise ValueError(f"Axis rotation must be 3x3, got {R.shape}")
    tokens = []
    used_axes = []
    axis_names = ("x", "y", "z")
    for column_index in range(3):
        column = R[:, column_index]
        axis_index = int(np.argmax(np.abs(column)))
        if not np.isclose(abs(column[axis_index]), 1.0, atol=atol):
            raise ValueError("Rotation is not a signed axis permutation.")
        residual = column.copy()
        residual[axis_index] = 0.0
        if not np.allclose(residual, 0.0, atol=atol):
            raise ValueError("Rotation is not a signed axis permutation.")
        used_axes.append(axis_index)
        sign = "+" if column[axis_index] > 0 else "-"
        tokens.append(sign + axis_names[axis_index])
    if len(set(used_axes)) != 3 or np.linalg.det(R) < 0.999:
        raise ValueError("Rotation is not a proper signed axis permutation.")
    return ",".join(tokens)


def describe_axis_rotation(R_object_mesh: np.ndarray) -> str:
    """Describe a signed permutation compactly, otherwise identify custom SO(3)."""
    try:
        return axis_map_from_rotation(R_object_mesh)
    except ValueError:
        return "custom_so3"


def proper_signed_axis_rotations() -> List[np.ndarray]:
    """Return all 24 proper rotations that permute/sign-flip XYZ axes."""
    rotations: List[np.ndarray] = []
    identity = np.eye(3, dtype=np.float64)
    for permutation in itertools.permutations(range(3)):
        basis = identity[:, permutation]
        for signs in itertools.product((-1.0, 1.0), repeat=3):
            candidate = basis @ np.diag(signs)
            if np.linalg.det(candidate) > 0.999:
                rotations.append(candidate)
    rotations.sort(key=axis_map_from_rotation)
    if len(rotations) != 24:
        raise AssertionError(f"Expected 24 proper axis rotations, got {len(rotations)}")
    return rotations


def calibrate_view_pose_branch_once(
    pred_pose_base_mesh_raw: np.ndarray,
    gt_pose_base_object: np.ndarray,
    T_object_mesh: np.ndarray,
    mode: str = "full",
    safe_branch_candidates: Optional[
        List[Tuple[str, np.ndarray]]
    ] = None,
    vertices_mesh_m: Optional[np.ndarray] = None,
    diameter_m: Optional[float] = None,
    add_weight: float = 0.8,
    adds_weight: float = 0.2,
    min_improvement_mm: float = 0.0,
) -> Tuple[np.ndarray, Dict[str, object]]:
    """Estimate one frozen local pose-branch correction for one camera.

    ``T_object_mesh`` is the single global object/mesh definition. It must not
    change from one camera to another. What can differ between independent
    MultiView registrations is the local pose branch returned for that same mesh.
    We model this explicitly:

        T_base_mesh_canonical = T_base_mesh_raw @ T_mesh_branch

    The correction is a proper rotation about the bbox-centred mesh origin and
    has zero translation. On the marked calibration frame, the desired
    canonical mesh orientation is:

        R_base_mesh_target = R_base_object_gt @ R_object_mesh

    ``full`` estimates the exact SO(3) correction. ``signed_permutation``
    selects one of the 24 proper signed XYZ rotations. When mesh vertices are
    supplied, candidates are ranked by weighted ADD/ADD-S (ADD is dominant so
    an end-for-end screwdriver flip is not hidden by nearest-neighbour ADD-S).
    ``geometry_safe`` searches only rotations independently verified as proxy
    symmetries. The selected correction applies to predictions only and is
    frozen for this view; no later GT pose is consulted.
    """
    pred = validate_rigid_transform(
        pred_pose_base_mesh_raw,
        "view-branch calibration raw T_base_mesh",
    )
    gt = validate_rigid_transform(
        gt_pose_base_object,
        "view-branch calibration GT T_base_object",
    )
    object_mesh = validate_rigid_transform(
        T_object_mesh,
        "view-branch calibration T_object_mesh",
    )
    mode = str(mode).strip().lower()
    if mode not in {
        "full", "signed_permutation", "geometry_safe",
    }:
        raise ValueError(
            "view-branch calibration mode must be 'full', "
            "'signed_permutation' or 'geometry_safe'"
        )

    target_pose_base_mesh = validate_rigid_transform(
        gt @ object_mesh,
        "view-branch calibration target T_base_mesh",
    )
    target_R_base_mesh = target_pose_base_mesh[:3, :3]
    direct_correction = project_rotation_to_so3(
        pred[:3, :3].T @ target_R_base_mesh
    )

    vertices: Optional[np.ndarray] = None
    if vertices_mesh_m is not None:
        vertices = np.asarray(
            vertices_mesh_m, dtype=np.float64
        ).reshape(-1, 3)
        if len(vertices) == 0 or not np.all(np.isfinite(vertices)):
            raise ValueError(
                "View-branch ADD calibration requires finite non-empty "
                "mesh vertices."
            )
    add_weight = float(add_weight)
    adds_weight = float(adds_weight)
    weight_sum = add_weight + adds_weight
    if add_weight < 0 or adds_weight < 0 or weight_sum <= 0:
        raise ValueError(
            "View-branch ADD/ADD-S weights must be non-negative and at "
            "least one must be positive."
        )
    add_weight /= weight_sum
    adds_weight /= weight_sum
    diameter = (
        max(float(diameter_m), 1e-12)
        if diameter_m is not None else float("nan")
    )

    if mode == "full":
        rotations_to_test = [
            (describe_axis_rotation(direct_correction), direct_correction)
        ]
        calibration_mode = "full_so3"
    elif mode == "geometry_safe":
        branch_candidates = (
            safe_branch_candidates
            if safe_branch_candidates
            else [("+x,+y,+z", np.eye(4, dtype=np.float64))]
        )
        rotations_to_test = []
        for name, transform in branch_candidates:
            correction = validate_pose_branch_correction(
                transform,
                f"geometry-safe calibration branch {name}",
            )
            rotations_to_test.append((
                str(name), correction[:3, :3].copy()
            ))
        calibration_mode = "geometry_safe_mesh_symmetries"
    else:
        rotations_to_test = [
            (axis_map_from_rotation(rotation), rotation)
            for rotation in proper_signed_axis_rotations()
        ]
        calibration_mode = "signed_permutation_add_adds"

    candidate_metrics: List[Dict[str, object]] = []
    candidate_rotations: List[np.ndarray] = []
    for name, candidate_R in rotations_to_test:
        candidate_correction = np.eye(4, dtype=np.float64)
        candidate_correction[:3, :3] = project_rotation_to_so3(
            candidate_R
        )
        aligned_pose = validate_rigid_transform(
            pred @ candidate_correction,
            f"view-branch candidate {name}",
        )
        rotation_error = rotation_distance_deg_from_matrices(
            aligned_pose[:3, :3],
            target_R_base_mesh,
        )
        add_mm = float("nan")
        adds_mm = float("nan")
        objective_mm = float("nan")
        if vertices is not None:
            add_mm = float(
                add_error_m(
                    vertices, aligned_pose, target_pose_base_mesh
                ) * 1000.0
            )
            adds_mm = float(
                add_s_error_m(
                    vertices, aligned_pose, target_pose_base_mesh
                ) * 1000.0
            )
            objective_mm = float(
                add_weight * add_mm + adds_weight * adds_mm
            )
        candidate_metrics.append({
            "candidate_index": int(len(candidate_metrics)),
            "branch_map": str(name),
            "rotation_error_deg": float(rotation_error),
            "ADD_mm": add_mm,
            "ADD-S_mm": adds_mm,
            "weighted_ADD_objective_mm": objective_mm,
            "ADD_over_diameter": (
                float(add_mm / (diameter * 1000.0))
                if np.isfinite(add_mm) and np.isfinite(diameter)
                else float("nan")
            ),
            "is_identity": bool(
                np.allclose(candidate_R, np.eye(3), atol=1e-10)
            ),
        })
        candidate_rotations.append(candidate_correction[:3, :3])

    use_add_ranking = bool(
        vertices is not None
        and len(candidate_metrics) > 1
        and all(
            np.isfinite(row["weighted_ADD_objective_mm"])
            for row in candidate_metrics
        )
    )
    if use_add_ranking:
        ranking = sorted(
            range(len(candidate_metrics)),
            key=lambda index: (
                candidate_metrics[index][
                    "weighted_ADD_objective_mm"
                ],
                candidate_metrics[index]["ADD_mm"],
                candidate_metrics[index]["rotation_error_deg"],
                candidate_metrics[index]["branch_map"],
            ),
        )
    else:
        ranking = sorted(
            range(len(candidate_metrics)),
            key=lambda index: (
                candidate_metrics[index]["rotation_error_deg"],
                candidate_metrics[index]["branch_map"],
            ),
        )
    best_index = int(ranking[0])
    selected_index = best_index
    objective_margin_mm = (
        float(
            candidate_metrics[ranking[1]][
                "weighted_ADD_objective_mm"
            ]
            - candidate_metrics[ranking[0]][
                "weighted_ADD_objective_mm"
            ]
        )
        if use_add_ranking and len(ranking) > 1
        else float("nan")
    )
    identity_index = next(
        (
            index
            for index, row in enumerate(candidate_metrics)
            if bool(row["is_identity"])
        ),
        None,
    )
    improvement_mm = float("nan")
    if use_add_ranking and identity_index is not None:
        improvement_mm = float(
            candidate_metrics[identity_index][
                "weighted_ADD_objective_mm"
            ]
            - candidate_metrics[best_index][
                "weighted_ADD_objective_mm"
            ]
        )
        if improvement_mm < float(min_improvement_mm):
            selected_index = int(identity_index)
    for index, row in enumerate(candidate_metrics):
        row["is_objective_best"] = bool(index == best_index)
        row["is_selected"] = bool(index == selected_index)

    correction_R = candidate_rotations[selected_index]
    correction_name = str(
        candidate_metrics[selected_index]["branch_map"]
    )
    rotation_rank = sorted(
        float(row["rotation_error_deg"])
        for row in candidate_metrics
    )
    second_error = (
        float(rotation_rank[1])
        if len(rotation_rank) > 1 else float("nan")
    )
    margin = (
        float(second_error - rotation_rank[0])
        if np.isfinite(second_error) else float("nan")
    )

    correction = np.eye(4, dtype=np.float64)
    correction[:3, :3] = project_rotation_to_so3(correction_R)
    aligned_pose = pred @ correction
    aligned_object_pose = aligned_pose @ np.linalg.inv(object_mesh)
    residual_deg = rotation_distance_deg_from_matrices(
        aligned_object_pose[:3, :3],
        gt[:3, :3],
    )
    rotvec = SciPyRotation.from_matrix(correction[:3, :3]).as_rotvec()
    angle_rad = float(np.linalg.norm(rotvec))
    if angle_rad > 1e-12:
        axis = (rotvec / angle_rad).tolist()
    else:
        axis = [0.0, 0.0, 1.0]

    diagnostics = {
        "calibration_mode": calibration_mode,
        "branch_map": correction_name,
        "branch_rotation_deg": float(np.degrees(angle_rad)),
        "branch_axis_mesh": [float(value) for value in axis],
        "best_rotation_error_deg": float(residual_deg),
        "second_best_rotation_error_deg": float(second_error),
        "selection_margin_deg": float(margin),
        "candidate_count": (
            len(candidate_metrics)
        ),
        "candidate_metrics": candidate_metrics,
        "selection_uses_add_adds": use_add_ranking,
        "ADD_ADD-S_used_only_for_diagnostics": bool(
            vertices is not None and not use_add_ranking
        ),
        "add_weight": add_weight,
        "adds_weight": adds_weight,
        "min_improvement_mm": float(min_improvement_mm),
        "identity_candidate_index": identity_index,
        "objective_best_candidate_index": best_index,
        "selected_candidate_index": selected_index,
        "improvement_over_identity_mm": improvement_mm,
        "best_vs_second_objective_margin_mm": objective_margin_mm,
        "selected_ADD_mm": float(
            candidate_metrics[selected_index]["ADD_mm"]
        ),
        "selected_ADD-S_mm": float(
            candidate_metrics[selected_index]["ADD-S_mm"]
        ),
        "selected_weighted_objective_mm": float(
            candidate_metrics[selected_index][
                "weighted_ADD_objective_mm"
            ]
        ),
        "selected_ADD_over_diameter": float(
            candidate_metrics[selected_index]["ADD_over_diameter"]
        ),
        "uses_gt_on_marked_calibration_frame": True,
        "translation_is_zero": True,
        "global_model_transform_changed": False,
        "scale_was_optimized_with_add": False,
        "scale_note": (
            "Scale is selected only from object_pose_setup and metadata bbox "
            "agreement. Optimizing scale with ADD/ADD-S is forbidden because "
            "shrinking a proxy model can reduce both errors artificially."
        ),
        "T_mesh_branch": correction.tolist(),
    }
    return correction, diagnostics


def validate_pose_branch_correction(
    transform: np.ndarray,
    label: str,
) -> np.ndarray:
    """Validate a local right-multiplied branch correction.

    A view branch may rotate the bbox-centred mesh frame but may not translate
    it. Allowing a translation here would silently calibrate position error and
    would make ADD/translation metrics invalid.
    """
    correction = validate_rigid_transform(transform, label).copy()
    translation_norm = float(np.linalg.norm(correction[:3, 3]))
    if translation_norm > 1e-9:
        raise ValueError(
            f"{label} must have zero translation; got "
            f"{translation_norm * 1000.0:.6f} mm"
        )
    correction[:3, 3] = 0.0
    return correction


def apply_pose_branch_correction(
    raw_pose_parent_mesh: np.ndarray,
    branch_correction: np.ndarray,
    label: str,
) -> np.ndarray:
    raw = ensure_pose_matrix(raw_pose_parent_mesh, f"{label} raw pose")
    correction = validate_pose_branch_correction(
        branch_correction,
        f"{label} branch correction",
    )
    return validate_rigid_transform(
        raw @ correction,
        f"{label} canonical pose",
    )


def branch_geometry_discrepancy(
    vertices_mesh: np.ndarray,
    branch_correction: np.ndarray,
) -> Dict[str, float]:
    """Report how nearly the correction is a symmetry of the proxy mesh."""
    vertices = np.asarray(vertices_mesh, dtype=np.float64).reshape(-1, 3)
    if len(vertices) == 0:
        return {
            "mean_bidirectional_mm": float("nan"),
            "max_bidirectional_mm": float("nan"),
        }
    correction = validate_pose_branch_correction(
        branch_correction,
        "branch geometry correction",
    )
    transformed = transform_points(vertices, correction)
    original_tree = cKDTree(vertices)
    transformed_tree = cKDTree(transformed)
    forward, _ = original_tree.query(transformed, k=1)
    backward, _ = transformed_tree.query(vertices, k=1)
    combined = np.concatenate([forward, backward])
    return {
        "mean_bidirectional_mm": float(np.mean(combined) * 1000.0),
        "max_bidirectional_mm": float(np.max(combined) * 1000.0),
    }


def detect_prediction_branch_symmetries(
    vertices_mesh: np.ndarray,
    diameter_m: float,
    args,
) -> Tuple[List[Tuple[str, np.ndarray]], Dict[str, object]]:
    """Detect safe discrete mesh symmetries for temporal branch recovery.

    This is prediction-side stabilization only. It never changes GT and it
    does not declare extra BOP metric symmetries. A signed local rotation is
    accepted only when its bidirectional model discrepancy is small relative
    to object diameter.
    """
    vertices = np.asarray(vertices_mesh, dtype=np.float64)
    mean_limit_mm = max(
        float(args.temporal_branch_symmetry_min_mean_mm),
        float(args.temporal_branch_symmetry_mean_ratio)
        * float(diameter_m) * 1000.0,
    )
    max_limit_mm = max(
        float(args.temporal_branch_symmetry_min_max_mm),
        float(args.temporal_branch_symmetry_max_ratio)
        * float(diameter_m) * 1000.0,
    )
    accepted: List[Tuple[str, np.ndarray]] = []
    diagnostics = []
    for rotation in proper_signed_axis_rotations():
        transform = np.eye(4, dtype=np.float64)
        transform[:3, :3] = rotation
        discrepancy = branch_geometry_discrepancy(vertices, transform)
        is_identity = np.allclose(rotation, np.eye(3), atol=1e-9)
        is_accepted = bool(
            is_identity
            or (
                discrepancy["mean_bidirectional_mm"] <= mean_limit_mm
                and discrepancy["max_bidirectional_mm"] <= max_limit_mm
            )
        )
        name = axis_map_from_rotation(rotation)
        diagnostics.append({
            "branch": name,
            **discrepancy,
            "accepted": is_accepted,
        })
        if is_accepted:
            accepted.append((name, transform))
    accepted.sort(key=lambda item: (item[0] != "+x,+y,+z", item[0]))
    return accepted, {
        "mean_limit_mm": float(mean_limit_mm),
        "max_limit_mm": float(max_limit_mm),
        "accepted_count": int(len(accepted)),
        "accepted_branches": [name for name, _ in accepted],
        "candidates": diagnostics,
        "note": (
            "Used only to keep prediction branches temporally consistent; "
            "does not alter GT or BOP symmetry declarations."
        ),
    }


def canonicalize_prediction_discrete_branch(
    raw_pose_parent_mesh: np.ndarray,
    reference_pose_parent_mesh: Optional[np.ndarray],
    branch_symmetries: List[Tuple[str, np.ndarray]],
    min_improvement_deg: float = 0.0,
) -> Tuple[np.ndarray, str, float, float]:
    """Select the geometry-equivalent prediction closest to a temporal prior."""
    raw_pose = ensure_pose_matrix(
        raw_pose_parent_mesh, "raw temporal-branch pose"
    )
    if reference_pose_parent_mesh is None or not branch_symmetries:
        return raw_pose.copy(), "+x,+y,+z", 0.0, 0.0
    reference = ensure_pose_matrix(
        reference_pose_parent_mesh, "temporal-branch reference"
    )
    identity_residual = rotation_distance_deg_from_matrices(
        raw_pose[:3, :3], reference[:3, :3]
    )
    candidates = []
    for name, transform in branch_symmetries:
        candidate = validate_rigid_transform(
            raw_pose @ transform,
            f"temporal prediction branch {name}",
        )
        residual = rotation_distance_deg_from_matrices(
            candidate[:3, :3], reference[:3, :3]
        )
        correction_angle = rotation_distance_deg_from_matrices(
            transform[:3, :3], np.eye(3)
        )
        candidates.append((
            float(residual),
            float(correction_angle),
            name,
            candidate,
        ))
    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    residual, correction_angle, name, selected = candidates[0]
    improvement = float(identity_residual - residual)
    if (
        name != "+x,+y,+z"
        and improvement < float(min_improvement_deg)
    ):
        return (
            raw_pose.copy(),
            "+x,+y,+z",
            0.0,
            float(identity_residual),
        )
    return selected, name, correction_angle, residual


def default_view_branch_output_path(
    args,
    save_root: str,
    view: str,
) -> str:
    """Resolve the deterministic output path for one calibrated view branch."""
    if args.proxy_calibration_output:
        stem, _ = os.path.splitext(
            os.path.abspath(args.proxy_calibration_output)
        )
        return f"{stem}_{view}_branch.txt"
    if args.axis_calibration_output:
        stem, _ = os.path.splitext(
            os.path.abspath(args.axis_calibration_output)
        )
        return f"{stem}_{view}_branch.txt"
    return os.path.join(
        save_root,
        f"{args.object_name}_axis_calibration",
        f"T_pose_branch_{view}.txt",
    )


def explicit_view_branch_path(args, view: str) -> Optional[str]:
    value = getattr(args, f"{view}_pose_branch_correction", None)
    if value is None or not str(value).strip():
        return None
    return os.path.abspath(str(value))


def resolve_T_base_cam(args, view: str) -> Tuple[np.ndarray, str]:
    """Load the explicitly enabled legacy fixed eye-to-hand fallback.

    Current datasets must instead use per-frame meta/T_cam_base. The legacy
    project convention is:
      right -> camera 1 -> T_base_cam1
      left  -> camera 2 -> T_base_cam2
    """
    if view == "right":
        path = args.T_base_cam1 or os.path.join(args.dataset_root, "cam", "T_base_cam1.txt")
        label = "T_base_cam1 for right view"
    elif view == "left":
        path = args.T_base_cam2 or os.path.join(args.dataset_root, "cam", "T_base_cam2.txt")
        label = "T_base_cam2 for left view"
    else:
        raise ValueError(f"Unsupported view for camera extrinsic: {view}")
    return load_transform_txt(path, label, args.camera_translation_scale), os.path.abspath(path)


def load_T_object_mesh(args) -> Tuple[np.ndarray, Optional[str]]:
    """Load the explicit transform from bbox-centred mesh frame to robot GT object frame.

    Definition:
        p_object = T_object_mesh @ p_mesh

    Identity is used only when the user states that both object frames are the same.
    """
    axis_map = getattr(args, "axis_map", None)
    if args.T_object_mesh is not None and axis_map not in (None, "", "none"):
        raise ValueError(
            "Use either --T_object_mesh or --axis_map, not both. A full "
            "T_object_mesh can include both an axis rotation and an origin offset."
        )
    if args.T_object_mesh is None and axis_map not in (None, "", "none"):
        R_object_mesh, normalized = parse_axis_map(axis_map)
        transform = np.eye(4, dtype=np.float64)
        transform[:3, :3] = R_object_mesh
        return transform, f"axis_map:{normalized}"
    if args.T_object_mesh is None:
        return np.eye(4, dtype=np.float64), None
    path = os.path.abspath(args.T_object_mesh)
    return load_transform_txt(
        path,
        "T_object_mesh",
        translation_scale=args.object_mesh_translation_scale,
    ), path


def validate_real_scene_units(
    reader: RealSceneMultiViewReader,
    mesh_diameter_m: float,
    T_object_mesh: np.ndarray,
    resolved_mesh_scale: float,
    args,
) -> None:
    """Fail early when GT, camera extrinsic, depth, or mesh units are inconsistent."""
    first_id = reader.frame_ids[0]
    T_base_object = reader.get_gt_pose_base(first_id)
    raw_t = reader.get_raw_gt_translation(first_id)
    T_cam_base, T_cam_base_source = reader.get_T_cam_base(first_id)
    T_base_mesh = T_base_object @ T_object_mesh
    T_cam_object = T_cam_base @ T_base_object
    T_cam_mesh = T_cam_object @ T_object_mesh

    validate_rigid_transform(T_base_object, f"T_base_object frame {first_id:04d}")
    validate_rigid_transform(
        T_cam_base,
        f"T_cam_base metadata frame {first_id:04d}",
    )
    validate_rigid_transform(T_object_mesh, "T_object_mesh")
    validate_rigid_transform(T_cam_mesh, f"converted T_cam_mesh GT frame {first_id:04d}")

    z_cam = float(T_cam_mesh[2, 3])
    norm_cam = float(np.linalg.norm(T_cam_mesh[:3, 3]))
    if not (args.min_valid_gt_z_m <= z_cam <= args.max_valid_gt_z_m):
        raise ValueError(
            f"Converted first-frame GT camera z is {z_cam:.6f} m. "
            f"Raw robot-base translation={raw_t.tolist()}, "
            f"GT translation is read directly in metres. Expected camera z in "
            f"[{args.min_valid_gt_z_m}, {args.max_valid_gt_z_m}] m. "
            f"Check metadata T_cam_base in {T_cam_base_source}."
        )
    if norm_cam > args.max_valid_gt_norm_m:
        raise ValueError(
            f"Converted first-frame GT camera translation norm is {norm_cam:.6f} m. "
            "Check metadata T_cam_base units/convention and the fixed model transform."
        )
    if not (args.min_mesh_diameter_m <= mesh_diameter_m <= args.max_mesh_diameter_m):
        raise ValueError(
            f"Mesh diameter is {mesh_diameter_m:.6f} m after resolved mesh "
            f"scale {resolved_mesh_scale:g}. "
            f"Expected [{args.min_mesh_diameter_m}, {args.max_mesh_diameter_m}] m. "
            "Check object_pose_setup.object_model_unit, --obj_mesh and "
            "--model_frame_mode."
        )
    depth = reader.get_depth(first_id)
    valid = np.isfinite(depth) & (depth > 0)
    if not valid.any():
        raise ValueError(f"First-frame depth contains no valid positive pixels: frame {first_id:04d}")
    median_depth = float(np.median(depth[valid]))
    if not (args.min_valid_depth_m <= median_depth <= args.max_valid_depth_m):
        raise ValueError(
            f"First-frame median depth is {median_depth:.6f} m. Expected "
            f"[{args.min_valid_depth_m}, {args.max_valid_depth_m}] m. "
            "Check --depth_npy_scale or --depth_png_scale."
        )


def pose_projection_state(pose: np.ndarray, bbox_corners: np.ndarray, K: np.ndarray,
                          image_shape: Tuple[int, int, int]) -> str:
    if pose is None or np.asarray(pose).shape != (4, 4) or not np.all(np.isfinite(pose)):
        return "invalid"
    z = float(pose[2, 3])
    if z <= 1e-6:
        return f"behind(z={z:.3f})"
    uv = project_points(transform_points(bbox_corners, pose), K)
    valid = np.isfinite(uv).all(axis=1)
    if valid.sum() < 2:
        return "not_projectable"
    h, w = image_shape[:2]
    inside = (
        (uv[:, 0] >= 0) & (uv[:, 0] < w) &
        (uv[:, 1] >= 0) & (uv[:, 1] < h) & valid
    )
    return "visible" if inside.any() else "off_image"


def ensure_pose_matrix(pose, label: str) -> np.ndarray:
    pose = np.asarray(pose, dtype=np.float64)
    if pose.shape != (4, 4):
        raise ValueError(f"{label} must be 4x4, got {pose.shape}")
    if not np.all(np.isfinite(pose)):
        raise ValueError(f"{label} contains NaN/Inf")
    return pose



def project_rotation_to_so3(R: np.ndarray) -> np.ndarray:
    """Project a near-rotation matrix onto SO(3)."""
    U, _, Vt = np.linalg.svd(np.asarray(R, dtype=np.float64))
    Rn = U @ Vt
    if np.linalg.det(Rn) < 0:
        U[:, -1] *= -1.0
        Rn = U @ Vt
    return Rn


def local_axis_rotation(axis_index: int, angle_rad: float) -> np.ndarray:
    """Return a 3x3 rotation around a local mesh axis."""
    c = float(np.cos(angle_rad))
    s = float(np.sin(angle_rad))
    if axis_index == 0:  # x
        return np.array([[1.0, 0.0, 0.0], [0.0, c, -s], [0.0, s, c]], dtype=np.float64)
    if axis_index == 1:  # y
        return np.array([[c, 0.0, s], [0.0, 1.0, 0.0], [-s, 0.0, c]], dtype=np.float64)
    if axis_index == 2:  # z
        return np.array([[c, -s, 0.0], [s, c, 0.0], [0.0, 0.0, 1.0]], dtype=np.float64)
    raise ValueError(f"Invalid axis index: {axis_index}")


def local_axis_flip_rotation(axis_index: int) -> np.ndarray:
    """Return a 180-degree local rotation that reverses the symmetry-axis direction."""
    # Rotate around a deterministic axis perpendicular to the cylinder axis.
    perpendicular_axis = 1 if axis_index == 0 else 0
    return local_axis_rotation(perpendicular_axis, np.pi)


def rotation_distance_deg_from_matrices(R_a: np.ndarray, R_b: np.ndarray) -> float:
    """Geodesic SO(3) distance between two rotation matrices."""
    R_a = project_rotation_to_so3(R_a)
    R_b = project_rotation_to_so3(R_b)
    R_delta = R_a @ R_b.T
    cos_theta = np.clip((np.trace(R_delta) - 1.0) * 0.5, -1.0, 1.0)
    return float(np.degrees(np.arccos(cos_theta)))


def resolve_symmetry_axis(vertices: np.ndarray, requested_axis: str) -> Tuple[int, str, np.ndarray]:
    """Resolve the cylinder axis in the bbox-centred mesh frame.

    `auto` deterministically selects the mesh axis with the largest bbox extent.
    Pass x/y/z explicitly when the OBJ axis convention is known.
    """
    axis_names = ["x", "y", "z"]
    requested_axis = requested_axis.lower().strip()
    extents = np.ptp(np.asarray(vertices, dtype=np.float64), axis=0)
    if requested_axis == "auto":
        axis_index = int(np.argmax(extents))
    else:
        if requested_axis not in axis_names:
            raise ValueError("--symmetry_axis must be auto, x, y, or z")
        axis_index = axis_names.index(requested_axis)
    return axis_index, axis_names[axis_index], extents


def canonicalize_axial_symmetric_pose(
    raw_pose_parent_mesh: np.ndarray,
    reference_pose_parent_mesh: np.ndarray,
    axis_index: int,
    samples: int,
    allow_axis_flip: bool,
) -> Tuple[np.ndarray, float, bool, float]:
    """Choose the symmetry-equivalent pose closest to a reference orientation.

    The symmetry operation is right-multiplied because it is defined in the local
    mesh frame. The translation is preserved exactly. Returned values are:
      canonical pose, selected axial angle in degrees, axis-flip flag,
      residual rotation distance to the reference in degrees.
    """
    raw_pose = ensure_pose_matrix(raw_pose_parent_mesh, "raw symmetric pose")
    ref_pose = ensure_pose_matrix(reference_pose_parent_mesh, "symmetry reference pose")
    if samples < 4:
        raise ValueError("symmetry samples must be >= 4")

    R_raw = project_rotation_to_so3(raw_pose[:3, :3])
    R_ref = project_rotation_to_so3(ref_pose[:3, :3])
    flip_options = [False, True] if allow_axis_flip else [False]
    flip_R = local_axis_flip_rotation(axis_index)

    best_R = R_raw
    best_angle_deg = 0.0
    best_flipped = False
    best_distance = float("inf")

    for flipped in flip_options:
        prefix = flip_R if flipped else np.eye(3, dtype=np.float64)
        for k in range(samples):
            angle = 2.0 * np.pi * float(k) / float(samples)
            symmetry_R = prefix @ local_axis_rotation(axis_index, angle)
            candidate_R = project_rotation_to_so3(R_raw @ symmetry_R)
            distance = rotation_distance_deg_from_matrices(candidate_R, R_ref)
            if distance < best_distance:
                best_distance = distance
                best_R = candidate_R
                best_angle_deg = float(np.degrees(angle))
                best_flipped = bool(flipped)

    result = raw_pose.copy()
    result[:3, :3] = best_R
    # For a bbox-centred mesh, local symmetry rotation must not change translation.
    result[:3, 3] = raw_pose[:3, 3]
    return result, best_angle_deg, best_flipped, best_distance


def _homogeneous_from_symmetry(sym: Dict[str, np.ndarray]) -> np.ndarray:
    T = np.eye(4, dtype=np.float64)
    T[:3, :3] = np.asarray(sym["R"], dtype=np.float64).reshape(3, 3)
    T[:3, 3] = np.asarray(sym["t"], dtype=np.float64).reshape(3)
    return T


def _symmetry_from_homogeneous(T: np.ndarray) -> Dict[str, np.ndarray]:
    T = validate_rigid_transform(T, "symmetry transform")
    return {
        "R": T[:3, :3].copy(),
        "t": T[:3, 3].reshape(3, 1).copy(),
    }


def transform_symmetries_to_object_frame(
    symmetries_mesh: List[Dict[str, np.ndarray]],
    T_object_mesh: np.ndarray,
) -> List[Dict[str, np.ndarray]]:
    """Express mesh-frame symmetry transforms in the fixed GT object frame."""
    object_mesh = validate_rigid_transform(
        T_object_mesh, "symmetry frame T_object_mesh"
    )
    mesh_object = np.linalg.inv(object_mesh)
    transformed = []
    for symmetry in symmetries_mesh:
        symmetry_mesh = _homogeneous_from_symmetry(symmetry)
        symmetry_object = (
            object_mesh @ symmetry_mesh @ mesh_object
        )
        transformed.append(
            _symmetry_from_homogeneous(symmetry_object)
        )
    return transformed


def _axis_angle_rotation(axis: np.ndarray, angle: float) -> np.ndarray:
    axis = np.asarray(axis, dtype=np.float64).reshape(3)
    norm = float(np.linalg.norm(axis))
    if norm < 1e-12:
        raise ValueError("Continuous-symmetry axis must be non-zero")
    x, y, z = axis / norm
    c, s = float(np.cos(angle)), float(np.sin(angle))
    C = 1.0 - c
    return np.array([
        [c + x * x * C, x * y * C - z * s, x * z * C + y * s],
        [y * x * C + z * s, c + y * y * C, y * z * C - x * s],
        [z * x * C - y * s, z * y * C + x * s, c + z * z * C],
    ], dtype=np.float64)


def discretize_continuous_symmetry(
    axis: np.ndarray,
    offset_m: np.ndarray,
    vertices_m: np.ndarray,
    diameter_m: float,
    max_sym_disc_step: float,
    max_count: int,
) -> List[np.ndarray]:
    """Discretize a BOP continuous symmetry in the mesh frame.

    The angular step is selected so that the maximum surface displacement
    between adjacent samples is at most max_sym_disc_step * diameter.
    """
    axis = np.asarray(axis, dtype=np.float64).reshape(3)
    axis /= max(float(np.linalg.norm(axis)), 1e-12)
    offset = np.asarray(offset_m, dtype=np.float64).reshape(3)
    rel = np.asarray(vertices_m, dtype=np.float64) - offset[None, :]
    radial = rel - (rel @ axis)[:, None] * axis[None, :]
    radii = np.linalg.norm(radial, axis=1)
    max_radius = float(radii.max()) if radii.size else 0.0
    max_displacement = max(float(max_sym_disc_step) * float(diameter_m), 1e-9)
    if max_radius <= 1e-12:
        count = 1
    else:
        ratio = min(1.0, max_displacement / (2.0 * max_radius))
        angular_step = max(2.0 * math.asin(ratio), 1e-6)
        count = int(math.ceil(2.0 * math.pi / angular_step))
    count = max(1, min(int(max_count), count))

    transforms = []
    for k in range(count):
        angle = 2.0 * math.pi * float(k) / float(count)
        R = _axis_angle_rotation(axis, angle)
        t = offset - R @ offset
        T = np.eye(4, dtype=np.float64)
        T[:3, :3] = R
        T[:3, 3] = t
        transforms.append(T)
    return transforms


def load_models_info_entry(args) -> Optional[Dict]:
    if args.models_info is None:
        return None
    path = os.path.abspath(args.models_info)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"--models_info does not exist: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if args.object_id is None:
        if any(k in data for k in ("diameter", "symmetries_discrete", "symmetries_continuous")):
            return data
        raise ValueError("--object_id is required when --models_info contains multiple objects")
    candidates = [str(args.object_id)]
    try:
        candidates.append(str(int(args.object_id)))
    except Exception:
        pass
    for key in candidates:
        if key in data:
            return data[key]
    raise KeyError(f"Object id {args.object_id!r} was not found in {path}")


def resolve_model_diameter_m(args, fallback_diameter_m: float) -> Tuple[float, str]:
    if args.model_diameter_m is not None:
        diameter = float(args.model_diameter_m)
        if diameter <= 0:
            raise ValueError("--model_diameter_m must be positive")
        return diameter, "cli"
    entry = load_models_info_entry(args)
    if entry is not None and "diameter" in entry:
        diameter = float(entry["diameter"]) * float(args.models_info_translation_scale)
        if diameter <= 0:
            raise ValueError("models_info diameter must be positive")
        return diameter, "models_info"
    return float(fallback_diameter_m), "computed_from_mesh"


def build_metric_symmetries(
    args,
    vertices_m: np.ndarray,
    diameter_m: float,
    resolved_axis_index: int,
) -> Tuple[List[Dict[str, np.ndarray]], str]:
    """Build the exact symmetry set used by MSSD/MSPD and symmetry-aware RE."""
    identity = np.eye(4, dtype=np.float64)
    mode = args.metric_symmetry
    if mode == "none":
        return [_symmetry_from_homogeneous(identity)], "identity_only"

    if mode == "axial":
        transforms = []
        flip_options = [False, True] if args.symmetry_allow_axis_flip else [False]
        flip_R = local_axis_flip_rotation(resolved_axis_index)
        for flipped in flip_options:
            prefix = flip_R if flipped else np.eye(3, dtype=np.float64)
            for k in range(args.symmetry_samples):
                angle = 2.0 * np.pi * float(k) / float(args.symmetry_samples)
                T = np.eye(4, dtype=np.float64)
                T[:3, :3] = prefix @ local_axis_rotation(resolved_axis_index, angle)
                transforms.append(T)
        return [_symmetry_from_homogeneous(T) for T in transforms], (
            f"axial_{args.symmetry_axis}_{args.symmetry_samples}"
            + ("_with_axis_flip" if args.symmetry_allow_axis_flip else "")
        )

    if mode != "models_info":
        raise ValueError(f"Unsupported metric symmetry mode: {mode}")

    entry = load_models_info_entry(args)
    if entry is None:
        raise ValueError("--metric_symmetry models_info requires --models_info")
    scale = float(args.models_info_translation_scale)

    discrete = [identity]
    for flat in entry.get("symmetries_discrete", []):
        T = np.asarray(flat, dtype=np.float64).reshape(4, 4).copy()
        T[:3, 3] *= scale
        discrete.append(validate_rigid_transform(T, "models_info discrete symmetry"))

    continuous_sets: List[List[np.ndarray]] = []
    for sym in entry.get("symmetries_continuous", []):
        axis = np.asarray(sym["axis"], dtype=np.float64)
        offset = np.asarray(sym.get("offset", [0.0, 0.0, 0.0]), dtype=np.float64) * scale
        continuous_sets.append(discretize_continuous_symmetry(
            axis=axis,
            offset_m=offset,
            vertices_m=vertices_m,
            diameter_m=diameter_m,
            max_sym_disc_step=args.max_sym_disc_step,
            max_count=args.max_symmetry_transforms,
        ))

    combined = discrete
    for continuous in continuous_sets:
        combined = [Td @ Tc for Td in combined for Tc in continuous]
        if len(combined) > args.max_symmetry_transforms:
            idx = np.linspace(
                0, len(combined) - 1,
                num=args.max_symmetry_transforms,
                dtype=int,
            )
            combined = [combined[i] for i in idx]

    # Remove numerical duplicates while preserving order.
    unique = []
    seen = set()
    for T in combined:
        key = tuple(np.round(T.reshape(-1), decimals=10))
        if key not in seen:
            seen.add(key)
            unique.append(T)
    return [_symmetry_from_homogeneous(T) for T in unique], (
        f"models_info:{os.path.abspath(args.models_info)}:object_id={args.object_id}"
    )


class BOPMetricEvaluator:
    """Adapter around the BOP implementation bundled with MultiView."""

    def __init__(
        self,
        mesh: trimesh.Trimesh,
        image_shape: Tuple[int, int, int],
        diameter_m: float,
        symmetries: List[Dict[str, np.ndarray]],
        args,
    ):
        try:
            from bop_toolkit_lib.pose_error_custom import (
                mssd as bop_mssd,
                mspd as bop_mspd,
                vsd as bop_vsd,
            )
            from bop_toolkit_lib.renderer_vispy import RendererVispy
        except Exception as exc:
            raise ImportError(
                "The MultiView-bundled bop_toolkit_lib is required for VSD/MSSD/MSPD. "
                "Run this script from the MultiView repository/environment, or add the "
                "repository to PYTHONPATH. Use --disable_bop_metrics only for a "
                "tracking-only diagnostic run."
            ) from exc

        self.mssd_fn = bop_mssd
        self.mspd_fn = bop_mspd
        self.vsd_fn = bop_vsd
        self.diameter_m = float(diameter_m)
        self.symmetries = symmetries
        self.vsd_taus = parse_float_list(args.vsd_taus, "--vsd_taus")
        self.vsd_correctness_thresholds = parse_float_list(
            args.vsd_correctness_thresholds,
            "--vsd_correctness_thresholds",
        )
        self.mssd_thresholds_d = parse_float_list(
            args.mssd_thresholds_d,
            "--mssd_thresholds_d",
        )
        self.mspd_thresholds_px = parse_float_list(
            args.mspd_thresholds_px,
            "--mspd_thresholds_px",
        )
        self.vsd_delta_mm = float(args.vsd_delta_mm)
        if self.vsd_delta_mm <= 0:
            raise ValueError("--vsd_delta_mm must be positive")

        h, w = image_shape[:2]
        try:
            self.renderer = RendererVispy(w, h, mode="depth")
            metric_mesh = {
                # MultiView's RendererVispy follows the BOP convention of millimetres.
                "pts": np.asarray(mesh.vertices, dtype=np.float64) * 1000.0,
                "normals": np.asarray(mesh.face_normals, dtype=np.float64),
                "faces": np.asarray(mesh.faces, dtype=np.int32),
            }
            self.renderer_object_id = 1
            self.renderer.my_add_object(metric_mesh, self.renderer_object_id)
        except Exception as exc:
            raise RuntimeError(
                "Failed to initialize the MultiView/BOP Vispy depth renderer. "
                "Verify that the MultiView environment has a working headless OpenGL "
                "backend before enabling VSD."
            ) from exc

    def evaluate(
        self,
        pred_pose_cam_mesh: np.ndarray,
        gt_pose_cam_mesh: np.ndarray,
        depth_m: np.ndarray,
        K: np.ndarray,
        metric_vertices_m: np.ndarray,
    ) -> Dict[str, object]:
        pred = ensure_pose_matrix(pred_pose_cam_mesh, "BOP predicted pose")
        gt = ensure_pose_matrix(gt_pose_cam_mesh, "BOP GT pose")
        depth = np.asarray(depth_m, dtype=np.float64)
        if depth.ndim != 2:
            raise ValueError(f"VSD depth image must be HxW, got {depth.shape}")
        depth = depth.copy()
        depth[~np.isfinite(depth) | (depth <= 0)] = 0.0

        mssd_m = float(self.mssd_fn(
            pose_est=pred,
            pose_gt=gt,
            pts=np.asarray(metric_vertices_m, dtype=np.float64),
            syms=self.symmetries,
        ))
        mspd_px = float(self.mspd_fn(
            pose_est=pred,
            pose_gt=gt,
            pts=np.asarray(metric_vertices_m, dtype=np.float64),
            K=np.asarray(K, dtype=np.float64).reshape(3, 3),
            syms=self.symmetries,
        ))

        pred_R = pred[:3, :3]
        pred_t_mm = pred[:3, 3].reshape(3, 1) * 1000.0
        gt_R = gt[:3, :3]
        gt_t_mm = gt[:3, 3].reshape(3, 1) * 1000.0
        vsd_errors = np.asarray(self.vsd_fn(
            pred_R,
            pred_t_mm,
            gt_R,
            gt_t_mm,
            depth * 1000.0,
            np.asarray(K, dtype=np.float64).reshape(3, 3),
            self.vsd_delta_mm,
            self.vsd_taus.tolist(),
            True,
            self.diameter_m * 1000.0,
            self.renderer,
            self.renderer_object_id,
        ), dtype=np.float64)

        mssd_ar = threshold_recall_contribution(
            mssd_m, self.mssd_thresholds_d * self.diameter_m
        )
        mspd_ar = threshold_recall_contribution(
            mspd_px, self.mspd_thresholds_px
        )
        vsd_ar = vsd_ar_contribution(
            vsd_errors, self.vsd_correctness_thresholds
        )
        bop_ar = float(np.mean([vsd_ar, mssd_ar, mspd_ar]))
        return {
            "VSD_errors": vsd_errors,
            "VSD_error_mean": float(np.mean(vsd_errors)),
            "VSD_AR_contribution": vsd_ar,
            "MSSD_m": mssd_m,
            "MSSD_norm_d": float(mssd_m / max(self.diameter_m, 1e-12)),
            "MSSD_AR_contribution": mssd_ar,
            "MSPD_px": mspd_px,
            "MSPD_AR_contribution": mspd_ar,
            "BOP_AR_contribution": bop_ar,
        }


# -----------------------------------------------------------------------------
# True two-view geometry fusion
# -----------------------------------------------------------------------------

def sample_mesh_surface_points(
    mesh: trimesh.Trimesh,
    max_points: int,
    seed: int = 31,
) -> np.ndarray:
    """Deterministically sample mesh surfaces in the bbox-centred mesh frame."""
    vertices = np.asarray(mesh.vertices, dtype=np.float64)
    faces = np.asarray(mesh.faces, dtype=np.int64)
    if max_points <= 0:
        max_points = max(len(vertices), 6000)
    if len(vertices) == 0:
        raise ValueError("Cannot sample fusion points from an empty mesh")
    if len(faces) == 0:
        return sample_vertices(vertices, max_points, seed=seed).astype(np.float64)

    tri = vertices[faces]
    areas = 0.5 * np.linalg.norm(
        np.cross(tri[:, 1] - tri[:, 0], tri[:, 2] - tri[:, 0]),
        axis=1,
    )
    valid_faces = np.isfinite(areas) & (areas > 1e-16)
    if not valid_faces.any():
        return sample_vertices(vertices, max_points, seed=seed).astype(np.float64)
    tri = tri[valid_faces]
    areas = areas[valid_faces]
    probabilities = areas / areas.sum()

    keep_vertices = min(len(vertices), max(64, max_points // 5))
    vertex_points = sample_vertices(vertices, keep_vertices, seed=seed + 1)
    n_surface = max(0, max_points - len(vertex_points))
    if n_surface == 0:
        return np.asarray(vertex_points, dtype=np.float64)

    rng = np.random.default_rng(seed)
    face_ids = rng.choice(len(tri), size=n_surface, replace=True, p=probabilities)
    selected = tri[face_ids]
    u = rng.random(n_surface)
    v = rng.random(n_surface)
    sqrt_u = np.sqrt(u)
    surface_points = (
        (1.0 - sqrt_u)[:, None] * selected[:, 0]
        + (sqrt_u * (1.0 - v))[:, None] * selected[:, 1]
        + (sqrt_u * v)[:, None] * selected[:, 2]
    )
    return np.vstack([vertex_points, surface_points]).astype(np.float64)


def align_pose_to_reference_symmetry(
    reference_pose: np.ndarray,
    candidate_pose: np.ndarray,
    symmetries: List[Dict[str, np.ndarray]],
    diameter_m: float,
) -> Tuple[np.ndarray, int, float, float]:
    """Return the symmetry-equivalent candidate closest to the reference pose."""
    reference = ensure_pose_matrix(reference_pose, "fusion reference pose")
    candidate = ensure_pose_matrix(candidate_pose, "fusion candidate pose")
    options = symmetries or [{"R": np.eye(3), "t": np.zeros((3, 1))}]
    best_pose = candidate.copy()
    best_index = 0
    best_rotation_deg = float("inf")
    best_translation_mm = float("inf")
    best_score = float("inf")
    diameter = max(float(diameter_m), 1e-9)
    for index, symmetry in enumerate(options):
        aligned = candidate @ _homogeneous_from_symmetry(symmetry)
        rotation_deg = rotation_distance_deg_from_matrices(
            reference[:3, :3], aligned[:3, :3]
        )
        translation_m = float(
            np.linalg.norm(reference[:3, 3] - aligned[:3, 3])
        )
        score = rotation_deg / 180.0 + translation_m / diameter
        if score < best_score:
            best_score = score
            best_pose = aligned
            best_index = index
            best_rotation_deg = rotation_deg
            best_translation_mm = translation_m * 1000.0
    return best_pose, best_index, best_rotation_deg, best_translation_mm


def blend_pose_on_se3(
    pose_a: np.ndarray,
    pose_b: np.ndarray,
    weight_b: float,
) -> np.ndarray:
    """Geodesically interpolate two already symmetry-aligned SE(3) poses."""
    a = ensure_pose_matrix(pose_a, "fusion pose A")
    b = ensure_pose_matrix(pose_b, "fusion pose B")
    alpha = float(np.clip(weight_b, 0.0, 1.0))
    relative_rotation = project_rotation_to_so3(a[:3, :3].T @ b[:3, :3])
    rotvec = SciPyRotation.from_matrix(relative_rotation).as_rotvec()
    result = np.eye(4, dtype=np.float64)
    result[:3, :3] = project_rotation_to_so3(
        a[:3, :3] @ SciPyRotation.from_rotvec(alpha * rotvec).as_matrix()
    )
    result[:3, 3] = (
        (1.0 - alpha) * a[:3, 3] + alpha * b[:3, 3]
    )
    return result


def _read_binary_mask(path: str, expected_shape: Tuple[int, int]) -> Optional[np.ndarray]:
    if not os.path.isfile(path):
        return None
    image = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    if image is None:
        return None
    if image.shape != expected_shape:
        raise ValueError(
            f"Fusion mask shape {image.shape} does not match depth/RGB shape "
            f"{expected_shape}: {path}"
        )
    return image > 0


def load_fusion_mask(
    reader: RealSceneMultiViewReader,
    frame_id: int,
    view_root: str,
    expected_shape: Tuple[int, int],
) -> Tuple[Optional[np.ndarray], str]:
    """Load an observed per-frame mask without using the GT pose."""
    try:
        mask = reader.get_mask(frame_id, fallback_to_first=False)
        if mask.shape != expected_shape:
            raise ValueError(
                f"Dataset mask shape {mask.shape} does not match {expected_shape}"
            )
        return mask.astype(np.bool_), "dataset_mask"
    except FileNotFoundError:
        pass

    sam_path = os.path.join(view_root, "sam2_masks", f"{frame_id:04d}.png")
    sam_mask = _read_binary_mask(sam_path, expected_shape)
    if sam_mask is not None:
        return sam_mask, "saved_sam2_mask"
    return None, "no_per_frame_mask"


def depth_mask_to_base_points(
    depth_m: np.ndarray,
    K: np.ndarray,
    mask: Optional[np.ndarray],
    T_base_cam: np.ndarray,
    max_points: int,
) -> np.ndarray:
    """Back-project observed masked depth into the robot-base frame."""
    if mask is None:
        return np.empty((0, 3), dtype=np.float64)
    depth = np.asarray(depth_m, dtype=np.float64)
    valid = (
        np.asarray(mask, dtype=bool)
        & np.isfinite(depth)
        & (depth > 0)
    )
    ys, xs = np.where(valid)
    if len(xs) == 0:
        return np.empty((0, 3), dtype=np.float64)
    if max_points > 0 and len(xs) > max_points:
        indices = np.linspace(0, len(xs) - 1, max_points, dtype=np.int64)
        xs = xs[indices]
        ys = ys[indices]
    z = depth[ys, xs]
    fx, fy = float(K[0, 0]), float(K[1, 1])
    cx, cy = float(K[0, 2]), float(K[1, 2])
    points_cam = np.column_stack([
        (xs.astype(np.float64) - cx) * z / fx,
        (ys.astype(np.float64) - cy) * z / fy,
        z,
    ])
    return transform_points(points_cam, T_base_cam)


def make_fusion_observation(
    reader: RealSceneMultiViewReader,
    frame_id: int,
    view_root: str,
    args,
) -> Dict[str, object]:
    rgb = reader.get_rgb(frame_id)
    depth = reader.get_depth(frame_id)
    if depth.shape != rgb.shape[:2]:
        raise ValueError(
            f"Fusion frame {frame_id:04d} {reader.view}: depth shape "
            f"{depth.shape} does not match RGB shape {rgb.shape[:2]}"
        )
    mask, mask_source = load_fusion_mask(
        reader=reader,
        frame_id=frame_id,
        view_root=view_root,
        expected_shape=depth.shape,
    )
    T_cam_base, T_cam_base_source = reader.get_T_cam_base(frame_id)
    T_base_cam = np.linalg.inv(T_cam_base)
    observed_points_base = depth_mask_to_base_points(
        depth_m=depth,
        K=reader.K,
        mask=mask,
        T_base_cam=T_base_cam,
        max_points=args.fusion_max_observed_points,
    )
    return {
        "view": reader.view,
        "rgb": rgb,
        "depth": depth,
        "mask": mask,
        "mask_source": mask_source,
        "mask_bbox": mask_bbox_xyxy(mask) if mask is not None else None,
        "observed_points_base": observed_points_base,
        "K": np.asarray(reader.K, dtype=np.float64),
        "T_base_cam": np.asarray(T_base_cam, dtype=np.float64),
        "T_cam_base": np.asarray(T_cam_base, dtype=np.float64),
        "T_cam_base_source": T_cam_base_source,
        "image_shape": rgb.shape,
    }


def _zbuffer_projected_samples(
    points_mesh: np.ndarray,
    T_cam_mesh: np.ndarray,
    K: np.ndarray,
    image_shape: Tuple[int, int, int],
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Project sampled surfaces and keep the nearest model depth per pixel."""
    h, w = image_shape[:2]
    points_cam = transform_points(points_mesh, T_cam_mesh)
    uv = project_points(points_cam, K)
    valid = (
        np.isfinite(uv).all(axis=1)
        & np.isfinite(points_cam[:, 2])
        & (points_cam[:, 2] > 1e-6)
    )
    if not valid.any():
        empty_i = np.empty(0, dtype=np.int64)
        empty_f = np.empty(0, dtype=np.float64)
        return empty_i, empty_i.copy(), empty_f
    uv = uv[valid]
    z = points_cam[valid, 2]
    x = np.rint(uv[:, 0]).astype(np.int64)
    y = np.rint(uv[:, 1]).astype(np.int64)
    inside = (x >= 0) & (x < w) & (y >= 0) & (y < h)
    x = x[inside]
    y = y[inside]
    z = z[inside]
    if len(x) == 0:
        empty_i = np.empty(0, dtype=np.int64)
        empty_f = np.empty(0, dtype=np.float64)
        return empty_i, empty_i.copy(), empty_f
    linear = y * w + x
    order = np.lexsort((z, linear))
    linear_sorted = linear[order]
    first = np.empty(len(order), dtype=bool)
    first[0] = True
    first[1:] = linear_sorted[1:] != linear_sorted[:-1]
    keep = order[first]
    return x[keep], y[keep], z[keep]


def _fusion_view_gate(
    diagnostics: Dict[str, object],
    args,
    threshold_factor: float = 1.0,
) -> bool:
    return not _fusion_view_gate_failure_reasons(
        diagnostics=diagnostics,
        args=args,
        threshold_factor=threshold_factor,
    )


def _fusion_view_gate_failure_reasons(
    diagnostics: Dict[str, object],
    args,
    threshold_factor: float = 1.0,
) -> List[str]:
    """Return rejection reasons for one candidate/view pair.

    ``strict`` reproduces the original all-cues-must-pass rule. ``balanced``
    still requires a valid projection and a bounded total geometry cost, but
    accepts a view when enough independent cue groups agree. This is more
    robust for thin proxy meshes, where one silhouette or point-cloud cue can
    be noisy even though the overall pose explains both cameras.
    """
    failures: List[str] = []
    if not bool(diagnostics.get("valid", False)):
        reason = str(diagnostics.get("reason", "invalid_geometry"))
        failures.append(f"invalid:{reason}")
        return failures

    if float(diagnostics.get("score", float("inf"))) > (
        args.fusion_view_pass_threshold * threshold_factor
    ):
        failures.append("geometry_cost")

    depth_ok = (
        float(diagnostics.get("depth_inlier_ratio", 0.0))
        >= args.fusion_min_depth_inlier_ratio
    )
    cue_results: List[Tuple[str, bool]] = [("depth", bool(depth_ok))]

    mask_available = bool(diagnostics.get("mask_available", False))
    if mask_available:
        bbox_ok = (
            float(diagnostics.get("mask_bbox_iou", 0.0))
            >= args.fusion_min_mask_bbox_iou
        )
        inside_ok = (
            float(diagnostics.get("projected_inside_mask_ratio", 0.0))
            >= args.fusion_min_projected_inside_mask_ratio
        )
        # A thin object can have a poor box IoU while most rendered samples
        # still lie inside the observed mask (or vice versa). Treat these as
        # two measurements of one silhouette cue, not two mandatory gates.
        cue_results.append(("mask", bool(bbox_ok or inside_ok)))
    else:
        bbox_ok = False
        inside_ok = False

    point_available = bool(
        diagnostics.get("point_term_available", False)
    )
    if point_available:
        point_ok = (
            float(diagnostics.get("point_inlier_ratio", 0.0))
            >= args.fusion_min_point_inlier_ratio
        )
        cue_results.append(("point", bool(point_ok)))
    else:
        point_ok = False

    passed_cues = [name for name, passed in cue_results if passed]
    failed_cues = [name for name, passed in cue_results if not passed]
    required_cues = min(
        max(1, int(args.fusion_min_consistency_cues)),
        len(cue_results),
    )
    diagnostics["fusion_gate_mode"] = str(args.fusion_gate_mode)
    diagnostics["consistency_cues_available"] = int(len(cue_results))
    diagnostics["consistency_cues_passed"] = int(len(passed_cues))
    diagnostics["consistency_cues_required"] = int(required_cues)
    diagnostics["passed_cues"] = "|".join(passed_cues)
    diagnostics["failed_cues"] = "|".join(failed_cues)

    if args.fusion_gate_mode == "strict":
        if not depth_ok:
            failures.append("depth_inlier_ratio")
        if mask_available:
            if not bbox_ok:
                failures.append("mask_bbox_iou")
            if not inside_ok:
                failures.append("projected_inside_mask_ratio")
        if point_available and not point_ok:
            failures.append("point_inlier_ratio")
    elif len(passed_cues) < required_cues:
        failures.append(
            f"consistency_cues_{len(passed_cues)}_of_{required_cues}"
        )
    return failures


def fusion_view_consistency(
    T_base_mesh: np.ndarray,
    observation: Dict[str, object],
    surface_points_mesh: np.ndarray,
    bbox_corners_mesh: np.ndarray,
    args,
    model_points_base: Optional[np.ndarray] = None,
    model_tree: Optional[cKDTree] = None,
) -> Dict[str, object]:
    """Evaluate whether one base-frame pose explains one RGB-D view."""
    try:
        pose_base = ensure_pose_matrix(T_base_mesh, "fusion base mesh pose")
        T_cam_mesh = observation["T_cam_base"] @ pose_base
    except Exception as exc:
        return {
            "valid": False,
            "passes": False,
            "score": float("inf"),
            "reason": f"invalid_pose:{type(exc).__name__}",
        }

    image_shape = observation["image_shape"]
    K = observation["K"]
    depth = observation["depth"]
    mask = observation["mask"]
    x, y, model_z = _zbuffer_projected_samples(
        points_mesh=surface_points_mesh,
        T_cam_mesh=T_cam_mesh,
        K=K,
        image_shape=image_shape,
    )
    projected_count = int(len(x))
    if projected_count < args.fusion_min_projected_points:
        return {
            "valid": False,
            "passes": False,
            "score": float("inf"),
            "reason": "too_few_projected_points",
            "projected_points": projected_count,
            "mask_available": mask is not None,
        }

    observed_z = np.asarray(depth, dtype=np.float64)[y, x]
    valid_observed_depth = np.isfinite(observed_z) & (observed_z > 0)
    signed_depth_residual = model_z - observed_z
    # A model point behind a closer observed surface can be physically occluded.
    # It is excluded rather than treated as a false free-space violation.
    non_occluded = valid_observed_depth & (
        signed_depth_residual <= args.fusion_occlusion_tolerance_m
    )
    compared_count = int(non_occluded.sum())
    valid_depth_count = int(valid_observed_depth.sum())
    occluded_count = int((valid_observed_depth & ~non_occluded).sum())

    if compared_count > 0:
        absolute_depth_residual = np.abs(signed_depth_residual[non_occluded])
        clipped_depth = np.minimum(
            absolute_depth_residual / args.fusion_depth_tolerance_m,
            1.0,
        )
        depth_error_norm = float(np.mean(clipped_depth))
        depth_inlier_ratio = float(
            np.mean(absolute_depth_residual <= args.fusion_depth_tolerance_m)
        )
        depth_rmse_mm = float(
            np.sqrt(np.mean(absolute_depth_residual ** 2)) * 1000.0
        )
    else:
        depth_error_norm = 1.0
        depth_inlier_ratio = 0.0
        depth_rmse_mm = float("nan")

    free_space_fraction = float(
        np.mean(
            signed_depth_residual[valid_observed_depth]
            < -args.fusion_depth_tolerance_m
        )
    ) if valid_depth_count else 1.0
    depth_coverage = float(compared_count / max(projected_count, 1))
    occluded_fraction = float(occluded_count / max(valid_depth_count, 1))

    pred_bbox = projected_bbox_xyxy(
        bbox_corners_mesh, T_cam_mesh, K
    )
    observed_bbox = observation["mask_bbox"]
    mask_available = mask is not None and observed_bbox is not None
    projected_axis, projected_axis_anisotropy = principal_axis_2d(
        np.column_stack([x, y])
    )
    if mask_available:
        projected_inside_mask_ratio = float(np.mean(mask[y, x]))
        mask_bbox_iou = bbox_iou_xyxy(pred_bbox, observed_bbox)
        silhouette_agreement = 0.5 * (
            projected_inside_mask_ratio + mask_bbox_iou
        )
        silhouette_error_norm = float(1.0 - silhouette_agreement)
        mask_y, mask_x = np.where(mask)
        mask_axis, mask_axis_anisotropy = principal_axis_2d(
            np.column_stack([mask_x, mask_y])
        )
        silhouette_axis_difference_deg = undirected_axis_angle_deg(
            projected_axis, mask_axis
        )
    else:
        projected_inside_mask_ratio = float("nan")
        mask_bbox_iou = float("nan")
        silhouette_error_norm = float("nan")
        mask_axis_anisotropy = float("nan")
        silhouette_axis_difference_deg = float("nan")

    observed_points_base = np.asarray(
        observation["observed_points_base"], dtype=np.float64
    )
    point_term_available = len(observed_points_base) >= args.fusion_min_observed_points
    if point_term_available:
        if model_points_base is None:
            model_points_base = transform_points(
                surface_points_mesh, pose_base
            )
        if model_tree is None:
            model_tree = cKDTree(model_points_base)
        distances_m, _ = model_tree.query(observed_points_base, k=1)
        clipped_point = np.minimum(
            distances_m / args.fusion_point_tolerance_m,
            1.0,
        )
        point_error_norm = float(np.mean(clipped_point))
        point_inlier_ratio = float(
            np.mean(distances_m <= args.fusion_point_tolerance_m)
        )
        point_rmse_mm = float(np.sqrt(np.mean(distances_m ** 2)) * 1000.0)
    else:
        point_error_norm = float("nan")
        point_inlier_ratio = float("nan")
        point_rmse_mm = float("nan")

    weighted_terms: List[Tuple[float, float]] = [
        (float(args.fusion_depth_weight), depth_error_norm),
        (float(args.fusion_inlier_weight), 1.0 - depth_inlier_ratio),
    ]
    if mask_available:
        weighted_terms.append(
            (float(args.fusion_mask_weight), silhouette_error_norm)
        )
    if point_term_available:
        weighted_terms.append(
            (float(args.fusion_point_weight), point_error_norm)
        )
    weight_sum = sum(weight for weight, _ in weighted_terms if weight > 0)
    if weight_sum <= 0:
        raise ValueError("At least one fusion geometry weight must be positive")
    score = float(
        sum(weight * value for weight, value in weighted_terms if weight > 0)
        / weight_sum
    )

    valid = (
        projected_count >= args.fusion_min_projected_points
        and compared_count >= args.fusion_min_compared_points
    )
    diagnostics = {
        "valid": bool(valid),
        "score": score,
        "reason": "ok" if valid else "too_few_depth_comparisons",
        "projected_points": projected_count,
        "valid_observed_depth_points": valid_depth_count,
        "compared_depth_points": compared_count,
        "depth_coverage": depth_coverage,
        "depth_error_norm": depth_error_norm,
        "depth_inlier_ratio": depth_inlier_ratio,
        "depth_rmse_mm": depth_rmse_mm,
        "free_space_fraction": free_space_fraction,
        "occluded_fraction": occluded_fraction,
        "mask_available": bool(mask_available),
        "mask_source": observation["mask_source"],
        "mask_bbox_iou": mask_bbox_iou,
        "projected_inside_mask_ratio": projected_inside_mask_ratio,
        "silhouette_error_norm": silhouette_error_norm,
        "projected_axis_anisotropy": projected_axis_anisotropy,
        "mask_axis_anisotropy": mask_axis_anisotropy,
        "silhouette_axis_difference_deg": (
            silhouette_axis_difference_deg
        ),
        "point_term_available": bool(point_term_available),
        "point_error_norm": point_error_norm,
        "point_inlier_ratio": point_inlier_ratio,
        "point_rmse_mm": point_rmse_mm,
        "pred_projection_state": pose_projection_state(
            T_cam_mesh, bbox_corners_mesh, K, image_shape
        ),
    }
    gate_failures = _fusion_view_gate_failure_reasons(diagnostics, args)
    diagnostics["gate_fail_reasons"] = "|".join(gate_failures)
    diagnostics["passes"] = not gate_failures
    return diagnostics


def evaluate_fusion_candidate(
    T_base_mesh: np.ndarray,
    observations: Dict[str, Dict[str, object]],
    surface_points_mesh: np.ndarray,
    bbox_corners_mesh: np.ndarray,
    args,
) -> Dict[str, object]:
    pose_base = ensure_pose_matrix(T_base_mesh, "fusion candidate pose")
    model_points_base = transform_points(surface_points_mesh, pose_base)
    model_tree = cKDTree(model_points_base)
    view_results = {}
    for view in ("right", "left"):
        view_results[view] = fusion_view_consistency(
            T_base_mesh=pose_base,
            observation=observations[view],
            surface_points_mesh=surface_points_mesh,
            bbox_corners_mesh=bbox_corners_mesh,
            args=args,
            model_points_base=model_points_base,
            model_tree=model_tree,
        )
    valid = all(bool(view_results[v].get("valid", False)) for v in ("right", "left"))
    passes = all(bool(view_results[v].get("passes", False)) for v in ("right", "left"))
    costs = [
        float(view_results[v].get("score", float("inf")))
        for v in ("right", "left")
    ]
    cost = float(np.mean(costs)) if np.isfinite(costs).all() else float("inf")
    return {
        "valid": bool(valid),
        "passes": bool(passes),
        "cost": cost,
        "views": view_results,
    }


def _fusion_candidate_passes(
    evaluation: Optional[Dict[str, object]],
    args,
    threshold_factor: float = 1.0,
) -> bool:
    if evaluation is None or not bool(evaluation.get("valid", False)):
        return False
    return all(
        _fusion_view_gate(
            evaluation["views"][view],
            args=args,
            threshold_factor=threshold_factor,
        )
        for view in ("right", "left")
    )


def principal_axis_2d(
    points_xy: np.ndarray,
    max_points: int = 5000,
) -> Tuple[Optional[np.ndarray], float]:
    """Return an undirected 2-D PCA axis and its anisotropy in ``[0, 1]``."""
    points = np.asarray(points_xy, dtype=np.float64).reshape(-1, 2)
    finite = np.isfinite(points).all(axis=1)
    points = points[finite]
    if len(points) < 3:
        return None, float("nan")
    if max_points > 0 and len(points) > max_points:
        indices = np.linspace(
            0, len(points) - 1, max_points, dtype=np.int64
        )
        points = points[indices]
    centered = points - points.mean(axis=0, keepdims=True)
    covariance = centered.T @ centered / max(len(centered) - 1, 1)
    eigenvalues, eigenvectors = np.linalg.eigh(covariance)
    major_index = int(np.argmax(eigenvalues))
    major_value = float(eigenvalues[major_index])
    minor_value = float(eigenvalues[1 - major_index])
    if not np.isfinite(major_value) or major_value <= 1e-12:
        return None, float("nan")
    axis = np.asarray(
        eigenvectors[:, major_index], dtype=np.float64
    )
    axis /= max(float(np.linalg.norm(axis)), 1e-12)
    anisotropy = float(
        np.clip((major_value - minor_value) / major_value, 0.0, 1.0)
    )
    return axis, anisotropy


def undirected_axis_angle_deg(
    axis_a: Optional[np.ndarray],
    axis_b: Optional[np.ndarray],
) -> float:
    """Smallest angle between two undirected 2-D axes, in ``[0, 90]``."""
    if axis_a is None or axis_b is None:
        return float("nan")
    a = np.asarray(axis_a, dtype=np.float64).reshape(2)
    b = np.asarray(axis_b, dtype=np.float64).reshape(2)
    denominator = float(np.linalg.norm(a) * np.linalg.norm(b))
    if denominator <= 1e-12:
        return float("nan")
    cosine = float(np.clip(abs(np.dot(a, b)) / denominator, 0.0, 1.0))
    return float(np.degrees(np.arccos(cosine)))


def make_tracking_observation_from_frame(
    reader: RealSceneMultiViewReader,
    frame_id: int,
    view_root: str,
    rgb: np.ndarray,
    depth: np.ndarray,
    T_cam_base: np.ndarray,
    T_base_cam: np.ndarray,
    args,
    mask_override: Optional[np.ndarray] = None,
) -> Dict[str, object]:
    """Build a GT-free single-view observation for tracking health checks."""
    if mask_override is not None:
        mask = np.asarray(mask_override, dtype=bool)
        mask_source = "current_registration_mask"
    else:
        mask, mask_source = load_fusion_mask(
            reader=reader,
            frame_id=frame_id,
            view_root=view_root,
            expected_shape=depth.shape,
        )
    observed_points_base = depth_mask_to_base_points(
        depth_m=depth,
        K=reader.K,
        mask=mask,
        T_base_cam=T_base_cam,
        max_points=args.fusion_max_observed_points,
    )
    return {
        "view": reader.view,
        "rgb": rgb,
        "depth": depth,
        "mask": mask,
        "mask_source": mask_source,
        "mask_bbox": (
            mask_bbox_xyxy(mask) if mask is not None else None
        ),
        "observed_points_base": observed_points_base,
        "K": np.asarray(reader.K, dtype=np.float64),
        "T_base_cam": np.asarray(T_base_cam, dtype=np.float64),
        "T_cam_base": np.asarray(T_cam_base, dtype=np.float64),
        "image_shape": rgb.shape,
    }


def evaluate_tracking_candidate_health(
    pose_base_mesh: np.ndarray,
    temporal_reference_base_mesh: Optional[np.ndarray],
    observation: Dict[str, object],
    surface_points_mesh: np.ndarray,
    bbox_corners_mesh: np.ndarray,
    geometry_args,
    args,
    branch_symmetries: Optional[
        List[Tuple[str, np.ndarray]]
    ] = None,
) -> Dict[str, object]:
    """Check one tracking pose without consulting GT.

    Thin tools often return a finite but incorrect pose, so exception-driven
    recovery is insufficient. This gate combines current RGB-D/mask agreement
    with a bounded constant-velocity innovation. Mask and depth are treated as
    alternative observability cues because metallic or transparent shafts can
    contain very sparse valid depth.
    """
    pose = ensure_pose_matrix(
        pose_base_mesh, "tracking-health candidate"
    )
    geometry = fusion_view_consistency(
        T_base_mesh=pose,
        observation=observation,
        surface_points_mesh=surface_points_mesh,
        bbox_corners_mesh=bbox_corners_mesh,
        args=geometry_args,
    )
    reference = (
        None
        if temporal_reference_base_mesh is None
        else ensure_pose_matrix(
            temporal_reference_base_mesh,
            "tracking-health temporal reference",
        )
    )
    if reference is None:
        translation_jump_mm = 0.0
        rotation_jump_deg = 0.0
        temporal_branch_for_jump = "+x,+y,+z"
    else:
        translation_jump_mm = float(
            np.linalg.norm(
                pose[:3, 3] - reference[:3, 3]
            ) * 1000.0
        )
        jump_candidates = [(
            rotation_distance_deg_from_matrices(
                pose[:3, :3], reference[:3, :3]
            ),
            "+x,+y,+z",
        )]
        for branch_name, branch_transform in (
            branch_symmetries or []
        ):
            branch = validate_pose_branch_correction(
                branch_transform,
                f"tracking-health temporal branch {branch_name}",
            )
            branch_pose = pose @ branch
            jump_candidates.append((
                rotation_distance_deg_from_matrices(
                    branch_pose[:3, :3], reference[:3, :3]
                ),
                str(branch_name),
            ))
        rotation_jump_deg, temporal_branch_for_jump = min(
            jump_candidates, key=lambda item: (item[0], item[1])
        )
        rotation_jump_deg = float(rotation_jump_deg)

    score = float(geometry.get("score", float("inf")))
    depth_inlier = float(
        geometry.get("depth_inlier_ratio", 0.0)
    )
    mask_available = bool(geometry.get("mask_available", False))
    mask_inside = float(
        geometry.get("projected_inside_mask_ratio", float("nan"))
    )
    mask_iou = float(
        geometry.get("mask_bbox_iou", float("nan"))
    )
    projected_axis_anisotropy = float(
        geometry.get("projected_axis_anisotropy", float("nan"))
    )
    mask_axis_anisotropy = float(
        geometry.get("mask_axis_anisotropy", float("nan"))
    )
    silhouette_axis_difference_deg = float(
        geometry.get(
            "silhouette_axis_difference_deg", float("nan")
        )
    )
    depth_cue_ok = bool(
        np.isfinite(depth_inlier)
        and depth_inlier
        >= float(args.tracking_health_min_depth_inlier_ratio)
    )
    mask_cue_ok = bool(
        mask_available
        and (
            (
                np.isfinite(mask_inside)
                and mask_inside
                >= float(
                    args.tracking_health_min_inside_mask_ratio
                )
            )
            or (
                np.isfinite(mask_iou)
                and mask_iou
                >= float(args.tracking_health_min_mask_bbox_iou)
            )
        )
    )
    projected_count = int(geometry.get("projected_points", 0))
    # Thin metallic shafts often have only a handful of valid depth pixels.
    # In the slender-tool profile, a sufficiently sampled silhouette is
    # therefore allowed to make the geometry observable even when the generic
    # fusion helper reports too few depth comparisons. The score and explicit
    # silhouette thresholds still have to pass, so this is not a blanket
    # acceptance of depthless poses.
    mask_only_fallback = bool(
        str(args.tracking_profile) == "slender_tool"
        and projected_count
        >= int(args.tracking_health_min_projected_points)
        and mask_cue_ok
    )
    geometry_observable = bool(
        geometry.get("valid", False) or mask_only_fallback
    )
    axis_cue_available = bool(
        str(args.tracking_profile) == "slender_tool"
        and mask_available
        and np.isfinite(projected_axis_anisotropy)
        and np.isfinite(mask_axis_anisotropy)
        and np.isfinite(silhouette_axis_difference_deg)
        and projected_axis_anisotropy
        >= float(args.tracking_health_axis_min_anisotropy)
        and mask_axis_anisotropy
        >= float(args.tracking_health_axis_min_anisotropy)
    )
    axis_cue_ok = bool(
        not axis_cue_available
        or silhouette_axis_difference_deg
        <= float(args.tracking_health_max_axis_difference_deg)
    )
    geometry_ok = bool(
        geometry_observable
        and np.isfinite(score)
        and score <= float(args.tracking_health_max_geometry_cost)
        and (depth_cue_ok or mask_cue_ok)
        and axis_cue_ok
    )
    temporal_ok = bool(
        reference is None
        or (
            translation_jump_mm
            <= float(args.tracking_health_max_translation_jump_mm)
            and rotation_jump_deg
            <= float(args.tracking_health_max_rotation_jump_deg)
        )
    )
    reasons = []
    if not geometry_observable:
        reasons.append(str(geometry.get("reason", "invalid_geometry")))
    if not np.isfinite(score):
        reasons.append("non_finite_geometry_cost")
    elif score > float(args.tracking_health_max_geometry_cost):
        reasons.append("geometry_cost")
    if not (depth_cue_ok or mask_cue_ok):
        reasons.append("no_observation_cue")
    if not axis_cue_ok:
        reasons.append("silhouette_axis_mismatch")
    if not temporal_ok:
        if (
            translation_jump_mm
            > float(args.tracking_health_max_translation_jump_mm)
        ):
            reasons.append("translation_jump")
        if (
            rotation_jump_deg
            > float(args.tracking_health_max_rotation_jump_deg)
        ):
            reasons.append("rotation_jump")

    normalized_translation = (
        translation_jump_mm
        / max(
            float(args.tracking_health_max_translation_jump_mm),
            1e-9,
        )
    )
    normalized_rotation = (
        rotation_jump_deg
        / max(
            float(args.tracking_health_max_rotation_jump_deg),
            1e-9,
        )
    )
    temporal_penalty = 0.5 * (
        normalized_translation + normalized_rotation
    )
    rank_score = (
        score
        if np.isfinite(score)
        else float(args.tracking_health_invalid_penalty)
    ) + float(args.tracking_health_temporal_weight) * temporal_penalty
    return {
        "passes": bool(geometry_ok and temporal_ok),
        "geometry_ok": geometry_ok,
        "temporal_ok": temporal_ok,
        "reason": "ok" if not reasons else "|".join(reasons),
        "rank_score": float(rank_score),
        "translation_jump_mm": float(translation_jump_mm),
        "rotation_jump_deg": float(rotation_jump_deg),
        "temporal_branch_for_jump": temporal_branch_for_jump,
        "depth_cue_ok": depth_cue_ok,
        "mask_cue_ok": mask_cue_ok,
        "mask_only_fallback": mask_only_fallback,
        "geometry_observable": geometry_observable,
        "axis_cue_available": axis_cue_available,
        "axis_cue_ok": axis_cue_ok,
        "silhouette_axis_difference_deg": (
            silhouette_axis_difference_deg
        ),
        "geometry": geometry,
    }


def synchronize_multiview_pose_state(
    estimator,
    raw_pose_cam_mesh: np.ndarray,
) -> str:
    """Synchronize a recovered raw mesh pose with MultiView's tracker state."""
    pose = ensure_pose_matrix(
        raw_pose_cam_mesh, "MultiView synchronized raw camera pose"
    )
    if not hasattr(estimator, "pose_last"):
        return "pose_last_not_exposed"
    current = getattr(estimator, "pose_last")
    try:
        if torch.is_tensor(current):
            value = torch.as_tensor(
                pose,
                dtype=current.dtype,
                device=current.device,
            )
            setattr(estimator, "pose_last", value)
        else:
            setattr(estimator, "pose_last", pose.copy())
        return "pose_last_synchronized"
    except Exception as exc:
        return f"pose_last_sync_failed:{type(exc).__name__}"


def optimize_joint_fusion_pose(
    initial_pose: np.ndarray,
    observations: Dict[str, Dict[str, object]],
    surface_points_mesh: np.ndarray,
    bbox_corners_mesh: np.ndarray,
    args,
) -> Tuple[np.ndarray, Dict[str, object], Dict[str, object]]:
    """Refine one common base-frame pose against both RGB-D observations."""
    initial = ensure_pose_matrix(initial_pose, "joint fusion initializer")
    translation_bound_mm = float(args.fusion_max_optimization_translation_mm)
    rotation_bound_rad = float(
        np.deg2rad(args.fusion_max_optimization_rotation_deg)
    )
    bounds = [
        (-translation_bound_mm, translation_bound_mm),
        (-translation_bound_mm, translation_bound_mm),
        (-translation_bound_mm, translation_bound_mm),
        (-rotation_bound_rad, rotation_bound_rad),
        (-rotation_bound_rad, rotation_bound_rad),
        (-rotation_bound_rad, rotation_bound_rad),
    ]

    cache: Dict[Tuple[float, ...], Tuple[np.ndarray, Dict[str, object], float]] = {}

    def pose_from_delta(delta: np.ndarray) -> np.ndarray:
        result = initial.copy()
        result[:3, :3] = project_rotation_to_so3(
            SciPyRotation.from_rotvec(delta[3:6]).as_matrix()
            @ initial[:3, :3]
        )
        result[:3, 3] = initial[:3, 3] + delta[:3] * 0.001
        return result

    def objective(delta: np.ndarray) -> float:
        key = tuple(np.round(np.asarray(delta, dtype=np.float64), 10))
        if key in cache:
            return cache[key][2]
        pose = pose_from_delta(np.asarray(delta, dtype=np.float64))
        evaluation = evaluate_fusion_candidate(
            T_base_mesh=pose,
            observations=observations,
            surface_points_mesh=surface_points_mesh,
            bbox_corners_mesh=bbox_corners_mesh,
            args=args,
        )
        geometry_cost = float(evaluation["cost"])
        if not np.isfinite(geometry_cost):
            geometry_cost = 10.0
        trans_norm = np.linalg.norm(delta[:3]) / max(translation_bound_mm, 1e-9)
        rot_norm = np.linalg.norm(delta[3:6]) / max(rotation_bound_rad, 1e-9)
        regularization = args.fusion_optimization_regularization * (
            trans_norm ** 2 + rot_norm ** 2
        )
        value = float(geometry_cost + regularization)
        cache[key] = (pose, evaluation, value)
        return value

    result = minimize(
        objective,
        x0=np.zeros(6, dtype=np.float64),
        method="Powell",
        bounds=bounds,
        options={
            "maxiter": int(args.fusion_optimization_max_iterations),
            "maxfev": int(args.fusion_optimization_max_evaluations),
            "xtol": float(args.fusion_optimization_xtol),
            "ftol": float(args.fusion_optimization_ftol),
            "disp": False,
        },
    )
    optimized_pose = pose_from_delta(np.asarray(result.x, dtype=np.float64))
    optimized_evaluation = evaluate_fusion_candidate(
        T_base_mesh=optimized_pose,
        observations=observations,
        surface_points_mesh=surface_points_mesh,
        bbox_corners_mesh=bbox_corners_mesh,
        args=args,
    )
    optimizer_info = {
        "success": bool(result.success),
        "status": int(result.status),
        "message": str(result.message),
        "nfev": int(result.nfev),
        "nit": int(getattr(result, "nit", -1)),
        "objective": float(result.fun),
        "delta_tx_mm": float(result.x[0]),
        "delta_ty_mm": float(result.x[1]),
        "delta_tz_mm": float(result.x[2]),
        "delta_rx_deg": float(np.degrees(result.x[3])),
        "delta_ry_deg": float(np.degrees(result.x[4])),
        "delta_rz_deg": float(np.degrees(result.x[5])),
    }
    return optimized_pose, optimized_evaluation, optimizer_info


def compute_view_softmax_contributions(
    right_energy: Optional[float],
    left_energy: Optional[float],
    tau: float,
) -> Tuple[float, float]:
    """Compute the paper's numerically stable two-view contribution weights.

    omega_v = exp(-(E_v - E_min) / tau)
              / sum_{v' in {L,R}} exp(-(E_v' - E_min) / tau)

    If only one view has a finite energy, that available view receives weight
    one. If neither energy is finite, both contributions are NaN.
    """
    if tau <= 0:
        raise ValueError("Fusion contribution temperature tau must be positive")
    energies = np.asarray([
        float("nan") if right_energy is None else float(right_energy),
        float("nan") if left_energy is None else float(left_energy),
    ], dtype=np.float64)
    finite = np.isfinite(energies)
    if not finite.any():
        return float("nan"), float("nan")
    if finite.sum() == 1:
        return (
            (1.0, 0.0) if finite[0] else (0.0, 1.0)
        )

    E_min = float(np.min(energies))
    qualities = np.exp(-(energies - E_min) / float(tau))
    denominator = float(np.sum(qualities))
    weights = qualities / max(denominator, np.finfo(np.float64).tiny)
    return float(weights[0]), float(weights[1])


def candidate_reporting_energy(
    evaluation: Optional[Dict[str, object]],
    invalid_view_penalty: float = 1.0,
) -> Optional[float]:
    """Return a finite paper-formula energy for an available pose candidate.

    Fusion component errors are normalized to approximately ``[0, 1]``. When a
    candidate cannot be evaluated in one camera (for example, too few projected
    depth samples), that camera receives the maximum normalized penalty instead
    of making the contribution weights NaN. This energy is used for the
    exported right/left contributions and for ranking current single-view
    fallbacks; it never makes an invalid candidate pass a hard fusion gate.
    """
    if evaluation is None:
        return None
    penalty = float(invalid_view_penalty)
    if not np.isfinite(penalty) or penalty <= 0:
        raise ValueError("invalid_view_penalty must be finite and positive")

    view_energies: List[float] = []
    for view in ("right", "left"):
        view_diagnostics = evaluation.get("views", {}).get(view, {})
        score = float(view_diagnostics.get("score", float("nan")))
        if (
            bool(view_diagnostics.get("valid", False))
            and np.isfinite(score)
        ):
            view_energies.append(float(np.clip(score, 0.0, penalty)))
        else:
            view_energies.append(penalty)
    return float(np.mean(view_energies))


def fuse_pose_candidates(
    right_pose_base_mesh: Optional[np.ndarray],
    left_pose_base_mesh: Optional[np.ndarray],
    previous_fused_pose_base_mesh: Optional[np.ndarray],
    observations: Dict[str, Dict[str, object]],
    surface_points_mesh: np.ndarray,
    bbox_corners_mesh: np.ndarray,
    symmetries: List[Dict[str, np.ndarray]],
    diameter_m: float,
    args,
    previous_fused_age_frames: int = 0,
) -> Tuple[Optional[np.ndarray], Dict[str, object]]:
    """Apply safe two-view fusion without reading ground truth.

    Passing, mutually consistent candidates are jointly optimized. If joint
    refinement is unsafe, or if the strict fusion conditions are not met, the
    best currently available canonical single-view pose is returned using only
    RGB-D geometry plus a bounded temporal prior. A previous pose is considered
    only when no current candidate exists, and only for a bounded number of
    frames.
    """
    right_eval = None
    left_eval = None
    if right_pose_base_mesh is not None:
        right_eval = evaluate_fusion_candidate(
            right_pose_base_mesh, observations, surface_points_mesh,
            bbox_corners_mesh, args,
        )
    if left_pose_base_mesh is not None:
        left_eval = evaluate_fusion_candidate(
            left_pose_base_mesh, observations, surface_points_mesh,
            bbox_corners_mesh, args,
        )

    right_pass = _fusion_candidate_passes(right_eval, args)
    left_pass = _fusion_candidate_passes(left_eval, args)
    diagnostics: Dict[str, object] = {
        "right_evaluation": right_eval,
        "left_evaluation": left_eval,
        "right_pass": bool(right_pass),
        "left_pass": bool(left_pass),
        "mutually_consistent": False,
        "candidate_translation_diff_mm": float("nan"),
        "candidate_sym_rotation_diff_deg": float("nan"),
        "left_alignment_symmetry_index": -1,
        "joint_attempted": False,
        "joint_accepted": False,
        "joint_deviation_translation_mm": float("nan"),
        "joint_deviation_rotation_deg": float("nan"),
        "joint_evaluation": None,
        "selected_evaluation": None,
        "optimizer": None,
        "recovery_evaluation": None,
        "selected_source": "failed",
        "fusion_status": "failed",
        "Right_view_contribution": float("nan"),
        "Left_view_contribution": float("nan"),
        "right_selection_score": float("nan"),
        "left_selection_score": float("nan"),
        "previous_fused_age_frames": int(previous_fused_age_frames),
    }

    right_energy = candidate_reporting_energy(right_eval)
    left_energy = candidate_reporting_energy(left_eval)
    contribution_right, contribution_left = (
        compute_view_softmax_contributions(
            right_energy=right_energy,
            left_energy=left_energy,
            tau=float(args.fusion_weight_temperature),
        )
    )
    diagnostics["Right_view_contribution"] = contribution_right
    diagnostics["Left_view_contribution"] = contribution_left
    diagnostics["right_candidate_reporting_energy"] = right_energy
    diagnostics["left_candidate_reporting_energy"] = left_energy
    diagnostics["initializer_weight_right"] = contribution_right
    diagnostics["initializer_weight_left"] = contribution_left

    aligned_left = left_pose_base_mesh
    if right_pose_base_mesh is not None and left_pose_base_mesh is not None:
        (
            aligned_left,
            symmetry_index,
            rotation_diff_deg,
            translation_diff_mm,
        ) = align_pose_to_reference_symmetry(
            reference_pose=right_pose_base_mesh,
            candidate_pose=left_pose_base_mesh,
            symmetries=symmetries,
            diameter_m=diameter_m,
        )
        mutually_consistent = (
            translation_diff_mm <= args.fusion_max_candidate_translation_mm
            and rotation_diff_deg <= args.fusion_max_candidate_rotation_deg
        )
        diagnostics.update({
            "mutually_consistent": bool(mutually_consistent),
            "candidate_translation_diff_mm": float(translation_diff_mm),
            "candidate_sym_rotation_diff_deg": float(rotation_diff_deg),
            "left_alignment_symmetry_index": int(symmetry_index),
        })

    def candidate_selection_score(
        source: str,
        pose: np.ndarray,
        evaluation: Dict[str, object],
    ) -> float:
        """Rank a current candidate without any GT-derived quantity."""
        reporting_energy = candidate_reporting_energy(evaluation)
        if reporting_energy is None or not np.isfinite(reporting_energy):
            reporting_energy = float(args.fusion_fallback_invalid_penalty)
        own_diag = evaluation.get("views", {}).get(source, {})
        own_score = float(own_diag.get("score", float("nan")))
        if not (
            bool(own_diag.get("valid", False))
            and np.isfinite(own_score)
        ):
            own_score = float(args.fusion_fallback_invalid_penalty)
        geometry_score = (
            (1.0 - args.fusion_fallback_own_view_weight)
            * float(reporting_energy)
            + args.fusion_fallback_own_view_weight * own_score
        )

        temporal_score = 0.0
        if previous_fused_pose_base_mesh is not None:
            previous = ensure_pose_matrix(
                previous_fused_pose_base_mesh,
                "previous fused pose for fallback ranking",
            )
            translation_mm = float(
                np.linalg.norm(
                    pose[:3, 3] - previous[:3, 3]
                ) * 1000.0
            )
            rotation_deg = rotation_distance_deg_from_matrices(
                pose[:3, :3],
                previous[:3, :3],
            )
            translation_norm = min(
                translation_mm
                / max(args.fusion_fallback_translation_scale_mm, 1e-9),
                3.0,
            )
            rotation_norm = min(
                rotation_deg
                / max(args.fusion_fallback_rotation_scale_deg, 1e-9),
                3.0,
            )
            temporal_score = 0.5 * (
                translation_norm + rotation_norm
            )
        return float(
            geometry_score
            + args.fusion_fallback_temporal_weight * temporal_score
        )

    selection_scores: Dict[str, float] = {}
    for source, pose, evaluation in (
        ("right", right_pose_base_mesh, right_eval),
        ("left", left_pose_base_mesh, left_eval),
    ):
        if pose is None or evaluation is None:
            continue
        selection_scores[source] = candidate_selection_score(
            source=source,
            pose=pose,
            evaluation=evaluation,
        )
        diagnostics[f"{source}_selection_score"] = float(
            selection_scores[source]
        )

    def choose_best_single(
        require_pass: bool,
    ) -> Tuple[np.ndarray, str, Dict[str, object]]:
        candidates = []
        for source, pose, evaluation, passed in (
            ("right", right_pose_base_mesh, right_eval, right_pass),
            ("left", left_pose_base_mesh, left_eval, left_pass),
        ):
            if pose is None or evaluation is None:
                continue
            if require_pass and not passed:
                continue
            score = selection_scores.get(source, float("inf"))
            own_score = float(
                evaluation.get("views", {})
                .get(source, {})
                .get("score", float("inf"))
            )
            # Deterministic last-resort tie break. It has no effect unless all
            # geometry and temporal terms are numerically equal.
            tie_rank = 0 if source == args.fusion_tie_break_view else 1
            candidates.append((
                float(score),
                own_score,
                tie_rank,
                pose,
                source,
                evaluation,
            ))
        if not candidates:
            requirement = "passing" if require_pass else "available"
            raise RuntimeError(
                f"No {requirement} single-view fusion candidate"
            )
        candidates.sort(key=lambda item: item[:3])
        _, _, _, pose, source, evaluation = candidates[0]
        return pose.copy(), source, evaluation

    if right_pass and left_pass and diagnostics["mutually_consistent"]:
        diagnostics["joint_attempted"] = True
        weight_left = float(diagnostics["Left_view_contribution"])

        blended_pose = blend_pose_on_se3(
            right_pose_base_mesh, aligned_left, weight_left
        )
        blended_eval = evaluate_fusion_candidate(
            blended_pose, observations, surface_points_mesh,
            bbox_corners_mesh, args,
        )
        diagnostics["blended_initializer_evaluation"] = blended_eval

        aligned_left_eval = left_eval
        if not np.allclose(
            aligned_left,
            left_pose_base_mesh,
            atol=1e-10,
            rtol=1e-10,
        ):
            aligned_left_eval = evaluate_fusion_candidate(
                aligned_left,
                observations,
                surface_points_mesh,
                bbox_corners_mesh,
                args,
            )
        initializers = [
            (float(right_eval["cost"]), right_pose_base_mesh),
            (float(aligned_left_eval["cost"]), aligned_left),
        ]
        if np.isfinite(float(blended_eval["cost"])):
            initializers.append((float(blended_eval["cost"]), blended_pose))
        initializers.sort(key=lambda item: item[0])
        initial_pose = initializers[0][1]

        optimized_pose, joint_eval, optimizer_info = optimize_joint_fusion_pose(
            initial_pose=initial_pose,
            observations=observations,
            surface_points_mesh=surface_points_mesh,
            bbox_corners_mesh=bbox_corners_mesh,
            args=args,
        )
        diagnostics["joint_evaluation"] = joint_eval
        diagnostics["optimizer"] = optimizer_info

        best_single_pose, best_source, best_single_eval = (
            choose_best_single(require_pass=True)
        )
        best_single_cost = float(best_single_eval["cost"])
        (
            joint_pose_for_acceptance,
            joint_alignment_symmetry_index,
            joint_deviation_rotation_deg,
            joint_deviation_translation_mm,
        ) = align_pose_to_reference_symmetry(
            reference_pose=best_single_pose,
            candidate_pose=optimized_pose,
            symmetries=symmetries,
            diameter_m=diameter_m,
        )
        if not np.allclose(
            joint_pose_for_acceptance,
            optimized_pose,
            atol=1e-10,
            rtol=1e-10,
        ):
            joint_eval_for_acceptance = evaluate_fusion_candidate(
                joint_pose_for_acceptance,
                observations,
                surface_points_mesh,
                bbox_corners_mesh,
                args,
            )
        else:
            joint_eval_for_acceptance = joint_eval
        acceptance_limit = (
            best_single_cost * (1.0 + args.fusion_joint_accept_relative_margin)
            + args.fusion_joint_accept_absolute_margin
        )
        joint_acceptable = (
            _fusion_candidate_passes(joint_eval_for_acceptance, args)
            and float(joint_eval_for_acceptance["cost"])
            <= acceptance_limit
            and joint_deviation_translation_mm
            <= args.fusion_joint_max_deviation_translation_mm
            and joint_deviation_rotation_deg
            <= args.fusion_joint_max_deviation_rotation_deg
        )
        diagnostics["best_single_source"] = best_source
        diagnostics["best_single_cost"] = best_single_cost
        diagnostics["joint_acceptance_limit"] = float(acceptance_limit)
        diagnostics["joint_alignment_symmetry_index"] = int(
            joint_alignment_symmetry_index
        )
        diagnostics["joint_deviation_translation_mm"] = float(
            joint_deviation_translation_mm
        )
        diagnostics["joint_deviation_rotation_deg"] = float(
            joint_deviation_rotation_deg
        )
        diagnostics["joint_accepted"] = bool(joint_acceptable)
        if joint_acceptable:
            diagnostics["selected_source"] = "joint_optimized"
            diagnostics["fusion_status"] = "valid"
            diagnostics["selected_evaluation"] = (
                joint_eval_for_acceptance
            )
            return joint_pose_for_acceptance, diagnostics
        diagnostics["selected_source"] = f"{best_source}_joint_rejected"
        diagnostics["fusion_status"] = "valid"
        diagnostics["selected_evaluation"] = best_single_eval
        return best_single_pose, diagnostics

    if right_pass and left_pass:
        best_pose, best_source, best_evaluation = choose_best_single(
            require_pass=True
        )
        diagnostics["selected_source"] = f"{best_source}_candidates_inconsistent"
        diagnostics["fusion_status"] = "valid"
        diagnostics["selected_evaluation"] = best_evaluation
        return best_pose, diagnostics
    if right_pass:
        diagnostics["selected_source"] = "right_only_passed"
        diagnostics["fusion_status"] = "valid"
        diagnostics["selected_evaluation"] = right_eval
        return right_pose_base_mesh.copy(), diagnostics
    if left_pass:
        diagnostics["selected_source"] = "left_only_passed"
        diagnostics["fusion_status"] = "valid"
        diagnostics["selected_evaluation"] = left_eval
        return left_pose_base_mesh.copy(), diagnostics

    current_candidates_available = bool(
        right_pose_base_mesh is not None
        or left_pose_base_mesh is not None
    )
    if (
        current_candidates_available
        and args.fusion_no_pass_fallback == "best_single"
    ):
        best_pose, best_source, best_evaluation = choose_best_single(
            require_pass=False
        )
        diagnostics["best_single_source"] = best_source
        diagnostics["best_single_cost"] = float(
            best_evaluation.get("cost", float("inf"))
        )
        diagnostics["selected_source"] = (
            f"{best_source}_best_available_no_gate"
        )
        diagnostics["fusion_status"] = "fallback"
        diagnostics["selected_evaluation"] = best_evaluation
        return best_pose, diagnostics

    if (
        not current_candidates_available
        and args.fusion_failure_policy == "previous"
        and previous_fused_pose_base_mesh is not None
        and int(previous_fused_age_frames)
        < int(args.fusion_max_recovery_frames)
    ):
        recovery_eval = evaluate_fusion_candidate(
            previous_fused_pose_base_mesh,
            observations,
            surface_points_mesh,
            bbox_corners_mesh,
            args,
        )
        diagnostics["recovery_evaluation"] = recovery_eval
        if _fusion_candidate_passes(
            recovery_eval,
            args,
            threshold_factor=args.fusion_recovery_threshold_factor,
        ):
            diagnostics["selected_source"] = "previous_pose_recovery"
            diagnostics["fusion_status"] = "recovered"
            diagnostics["selected_evaluation"] = recovery_eval
            return previous_fused_pose_base_mesh.copy(), diagnostics

    diagnostics["selected_source"] = "fusion_failed_no_valid_candidate"
    diagnostics["fusion_status"] = "failed"
    return None, diagnostics


def _blank_bop_result(num_taus: int, ar_failure: float) -> Dict[str, object]:
    return {
        "VSD_errors": np.full(num_taus, np.nan, dtype=np.float64),
        "VSD_error_mean": float("nan"),
        "VSD_AR_contribution": ar_failure,
        "MSSD_m": float("nan"),
        "MSSD_norm_d": float("nan"),
        "MSSD_AR_contribution": ar_failure,
        "MSPD_px": float("nan"),
        "MSPD_AR_contribution": ar_failure,
        "BOP_AR_contribution": ar_failure,
    }


def _mean_finite(values: List[float]) -> float:
    array = np.asarray(values, dtype=np.float64)
    finite = np.isfinite(array)
    return float(np.mean(array[finite])) if finite.any() else float("nan")


def _flatten_fusion_evaluation(
    output: Dict[str, object],
    prefix: str,
    evaluation: Optional[Dict[str, object]],
) -> None:
    output[f"{prefix}_cost"] = (
        float(evaluation["cost"]) if evaluation is not None else float("nan")
    )
    output[f"{prefix}_passes"] = bool(
        evaluation is not None and evaluation.get("passes", False)
    )
    for view in ("right", "left"):
        view_diag = (
            evaluation.get("views", {}).get(view, {})
            if evaluation is not None else {}
        )
        stem = f"{prefix}_on_{view}"
        for key in (
            "score",
            "depth_error_norm",
            "depth_inlier_ratio",
            "depth_rmse_mm",
            "depth_coverage",
            "free_space_fraction",
            "occluded_fraction",
            "mask_bbox_iou",
            "projected_inside_mask_ratio",
            "point_error_norm",
            "point_inlier_ratio",
            "point_rmse_mm",
        ):
            output[f"{stem}_{key}"] = float(
                view_diag.get(key, float("nan"))
            )
        output[f"{stem}_valid"] = bool(view_diag.get("valid", False))
        output[f"{stem}_passes"] = bool(view_diag.get("passes", False))
        output[f"{stem}_reason"] = str(view_diag.get("reason", ""))
        output[f"{stem}_gate_fail_reasons"] = str(
            view_diag.get("gate_fail_reasons", "")
        )
        output[f"{stem}_fusion_gate_mode"] = str(
            view_diag.get("fusion_gate_mode", "")
        )
        for key in (
            "consistency_cues_available",
            "consistency_cues_passed",
            "consistency_cues_required",
        ):
            output[f"{stem}_{key}"] = int(view_diag.get(key, 0))
        output[f"{stem}_passed_cues"] = str(
            view_diag.get("passed_cues", "")
        )
        output[f"{stem}_failed_cues"] = str(
            view_diag.get("failed_cues", "")
        )
        output[f"{stem}_mask_source"] = str(
            view_diag.get("mask_source", "")
        )


def print_fusion_decision_summary(records: List[Dict[str, object]]) -> None:
    """Print failure diagnostics without adding unwanted Excel columns."""
    if not records:
        return
    status_counts = Counter(
        str(record.get("fusion_status", "unknown")) for record in records
    )
    print(
        "[Fusion summary] "
        + ", ".join(
            f"{status}={status_counts.get(status, 0)}"
            for status in ("valid", "fallback", "recovered", "failed")
        )
        + f", total={len(records)}"
    )

    failure_reasons: Counter = Counter()
    finite_contribution_rows = 0
    for record in records:
        right_weight = float(
            record.get("Right_view_contribution", float("nan"))
        )
        left_weight = float(
            record.get("Left_view_contribution", float("nan"))
        )
        if np.isfinite(right_weight) and np.isfinite(left_weight):
            finite_contribution_rows += 1
        if str(record.get("fusion_status", "")) != "failed":
            continue
        for candidate in ("right_candidate", "left_candidate"):
            available_key = (
                "right_candidate_available"
                if candidate.startswith("right")
                else "left_candidate_available"
            )
            if not bool(record.get(available_key, False)):
                failure_reasons[f"{candidate}:missing_pose"] += 1
                continue
            for view in ("right", "left"):
                raw = str(
                    record.get(
                        f"{candidate}_on_{view}_gate_fail_reasons", ""
                    )
                ).strip()
                if not raw:
                    continue
                for reason in raw.split("|"):
                    reason = reason.strip()
                    if reason:
                        failure_reasons[
                            f"{candidate}_on_{view}:{reason}"
                        ] += 1
                failed_cues = str(
                    record.get(
                        f"{candidate}_on_{view}_failed_cues", ""
                    )
                ).strip()
                for cue in failed_cues.split("|"):
                    cue = cue.strip()
                    if cue:
                        failure_reasons[
                            f"{candidate}_on_{view}:failed_{cue}_cue"
                        ] += 1

    print(
        "[Fusion contribution] finite R/L weights for "
        f"{finite_contribution_rows}/{len(records)} rows"
    )
    if failure_reasons:
        details = ", ".join(
            f"{reason}={count}"
            for reason, count in failure_reasons.most_common(12)
        )
        print(f"[Fusion failure gates] {details}")


def default_proxy_calibration_output_path(
    args,
    save_root: str,
) -> str:
    if args.proxy_calibration_output:
        return os.path.abspath(args.proxy_calibration_output)
    return os.path.join(
        save_root,
        f"{args.object_name}_proxy_calibration",
        "T_object_mesh_proxy_calibrated.txt",
    )


def _disabled_global_proxy_calibration_legacy(
    args,
    view: str,
    save_root: str,
) -> Dict[str, object]:
    """Removed v10 experiment retained only to make old traces recognizable.

    A first-frame prediction must never rewrite the global T_object_mesh.
    Proxy calibration is now performed independently on each view as a
    prediction-only zero-translation branch in process_one_view().
    """
    raise RuntimeError(
        "Global prediction-derived T_object_mesh calibration is disabled. "
        "Use --calibrate_proxy_model_from_first_frame, which calibrates a "
        "prediction-only branch for each requested view."
    )
    """
    Historical v10 implementation (intentionally inert):
    reader = RealSceneMultiViewReader(
        dataset_root=args.dataset_root,
        view=view,
        start_frame=args.start_frame,
        stride=args.running_stride,
        args=args,
    )
    setup = reader.get_reference_object_pose_setup()
    if not setup:
        raise ValueError(
            "Pre-tracking proxy calibration requires object_pose_setup with "
            "model unit, bbox, axis and origin metadata."
        )
    model_context = load_mesh_from_object_pose_setup(
        args=args,
        setup=setup,
        seq_dir=reader.seq_dir,
    )
    mesh = model_context["mesh"]
    diameter_m, diameter_source = resolve_model_diameter_m(
        args,
        fallback_diameter_m=float(model_context["computed_diameter_m"]),
    )
    base_T_object_mesh = np.asarray(
        model_context["T_object_mesh"], dtype=np.float64
    ).copy()
    validate_real_scene_units(
        reader=reader,
        mesh_diameter_m=diameter_m,
        T_object_mesh=base_T_object_mesh,
        resolved_mesh_scale=float(model_context["mesh_scale"]),
        args=args,
    )

    frame_id = int(reader.frame_ids[0])
    rgb = reader.get_rgb(frame_id)
    depth = reader.get_depth(frame_id)
    try:
        mask = reader.get_mask(
            frame_id, fallback_to_first=False
        )
    except Exception as exc:
        raise RuntimeError(
            "Proxy calibration deliberately uses only the saved first-frame "
            f"mask, but mask/{frame_id:04d} could not be loaded for {view}: "
            f"{exc}. Supply one correct mask instead of enabling repeated SAM."
        ) from exc
    mask = keep_largest_connected_component(mask).astype(np.bool_)
    if int(mask.sum()) < int(args.min_mask_pixels):
        raise RuntimeError(
            f"Saved calibration mask for {view} frame {frame_id:04d} has "
            f"only {int(mask.sum())} pixels; at least "
            f"{int(args.min_mask_pixels)} are required."
        )

    calibration_root = os.path.join(
        save_root, f"{args.object_name}_proxy_calibration"
    )
    os.makedirs(calibration_root, exist_ok=True)
    cv2.imwrite(
        os.path.join(
            calibration_root,
            f"{view}_{frame_id:04d}_saved_mask_used.png",
        ),
        mask.astype(np.uint8) * 255,
    )
    estimator = init_multiview(
        mesh=mesh,
        save_root=calibration_root,
        debug=args.debug,
    )
    pred_pose_cam_mesh = ensure_pose_matrix(
        estimator.register(
            K=reader.K,
            rgb=rgb,
            depth=depth,
            ob_mask=mask,
            iteration=int(args.proxy_calibration_iteration),
            name=f"{reader.seq_name}_proxy_calibration",
        ),
        "pre-tracking proxy calibration prediction",
    )
    T_cam_base, T_cam_base_source = reader.get_T_cam_base(frame_id)
    T_base_cam = np.linalg.inv(T_cam_base)
    pred_pose_base_mesh = T_base_cam @ pred_pose_cam_mesh
    gt_pose_base_object = reader.get_gt_pose_base(frame_id)
    calibration_vertices = sample_vertices(
        np.asarray(mesh.vertices, dtype=np.float64),
        int(args.proxy_calibration_max_points),
        seed=71,
    )
    (
        selected_T_object_mesh,
        diagnostics,
    ) = select_proxy_model_axis_mapping(
        pred_pose_base_mesh=pred_pose_base_mesh,
        gt_pose_base_object=gt_pose_base_object,
        base_T_object_mesh=base_T_object_mesh,
        vertices_mesh_m=calibration_vertices,
        diameter_m=diameter_m,
        add_weight=float(args.proxy_calibration_add_weight),
        adds_weight=float(args.proxy_calibration_adds_weight),
        min_improvement_mm=float(
            args.proxy_calibration_min_improvement_mm
        ),
    )
    selected_ratio = float(
        diagnostics["selected_ADD_over_diameter"]
    )
    reliable = bool(
        selected_ratio
        <= float(args.proxy_calibration_max_add_ratio)
    )
    diagnostics.update({
        "build_id": SCRIPT_BUILD_ID,
        "object_name": args.object_name,
        "calibration_view": view,
        "calibration_frame_id": frame_id,
        "calibration_frame_must_be_excluded_from_unbiased_aggregate": True,
        "saved_mask_pixels": int(mask.sum()),
        "T_cam_base_source": T_cam_base_source,
        "T_cam_base": T_cam_base.tolist(),
        "pred_pose_cam_mesh": pred_pose_cam_mesh.tolist(),
        "pred_pose_base_mesh": pred_pose_base_mesh.tolist(),
        "gt_pose_base_object": gt_pose_base_object.tolist(),
        "mesh_scale_resolved": float(model_context["mesh_scale"]),
        "mesh_scale_source": model_context["mesh_scale_source"],
        "bbox_diagonal_ratio_model_to_metadata": float(
            model_context["bbox_diagonal_ratio_model_to_metadata"]
        ),
        "scale_candidates_audit": model_context.get(
            "scale_candidates_audit", []
        ),
        "model_frame_mode": model_context["model_frame_mode"],
        "T_object_mesh_source_before_calibration": model_context[
            "T_object_mesh_source"
        ],
        "mesh_diameter_m": float(diameter_m),
        "mesh_diameter_source": diameter_source,
        "calibration_reliable": reliable,
        "max_allowed_ADD_over_diameter": float(
            args.proxy_calibration_max_add_ratio
        ),
        "tracking_policy_after_calibration": (
            "register saved first-frame mask once, then track_one"
            if args.register_once_then_track
            else "configured by tracking arguments"
        ),
    })
    output_path = default_proxy_calibration_output_path(
        args, save_root
    )
    os.makedirs(
        os.path.dirname(os.path.abspath(output_path)),
        exist_ok=True,
    )
    np.savetxt(
        output_path, selected_T_object_mesh, fmt="%.12g"
    )
    report_path = os.path.splitext(output_path)[0] + "_report.json"
    with open(report_path, "w", encoding="utf-8") as file:
        json.dump(
            to_json_safe(diagnostics),
            file,
            ensure_ascii=False,
            indent=2,
        )
    candidate_rows = []
    for row in diagnostics["candidates"]:
        candidate_rows.append({
            key: value
            for key, value in row.items()
            if key != "T_object_mesh"
        })
    candidate_csv_path = (
        os.path.splitext(output_path)[0] + "_candidates.csv"
    )
    pd.DataFrame(candidate_rows).to_csv(
        candidate_csv_path, index=False
    )
    np.savetxt(
        os.path.join(
            calibration_root,
            f"{view}_{frame_id:04d}_pred_pose_base_mesh.txt",
        ),
        pred_pose_base_mesh,
    )
    np.savetxt(
        os.path.join(
            calibration_root,
            f"{view}_{frame_id:04d}_gt_pose_base_object.txt",
        ),
        gt_pose_base_object,
    )

    if not reliable:
        message = (
            "[Proxy calibration warning] The best axis candidate still has "
            f"ADD/d={selected_ratio:.3f}, above "
            f"{float(args.proxy_calibration_max_add_ratio):.3f}. This means "
            "the first registration is not merely a horizontal/vertical axis "
            "error; inspect the one saved mask, proxy geometry and depth."
        )
        if args.proxy_calibration_fail_unreliable:
            raise RuntimeError(message)
        print(message)
    print(
        "[Proxy calibration] "
        f"view={view}, frame={frame_id:04d}, branch="
        f"{diagnostics['selected_local_axis_branch']}, "
        f"ADD={diagnostics['selected_ADD_mm']:.3f} mm, "
        f"ADD-S={diagnostics['selected_ADD-S_mm']:.3f} mm, "
        f"improvement="
        f"{diagnostics['improvement_over_metadata_mm']:.3f} mm"
    )
    print(
        "[Proxy scale audit] resolved scale="
        f"{float(model_context['mesh_scale']):g}, source="
        f"{model_context['mesh_scale_source']}, model/metadata bbox ratio="
        f"{float(model_context['bbox_diagonal_ratio_model_to_metadata']):.6f}. "
        "Scale was not minimized with ADD/ADD-S."
    )
    print(
        "[Proxy calibration] fixed transform shared by both views: "
        f"{output_path}"
    )
    print(
        "[Proxy calibration] candidate table: "
        f"{candidate_csv_path}"
    )

    # Feed the selected matrix back through the ordinary immutable model path.
    # This keeps process_one_view() and fusion on exactly the same definition.
    args.T_object_mesh = output_path
    args.axis_map = None
    args.object_mesh_translation_scale = 1.0
    # Freeze the scale already selected by metadata/bbox auditing in this
    # pre-pass so the later right/left loaders cannot resolve a different unit.
    args.mesh_scale = float(model_context["mesh_scale"])
    setattr(args, "_proxy_model_calibration_result", diagnostics)
    setattr(args, "_proxy_calibration_frame_id", frame_id)
    setattr(args, "_proxy_calibration_view_resolved", view)
    setattr(args, "_proxy_calibration_output_resolved", output_path)
    return diagnostics
    """


def process_one_view(args, view: str, save_root: str) -> List[Dict]:
    reader = RealSceneMultiViewReader(
        dataset_root=args.dataset_root,
        view=view,
        start_frame=args.start_frame,
        stride=args.running_stride,
        args=args,
    )

    object_pose_setup = reader.get_reference_object_pose_setup()
    if object_pose_setup:
        model_context = load_mesh_from_object_pose_setup(
            args=args,
            setup=object_pose_setup,
            seq_dir=reader.seq_dir,
        )
    else:
        if args.mesh_scale is None:
            raise ValueError(
                "Legacy object setup requires an explicit --mesh_scale."
            )
        mesh_legacy, center_legacy, diameter_legacy = (
            load_bbox_centered_mesh(args.obj_mesh, args.mesh_scale)
        )
        T_object_mesh_legacy, T_object_mesh_legacy_path = (
            load_T_object_mesh(args)
        )
        legacy_vertices_object = transform_points(
            np.asarray(mesh_legacy.vertices, dtype=np.float64),
            T_object_mesh_legacy,
        )
        model_context = {
            "mesh": mesh_legacy,
            "mesh_visual_source": "legacy_checked_visual",
            "mesh_path": os.path.abspath(args.obj_mesh),
            "dataset_model_path": None,
            "source_model_path": os.path.abspath(args.obj_mesh),
            "model_frame_mode": "legacy_bbox_centered",
            "object_model_unit": "legacy_cli",
            "mesh_scale": float(args.mesh_scale),
            "mesh_scale_source": "legacy_cli",
            "mesh_center_offset_after_scale_m": center_legacy,
            "T_object_mesh": T_object_mesh_legacy,
            "T_object_mesh_source": (
                T_object_mesh_legacy_path or "legacy_identity"
            ),
            "object_model_axis_alignment": None,
            "object_model_axis_alignment_source": "legacy_cli",
            "object_model_origin_mode": "legacy_bbox_centered",
            "computed_diameter_m": diameter_legacy,
            "bbox_min_object_m": legacy_vertices_object.min(axis=0),
            "bbox_max_object_m": legacy_vertices_object.max(axis=0),
            "bbox_extent_object_m": np.ptp(
                legacy_vertices_object, axis=0
            ),
            "bbox_diagonal_ratio_model_to_metadata": 1.0,
            "model_bbox_extent_object_m": np.ptp(
                legacy_vertices_object, axis=0
            ),
            "object_pose_setup": {},
        }

    mesh = model_context["mesh"]
    mesh_center_offset = np.asarray(
        model_context["mesh_center_offset_after_scale_m"],
        dtype=np.float64,
    )
    resolved_mesh_scale = float(model_context["mesh_scale"])
    computed_diameter_m = float(model_context["computed_diameter_m"])
    diameter_m, diameter_source = resolve_model_diameter_m(
        args, fallback_diameter_m=computed_diameter_m
    )

    # T_object_mesh is immutable throughout tracking and shared by both views.
    # It is metadata-derived (or an explicit offline input) and is never
    # rewritten from a first-frame prediction. Proxy calibration below changes
    # only a zero-translation prediction branch for each camera independently.
    T_object_mesh = np.asarray(
        model_context["T_object_mesh"], dtype=np.float64
    ).copy()
    T_object_mesh_path = (
        os.path.abspath(args.T_object_mesh)
        if args.T_object_mesh is not None else None
    )
    axis_alignment_source = str(
        model_context["T_object_mesh_source"]
    )
    axis_map_resolved = describe_axis_rotation(T_object_mesh[:3, :3])

    fixed_definition = getattr(args, "_fixed_model_definition", None)
    current_definition = {
        "T_object_mesh": T_object_mesh.copy(),
        "bbox_min_object_m": np.asarray(
            model_context["bbox_min_object_m"], dtype=np.float64
        ).copy(),
        "bbox_max_object_m": np.asarray(
            model_context["bbox_max_object_m"], dtype=np.float64
        ).copy(),
        "mesh_center_offset_after_scale_m": mesh_center_offset.copy(),
        "mesh_scale": resolved_mesh_scale,
        "model_frame_mode": str(model_context["model_frame_mode"]),
    }
    if fixed_definition is None:
        setattr(args, "_fixed_model_definition", current_definition)
    else:
        for matrix_key in (
            "T_object_mesh",
            "bbox_min_object_m",
            "bbox_max_object_m",
            "mesh_center_offset_after_scale_m",
        ):
            if not np.allclose(
                np.asarray(fixed_definition[matrix_key]),
                np.asarray(current_definition[matrix_key]),
                atol=1e-9,
                rtol=1e-7,
            ):
                raise ValueError(
                    f"Right/left metadata disagree on fixed {matrix_key}. "
                    "Fusion requires both predictions in one GT object frame."
                )
        if not np.isclose(
            float(fixed_definition["mesh_scale"]),
            resolved_mesh_scale,
            atol=1e-15,
            rtol=1e-12,
        ):
            raise ValueError(
                "Right/left metadata resolved different model scales."
            )
        if (
            str(fixed_definition["model_frame_mode"])
            != str(current_definition["model_frame_mode"])
        ):
            raise ValueError(
                "Right/left metadata resolved different model-frame modes: "
                f"{fixed_definition['model_frame_mode']!r} vs "
                f"{current_definition['model_frame_mode']!r}."
            )
    view_branch_corrections = getattr(
        args, "_view_branch_corrections", None
    )
    if view_branch_corrections is None:
        view_branch_corrections = {}
        setattr(args, "_view_branch_corrections", view_branch_corrections)
    view_branch_diagnostics = getattr(
        args, "_view_branch_diagnostics", None
    )
    if view_branch_diagnostics is None:
        view_branch_diagnostics = {}
        setattr(args, "_view_branch_diagnostics", view_branch_diagnostics)
    view_branch_paths = getattr(args, "_view_branch_paths", None)
    if view_branch_paths is None:
        view_branch_paths = {}
        setattr(args, "_view_branch_paths", view_branch_paths)

    explicit_branch_path = explicit_view_branch_path(args, view)
    view_branch_path: Optional[str] = None
    if explicit_branch_path is not None:
        view_branch_correction = validate_pose_branch_correction(
            load_transform_txt(
                explicit_branch_path,
                f"{view} pose-branch correction",
                translation_scale=1.0,
            ),
            f"{view} pose-branch correction",
        )
        view_branch_path = explicit_branch_path
        view_branch_paths[view] = view_branch_path
        view_branch_corrections[view] = view_branch_correction.copy()
        loaded_branch_angle_deg = float(np.degrees(np.linalg.norm(
            SciPyRotation.from_matrix(
                view_branch_correction[:3, :3]
            ).as_rotvec()
        )))
        view_branch_diagnostics.setdefault(view, {
            "source": f"loaded:{explicit_branch_path}",
            "uses_gt_on_marked_calibration_frame": False,
            "branch_map": describe_axis_rotation(
                view_branch_correction[:3, :3]
            ),
            "branch_rotation_deg": loaded_branch_angle_deg,
            "T_mesh_branch": view_branch_correction.tolist(),
        })
    elif view in view_branch_corrections:
        view_branch_correction = validate_pose_branch_correction(
            view_branch_corrections[view],
            f"cached {view} pose-branch correction",
        )
        view_branch_path = default_view_branch_output_path(
            args, save_root, view
        )
    else:
        view_branch_correction = np.eye(4, dtype=np.float64)

    view_branch_calibration_pending = bool(
        (
            args.calibrate_axis_from_first_frame
            or args.calibrate_proxy_model_from_first_frame
        )
        and explicit_branch_path is None
        and view not in view_branch_diagnostics
    )
    view_branch_calibration_frame_id: Optional[int] = None

    validate_real_scene_units(
        reader=reader,
        mesh_diameter_m=diameter_m,
        T_object_mesh=T_object_mesh,
        resolved_mesh_scale=resolved_mesh_scale,
        args=args,
    )

    vertices = np.asarray(mesh.vertices, dtype=np.float32)
    bbox_corners = make_bbox_corners(vertices)
    object_bbox_corners = make_bbox_corners_from_bounds(
        model_context["bbox_min_object_m"],
        model_context["bbox_max_object_m"],
    )
    add_vertices = sample_vertices(vertices, args.max_add_points)
    bop_vertices = sample_vertices(vertices, args.max_bop_points, seed=17)
    axis_length = max(diameter_m * args.axis_ratio, args.min_axis_length_m)
    T_mesh_object = np.linalg.inv(T_object_mesh)
    symmetry_axis_index, symmetry_axis_name, mesh_extents_m = resolve_symmetry_axis(
        vertices, args.symmetry_axis
    )
    metric_symmetries, metric_symmetry_source = build_metric_symmetries(
        args=args,
        vertices_m=vertices,
        diameter_m=diameter_m,
        resolved_axis_index=symmetry_axis_index,
    )
    (
        prediction_branch_symmetries,
        prediction_branch_symmetry_diagnostics,
    ) = detect_prediction_branch_symmetries(
        sample_vertices(vertices, 5000, seed=37),
        diameter_m=diameter_m,
        args=args,
    )
    tracking_profile = str(args.tracking_profile)
    register_once_then_track = bool(args.register_once_then_track)
    effective_view_branch_calibration_mode = str(
        args.view_branch_calibration_mode
    )
    if args.calibrate_proxy_model_from_first_frame:
        # Keep this deliberately simple: directly measure the constant
        # first-frame SO(3) offset in each view. This aligns the predicted
        # object axes with GT exactly on the marked calibration frame without
        # asking inaccurate proxy geometry or ADD/ADD-S to choose an axis.
        # Only predictions are corrected; T_object_mesh and GT stay fixed.
        effective_view_branch_calibration_mode = "full"
    elif (
        tracking_profile == "slender_tool"
        and effective_view_branch_calibration_mode == "full"
    ):
        effective_view_branch_calibration_mode = "geometry_safe"
    tracking_health_enabled = bool(
        not register_once_then_track
        and (
            args.enable_tracking_health_check
            or tracking_profile == "slender_tool"
        )
    )
    effective_reinit_interval = int(args.reinit_interval)
    effective_iteration = int(args.iteration)
    if args.calibrate_proxy_model_from_first_frame:
        effective_iteration = max(
            effective_iteration,
            int(args.proxy_calibration_iteration),
        )
    effective_failure_policy = str(args.failure_policy)
    if register_once_then_track:
        effective_reinit_interval = 0
        effective_failure_policy = "freeze"
    elif tracking_profile == "slender_tool":
        if effective_reinit_interval == 0:
            effective_reinit_interval = int(
                args.slender_tool_reinit_interval
            )
        effective_iteration = max(
            effective_iteration, int(args.slender_tool_iteration)
        )
    tracking_geometry_args = copy.copy(args)
    tracking_geometry_args.fusion_min_projected_points = int(
        args.tracking_health_min_projected_points
    )
    tracking_geometry_args.fusion_min_compared_points = int(
        args.tracking_health_min_compared_points
    )
    if tracking_profile == "slender_tool":
        # For a thin metallic/transparent shaft, silhouette agreement is more
        # reliable than requiring dense depth everywhere.
        tracking_geometry_args.fusion_depth_weight = 0.20
        tracking_geometry_args.fusion_inlier_weight = 0.15
        tracking_geometry_args.fusion_mask_weight = 0.45
        tracking_geometry_args.fusion_point_weight = 0.20
    tracking_surface_points = (
        sample_mesh_surface_points(
            mesh,
            max_points=args.tracking_health_max_surface_points,
            seed=43,
        )
        if tracking_health_enabled
        else np.empty((0, 3), dtype=np.float64)
    )
    positive_extents = np.maximum(
        np.asarray(mesh_extents_m, dtype=np.float64), 1e-9
    )
    mesh_aspect_ratio = float(
        np.max(positive_extents) / np.min(positive_extents)
    )

    view_root = os.path.join(save_root, reader.seq_name)
    pred_cam_pose_dir = os.path.join(view_root, "poses_pred_camera_stable_mesh")
    pred_cam_object_pose_dir = os.path.join(
        view_root, "poses_pred_camera_stable_object"
    )
    pred_base_pose_dir = os.path.join(view_root, "poses_pred_base_stable_mesh")
    pred_base_object_pose_dir = os.path.join(view_root, "poses_pred_base_object")
    pred_cam_raw_pose_dir = os.path.join(view_root, "poses_pred_camera_raw_mesh")
    pred_cam_raw_object_pose_dir = os.path.join(
        view_root, "poses_pred_camera_raw_object"
    )
    pred_base_raw_pose_dir = os.path.join(view_root, "poses_pred_base_raw_mesh")
    pred_base_raw_object_pose_dir = os.path.join(view_root, "poses_pred_base_raw_object")
    gt_cam_pose_dir = os.path.join(view_root, "poses_gt_camera_mesh")
    gt_cam_object_pose_dir = os.path.join(view_root, "poses_gt_camera_object")
    gt_base_pose_dir = os.path.join(view_root, "poses_gt_base_mesh")
    gt_base_object_input_dir = os.path.join(view_root, "poses_gt_base_object_input")
    vis_dir = os.path.join(view_root, "vis")
    for d in (
        pred_cam_pose_dir, pred_cam_object_pose_dir,
        pred_base_pose_dir, pred_base_object_pose_dir,
        pred_cam_raw_pose_dir, pred_cam_raw_object_pose_dir,
        pred_base_raw_pose_dir, pred_base_raw_object_pose_dir,
        gt_cam_pose_dir, gt_cam_object_pose_dir,
        gt_base_pose_dir, gt_base_object_input_dir,
    ):
        os.makedirs(d, exist_ok=True)
    if args.save_images:
        os.makedirs(vis_dir, exist_ok=True)

    first_id = reader.frame_ids[0]
    first_T_cam_base, first_T_cam_base_source = (
        reader.get_T_cam_base(first_id)
    )
    first_T_base_cam = np.linalg.inv(first_T_cam_base)

    # Compatibility: keep accepting --axis_calibration_output, but save the
    # immutable metadata-derived model mapping. A prediction is never allowed
    # to rewrite this matrix. Per-view learned rotations go to separate
    # *_right_branch.txt and *_left_branch.txt files.
    if (
        args.calibrate_axis_from_first_frame
        and args.axis_calibration_output
        and not getattr(args, "_fixed_model_transform_saved", False)
    ):
        fixed_output_path = os.path.abspath(args.axis_calibration_output)
        os.makedirs(
            os.path.dirname(fixed_output_path), exist_ok=True
        )
        np.savetxt(
            fixed_output_path, T_object_mesh, fmt="%.12g"
        )
        fixed_report = {
            "build_id": SCRIPT_BUILD_ID,
            "source": axis_alignment_source,
            "model_frame_mode": model_context["model_frame_mode"],
            "object_model_axis_alignment": model_context[
                "object_model_axis_alignment"
            ],
            "object_model_axis_alignment_source": model_context[
                "object_model_axis_alignment_source"
            ],
            "object_model_origin_mode": model_context[
                "object_model_origin_mode"
            ],
            "mesh_scale": resolved_mesh_scale,
            "prediction_was_used": False,
            "T_object_mesh": T_object_mesh.tolist(),
            "definition": (
                "p_GT_object = T_object_mesh @ p_estimator_mesh"
            ),
            "note": (
                "This is the immutable metadata-derived model mapping. "
                "First-frame GT may calibrate only separate per-view "
                "prediction branch files."
            ),
        }
        with open(
            os.path.splitext(fixed_output_path)[0]
            + "_fixed_metadata_report.json",
            "w",
            encoding="utf-8",
        ) as file:
            json.dump(
                to_json_safe(fixed_report),
                file,
                ensure_ascii=False,
                indent=2,
            )
        setattr(args, "_fixed_model_transform_saved", True)
        print(
            "[Fixed model transform] saved metadata-derived "
            f"T_object_mesh to {fixed_output_path}; prediction was not used."
        )
    if view_branch_calibration_pending:
        view_branch_calibration_frame_id = int(first_id)
        calibration_frames = getattr(
            args, "_view_branch_calibration_frames", None
        )
        if calibration_frames is None:
            calibration_frames = {}
            setattr(
                args,
                "_view_branch_calibration_frames",
                calibration_frames,
            )
        calibration_frames[view] = view_branch_calibration_frame_id
    first_rgb = reader.get_rgb(first_id)
    bop_evaluator = None
    if not args.disable_bop_metrics:
        bop_evaluator = BOPMetricEvaluator(
            mesh=mesh,
            image_shape=first_rgb.shape,
            diameter_m=diameter_m,
            symmetries=metric_symmetries,
            args=args,
        )
    vsd_taus_for_columns = (
        bop_evaluator.vsd_taus
        if bop_evaluator is not None
        else parse_float_list(args.vsd_taus, "--vsd_taus")
    )

    run_config_path = os.path.join(view_root, "run_config.json")
    run_config = {
            **argparse_config_dict(args),
            "view": view,
            "seq_dir": reader.seq_dir,
            "K": reader.K.tolist(),
            "mesh_bbox_center_offset_after_scale_m": mesh_center_offset.tolist(),
            "mesh_scale_resolved": resolved_mesh_scale,
            "mesh_scale_source": model_context["mesh_scale_source"],
            "mesh_visual_source": model_context[
                "mesh_visual_source"
            ],
            "model_frame_mode": model_context["model_frame_mode"],
            "object_model_axis_alignment": model_context[
                "object_model_axis_alignment"
            ],
            "object_model_axis_alignment_source": model_context[
                "object_model_axis_alignment_source"
            ],
            "object_model_origin_mode": model_context[
                "object_model_origin_mode"
            ],
            "dataset_model_path": model_context["dataset_model_path"],
            "source_model_path": model_context["source_model_path"],
            "mesh_diameter_m": diameter_m,
            "mesh_diameter_source": diameter_source,
            "computed_mesh_diameter_m": computed_diameter_m,
            "metadata_object_bbox_min_m": np.asarray(
                model_context["bbox_min_object_m"]
            ).tolist(),
            "metadata_object_bbox_max_m": np.asarray(
                model_context["bbox_max_object_m"]
            ).tolist(),
            "model_bbox_extent_in_object_frame_m": np.asarray(
                model_context["model_bbox_extent_object_m"]
            ).tolist(),
            "bbox_diagonal_ratio_model_to_metadata": float(
                model_context[
                    "bbox_diagonal_ratio_model_to_metadata"
                ]
            ),
            "object_pose_setup": object_pose_setup,
            "first_gt_translation_m_direct": reader.get_raw_gt_translation(first_id).tolist(),
            "gt_translation_span_m": reader.gt_translation_span_m.tolist(),
            "first_gt_base_object_translation_m": reader.get_gt_pose_base(first_id)[:3, 3].tolist(),
            "camera_extrinsics": reader.camera_extrinsic_summary(),
            "first_T_cam_base_source": first_T_cam_base_source,
            "first_T_cam_base": first_T_cam_base.tolist(),
            "first_T_base_cam": first_T_base_cam.tolist(),
            "T_object_mesh_path": T_object_mesh_path,
            "T_object_mesh": T_object_mesh.tolist(),
            "axis_alignment_source": axis_alignment_source,
            "axis_map_mesh_to_object": axis_map_resolved,
            "global_model_calibration_from_prediction": False,
            "proxy_prediction_branch_calibration": bool(
                args.calibrate_proxy_model_from_first_frame
            ),
            "proxy_calibration_processing_order_view": (
                args.proxy_calibration_view
            ),
            "proxy_calibration_output_stem": (
                args.proxy_calibration_output
            ),
            "view_branch_calibration_pending": (
                view_branch_calibration_pending
            ),
            "view_branch_correction_path": view_branch_path,
            "T_pose_branch": view_branch_correction.tolist(),
            "fixed_mesh_frame": str(model_context["model_frame_mode"]),
            "fixed_gt_frame": "T_base_object_O",
            "converted_gt_object_frame": (
                "T_cam_object_gt = T_cam_base_meta(frame) @ "
                "T_base_object_gt(frame)"
            ),
            "converted_gt_mesh_frame": (
                "T_cam_mesh_gt = T_cam_object_gt @ immutable_T_object_mesh"
            ),
            "prediction_frame_raw": "direct_MultiView_T_cam_mesh_raw",
            "prediction_frame_stable": (
                "direct MultiView T_cam_mesh right-multiplied by one frozen "
                "per-view prediction-only proper-SO(3) branch correction; "
                "the fixed GT/model transform is never changed"
            ),
            "metric_prediction_frame": (
                "ADD/ADD-S/VSD/MSSD/MSPD, translation, rotation, red mesh, "
                "saved canonical pose and fusion all use the same "
                "branch-canonical mesh pose"
            ),
            "symmetry_axis_resolved": symmetry_axis_name,
            "mesh_extents_m": mesh_extents_m.tolist(),
            "metric_symmetry_source": metric_symmetry_source,
            "metric_symmetry_count": len(metric_symmetries),
            "prediction_branch_symmetry_detection": (
                prediction_branch_symmetry_diagnostics
            ),
            "tracking_profile_resolved": tracking_profile,
            "register_once_then_track_resolved": (
                register_once_then_track
            ),
            "view_branch_calibration_mode_resolved": (
                effective_view_branch_calibration_mode
            ),
            "tracking_health_enabled": tracking_health_enabled,
            "tracking_effective_reinit_interval": (
                effective_reinit_interval
            ),
            "tracking_effective_iteration": effective_iteration,
            "tracking_effective_failure_policy": (
                effective_failure_policy
            ),
            "tracking_mesh_aspect_ratio": mesh_aspect_ratio,
            "tracking_health_surface_point_count": int(
                len(tracking_surface_points)
            ),
            "bop_metric_vertex_count": int(len(bop_vertices)),
        }
    with open(run_config_path, "w", encoding="utf-8") as f:
        json.dump(to_json_safe(run_config), f, ensure_ascii=False, indent=2)

    est = init_multiview(mesh=mesh, save_root=view_root, debug=args.debug)

    need_sam2 = args.sam2_enable and (
        args.use_manual_first_bbox
        or (
            not register_once_then_track
            and (
                effective_reinit_interval > 0
                or tracking_health_enabled
                or effective_failure_policy == "reregister"
            )
        )
    )
    sam2_segmenter = None
    if need_sam2:
        sam2_segmenter = SAM2BoxSegmenter(
            checkpoint=args.sam2_checkpoint,
            model_cfg=args.sam2_model_cfg,
            device=args.sam2_device,
        )

    tracking_initialized = False
    # Canonical history is used for masks/motion prediction. Raw history is
    # retained separately so a tracker failure cannot apply the fixed view
    # branch twice when the previous direct MultiView pose is reused.
    pred_pose_q_prev: Optional[np.ndarray] = None
    pred_pose_q_prev2: Optional[np.ndarray] = None
    pred_pose_q_prev_raw: Optional[np.ndarray] = None
    pred_pose_q_prev_raw2: Optional[np.ndarray] = None
    temporal_mesh_pose_prev: Optional[np.ndarray] = None
    temporal_mesh_pose_prev2: Optional[np.ndarray] = None
    tracking_recovery_age = 0
    force_reregister_next_frame = False
    # In register-once mode, a calibrated/loaded prediction branch is applied
    # to the first registered pose and then written into MultiView pose_last. Later
    # track_one() outputs are already on that corrected branch, so the fixed
    # rotation must not be multiplied a second time.
    tracker_state_contains_view_branch = False
    tracker_branch_state_sync = "not_requested"
    axis_pose_prev: Optional[np.ndarray] = None
    axis_pose_prev2: Optional[np.ndarray] = None
    metric_prev_pred_raw: Optional[np.ndarray] = None
    metric_prev_gt: Optional[np.ndarray] = None
    vertices_for_prompt = sample_vertices(vertices, args.sam_prompt_vertices, seed=13)
    records: List[Dict] = []
    tracking_health_rows: List[Dict[str, object]] = []

    def append_tracking_health_row(
        frame_id: int,
        pose_source_value: str,
        diagnostics: Dict[str, object],
    ) -> None:
        if not tracking_health_enabled:
            return
        initial = diagnostics.get("initial") or {}
        retry = diagnostics.get("reregister") or {}
        selected = diagnostics.get("selected") or {}
        initial_geometry = initial.get("geometry") or {}
        retry_geometry = retry.get("geometry") or {}
        selected_geometry = selected.get("geometry") or {}
        tracking_health_rows.append({
            "frame_idx": int(frame_id),
            "view": view,
            "status": diagnostics.get("status", ""),
            "pose_source": str(pose_source_value),
            "initial_pass": bool(initial.get("passes", False)),
            "initial_reason": initial.get("reason", ""),
            "initial_geometry_cost": initial_geometry.get(
                "score", float("nan")
            ),
            "initial_translation_jump_mm": initial.get(
                "translation_jump_mm", float("nan")
            ),
            "initial_rotation_jump_deg": initial.get(
                "rotation_jump_deg", float("nan")
            ),
            "reregister_attempted": bool(retry),
            "reregister_pass": bool(retry.get("passes", False)),
            "reregister_reason": retry.get("reason", ""),
            "reregister_geometry_cost": retry_geometry.get(
                "score", float("nan")
            ),
            "selected_pass": bool(selected.get("passes", False)),
            "selected_reason": selected.get("reason", ""),
            "selected_mask_only_fallback": bool(
                selected.get("mask_only_fallback", False)
            ),
            "selected_axis_cue_available": bool(
                selected.get("axis_cue_available", False)
            ),
            "selected_axis_difference_deg": selected.get(
                "silhouette_axis_difference_deg", float("nan")
            ),
            "selected_geometry_cost": selected_geometry.get(
                "score", float("nan")
            ),
            "selected_mask_source": selected_geometry.get(
                "mask_source", ""
            ),
            "selected_projected_points": selected_geometry.get(
                "projected_points", 0
            ),
            "selected_compared_depth_points": selected_geometry.get(
                "compared_depth_points", 0
            ),
            "selected_depth_inlier_ratio": selected_geometry.get(
                "depth_inlier_ratio", float("nan")
            ),
            "selected_mask_inside_ratio": selected_geometry.get(
                "projected_inside_mask_ratio", float("nan")
            ),
            "selected_translation_jump_mm": selected.get(
                "translation_jump_mm", float("nan")
            ),
            "selected_rotation_jump_deg": selected.get(
                "rotation_jump_deg", float("nan")
            ),
            "estimator_state_sync": diagnostics.get(
                "estimator_state_sync", "not_requested"
            ),
        })

    writer = None
    video_path = os.path.join(view_root, f"{reader.seq_name}_multiview_real_pose_track.mp4")

    print(f"\n[View] {reader.seq_name}")
    print(f"[Data] {reader.seq_dir}")
    print(f"[Mesh] {args.obj_mesh}")
    print(
        f"[Mesh] scale={resolved_mesh_scale:g} "
        f"({model_context['mesh_scale_source']}), "
        f"fixed_frame={model_context['model_frame_mode']}, "
        f"diameter={diameter_m:.6f} m ({diameter_source})"
    )
    print(
        f"[Mesh visual] {model_context['mesh_visual_source']} | "
        f"aspect_ratio={mesh_aspect_ratio:.3f}"
    )
    print(
        "[Mesh scale audit] model/metadata bbox diagonal ratio="
        f"{float(model_context['bbox_diagonal_ratio_model_to_metadata']):.6f} "
        "(approximately 1 means --mesh_scale is not the tracking cause)."
    )
    print(f"[Mesh] removed source bbox center after scaling: {mesh_center_offset.tolist()} m")
    first_T_base_object = reader.get_gt_pose_base(first_id)
    first_T_base_mesh = first_T_base_object @ T_object_mesh
    first_T_cam_object = first_T_cam_base @ first_T_base_object
    first_T_cam_mesh = first_T_cam_object @ T_object_mesh
    print(f"[Camera extrinsic metadata] {first_T_cam_base_source}")
    print(f"[Camera mapping] {view} -> {'cam1' if view == 'right' else 'cam2'}")
    print(
        f"[T_cam_base first frame] "
        f"t={first_T_cam_base[:3,3].tolist()} m, "
        f"det(R)={np.linalg.det(first_T_cam_base[:3,:3]):.6f}"
    )
    print("[GT] stored as per-frame T_base_object; translation is read verbatim in metres")
    print(
        "[GT conversion] T_cam_object_gt(frame) = "
        "T_cam_base_meta(frame) @ T_base_object_gt(frame)"
    )
    print(f"[GT first file] {reader.get_gt_pose_path(first_id)}")
    print(f"[GT first base object] t={first_T_base_object[:3,3].tolist()} m")
    print(f"[GT trajectory translation span] {reader.gt_translation_span_m.tolist()} m")
    print(
        f"[Object/mesh] source={axis_alignment_source} | "
        f"mesh XYZ in object frame={axis_map_resolved}"
    )
    if args.calibrate_proxy_model_from_first_frame:
        print(
            "[Proxy scale audit] resolved scale="
            f"{resolved_mesh_scale:g}, source="
            f"{model_context['mesh_scale_source']}, "
            "model/metadata bbox diagonal ratio="
            f"{float(model_context['bbox_diagonal_ratio_model_to_metadata']):.6f}. "
            "Scale comes only from object_pose_setup/bbox agreement; "
            "ADD/ADD-S are not allowed to optimize scale."
        )
    if view_branch_calibration_pending:
        print(
            f"[View branch calibration] view={view}, frame={first_id:04d}; "
            f"mode={effective_view_branch_calibration_mode}; estimate one local "
            "prediction-only zero-translation rotation correction directly "
            "from the first-frame pred/GT orientation. GT/model coordinates "
            "remain unchanged."
        )
    else:
        branch_angle_deg = float(np.degrees(np.linalg.norm(
            SciPyRotation.from_matrix(
                view_branch_correction[:3, :3]
            ).as_rotvec()
        )))
        print(
            f"[View branch] view={view}, source="
            f"{view_branch_path or 'identity'}, "
            f"rotation={branch_angle_deg:.3f} degree"
        )
    print(f"[GT first camera mesh] t={first_T_cam_mesh[:3,3].tolist()} m")
    print(f"[Depth] npy_scale={args.depth_npy_scale}, png_scale={args.depth_png_scale}")
    print(
        f"[Frames] {len(reader)} | tracking_profile={tracking_profile} | "
        f"health_check={tracking_health_enabled} | "
        f"reinit_interval={effective_reinit_interval} | "
        f"iteration={effective_iteration} | "
        f"failure_policy={effective_failure_policy}"
    )
    if register_once_then_track:
        print(
            "[Tracking mode] register_once_then_track: use one saved first-"
            "frame mask, then track_one only; repeated SAM generation and "
            "tracking-health rejection are disabled."
        )
    if (
        effective_view_branch_calibration_mode
        != args.view_branch_calibration_mode
    ):
        print(
            "[Tracking profile] Replaced arbitrary full-SO(3) first-frame "
            "branch calibration with geometry_safe for this slender tool."
        )
    if (
        not register_once_then_track
        and
        tracking_profile == "slender_tool"
        and str(model_context["mesh_visual_source"]).startswith(
            "fallback_vertex_color:"
        )
    ):
        print(
            "[Warning] Slender-tool proxy has no usable texture. Scale may be "
            "correct while rotation remains weakly observable; periodic "
            "mask/depth recovery is enabled, but a real texture or accurate "
            "vertex color model will still improve tracking."
        )
    print(
        f"[Symmetry] enabled={args.symmetry_stabilization} | requested_axis={args.symmetry_axis} "
        f"| resolved_axis={symmetry_axis_name} | mesh_extents={mesh_extents_m.tolist()} m "
        f"| samples={args.symmetry_samples} | axis_flip={args.symmetry_allow_axis_flip} "
        f"| initial_reference={args.symmetry_initial_reference}"
    )
    print(
        f"[Temporal object axes] enabled={args.temporal_axis_stabilization} | "
        "immutable T_object_mesh + fixed per-view prediction branch + "
        f"{len(prediction_branch_symmetries)} auto-verified discrete mesh "
        "symmetry branch(es); GT is never changed"
    )
    print(
        f"[BOP metrics] enabled={not args.disable_bop_metrics} | "
        f"symmetry={metric_symmetry_source} | transforms={len(metric_symmetries)} | "
        f"model_points={len(bop_vertices)}"
    )

    def append_failed_frame(
        frame_id: int,
        local_idx: int,
        gt_pose_file: str,
        gt_pose_cam_mesh: np.ndarray,
        gt_pose_base_object: np.ndarray,
        reason: str,
    ) -> None:
        """Keep one Excel row for every requested frame, including failures."""
        ar_failure = float("nan") if args.disable_bop_metrics else 0.0
        is_global_axis_calibration_frame = bool(
            getattr(args, "_proxy_calibration_frame_id", None)
            == frame_id
        )
        is_view_branch_calibration_frame = bool(
            view_branch_calibration_frame_id == frame_id
        )
        is_axis_calibration_frame = bool(
            is_global_axis_calibration_frame
            or is_view_branch_calibration_frame
        )
        row = {
            "object_name": args.object_name,
            "object_id": "" if args.object_id is None else str(args.object_id),
            "camera_view": reader.view,
            "sequence": reader.seq_name,
            "frame_idx": int(frame_id),
            "local_idx": int(local_idx),
            "estimate_valid": False,
            "pose_source": reason,
            "metric_status": "not_evaluated_estimation_failure",
            "axis_alignment_source": axis_alignment_source,
            "axis_map_mesh_to_object": axis_map_resolved,
            "axis_calibration_frame": is_axis_calibration_frame,
            "global_axis_calibration_frame": (
                is_global_axis_calibration_frame
            ),
            "view_branch_calibration_frame": (
                is_view_branch_calibration_frame
            ),
            "axis_calibration_uses_gt": bool(
                args.calibrate_axis_from_first_frame
                or args.calibrate_proxy_model_from_first_frame
            ),
            "official_evaluation_frame": not is_axis_calibration_frame,
            "gt_pose_file": gt_pose_file,
            "mask_pixels": 0,
            "gt_projection_state": pose_projection_state(
                gt_pose_cam_mesh @ T_mesh_object,
                object_bbox_corners,
                reader.K,
                first_rgb.shape,
            ),
            "pred_projection_state": "invalid",
            "ADD_mm": float("nan"),
            "ADDS_mm": float("nan"),
            "ADD_0.1d_success": False,
            "ADDS_0.1d_success": False,
            "Rotation_error_deg": float("nan"),
            "Symmetry_aware_rotation_error_deg": float("nan"),
            "Translation_error_mm": float("nan"),
            "Translation_error_base_mm": float("nan"),
            "Translation_error_mesh_origin_mm": float("nan"),
            "Mean_projection_error_px": float("nan"),
            "5deg_5cm_success": False,
            "RPE_translation_mm": float("nan"),
            "RPE_rotation_deg": float("nan"),
            "VSD_error_mean": float("nan"),
            "VSD_AR_contribution": ar_failure,
            "MSSD_mm": float("nan"),
            "MSSD_norm_d": float("nan"),
            "MSSD_AR_contribution": ar_failure,
            "MSPD_px": float("nan"),
            "MSPD_AR_contribution": ar_failure,
            "BOP_AR_contribution": ar_failure,
            "Right_view_contribution": float("nan"),
            "Left_view_contribution": float("nan"),
            **{f"VSD_tau_{tau:.2f}": float("nan") for tau in vsd_taus_for_columns},
            "gt_cam_tx_m": float(gt_pose_cam_mesh[0, 3]),
            "gt_cam_ty_m": float(gt_pose_cam_mesh[1, 3]),
            "gt_cam_tz_m": float(gt_pose_cam_mesh[2, 3]),
            "gt_base_object_tx_m": float(gt_pose_base_object[0, 3]),
            "gt_base_object_ty_m": float(gt_pose_base_object[1, 3]),
            "gt_base_object_tz_m": float(gt_pose_base_object[2, 3]),
            "metric_symmetry_source": metric_symmetry_source,
            "metric_symmetry_count": int(len(metric_symmetries)),
            "mesh_scale": resolved_mesh_scale,
            "diameter_m": float(diameter_m),
            "diameter_source": diameter_source,
        }
        records.append(row)

    stop_requested = False
    for local_idx in tqdm(range(len(reader)), desc=f"Tracking {reader.seq_name}"):
        frame_id = reader.get_frame_id(local_idx)
        color_rgb = reader.get_rgb(frame_id)
        depth = reader.get_depth(frame_id)
        if depth.shape != color_rgb.shape[:2]:
            raise ValueError(
                f"Frame {frame_id:04d}: depth shape {depth.shape} does not match "
                f"RGB shape {color_rgb.shape[:2]}. VSD requires depth aligned to RGB/K."
            )
        # Per-frame metadata is authoritative for the camera extrinsic.
        # The GT object pose is never calibrated from a prediction:
        #   T_cam_object_gt = T_cam_base_meta @ T_base_object_gt.
        T_cam_base, T_cam_base_source = reader.get_T_cam_base(frame_id)
        T_base_cam = np.linalg.inv(T_cam_base)
        gt_pose_file = reader.get_gt_pose_path(frame_id)
        gt_pose_base_object = reader.get_gt_pose_base(frame_id)
        gt_pose_cam_object = T_cam_base @ gt_pose_base_object
        gt_pose_base_mesh = gt_pose_base_object @ T_object_mesh
        gt_pose_cam_mesh = gt_pose_cam_object @ T_object_mesh

        forced_reregister_this_frame = bool(
            force_reregister_next_frame
        )
        force_reregister_next_frame = False
        should_register = (
            not tracking_initialized
            or forced_reregister_this_frame
            or (
                effective_reinit_interval > 0
                and local_idx % effective_reinit_interval == 0
            )
        )

        mask = None
        pred_pose_q = None
        pose_source = None

        if should_register:
            mask, mask_source = make_register_mask(
                args=args,
                reader=reader,
                frame_id=frame_id,
                rgb=color_rgb,
                bbox_corners=bbox_corners,
                gt_pose_q=gt_pose_cam_mesh,
                pred_pose_q_prev=pred_pose_q_prev,
                sam2_segmenter=sam2_segmenter,
                view_root=view_root,
                pred_pose_q_prev2=pred_pose_q_prev2,
                vertices_for_prompt=vertices_for_prompt,
                depth=depth,
            )
            if mask is None or mask.sum() < args.min_mask_pixels:
                if not tracking_initialized:
                    append_failed_frame(
                        frame_id, local_idx, gt_pose_file,
                        gt_pose_cam_mesh, gt_pose_base_object,
                        f"initialization_failed:{mask_source}",
                    )
                    save_metrics_bundle(
                        records, view_root, "pose_metrics_per_frame", args,
                        metadata={"aborted": True, "reason": mask_source},
                    )
                    raise RuntimeError(
                        f"frame={frame_id:04d}: no valid first mask ({mask_source}). "
                        "Provide mask/<first_frame>.png or pass a correct manual bbox with --use_manual_first_bbox."
                    )
                print(f"[Warning] frame={frame_id:04d}: re-register skipped ({mask_source}); using track_one().")
                should_register = False
            else:
                try:
                    pred_pose_q = ensure_pose_matrix(
                        est.register(
                            K=reader.K,
                            rgb=color_rgb,
                            depth=depth,
                            ob_mask=mask,
                            iteration=effective_iteration,
                            name=reader.seq_name,
                        ),
                        f"register pose frame {frame_id:04d}",
                    )
                    tracking_initialized = True
                    pose_source = f"register:{mask_source}"
                except Exception as exc:
                    print(f"[Error] register failed at frame {frame_id:04d}: {exc!r}")
                    if (
                        effective_failure_policy == "raise"
                        or not tracking_initialized
                    ):
                        raise
                    should_register = False

        if not should_register:
            if not tracking_initialized:
                append_failed_frame(
                    frame_id, local_idx, gt_pose_file,
                    gt_pose_cam_mesh, gt_pose_base_object,
                    "tracking_not_initialized",
                )
                continue
            try:
                pred_pose_q = ensure_pose_matrix(
                    est.track_one(
                        rgb=color_rgb,
                        depth=depth,
                        K=reader.K,
                        iteration=effective_iteration,
                    ),
                    f"track pose frame {frame_id:04d}",
                )
                pose_source = "track_one"
            except Exception as exc:
                print(f"[Warning] track_one failed at frame {frame_id:04d}: {exc!r}")
                if effective_failure_policy == "raise":
                    raise
                if effective_failure_policy == "skip":
                    append_failed_frame(
                        frame_id, local_idx, gt_pose_file,
                        gt_pose_cam_mesh, gt_pose_base_object,
                        f"track_one_failed_skip:{type(exc).__name__}",
                    )
                    continue
                recovered = False
                if effective_failure_policy == "reregister":
                    try:
                        mask, mask_source = make_register_mask(
                            args=args,
                            reader=reader,
                            frame_id=frame_id,
                            rgb=color_rgb,
                            bbox_corners=bbox_corners,
                            gt_pose_q=gt_pose_cam_mesh,
                            pred_pose_q_prev=pred_pose_q_prev,
                            sam2_segmenter=sam2_segmenter,
                            view_root=view_root,
                            pred_pose_q_prev2=pred_pose_q_prev2,
                            vertices_for_prompt=vertices_for_prompt,
                            depth=depth,
                        )
                        if mask is not None and mask.sum() >= args.min_mask_pixels:
                            pred_pose_q = ensure_pose_matrix(
                                est.register(
                                    K=reader.K,
                                    rgb=color_rgb,
                                    depth=depth,
                                    ob_mask=mask,
                                    iteration=effective_iteration,
                                    name=f"{reader.seq_name}_recover",
                                ),
                                f"recovery pose frame {frame_id:04d}",
                            )
                            pose_source = f"recover_register:{mask_source}"
                            recovered = True
                    except Exception as reg_exc:
                        print(f"[Warning] recovery register failed at frame {frame_id:04d}: {reg_exc!r}")
                if not recovered:
                    if pred_pose_q_prev_raw is None:
                        append_failed_frame(
                            frame_id, local_idx, gt_pose_file,
                            gt_pose_cam_mesh, gt_pose_base_object,
                            "recovery_failed_no_previous_pose",
                        )
                        continue
                    if effective_failure_policy == "freeze":
                        pred_pose_q = pred_pose_q_prev_raw.copy()
                        pose_source = "raw_freeze_recovery"
                    else:
                        motion_recovery = (
                            predict_next_pose_constant_velocity(
                                pred_pose_q_prev_raw2,
                                pred_pose_q_prev_raw,
                            )
                        )
                        pred_pose_q = (
                            pred_pose_q_prev_raw.copy()
                            if motion_recovery is None
                            else motion_recovery
                        )
                        pose_source = "raw_motion_recovery"

        if pred_pose_q is None:
            append_failed_frame(
                frame_id, local_idx, gt_pose_file,
                gt_pose_cam_mesh, gt_pose_base_object,
                "prediction_none",
            )
            continue

        # MultiView directly predicts a raw T_cam_mesh. Cylindrical symmetry can make
        # its local axes jump although the projected object remains correct.
        pred_pose_cam_mesh_raw = pred_pose_q
        pred_pose_base_mesh_raw = T_base_cam @ pred_pose_cam_mesh_raw

        branch_calibrated_this_frame = False
        if view_branch_calibration_pending:
            # As above, record the first successful calibration frame rather
            # than assuming reader.frame_ids[0] always produced a pose.
            view_branch_calibration_frame_id = int(frame_id)
            calibration_frames = getattr(
                args,
                "_view_branch_calibration_frames",
                None,
            )
            if calibration_frames is None:
                calibration_frames = {}
                setattr(
                    args,
                    "_view_branch_calibration_frames",
                    calibration_frames,
                )
            calibration_frames[view] = view_branch_calibration_frame_id
            (
                view_branch_correction,
                current_branch_diagnostics,
            ) = calibrate_view_pose_branch_once(
                pred_pose_base_mesh_raw=pred_pose_base_mesh_raw,
                gt_pose_base_object=gt_pose_base_object,
                T_object_mesh=T_object_mesh,
                mode=effective_view_branch_calibration_mode,
                safe_branch_candidates=prediction_branch_symmetries,
                vertices_mesh_m=(
                    sample_vertices(
                        vertices,
                        int(args.proxy_calibration_max_points),
                        seed=71,
                    )
                    if args.calibrate_proxy_model_from_first_frame
                    else None
                ),
                diameter_m=diameter_m,
                add_weight=float(args.proxy_calibration_add_weight),
                adds_weight=float(
                    args.proxy_calibration_adds_weight
                ),
                min_improvement_mm=float(
                    args.proxy_calibration_min_improvement_mm
                ),
            )
            geometry_diag = branch_geometry_discrepancy(
                sample_vertices(vertices, 5000, seed=41),
                view_branch_correction,
            )
            current_branch_diagnostics.update({
                **geometry_diag,
                "object_name": args.object_name,
                "view": view,
                "calibration_frame_id": int(frame_id),
                "fixed_model_transform_source": axis_alignment_source,
                "global_model_transform_changed": False,
                "T_object_mesh": T_object_mesh.tolist(),
                "pred_pose_base_mesh_raw": (
                    pred_pose_base_mesh_raw.tolist()
                ),
                "gt_pose_base_object": gt_pose_base_object.tolist(),
                "coordinate_definition": (
                    "T_base_mesh_canonical = T_base_mesh_raw @ "
                    "T_mesh_branch; T_mesh_branch is frozen, proper SO(3), "
                    "and has zero translation"
                ),
                "evaluation_warning": (
                    "This marked calibration frame is used only to freeze one "
                    "view-specific pose branch. Reuse the saved branch TXT on "
                    "a separate evaluation sequence for label-independent "
                    "test metrics."
                ),
                "tracking_health_validated_before_save": False,
            })
            if args.calibrate_proxy_model_from_first_frame:
                selected_add_ratio = float(
                    current_branch_diagnostics.get(
                        "selected_ADD_over_diameter",
                        float("nan"),
                    )
                )
                calibration_reliable = bool(
                    np.isfinite(selected_add_ratio)
                    and selected_add_ratio
                    <= float(args.proxy_calibration_max_add_ratio)
                )
                current_branch_diagnostics.update({
                    "calibration_reliable": calibration_reliable,
                    "max_allowed_ADD_over_diameter": float(
                        args.proxy_calibration_max_add_ratio
                    ),
                    "proxy_scale_candidates_audit": model_context.get(
                        "scale_candidates_audit", []
                    ),
                    "mesh_scale_resolved": resolved_mesh_scale,
                    "mesh_scale_source": model_context[
                        "mesh_scale_source"
                    ],
                    "bbox_diagonal_ratio_model_to_metadata": float(
                        model_context[
                            "bbox_diagonal_ratio_model_to_metadata"
                        ]
                    ),
                })
                if not calibration_reliable:
                    message = (
                        f"[Proxy calibration warning] {view} frame "
                        f"{frame_id:04d}: after direct rotation alignment, "
                        f"ADD/d is still {selected_add_ratio:.3f}. Rotation "
                        "axes are aligned, so the remaining discrepancy comes "
                        "from translation, proxy shape/scale, depth or the "
                        "saved first-frame mask."
                    )
                    if args.proxy_calibration_fail_unreliable:
                        raise RuntimeError(message)
                    print(message)
            view_branch_path = default_view_branch_output_path(
                args, save_root, view
            )
            branch_calibrated_this_frame = True
            print(
                f"[View branch calibration] view={view} frame="
                f"{frame_id:04d}: provisional prediction branch will be "
                "saved only after the prediction passes the GT-free tracking "
                "health check."
            )

        branch_for_this_pose = (
            np.eye(4, dtype=np.float64)
            if tracker_state_contains_view_branch
            else view_branch_correction
        )
        pred_pose_base_mesh_canonical = apply_pose_branch_correction(
            raw_pose_parent_mesh=pred_pose_base_mesh_raw,
            branch_correction=branch_for_this_pose,
            label=f"{view} frame {frame_id:04d}",
        )
        if (
            register_once_then_track
            and not tracker_state_contains_view_branch
            and str(pose_source).startswith("register:")
        ):
            corrected_pose_cam_mesh = validate_rigid_transform(
                T_cam_base @ pred_pose_base_mesh_canonical,
                f"{view} corrected initial tracker pose",
            )
            tracker_branch_state_sync = synchronize_multiview_pose_state(
                est, corrected_pose_cam_mesh
            )
            if tracker_branch_state_sync == "pose_last_synchronized":
                tracker_state_contains_view_branch = True
                # From this point the estimator itself evolves the corrected
                # pose. Keep raw history on the same branch so freeze recovery
                # cannot resurrect the pre-calibration horizontal/vertical
                # orientation.
                pred_pose_q = corrected_pose_cam_mesh.copy()
                pred_pose_cam_mesh_raw = corrected_pose_cam_mesh.copy()
                pred_pose_base_mesh_raw = (
                    pred_pose_base_mesh_canonical.copy()
                )
                pose_source = (
                    f"{pose_source}:initial_branch_absorbed"
                )
            elif not np.allclose(
                branch_for_this_pose,
                np.eye(4),
                atol=1e-10,
            ):
                print(
                    f"[Proxy branch warning] {view}: selected initial "
                    "direction could not be written into MultiView tracker state "
                    f"({tracker_branch_state_sync}). The fixed correction "
                    "will still be applied to outputs, but track_one() starts "
                    "from the estimator's uncorrected internal pose."
                )
            if view_branch_calibration_pending:
                current_branch_diagnostics[
                    "initial_tracker_state_sync"
                ] = tracker_branch_state_sync
                current_branch_diagnostics[
                    "branch_absorbed_into_tracker_state"
                ] = bool(tracker_state_contains_view_branch)

        temporal_mesh_reference = predict_next_pose_constant_velocity(
            temporal_mesh_pose_prev2,
            temporal_mesh_pose_prev,
        )
        tracking_health_diagnostics: Dict[str, object] = {
            "enabled": tracking_health_enabled,
            "status": "disabled",
            "initial": None,
            "reregister": None,
            "selected": None,
            "estimator_state_sync": "not_requested",
        }
        if tracking_health_enabled:
            tracking_observation = make_tracking_observation_from_frame(
                reader=reader,
                frame_id=frame_id,
                view_root=view_root,
                rgb=color_rgb,
                depth=depth,
                T_cam_base=T_cam_base,
                T_base_cam=T_base_cam,
                args=args,
                mask_override=mask,
            )
            initial_health = evaluate_tracking_candidate_health(
                pose_base_mesh=pred_pose_base_mesh_canonical,
                temporal_reference_base_mesh=temporal_mesh_reference,
                observation=tracking_observation,
                surface_points_mesh=tracking_surface_points,
                bbox_corners_mesh=bbox_corners,
                geometry_args=tracking_geometry_args,
                args=args,
                branch_symmetries=prediction_branch_symmetries,
            )
            tracking_health_diagnostics["initial"] = initial_health
            selected_health = initial_health

            # A finite but unhealthy track result is the common failure for
            # screwdriver-like objects. Re-register on the current observed
            # mask instead of waiting for track_one() to throw an exception.
            can_retry_registration = bool(
                not str(pose_source).startswith("register:")
                and not str(pose_source).startswith("recover_register:")
            )
            if (
                not bool(initial_health["passes"])
                and can_retry_registration
            ):
                retry_mask = None
                retry_mask_source = "not_attempted"
                try:
                    retry_mask, retry_mask_source = make_register_mask(
                        args=args,
                        reader=reader,
                        frame_id=frame_id,
                        rgb=color_rgb,
                        bbox_corners=bbox_corners,
                        gt_pose_q=gt_pose_cam_mesh,
                        pred_pose_q_prev=pred_pose_q_prev,
                        sam2_segmenter=sam2_segmenter,
                        view_root=view_root,
                        pred_pose_q_prev2=pred_pose_q_prev2,
                        vertices_for_prompt=vertices_for_prompt,
                        depth=depth,
                    )
                    if (
                        retry_mask is not None
                        and retry_mask.sum() >= args.min_mask_pixels
                    ):
                        retry_pose_cam_raw = ensure_pose_matrix(
                            est.register(
                                K=reader.K,
                                rgb=color_rgb,
                                depth=depth,
                                ob_mask=retry_mask,
                                iteration=effective_iteration,
                                name=(
                                    f"{reader.seq_name}_health_recover"
                                ),
                            ),
                            (
                                "health recovery register pose frame "
                                f"{frame_id:04d}"
                            ),
                        )
                        retry_pose_base_raw = (
                            T_base_cam @ retry_pose_cam_raw
                        )
                        retry_pose_base_canonical = (
                            apply_pose_branch_correction(
                                raw_pose_parent_mesh=retry_pose_base_raw,
                                branch_correction=view_branch_correction,
                                label=(
                                    f"{view} health recovery frame "
                                    f"{frame_id:04d}"
                                ),
                            )
                        )
                        retry_observation = (
                            make_tracking_observation_from_frame(
                                reader=reader,
                                frame_id=frame_id,
                                view_root=view_root,
                                rgb=color_rgb,
                                depth=depth,
                                T_cam_base=T_cam_base,
                                T_base_cam=T_base_cam,
                                args=args,
                                mask_override=retry_mask,
                            )
                        )
                        retry_health = (
                            evaluate_tracking_candidate_health(
                                pose_base_mesh=(
                                    retry_pose_base_canonical
                                ),
                                temporal_reference_base_mesh=(
                                    temporal_mesh_reference
                                ),
                                observation=retry_observation,
                                surface_points_mesh=(
                                    tracking_surface_points
                                ),
                                bbox_corners_mesh=bbox_corners,
                                geometry_args=tracking_geometry_args,
                                args=args,
                                branch_symmetries=(
                                    prediction_branch_symmetries
                                ),
                            )
                        )
                        tracking_health_diagnostics[
                            "reregister"
                        ] = retry_health
                        if bool(retry_health["passes"]):
                            pred_pose_q = retry_pose_cam_raw
                            pred_pose_cam_mesh_raw = (
                                retry_pose_cam_raw
                            )
                            pred_pose_base_mesh_raw = (
                                retry_pose_base_raw
                            )
                            pred_pose_base_mesh_canonical = (
                                retry_pose_base_canonical
                            )
                            mask = retry_mask
                            pose_source = (
                                "health_reregister:"
                                f"{retry_mask_source}"
                            )
                            selected_health = retry_health
                            tracking_observation = retry_observation
                except Exception as retry_exc:
                    tracking_health_diagnostics[
                        "reregister_error"
                    ] = (
                        f"{type(retry_exc).__name__}:{retry_exc}"
                    )

            if not bool(selected_health["passes"]):
                temporal_recovery_health = None
                if (
                    temporal_mesh_reference is not None
                    and tracking_recovery_age
                    < int(args.tracking_health_max_recovery_frames)
                ):
                    temporal_recovery_health = (
                        evaluate_tracking_candidate_health(
                            pose_base_mesh=temporal_mesh_reference,
                            temporal_reference_base_mesh=None,
                            observation=tracking_observation,
                            surface_points_mesh=tracking_surface_points,
                            bbox_corners_mesh=bbox_corners,
                            geometry_args=tracking_geometry_args,
                            args=args,
                            branch_symmetries=(
                                prediction_branch_symmetries
                            ),
                        )
                    )
                    tracking_health_diagnostics[
                        "temporal_recovery"
                    ] = temporal_recovery_health
                if (
                    temporal_recovery_health is not None
                    and bool(temporal_recovery_health["passes"])
                ):
                    pred_pose_base_mesh_canonical = (
                        temporal_mesh_reference.copy()
                    )
                    pred_pose_base_mesh_raw = (
                        pred_pose_base_mesh_canonical
                        @ np.linalg.inv(view_branch_correction)
                    )
                    pred_pose_cam_mesh_raw = (
                        T_cam_base @ pred_pose_base_mesh_raw
                    )
                    pred_pose_q = pred_pose_cam_mesh_raw.copy()
                    pose_source = "tracking_health_temporal_recovery"
                    selected_health = temporal_recovery_health
                    tracking_recovery_age += 1
                    force_reregister_next_frame = True
                    print(
                        f"[Tracking health] {view} frame "
                        f"{frame_id:04d}: rejected current pose "
                        f"({initial_health['reason']}); using bounded "
                        "constant-velocity recovery and forcing next-frame "
                        "re-registration."
                    )
                else:
                    force_reregister_next_frame = True
                    tracking_health_diagnostics["status"] = "failed"
                    tracking_health_diagnostics[
                        "selected"
                    ] = selected_health
                    append_tracking_health_row(
                        frame_id,
                        str(pose_source),
                        tracking_health_diagnostics,
                    )
                    print(
                        f"[Tracking health] {view} frame "
                        f"{frame_id:04d}: pose rejected "
                        f"({selected_health['reason']}); frame marked "
                        "invalid and next frame will re-register."
                    )
                    append_failed_frame(
                        frame_id,
                        local_idx,
                        gt_pose_file,
                        gt_pose_cam_mesh,
                        gt_pose_base_object,
                        (
                            "tracking_health_failed:"
                            f"{selected_health['reason']}"
                        ),
                    )
                    continue
            else:
                tracking_recovery_age = 0

            tracking_health_diagnostics["status"] = (
                "recovered"
                if str(pose_source).startswith(
                    ("health_", "tracking_health_")
                )
                else "passed"
            )
            tracking_health_diagnostics["selected"] = selected_health

        branch_commit_allowed = bool(
            branch_calibrated_this_frame
            and (
                not tracking_health_enabled
                or (
                    tracking_health_diagnostics.get("status")
                    in {"passed", "recovered"}
                    and not str(pose_source).startswith(
                        (
                            "health_reregister:",
                            "tracking_health_temporal_recovery",
                        )
                    )
                )
            )
        )
        if branch_calibrated_this_frame and not branch_commit_allowed:
            print(
                f"[View branch calibration] view={view} frame="
                f"{frame_id:04d}: provisional branch was not saved because "
                "the accepted recovery pose was not the pose used to estimate "
                "that branch. Calibration remains pending for the next "
                "directly observed pose."
            )
        if branch_commit_allowed:
            current_branch_diagnostics[
                "tracking_health_validated_before_save"
            ] = bool(tracking_health_enabled)
            os.makedirs(
                os.path.dirname(os.path.abspath(view_branch_path)),
                exist_ok=True,
            )
            np.savetxt(
                view_branch_path,
                view_branch_correction,
                fmt="%.12g",
            )
            branch_report_path = os.path.splitext(
                view_branch_path
            )[0] + "_report.json"
            with open(
                branch_report_path, "w", encoding="utf-8"
            ) as file:
                json.dump(
                    to_json_safe(current_branch_diagnostics),
                    file,
                    ensure_ascii=False,
                    indent=2,
                )
            candidate_metrics = current_branch_diagnostics.get(
                "candidate_metrics", []
            )
            if len(candidate_metrics) > 1:
                pd.DataFrame(candidate_metrics).to_csv(
                    os.path.splitext(view_branch_path)[0]
                    + "_candidates.csv",
                    index=False,
                )
            view_branch_corrections[view] = (
                view_branch_correction.copy()
            )
            view_branch_paths[view] = view_branch_path
            view_branch_diagnostics[view] = dict(
                current_branch_diagnostics
            )
            view_branch_calibration_pending = False
            run_config.update({
                "view_branch_calibration_pending": False,
                "view_branch_correction_path": view_branch_path,
                "T_pose_branch": view_branch_correction.tolist(),
                "view_branch_calibration": (
                    current_branch_diagnostics
                ),
            })
            with open(
                run_config_path, "w", encoding="utf-8"
            ) as file:
                json.dump(
                    to_json_safe(run_config),
                    file,
                    ensure_ascii=False,
                    indent=2,
                )
            print(
                f"[View branch calibration] view={view} | "
                f"rotation="
                f"{current_branch_diagnostics['branch_rotation_deg']:.3f} "
                f"degree | residual="
                f"{current_branch_diagnostics['best_rotation_error_deg']:.6f} "
                f"degree | mesh symmetry mean/max="
                f"{geometry_diag['mean_bidirectional_mm']:.3f}/"
                f"{geometry_diag['max_bidirectional_mm']:.3f} mm | "
                f"tracking-health-checked={tracking_health_enabled} | "
                f"saved {view_branch_path}"
            )
            if current_branch_diagnostics.get(
                "selection_uses_add_adds", False
            ):
                print(
                    f"[Proxy branch score] view={view} | "
                    f"branch={current_branch_diagnostics['branch_map']} | "
                    f"ADD="
                    f"{current_branch_diagnostics['selected_ADD_mm']:.3f} "
                    f"mm | ADD-S="
                    f"{current_branch_diagnostics['selected_ADD-S_mm']:.3f} "
                    f"mm | objective="
                    f"{current_branch_diagnostics['selected_weighted_objective_mm']:.3f} "
                    "mm"
                )

        if args.temporal_axis_stabilization:
            (
                pred_pose_base_mesh_temporal,
                temporal_axis_map,
                temporal_axis_correction_deg,
                temporal_axis_reference_error_deg,
            ) = canonicalize_prediction_discrete_branch(
                raw_pose_parent_mesh=pred_pose_base_mesh_canonical,
                reference_pose_parent_mesh=temporal_mesh_reference,
                branch_symmetries=prediction_branch_symmetries,
                min_improvement_deg=(
                    args.temporal_axis_min_improvement_deg
                ),
            )
        else:
            pred_pose_base_mesh_temporal = (
                pred_pose_base_mesh_canonical.copy()
            )
            temporal_axis_map = "+x,+y,+z"
            temporal_axis_correction_deg = 0.0
            temporal_axis_reference_error_deg = (
                0.0 if temporal_mesh_reference is None else
                rotation_distance_deg_from_matrices(
                    pred_pose_base_mesh_temporal[:3, :3],
                    temporal_mesh_reference[:3, :3],
                )
            )

        symmetry_angle_deg = 0.0
        symmetry_axis_flipped = False
        symmetry_reference_error_deg = 0.0
        if args.symmetry_stabilization:
            symmetry_reference_base_mesh = (
                pred_pose_base_mesh_temporal
                if temporal_mesh_reference is None
                else temporal_mesh_reference
            )

            pred_pose_base_mesh, symmetry_angle_deg, symmetry_axis_flipped, symmetry_reference_error_deg = (
                canonicalize_axial_symmetric_pose(
                    raw_pose_parent_mesh=pred_pose_base_mesh_temporal,
                    reference_pose_parent_mesh=symmetry_reference_base_mesh,
                    axis_index=symmetry_axis_index,
                    samples=args.symmetry_samples,
                    allow_axis_flip=args.symmetry_allow_axis_flip,
                )
            )
            pred_pose_cam_mesh = T_cam_base @ pred_pose_base_mesh
        else:
            pred_pose_base_mesh = pred_pose_base_mesh_temporal.copy()
            pred_pose_cam_mesh = T_cam_base @ pred_pose_base_mesh

        tracker_state_pose_cam_raw = pred_pose_cam_mesh_raw.copy()
        if tracking_health_enabled:
            tracker_state_pose_base_raw = (
                pred_pose_base_mesh
                @ np.linalg.inv(view_branch_correction)
            )
            tracker_state_pose_cam_raw = (
                T_cam_base @ tracker_state_pose_base_raw
            )
            tracking_health_diagnostics[
                "estimator_state_sync"
            ] = synchronize_multiview_pose_state(
                est, tracker_state_pose_cam_raw
            )
            append_tracking_health_row(
                frame_id,
                str(pose_source),
                tracking_health_diagnostics,
            )

        temporal_mesh_pose_prev2 = (
            None if temporal_mesh_pose_prev is None
            else temporal_mesh_pose_prev.copy()
        )
        temporal_mesh_pose_prev = pred_pose_base_mesh.copy()

        # Convert both the direct estimator output and the branch-canonical
        # mesh pose to the one robot-defined object frame. Only the canonical
        # chain is used for reported metrics, visualization and fusion.
        pred_pose_base_object_raw = (
            pred_pose_base_mesh_raw @ T_mesh_object
        )
        pred_pose_base_object_candidate = (
            pred_pose_base_mesh @ T_mesh_object
        )

        # Keep the object pose rigidly tied to the stabilized mesh pose through
        # one immutable metadata-derived T_object_mesh. Every temporal branch
        # above was applied to the prediction mesh pose itself, so the axes,
        # projected geometry, metrics and fusion candidate remain coupled.
        pred_pose_base_object = pred_pose_base_object_candidate.copy()

        axis_pose_prev2 = (
            None if axis_pose_prev is None else axis_pose_prev.copy()
        )
        axis_pose_prev = pred_pose_base_object.copy()

        pred_pose_cam_object = T_cam_base @ pred_pose_base_object
        pred_pose_cam_object_raw = T_cam_base @ pred_pose_base_object_raw

        # Motion prompts use the canonical pose. Freeze recovery retains the
        # direct MultiView pose separately, preventing the fixed branch transform
        # from being applied twice.
        pred_pose_q_prev_raw2 = (
            None
            if pred_pose_q_prev_raw is None
            else pred_pose_q_prev_raw.copy()
        )
        pred_pose_q_prev_raw = tracker_state_pose_cam_raw.copy()
        pred_pose_q_prev2 = None if pred_pose_q_prev is None else pred_pose_q_prev.copy()
        pred_pose_q_prev = pred_pose_cam_mesh.copy()

        # Every requested metric uses the same branch-canonical pose that is
        # drawn, saved and passed to multiview fusion. The direct raw error is
        # retained only as an internal diagnostic.
        raw_r_err = rotation_error_deg(
            pred_pose_base_object_raw, gt_pose_base_object
        )
        r_err = rotation_error_deg(
            pred_pose_base_object, gt_pose_base_object
        )
        sym_r_err = symmetry_aware_rotation_error_deg(
            pred_pose_cam_mesh, gt_pose_cam_mesh, metric_symmetries
        )
        t_err_mm = float(
            np.linalg.norm(
                pred_pose_cam_object[:3, 3]
                - gt_pose_cam_object[:3, 3]
            ) * 1000.0
        )
        t_err_mesh_origin_mm = float(
            np.linalg.norm(
                pred_pose_cam_mesh[:3, 3]
                - gt_pose_cam_mesh[:3, 3]
            ) * 1000.0
        )
        t_err_base_mm = float(
            np.linalg.norm(
                pred_pose_base_object[:3, 3] - gt_pose_base_object[:3, 3]
            ) * 1000.0
        )
        add_m = add_error_m(
            add_vertices, pred_pose_cam_mesh, gt_pose_cam_mesh
        )
        adds_m = add_s_error_m(
            add_vertices, pred_pose_cam_mesh, gt_pose_cam_mesh
        )
        projection_mean_px = mean_projection_error_px(
            add_vertices, pred_pose_cam_mesh, gt_pose_cam_mesh, reader.K
        )
        add_success = bool(add_m < 0.1 * diameter_m)
        adds_success = bool(adds_m < 0.1 * diameter_m)
        success_5deg_5cm = bool(r_err < 5.0 and t_err_mm < 50.0)

        stable_r_err = r_err
        stable_add_m = add_error_m(add_vertices, pred_pose_cam_mesh, gt_pose_cam_mesh)
        stable_adds_m = add_s_error_m(add_vertices, pred_pose_cam_mesh, gt_pose_cam_mesh)

        delta_pose = (
            np.linalg.inv(gt_pose_cam_object) @ pred_pose_cam_object
        )
        signed_camera_t_mm = (
            pred_pose_cam_object[:3, 3] - gt_pose_cam_object[:3, 3]
        ) * 1000.0
        signed_base_t_mm = (
            pred_pose_base_object[:3, 3] - gt_pose_base_object[:3, 3]
        ) * 1000.0
        rpe_t_mm, rpe_r_deg = relative_pose_errors(
            metric_prev_pred_raw,
            pred_pose_base_object,
            metric_prev_gt,
            gt_pose_base_object,
        )
        metric_prev_pred_raw = pred_pose_base_object.copy()
        metric_prev_gt = gt_pose_base_object.copy()

        metric_status = "disabled"
        bop_result: Dict[str, object] = {
            "VSD_errors": np.full(len(vsd_taus_for_columns), np.nan, dtype=np.float64),
            "VSD_error_mean": float("nan"),
            "VSD_AR_contribution": float("nan"),
            "MSSD_m": float("nan"),
            "MSSD_norm_d": float("nan"),
            "MSSD_AR_contribution": float("nan"),
            "MSPD_px": float("nan"),
            "MSPD_AR_contribution": float("nan"),
            "BOP_AR_contribution": float("nan"),
        }
        if bop_evaluator is not None:
            try:
                bop_result = bop_evaluator.evaluate(
                    pred_pose_cam_mesh=pred_pose_cam_mesh,
                    gt_pose_cam_mesh=gt_pose_cam_mesh,
                    depth_m=depth,
                    K=reader.K,
                    metric_vertices_m=bop_vertices,
                )
                metric_status = "ok"
            except Exception as exc:
                metric_status = f"error:{type(exc).__name__}:{exc}"
                if args.metric_failure_policy == "raise":
                    raise
                print(f"[Warning] BOP metrics failed at frame {frame_id:04d}: {exc!r}")

        vsd_columns = {
            f"VSD_tau_{tau:.2f}": float(error)
            for tau, error in zip(
                vsd_taus_for_columns,
                np.asarray(bop_result["VSD_errors"]).reshape(-1),
            )
        }

        np.savetxt(os.path.join(pred_cam_raw_pose_dir, f"{frame_id:04d}.txt"), pred_pose_cam_mesh_raw)
        np.savetxt(
            os.path.join(
                pred_cam_raw_object_pose_dir, f"{frame_id:04d}.txt"
            ),
            pred_pose_cam_object_raw,
        )
        np.savetxt(os.path.join(pred_base_raw_pose_dir, f"{frame_id:04d}.txt"), pred_pose_base_mesh_raw)
        np.savetxt(os.path.join(pred_base_raw_object_pose_dir, f"{frame_id:04d}.txt"), pred_pose_base_object_raw)
        np.savetxt(os.path.join(pred_cam_pose_dir, f"{frame_id:04d}.txt"), pred_pose_cam_mesh)
        np.savetxt(
            os.path.join(pred_cam_object_pose_dir, f"{frame_id:04d}.txt"),
            pred_pose_cam_object,
        )
        np.savetxt(os.path.join(pred_base_pose_dir, f"{frame_id:04d}.txt"), pred_pose_base_mesh)
        np.savetxt(os.path.join(pred_base_object_pose_dir, f"{frame_id:04d}.txt"), pred_pose_base_object)
        np.savetxt(os.path.join(gt_cam_pose_dir, f"{frame_id:04d}.txt"), gt_pose_cam_mesh)
        np.savetxt(
            os.path.join(gt_cam_object_pose_dir, f"{frame_id:04d}.txt"),
            gt_pose_cam_object,
        )
        np.savetxt(os.path.join(gt_base_pose_dir, f"{frame_id:04d}.txt"), gt_pose_base_mesh)
        np.savetxt(os.path.join(gt_base_object_input_dir, f"{frame_id:04d}.txt"), gt_pose_base_object)

        vis = cv2.cvtColor(color_rgb, cv2.COLOR_RGB2BGR)
        if args.show_mask:
            try:
                vis_mask = mask if mask is not None else reader.get_mask(frame_id, fallback_to_first=False)
                draw_mask_contour(vis, vis_mask)
            except Exception:
                pass

        # Green GT cuboid uses the capture metadata bounds directly in the GT
        # object frame. It is independent of proxy-mesh scale/centering.
        gt_state = pose_projection_state(
            gt_pose_cam_object,
            object_bbox_corners,
            reader.K,
            vis.shape,
        )
        pred_state = pose_projection_state(
            pred_pose_cam_object,
            object_bbox_corners,
            reader.K,
            vis.shape,
        )

        draw_projected_bbox(
            vis,
            object_bbox_corners,
            gt_pose_cam_object,
            reader.K,
            color=(0, 255, 0),
            thickness=3,
        )
        draw_projected_bbox(
            vis, object_bbox_corners, pred_pose_cam_object, reader.K,
            color=(0, 0, 255), thickness=2,
        )
        if args.draw_pose_axes:
            # Both boxes and rays now use the same fixed GT object frame.
            draw_pose_rays_monocolor(vis, gt_pose_cam_object, reader.K, axis_length=axis_length, color=(0, 255, 0), thickness=3)
            draw_pose_rays_monocolor(vis, pred_pose_cam_object, reader.K, axis_length=axis_length, color=(0, 0, 255), thickness=2)
        if args.draw_origin_gap:
            draw_origin_gap(
                vis, gt_pose_cam_object, pred_pose_cam_object, reader.K
            )

        # Keep the RGB content unobstructed by default. Metrics are saved to the
        # compact per-frame workbook. Enable --show_overlay only for a compact
        # diagnostic label.
        if args.show_overlay:
            compact_lines = [
                f"frame: {frame_id:04d} | {pose_source}",
                f"green: GT | red: stable prediction",
                f"object axes: {axis_map_resolved}",
                f"sym axis: {symmetry_axis_name} | correction: {symmetry_angle_deg:.1f} deg"
                + (" | axis flip" if symmetry_axis_flipped else ""),
                (
                    f"temporal axis: {temporal_axis_map} | "
                    f"residual: {temporal_axis_reference_error_deg:.1f} deg"
                ),
            ]
            draw_text_panel(vis, compact_lines, origin=(12, 24), line_height=22)

        if args.save_video:
            frame_for_video = vis
            if abs(args.video_scale - 1.0) > 1e-9:
                frame_for_video = cv2.resize(
                    vis, None, fx=args.video_scale, fy=args.video_scale,
                    interpolation=cv2.INTER_LINEAR,
                )
            if writer is None:
                h, w = frame_for_video.shape[:2]
                writer = cv2.VideoWriter(
                    video_path,
                    cv2.VideoWriter_fourcc(*"mp4v"),
                    args.fps,
                    (w, h),
                )
                if not writer.isOpened():
                    raise RuntimeError(f"Cannot open video writer: {video_path}")
            writer.write(frame_for_video)

        if args.save_images:
            cv2.imwrite(os.path.join(vis_dir, f"{frame_id:04d}.png"), vis)

        if not args.no_display:
            display = vis
            if abs(args.display_scale - 1.0) > 1e-9:
                display = cv2.resize(
                    vis, None, fx=args.display_scale, fy=args.display_scale,
                    interpolation=cv2.INTER_LINEAR,
                )
            window_name = f"MultiView real-scene pose tracking - {reader.seq_name}"
            cv2.imshow(window_name, display)
            key = cv2.waitKey(args.wait_ms) & 0xFF
            if key in [ord("q"), 27]:
                print(f"\n[Stopped] User requested stop at frame {frame_id:04d}; saving collected metrics.")
                stop_requested = True
            if key == ord(" "):
                while True:
                    key2 = cv2.waitKey(0) & 0xFF
                    if key2 in [ord(" "), ord("q"), 27]:
                        break
                if key2 in [ord("q"), 27]:
                    print(f"\n[Stopped] User requested stop at frame {frame_id:04d}; saving collected metrics.")
                    stop_requested = True

        is_global_axis_calibration_frame = bool(
            getattr(args, "_proxy_calibration_frame_id", None)
            == frame_id
        )
        is_view_branch_calibration_frame = bool(
            view_branch_calibration_frame_id == frame_id
        )
        is_axis_calibration_frame = bool(
            is_global_axis_calibration_frame
            or is_view_branch_calibration_frame
        )
        calibration_diag_for_row = {}
        branch_diag_for_row = view_branch_diagnostics.get(view, {})
        gt_object_quaternion = SciPyRotation.from_matrix(
            gt_pose_base_object[:3, :3]
        ).as_quat()
        pred_object_raw_quaternion = SciPyRotation.from_matrix(
            pred_pose_base_object_raw[:3, :3]
        ).as_quat()
        pred_object_stable_quaternion = SciPyRotation.from_matrix(
            pred_pose_base_object[:3, :3]
        ).as_quat()
        selected_tracking_health = (
            tracking_health_diagnostics.get("selected") or {}
        )
        selected_tracking_geometry = (
            selected_tracking_health.get("geometry") or {}
        )
        record = {
            "object_name": args.object_name,
            "object_id": "" if args.object_id is None else str(args.object_id),
            "camera_view": reader.view,
            "sequence": reader.seq_name,
            "frame_idx": int(frame_id),
            "local_idx": int(local_idx),
            "estimate_valid": True,
            "pose_source": pose_source,
            "metric_status": metric_status,
            "axis_alignment_source": axis_alignment_source,
            "axis_map_mesh_to_object": axis_map_resolved,
            "axis_calibration_frame": is_axis_calibration_frame,
            "global_axis_calibration_frame": (
                is_global_axis_calibration_frame
            ),
            "view_branch_calibration_frame": (
                is_view_branch_calibration_frame
            ),
            "axis_calibration_uses_gt": bool(
                args.calibrate_axis_from_first_frame
                or args.calibrate_proxy_model_from_first_frame
            ),
            "axis_calibration_best_error_deg": float(
                calibration_diag_for_row.get(
                    "best_rotation_error_deg", float("nan")
                )
            ),
            "axis_calibration_margin_deg": float(
                calibration_diag_for_row.get(
                    "selection_margin_deg", float("nan")
                )
            ),
            "view_branch_correction_path": view_branch_path or "identity",
            "view_branch_rotation_deg": float(
                branch_diag_for_row.get(
                    "branch_rotation_deg", 0.0
                )
            ),
            "view_branch_residual_deg": float(
                branch_diag_for_row.get(
                    "best_rotation_error_deg", float("nan")
                )
            ),
            "official_evaluation_frame": not is_axis_calibration_frame,
            "gt_pose_file": gt_pose_file,
            "T_cam_base_meta_source": T_cam_base_source,
            "mask_pixels": int(mask.sum()) if mask is not None else 0,
            "gt_projection_state": gt_state,
            "pred_projection_state": pred_state,
            "tracking_profile": tracking_profile,
            "tracking_health_enabled": tracking_health_enabled,
            "tracking_health_status": tracking_health_diagnostics.get(
                "status", "disabled"
            ),
            "tracking_health_reason": selected_tracking_health.get(
                "reason", ""
            ),
            "tracking_health_geometry_cost": float(
                selected_tracking_geometry.get(
                    "score", float("nan")
                )
            ),
            "tracking_health_depth_inlier_ratio": float(
                selected_tracking_geometry.get(
                    "depth_inlier_ratio", float("nan")
                )
            ),
            "tracking_health_mask_inside_ratio": float(
                selected_tracking_geometry.get(
                    "projected_inside_mask_ratio", float("nan")
                )
            ),
            "tracking_health_translation_jump_mm": float(
                selected_tracking_health.get(
                    "translation_jump_mm", float("nan")
                )
            ),
            "tracking_health_rotation_jump_deg": float(
                selected_tracking_health.get(
                    "rotation_jump_deg", float("nan")
                )
            ),
            "tracking_health_estimator_state_sync": (
                tracking_health_diagnostics.get(
                    "estimator_state_sync", "not_requested"
                )
            ),

            # All compact metrics use the same fixed-branch canonical pose.
            "ADD_mm": float(add_m * 1000.0),
            "ADDS_mm": float(adds_m * 1000.0),
            "ADD_0.1d_success": add_success,
            "ADDS_0.1d_success": adds_success,
            "Rotation_error_deg": float(r_err),
            "Raw_rotation_error_deg": float(raw_r_err),
            "Symmetry_aware_rotation_error_deg": float(sym_r_err),
            "Translation_error_mm": float(t_err_mm),
            "Translation_error_base_mm": float(t_err_base_mm),
            "Translation_error_mesh_origin_mm": float(
                t_err_mesh_origin_mm
            ),
            "Mean_projection_error_px": float(projection_mean_px),
            "5deg_5cm_success": success_5deg_5cm,
            "RPE_translation_mm": float(rpe_t_mm),
            "RPE_rotation_deg": float(rpe_r_deg),
            "VSD_error_mean": float(bop_result["VSD_error_mean"]),
            "VSD_AR_contribution": float(bop_result["VSD_AR_contribution"]),
            "MSSD_mm": float(bop_result["MSSD_m"]) * 1000.0,
            "MSSD_norm_d": float(bop_result["MSSD_norm_d"]),
            "MSSD_AR_contribution": float(bop_result["MSSD_AR_contribution"]),
            "MSPD_px": float(bop_result["MSPD_px"]),
            "MSPD_AR_contribution": float(bop_result["MSPD_AR_contribution"]),
            "BOP_AR_contribution": float(bop_result["BOP_AR_contribution"]),
            "Right_view_contribution": float("nan"),
            "Left_view_contribution": float("nan"),
            **vsd_columns,

            # Signed translation residuals help locate axis/calibration problems.
            "Camera_dx_mm": float(signed_camera_t_mm[0]),
            "Camera_dy_mm": float(signed_camera_t_mm[1]),
            "Camera_dz_mm": float(signed_camera_t_mm[2]),
            "Base_dx_mm": float(signed_base_t_mm[0]),
            "Base_dy_mm": float(signed_base_t_mm[1]),
            "Base_dz_mm": float(signed_base_t_mm[2]),
            "Relative_delta_tx_mm": float(delta_pose[0, 3] * 1000.0),
            "Relative_delta_ty_mm": float(delta_pose[1, 3] * 1000.0),
            "Relative_delta_tz_mm": float(delta_pose[2, 3] * 1000.0),

            # Pose coordinates retained for frame-by-frame auditing.
            "gt_cam_tx_m": float(gt_pose_cam_mesh[0, 3]),
            "gt_cam_ty_m": float(gt_pose_cam_mesh[1, 3]),
            "gt_cam_tz_m": float(gt_pose_cam_mesh[2, 3]),
            "pred_cam_raw_tx_m": float(pred_pose_cam_mesh_raw[0, 3]),
            "pred_cam_raw_ty_m": float(pred_pose_cam_mesh_raw[1, 3]),
            "pred_cam_raw_tz_m": float(pred_pose_cam_mesh_raw[2, 3]),
            "pred_cam_stable_tx_m": float(pred_pose_cam_mesh[0, 3]),
            "pred_cam_stable_ty_m": float(pred_pose_cam_mesh[1, 3]),
            "pred_cam_stable_tz_m": float(pred_pose_cam_mesh[2, 3]),
            "gt_base_object_tx_m": float(gt_pose_base_object[0, 3]),
            "gt_base_object_ty_m": float(gt_pose_base_object[1, 3]),
            "gt_base_object_tz_m": float(gt_pose_base_object[2, 3]),
            "gt_base_object_qx": float(gt_object_quaternion[0]),
            "gt_base_object_qy": float(gt_object_quaternion[1]),
            "gt_base_object_qz": float(gt_object_quaternion[2]),
            "gt_base_object_qw": float(gt_object_quaternion[3]),
            "pred_base_object_raw_tx_m": float(pred_pose_base_object_raw[0, 3]),
            "pred_base_object_raw_ty_m": float(pred_pose_base_object_raw[1, 3]),
            "pred_base_object_raw_tz_m": float(pred_pose_base_object_raw[2, 3]),
            "pred_base_object_raw_qx": float(
                pred_object_raw_quaternion[0]
            ),
            "pred_base_object_raw_qy": float(
                pred_object_raw_quaternion[1]
            ),
            "pred_base_object_raw_qz": float(
                pred_object_raw_quaternion[2]
            ),
            "pred_base_object_raw_qw": float(
                pred_object_raw_quaternion[3]
            ),
            "pred_base_object_stable_tx_m": float(pred_pose_base_object[0, 3]),
            "pred_base_object_stable_ty_m": float(pred_pose_base_object[1, 3]),
            "pred_base_object_stable_tz_m": float(pred_pose_base_object[2, 3]),
            "pred_base_object_stable_qx": float(
                pred_object_stable_quaternion[0]
            ),
            "pred_base_object_stable_qy": float(
                pred_object_stable_quaternion[1]
            ),
            "pred_base_object_stable_qz": float(
                pred_object_stable_quaternion[2]
            ),
            "pred_base_object_stable_qw": float(
                pred_object_stable_quaternion[3]
            ),

            # Optional physical-symmetry canonicalization diagnostics.
            "Stabilized_rotation_error_deg": float(stable_r_err),
            "Stabilized_ADD_mm": float(stable_add_m * 1000.0),
            "Stabilized_ADDS_mm": float(stable_adds_m * 1000.0),
            "symmetry_axis": symmetry_axis_name,
            "stabilization_correction_deg": float(symmetry_angle_deg),
            "stabilization_axis_flipped": bool(symmetry_axis_flipped),
            "stabilization_reference_error_deg": float(symmetry_reference_error_deg),
            "temporal_axis_map": temporal_axis_map,
            "temporal_axis_correction_deg": float(
                temporal_axis_correction_deg
            ),
            "temporal_axis_reference_error_deg": float(
                temporal_axis_reference_error_deg
            ),
            "metric_symmetry_source": metric_symmetry_source,
            "metric_symmetry_count": int(len(metric_symmetries)),
            "mesh_scale": resolved_mesh_scale,
            "diameter_m": float(diameter_m),
            "diameter_source": diameter_source,
        }
        records.append(record)

        if (
            not getattr(args, "_defer_view_metric_export", False)
            and args.metrics_flush_interval > 0
            and len(records) % args.metrics_flush_interval == 0
        ):
            saved = save_metrics_bundle(
                records=records,
                output_dir=view_root,
                basename="pose_metrics_per_frame",
                args=args,
                metadata={
                    "object_name": args.object_name,
                    "camera_view": reader.view,
                    "camera_extrinsic_source": (
                        "per_frame_meta/T_cam_base"
                    ),
                    "T_object_mesh_path": T_object_mesh_path or "identity",
                    "axis_alignment_source": axis_alignment_source,
                    "axis_map_mesh_to_object": axis_map_resolved,
                    "view_branch_correction_path": (
                        view_branch_path or "identity"
                    ),
                    "metric_symmetry_source": metric_symmetry_source,
                },
            )
            print(f"[Checkpoint] {saved.get('xlsx', '')}")
        if stop_requested:
            break

    if writer is not None:
        writer.release()
        print(f"[Saved] video: {video_path}")

    if records and not getattr(
        args, "_defer_view_metric_export", False
    ):
        saved = save_metrics_bundle(
            records=records,
            output_dir=view_root,
            basename="pose_metrics_per_frame",
            args=args,
            metadata={
                "object_name": args.object_name,
                "camera_view": reader.view,
                "camera_extrinsic_source": (
                    "per_frame_meta/T_cam_base"
                ),
                "T_object_mesh_path": T_object_mesh_path or "identity",
                "axis_alignment_source": axis_alignment_source,
                "axis_map_mesh_to_object": axis_map_resolved,
                "axis_calibration_frame_id": None,
                "view_branch_correction_path": (
                    view_branch_path or "identity"
                ),
                "view_branch_calibration_frame_id": (
                    view_branch_calibration_frame_id
                ),
                "metric_symmetry_source": metric_symmetry_source,
                "metric_symmetry_count": len(metric_symmetries),
                "diameter_m": diameter_m,
                "diameter_source": diameter_source,
            },
        )
        print(f"[Saved] metrics CSV: {saved['csv']}")
        print(f"[Saved] metrics Excel: {saved['xlsx']}")

    if tracking_health_rows:
        tracking_health_path = os.path.join(
            view_root, "tracking_health_diagnostics.csv"
        )
        pd.DataFrame(tracking_health_rows).to_csv(
            tracking_health_path, index=False
        )
        print(
            "[Saved] tracking health diagnostics: "
            f"{tracking_health_path}"
        )

    return records


def process_two_view_fusion(
    args,
    save_root: str,
    view_records_by_view: Dict[str, List[Dict]],
) -> List[Dict]:
    """Fuse synchronized right/left estimates and evaluate the final pose."""
    readers = {
        view: RealSceneMultiViewReader(
            dataset_root=args.dataset_root,
            view=view,
            start_frame=args.start_frame,
            stride=args.running_stride,
            args=args,
        )
        for view in ("right", "left")
    }
    model_contexts = {
        view: load_mesh_from_object_pose_setup(
            args=args,
            setup=readers[view].get_reference_object_pose_setup(),
            seq_dir=readers[view].seq_dir,
        )
        for view in ("right", "left")
    }
    for key in (
        "T_object_mesh",
        "bbox_min_object_m",
        "bbox_max_object_m",
        "mesh_center_offset_after_scale_m",
    ):
        if not np.allclose(
            np.asarray(model_contexts["right"][key]),
            np.asarray(model_contexts["left"][key]),
            atol=1e-9,
            rtol=1e-7,
        ):
            raise ValueError(
                f"Right/left object_pose_setup resolve different {key}; "
                "fusion cannot mix object coordinate systems."
            )
    if not np.isclose(
        float(model_contexts["right"]["mesh_scale"]),
        float(model_contexts["left"]["mesh_scale"]),
        atol=1e-15,
        rtol=1e-12,
    ):
        raise ValueError(
            "Right/left object_pose_setup resolve different mesh scales."
        )
    if (
        str(model_contexts["right"]["model_frame_mode"])
        != str(model_contexts["left"]["model_frame_mode"])
    ):
        raise ValueError(
            "Right/left object_pose_setup resolve different model-frame "
            "modes; fusion cannot mix differently preprocessed proxy models."
        )
    right_vertices = np.asarray(
        model_contexts["right"]["mesh"].vertices, dtype=np.float64
    )
    left_vertices = np.asarray(
        model_contexts["left"]["mesh"].vertices, dtype=np.float64
    )
    if (
        right_vertices.shape != left_vertices.shape
        or not np.allclose(
            right_vertices, left_vertices, atol=1e-10, rtol=1e-8
        )
    ):
        raise ValueError(
            "Right/left proxy meshes differ after metadata unit/origin "
            "processing; both predictions must enter fusion in one fixed "
            "GT object coordinate system."
        )

    model_context = model_contexts["right"]
    mesh = model_context["mesh"]
    mesh_center_offset = np.asarray(
        model_context["mesh_center_offset_after_scale_m"],
        dtype=np.float64,
    )
    resolved_mesh_scale = float(model_context["mesh_scale"])
    computed_diameter_m = float(model_context["computed_diameter_m"])
    diameter_m, diameter_source = resolve_model_diameter_m(
        args, fallback_diameter_m=computed_diameter_m
    )
    T_object_mesh = np.asarray(
        model_context["T_object_mesh"], dtype=np.float64
    ).copy()
    T_object_mesh_path = (
        os.path.abspath(args.T_object_mesh)
        if args.T_object_mesh is not None else None
    )
    axis_alignment_source = str(
        model_context["T_object_mesh_source"]
    )
    axis_map_resolved = describe_axis_rotation(T_object_mesh[:3, :3])
    vertices = np.asarray(mesh.vertices, dtype=np.float64)
    bbox_corners = make_bbox_corners(vertices)
    vertices_object = transform_points(vertices, T_object_mesh)
    fusion_bbox_corners_object = make_bbox_corners(vertices_object)
    object_bbox_corners = make_bbox_corners_from_bounds(
        model_context["bbox_min_object_m"],
        model_context["bbox_max_object_m"],
    )
    add_vertices = sample_vertices(vertices, args.max_add_points)
    bop_vertices = sample_vertices(vertices, args.max_bop_points, seed=17)
    surface_points = sample_mesh_surface_points(
        mesh, args.fusion_max_surface_points, seed=31
    )
    surface_points_object = transform_points(
        surface_points, T_object_mesh
    )
    symmetry_axis_index, symmetry_axis_name, mesh_extents_m = resolve_symmetry_axis(
        vertices, args.symmetry_axis
    )
    metric_symmetries, metric_symmetry_source = build_metric_symmetries(
        args=args,
        vertices_m=vertices,
        diameter_m=diameter_m,
        resolved_axis_index=symmetry_axis_index,
    )
    fusion_symmetries_object = transform_symmetries_to_object_frame(
        metric_symmetries, T_object_mesh
    )

    fusion_root = os.path.join(save_root, f"{args.object_name}_fusion")
    view_roots = {
        view: os.path.join(save_root, readers[view].seq_name)
        for view in ("right", "left")
    }
    candidate_pose_dirs = {
        # Both candidates are explicitly represented in the same immutable GT
        # object frame before any cross-view gate or optimization.
        view: os.path.join(view_roots[view], "poses_pred_base_object")
        for view in ("right", "left")
    }
    fused_base_mesh_dir = os.path.join(fusion_root, "poses_fused_base_mesh")
    fused_base_object_dir = os.path.join(fusion_root, "poses_fused_base_object")
    fused_camera_dirs = {
        view: os.path.join(fusion_root, f"poses_fused_camera_{view}_mesh")
        for view in ("right", "left")
    }
    fused_camera_object_dirs = {
        view: os.path.join(
            fusion_root, f"poses_fused_camera_{view}_object"
        )
        for view in ("right", "left")
    }
    vis_dirs = {
        view: os.path.join(fusion_root, "vis", view)
        for view in ("right", "left")
    }
    for path in (
        fused_base_mesh_dir,
        fused_base_object_dir,
        *fused_camera_dirs.values(),
        *fused_camera_object_dirs.values(),
    ):
        os.makedirs(path, exist_ok=True)
    if args.save_images:
        for path in vis_dirs.values():
            os.makedirs(path, exist_ok=True)

    frame_sets = {
        view: set(readers[view].frame_ids)
        for view in ("right", "left")
    }
    all_frame_ids = sorted(frame_sets["right"] | frame_sets["left"])
    common_frame_ids = sorted(frame_sets["right"] & frame_sets["left"])
    if not common_frame_ids:
        raise ValueError(
            "The right and left sequences have no common frame ids; "
            "two-view fusion requires synchronized frame numbering."
        )

    bop_evaluators: Dict[str, Optional[BOPMetricEvaluator]] = {
        "right": None,
        "left": None,
    }
    if not args.disable_bop_metrics:
        for view in ("right", "left"):
            first_rgb = readers[view].get_rgb(common_frame_ids[0])
            bop_evaluators[view] = BOPMetricEvaluator(
                mesh=mesh,
                image_shape=first_rgb.shape,
                diameter_m=diameter_m,
                symmetries=metric_symmetries,
                args=args,
            )
    vsd_taus = (
        bop_evaluators["right"].vsd_taus
        if bop_evaluators["right"] is not None
        else parse_float_list(args.vsd_taus, "--vsd_taus")
    )

    fusion_config = {
        **argparse_config_dict(args),
        "object_name": args.object_name,
        "coordinate_frame": "robot_base",
        "camera_mapping": {"right": "cam1", "left": "cam2"},
        "camera_extrinsics": {
            view: readers[view].camera_extrinsic_summary()
            for view in ("right", "left")
        },
        "gt_camera_conversion": (
            "T_cam_object_gt(frame,view) = "
            "T_cam_base_meta(frame,view) @ T_base_object_gt(frame)"
        ),
        "T_object_mesh_path": T_object_mesh_path,
        "T_object_mesh": T_object_mesh.tolist(),
        "axis_alignment_source": axis_alignment_source,
        "axis_map_mesh_to_object": axis_map_resolved,
        "global_model_calibration_from_prediction": False,
        "proxy_prediction_branch_calibration": bool(
            args.calibrate_proxy_model_from_first_frame
        ),
        "proxy_calibration_processing_order_view": (
            args.proxy_calibration_view
        ),
        "proxy_calibration_output_stem": args.proxy_calibration_output,
        "view_branch_corrections": {
            view: np.asarray(transform, dtype=np.float64).tolist()
            for view, transform in getattr(
                args, "_view_branch_corrections", {}
            ).items()
        },
        "view_branch_diagnostics": getattr(
            args, "_view_branch_diagnostics", {}
        ),
        "view_branch_correction_paths": getattr(
            args, "_view_branch_paths", {}
        ),
        "fusion_candidate_pose_source": (
            "poses_pred_base_object (both views converted through the same "
            "immutable metadata-derived T_object_mesh before fusion)"
        ),
        "fusion_optimization_frame": "fixed_GT_object_frame",
        "mesh_bbox_center_offset_after_scale_m": mesh_center_offset.tolist(),
        "mesh_scale_resolved": resolved_mesh_scale,
        "mesh_scale_source": model_context["mesh_scale_source"],
        "model_frame_mode": model_context["model_frame_mode"],
        "object_model_axis_alignment": model_context[
            "object_model_axis_alignment"
        ],
        "object_model_axis_alignment_source": model_context[
            "object_model_axis_alignment_source"
        ],
        "object_model_origin_mode": model_context[
            "object_model_origin_mode"
        ],
        "metadata_object_bbox_min_m": np.asarray(
            model_context["bbox_min_object_m"]
        ).tolist(),
        "metadata_object_bbox_max_m": np.asarray(
            model_context["bbox_max_object_m"]
        ).tolist(),
        "mesh_diameter_m": diameter_m,
        "mesh_diameter_source": diameter_source,
        "mesh_extents_m": mesh_extents_m.tolist(),
        "symmetry_axis_resolved": symmetry_axis_name,
        "metric_symmetry_source": metric_symmetry_source,
        "metric_symmetry_count": len(metric_symmetries),
        "surface_point_count": int(len(surface_points)),
        "fusion_gate": {
            "mode": args.fusion_gate_mode,
            "minimum_consistency_cues": int(
                args.fusion_min_consistency_cues
            ),
            "view_cost_threshold": float(
                args.fusion_view_pass_threshold
            ),
            "no_pass_fallback": args.fusion_no_pass_fallback,
            "max_previous_recovery_frames": int(
                args.fusion_max_recovery_frames
            ),
            "fallback_own_view_weight": float(
                args.fusion_fallback_own_view_weight
            ),
            "fallback_temporal_weight": float(
                args.fusion_fallback_temporal_weight
            ),
        },
        "fusion_rule": (
            "fixed per-view branch canonicalization -> cross-view candidate "
            "validation -> symmetry-aware mutual consistency -> joint SE(3) "
            "optimization -> cost/deviation-gated acceptance; otherwise rank "
            "the current single-view candidates without GT"
        ),
        "gt_usage": (
            "GT is loaded only after pose fusion is finalized and is never "
            "used by candidate validation, optimization, acceptance or recovery. "
            "The fixed model coordinate mapping comes only from "
            "object_pose_setup. If requested, first-frame GT calibrates only "
            "separate prediction-side view branches; no fusion decision or "
            "GT/model transform is modified per frame."
        ),
    }
    os.makedirs(fusion_root, exist_ok=True)
    with open(
        os.path.join(fusion_root, "fusion_run_config.json"),
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            to_json_safe(fusion_config),
            file,
            ensure_ascii=True,
            indent=2,
        )

    def load_candidate_pose(view: str, frame_id: int) -> Optional[np.ndarray]:
        path = os.path.join(
            candidate_pose_dirs[view], f"{frame_id:04d}.txt"
        )
        if not os.path.isfile(path):
            return None
        return validate_rigid_transform(
            np.loadtxt(path).astype(np.float64),
            f"{view} fixed-GT-object fusion candidate "
            f"frame {frame_id:04d}",
        )

    def load_gt_after_fusion(frame_id: int) -> Tuple[np.ndarray, float, float, str]:
        available = [
            view for view in ("right", "left")
            if frame_id in frame_sets[view]
        ]
        if not available:
            raise KeyError(f"No GT is available for fusion frame {frame_id:04d}")
        gt_by_view = {
            view: readers[view].get_gt_pose_base(frame_id)
            for view in available
        }
        source_view = "right" if "right" in gt_by_view else available[0]
        gt = gt_by_view[source_view]
        gt_translation_disagreement_mm = float("nan")
        gt_rotation_disagreement_deg = float("nan")
        if "right" in gt_by_view and "left" in gt_by_view:
            gt_translation_disagreement_mm = float(
                np.linalg.norm(
                    gt_by_view["right"][:3, 3] - gt_by_view["left"][:3, 3]
                ) * 1000.0
            )
            gt_rotation_disagreement_deg = rotation_distance_deg_from_matrices(
                gt_by_view["right"][:3, :3],
                gt_by_view["left"][:3, :3],
            )
            if (
                gt_translation_disagreement_mm
                > args.fusion_gt_max_translation_disagreement_mm
                or gt_rotation_disagreement_deg
                > args.fusion_gt_max_rotation_disagreement_deg
            ):
                raise ValueError(
                    f"Frame {frame_id:04d}: right/left T_base_object GT files "
                    f"disagree by {gt_translation_disagreement_mm:.3f} mm and "
                    f"{gt_rotation_disagreement_deg:.3f} degree. The fusion "
                    "evaluator will not average inconsistent GT poses."
                )
        return (
            gt,
            gt_translation_disagreement_mm,
            gt_rotation_disagreement_deg,
            source_view,
        )

    def initialize_record(
        frame_id: int,
        local_idx: int,
        right_candidate: Optional[np.ndarray],
        left_candidate: Optional[np.ndarray],
        diagnostics: Dict[str, object],
        gt_base_object: np.ndarray,
        gt_translation_disagreement_mm: float,
        gt_rotation_disagreement_deg: float,
        gt_source_view: str,
    ) -> Dict[str, object]:
        gt_quaternion = SciPyRotation.from_matrix(
            gt_base_object[:3, :3]
        ).as_quat()
        branch_calibration_frames = set(
            getattr(
                args, "_view_branch_calibration_frames", {}
            ).values()
        )
        proxy_calibration_frame = getattr(
            args, "_proxy_calibration_frame_id", None
        )
        if proxy_calibration_frame is not None:
            branch_calibration_frames.add(
                int(proxy_calibration_frame)
            )
        is_axis_calibration_frame = bool(
            frame_id in branch_calibration_frames
        )
        calibration_diag = {}
        record: Dict[str, object] = {
            "object_name": args.object_name,
            "object_id": "" if args.object_id is None else str(args.object_id),
            "camera_view": "fusion",
            "sequence": f"{args.object_name}_fusion",
            "frame_idx": int(frame_id),
            "local_idx": int(local_idx),
            "pose_source": diagnostics.get("selected_source", "failed"),
            "fusion_status": diagnostics.get("fusion_status", "failed"),
            "axis_alignment_source": axis_alignment_source,
            "axis_map_mesh_to_object": axis_map_resolved,
            "axis_calibration_frame": is_axis_calibration_frame,
            "axis_calibration_uses_gt": bool(
                args.calibrate_axis_from_first_frame
                or args.calibrate_proxy_model_from_first_frame
            ),
            "axis_calibration_best_error_deg": float(
                calibration_diag.get(
                    "best_rotation_error_deg", float("nan")
                )
            ),
            "axis_calibration_margin_deg": float(
                calibration_diag.get(
                    "selection_margin_deg", float("nan")
                )
            ),
            "official_evaluation_frame": not is_axis_calibration_frame,
            "right_candidate_available": right_candidate is not None,
            "left_candidate_available": left_candidate is not None,
            "right_candidate_pass": bool(diagnostics.get("right_pass", False)),
            "left_candidate_pass": bool(diagnostics.get("left_pass", False)),
            "candidates_mutually_consistent": bool(
                diagnostics.get("mutually_consistent", False)
            ),
            "candidate_translation_diff_mm": float(
                diagnostics.get("candidate_translation_diff_mm", float("nan"))
            ),
            "candidate_sym_rotation_diff_deg": float(
                diagnostics.get("candidate_sym_rotation_diff_deg", float("nan"))
            ),
            "left_alignment_symmetry_index": int(
                diagnostics.get("left_alignment_symmetry_index", -1)
            ),
            "joint_attempted": bool(diagnostics.get("joint_attempted", False)),
            "joint_accepted": bool(diagnostics.get("joint_accepted", False)),
            "joint_deviation_translation_mm": float(
                diagnostics.get(
                    "joint_deviation_translation_mm", float("nan")
                )
            ),
            "joint_deviation_rotation_deg": float(
                diagnostics.get(
                    "joint_deviation_rotation_deg", float("nan")
                )
            ),
            "right_selection_score": float(
                diagnostics.get("right_selection_score", float("nan"))
            ),
            "left_selection_score": float(
                diagnostics.get("left_selection_score", float("nan"))
            ),
            "previous_fused_age_frames": int(
                diagnostics.get("previous_fused_age_frames", 0)
            ),
            "initializer_weight_right": float(
                diagnostics.get("initializer_weight_right", float("nan"))
            ),
            "initializer_weight_left": float(
                diagnostics.get("initializer_weight_left", float("nan"))
            ),
            "Right_view_contribution": float(
                diagnostics.get(
                    "Right_view_contribution", float("nan")
                )
            ),
            "Left_view_contribution": float(
                diagnostics.get(
                    "Left_view_contribution", float("nan")
                )
            ),
            "best_single_source": str(
                diagnostics.get("best_single_source", "")
            ),
            "best_single_cost": float(
                diagnostics.get("best_single_cost", float("nan"))
            ),
            "joint_acceptance_limit": float(
                diagnostics.get("joint_acceptance_limit", float("nan"))
            ),
            "gt_source_view": gt_source_view,
            "gt_right_left_translation_disagreement_mm": float(
                gt_translation_disagreement_mm
            ),
            "gt_right_left_rotation_disagreement_deg": float(
                gt_rotation_disagreement_deg
            ),
            "gt_base_object_tx_m": float(gt_base_object[0, 3]),
            "gt_base_object_ty_m": float(gt_base_object[1, 3]),
            "gt_base_object_tz_m": float(gt_base_object[2, 3]),
            "gt_base_object_qx": float(gt_quaternion[0]),
            "gt_base_object_qy": float(gt_quaternion[1]),
            "gt_base_object_qz": float(gt_quaternion[2]),
            "gt_base_object_qw": float(gt_quaternion[3]),
            "metric_symmetry_source": metric_symmetry_source,
            "metric_symmetry_count": int(len(metric_symmetries)),
            "mesh_scale": resolved_mesh_scale,
            "diameter_m": float(diameter_m),
            "diameter_source": diameter_source,
        }
        for view, candidate in (
            ("right", right_candidate),
            ("left", left_candidate),
        ):
            if candidate is None:
                continue
            candidate_quaternion = SciPyRotation.from_matrix(
                candidate[:3, :3]
            ).as_quat()
            record[f"{view}_candidate_base_object_tx_m"] = float(
                candidate[0, 3]
            )
            record[f"{view}_candidate_base_object_ty_m"] = float(
                candidate[1, 3]
            )
            record[f"{view}_candidate_base_object_tz_m"] = float(
                candidate[2, 3]
            )
            record[f"{view}_candidate_base_object_qx"] = float(
                candidate_quaternion[0]
            )
            record[f"{view}_candidate_base_object_qy"] = float(
                candidate_quaternion[1]
            )
            record[f"{view}_candidate_base_object_qz"] = float(
                candidate_quaternion[2]
            )
            record[f"{view}_candidate_base_object_qw"] = float(
                candidate_quaternion[3]
            )
        _flatten_fusion_evaluation(
            record, "right_candidate", diagnostics.get("right_evaluation")
        )
        _flatten_fusion_evaluation(
            record, "left_candidate", diagnostics.get("left_evaluation")
        )
        _flatten_fusion_evaluation(
            record, "joint_candidate", diagnostics.get("joint_evaluation")
        )
        _flatten_fusion_evaluation(
            record, "selected_pose", diagnostics.get("selected_evaluation")
        )
        _flatten_fusion_evaluation(
            record, "recovery_candidate", diagnostics.get("recovery_evaluation")
        )
        optimizer = diagnostics.get("optimizer") or {}
        for key in (
            "success", "status", "message", "nfev", "nit", "objective",
            "delta_tx_mm", "delta_ty_mm", "delta_tz_mm",
            "delta_rx_deg", "delta_ry_deg", "delta_rz_deg",
        ):
            record[f"optimizer_{key}"] = optimizer.get(
                key,
                False if key == "success" else (
                    "" if key == "message" else float("nan")
                ),
            )
        return record

    previous_fused_pose: Optional[np.ndarray] = None
    previous_fused_age_frames = 0
    previous_fused_axis_object_pose: Optional[np.ndarray] = None
    previous_fused_axis_object_pose2: Optional[np.ndarray] = None
    previous_metric_pred: Optional[np.ndarray] = None
    previous_metric_gt: Optional[np.ndarray] = None
    records: List[Dict] = []
    video_writer = None
    video_path = os.path.join(
        fusion_root, f"{args.object_name}_two_view_fusion.mp4"
    )
    stop_requested = False

    print(f"\n[Fusion] {args.object_name}: right/cam1 + left/cam2")
    print(
        f"[Fusion] frames union={len(all_frame_ids)}, common={len(common_frame_ids)}, "
        f"surface_points={len(surface_points)}"
    )
    print(
        f"[Fusion gates] mode={args.fusion_gate_mode}, "
        f"view_cost<={args.fusion_view_pass_threshold}, "
        f"cues>={args.fusion_min_consistency_cues}, "
        f"depth_tol={args.fusion_depth_tolerance_m * 1000.0:.1f} mm, "
        f"point_tol={args.fusion_point_tolerance_m * 1000.0:.1f} mm, "
        f"candidate_dt<={args.fusion_max_candidate_translation_mm} mm, "
        f"candidate_dR<={args.fusion_max_candidate_rotation_deg} degree"
    )
    print(
        "[Fusion contribution] "
        "omega_v=exp(-(E_v-E_min)/tau)/sum(exp(-(E-E_min)/tau)), "
        f"tau={args.fusion_weight_temperature:g}"
    )
    print(
        f"[Fusion fallback] no-pass={args.fusion_no_pass_fallback}, "
        f"previous_max_age={args.fusion_max_recovery_frames}, "
        f"own_view_weight={args.fusion_fallback_own_view_weight:g}, "
        f"temporal_weight={args.fusion_fallback_temporal_weight:g}"
    )

    for local_idx, frame_id in enumerate(
        tqdm(all_frame_ids, desc=f"Fusion {args.object_name}")
    ):
        right_candidate = load_candidate_pose("right", frame_id)
        left_candidate = load_candidate_pose("left", frame_id)
        observations: Optional[Dict[str, Dict[str, object]]] = None
        diagnostics: Dict[str, object]
        fused_pose_base_object_selected: Optional[np.ndarray]

        if frame_id not in frame_sets["right"] or frame_id not in frame_sets["left"]:
            missing_view = (
                "right" if frame_id not in frame_sets["right"] else "left"
            )
            diagnostics = {
                "right_pass": False,
                "left_pass": False,
                "mutually_consistent": False,
                "selected_source": f"unsynchronized_missing_{missing_view}",
                "fusion_status": "failed",
            }
            fused_pose_base_object_selected = None
        else:
            observations = {
                view: make_fusion_observation(
                    reader=readers[view],
                    frame_id=frame_id,
                    view_root=view_roots[view],
                    args=args,
                )
                for view in ("right", "left")
            }
            (
                fused_pose_base_object_selected,
                diagnostics,
            ) = fuse_pose_candidates(
                right_pose_base_mesh=right_candidate,
                left_pose_base_mesh=left_candidate,
                previous_fused_pose_base_mesh=previous_fused_pose,
                observations=observations,
                surface_points_mesh=surface_points_object,
                bbox_corners_mesh=fusion_bbox_corners_object,
                symmetries=fusion_symmetries_object,
                diameter_m=diameter_m,
                args=args,
                previous_fused_age_frames=previous_fused_age_frames,
            )

        selected_source = str(
            diagnostics.get("selected_source", "")
        )
        if selected_source == "previous_pose_recovery":
            previous_fused_age_frames += 1
        elif fused_pose_base_object_selected is not None:
            previous_fused_age_frames = 0
        else:
            previous_fused_age_frames += 1

        # GT is deliberately loaded only after the fusion decision above.
        (
            gt_base_object,
            gt_t_disagreement_mm,
            gt_r_disagreement_deg,
            gt_source_view,
        ) = load_gt_after_fusion(frame_id)
        gt_base_mesh = gt_base_object @ T_object_mesh
        record = initialize_record(
            frame_id=frame_id,
            local_idx=local_idx,
            right_candidate=right_candidate,
            left_candidate=left_candidate,
            diagnostics=diagnostics,
            gt_base_object=gt_base_object,
            gt_translation_disagreement_mm=gt_t_disagreement_mm,
            gt_rotation_disagreement_deg=gt_r_disagreement_deg,
            gt_source_view=gt_source_view,
        )

        if observations is not None:
            record["right_mask_source"] = observations["right"]["mask_source"]
            record["left_mask_source"] = observations["left"]["mask_source"]
            record["right_mask_pixels"] = int(
                observations["right"]["mask"].sum()
                if observations["right"]["mask"] is not None else 0
            )
            record["left_mask_pixels"] = int(
                observations["left"]["mask"].sum()
                if observations["left"]["mask"] is not None else 0
            )

        if fused_pose_base_object_selected is None:
            # Do not report a relative pose error across a failed-frame gap.
            previous_metric_pred = None
            previous_metric_gt = None
            ar_failure = float("nan") if args.disable_bop_metrics else 0.0
            record.update({
                "estimate_valid": False,
                "metric_status": "not_evaluated_fusion_failure",
                "ADD_mm": float("nan"),
                "ADDS_mm": float("nan"),
                "ADD_0.1d_success": False,
                "ADDS_0.1d_success": False,
                "Rotation_error_deg": float("nan"),
                "Symmetry_aware_rotation_error_deg": float("nan"),
                "Translation_error_mm": float("nan"),
                "Translation_error_base_mm": float("nan"),
                "Translation_error_mesh_origin_mm": float("nan"),
                "Mean_projection_error_px": float("nan"),
                "Mean_projection_error_right_px": float("nan"),
                "Mean_projection_error_left_px": float("nan"),
                "5deg_5cm_success": False,
                "RPE_translation_mm": float("nan"),
                "RPE_rotation_deg": float("nan"),
                "VSD_error_mean": float("nan"),
                "VSD_AR_contribution": ar_failure,
                "MSSD_mm": float("nan"),
                "MSSD_norm_d": float("nan"),
                "MSSD_AR_contribution": ar_failure,
                "MSPD_px": float("nan"),
                "MSPD_AR_contribution": ar_failure,
                "BOP_AR_contribution": ar_failure,
                **{
                    f"VSD_tau_{tau:.2f}": float("nan")
                    for tau in vsd_taus
                },
            })
            records.append(record)
            if (
                args.metrics_flush_interval > 0
                and len(records) % args.metrics_flush_interval == 0
            ):
                save_fusion_metrics_workbook(
                    right_records=view_records_by_view.get("right", []),
                    left_records=view_records_by_view.get("left", []),
                    fusion_records=records,
                    output_dir=fusion_root,
                )
            continue

        # Fusion is performed directly in the one fixed GT object coordinate
        # frame. Convert to the estimator mesh frame only after the decision,
        # for ADD/BOP rendering and mesh-pose file compatibility.
        fused_pose_base_object = ensure_pose_matrix(
            fused_pose_base_object_selected,
            f"fused GT-object pose frame {frame_id:04d}",
        )
        previous_fused_pose = fused_pose_base_object.copy()
        fused_pose_base_mesh = (
            fused_pose_base_object @ T_object_mesh
        )
        fused_pose_base_object_raw = fused_pose_base_object.copy()
        fused_axis_symmetry_angle_deg = 0.0
        fused_axis_symmetry_flipped = False
        fused_axis_symmetry_reference_error_deg = 0.0
        fused_axis_object_reference = predict_next_pose_constant_velocity(
            previous_fused_axis_object_pose2,
            previous_fused_axis_object_pose,
        )
        fused_temporal_axis_map = "fusion_in_fixed_GT_object_frame"
        fused_temporal_axis_correction_deg = 0.0
        fused_temporal_axis_reference_error_deg = (
            0.0 if fused_axis_object_reference is None else
            rotation_distance_deg_from_matrices(
                fused_pose_base_object[:3, :3],
                fused_axis_object_reference[:3, :3],
            )
        )
        previous_fused_axis_object_pose2 = (
            None if previous_fused_axis_object_pose is None
            else previous_fused_axis_object_pose.copy()
        )
        previous_fused_axis_object_pose = (
            fused_pose_base_object.copy()
        )

        T_cam_base_by_view = {
            view: np.asarray(
                observations[view]["T_cam_base"], dtype=np.float64
            )
            for view in ("right", "left")
        }
        fused_pose_cam = {
            view: T_cam_base_by_view[view] @ fused_pose_base_mesh
            for view in ("right", "left")
        }
        fused_pose_cam_object = {
            view: T_cam_base_by_view[view] @ fused_pose_base_object
            for view in ("right", "left")
        }
        # Required GT chain, evaluated independently for each camera/frame.
        gt_pose_cam_object = {
            view: T_cam_base_by_view[view] @ gt_base_object
            for view in ("right", "left")
        }
        gt_pose_cam = {
            view: gt_pose_cam_object[view] @ T_object_mesh
            for view in ("right", "left")
        }

        rotation_error = rotation_error_deg(
            fused_pose_base_object, gt_base_object
        )
        raw_rotation_error = rotation_error_deg(
            fused_pose_base_object_raw, gt_base_object
        )
        symmetry_rotation_error = symmetry_aware_rotation_error_deg(
            fused_pose_base_mesh, gt_base_mesh, metric_symmetries
        )
        translation_error_object_mm = float(
            np.linalg.norm(
                fused_pose_base_object[:3, 3] - gt_base_object[:3, 3]
            ) * 1000.0
        )
        translation_error_mesh_mm = float(
            np.linalg.norm(
                fused_pose_base_mesh[:3, 3] - gt_base_mesh[:3, 3]
            ) * 1000.0
        )
        add_m = add_error_m(
            add_vertices, fused_pose_base_mesh, gt_base_mesh
        )
        adds_m = add_s_error_m(
            add_vertices, fused_pose_base_mesh, gt_base_mesh
        )
        projection_errors = {
            view: mean_projection_error_px(
                add_vertices,
                fused_pose_cam[view],
                gt_pose_cam[view],
                readers[view].K,
            )
            for view in ("right", "left")
        }
        rpe_t_mm, rpe_r_deg = relative_pose_errors(
            previous_metric_pred,
            fused_pose_base_object,
            previous_metric_gt,
            gt_base_object,
        )
        previous_metric_pred = fused_pose_base_object.copy()
        previous_metric_gt = gt_base_object.copy()

        ar_failure = float("nan") if args.disable_bop_metrics else 0.0
        bop_results = {
            view: _blank_bop_result(len(vsd_taus), ar_failure)
            for view in ("right", "left")
        }
        metric_status_by_view = {"right": "disabled", "left": "disabled"}
        if not args.disable_bop_metrics:
            for view in ("right", "left"):
                try:
                    bop_results[view] = bop_evaluators[view].evaluate(
                        pred_pose_cam_mesh=fused_pose_cam[view],
                        gt_pose_cam_mesh=gt_pose_cam[view],
                        depth_m=observations[view]["depth"],
                        K=readers[view].K,
                        metric_vertices_m=bop_vertices,
                    )
                    metric_status_by_view[view] = "ok"
                except Exception as exc:
                    metric_status_by_view[view] = (
                        f"error:{type(exc).__name__}:{exc}"
                    )
                    if args.metric_failure_policy == "raise":
                        raise
                    print(
                        f"[Warning] Fused BOP metrics failed for {view} "
                        f"frame {frame_id:04d}: {exc!r}"
                    )

        vsd_errors_by_view = {
            view: np.asarray(
                bop_results[view]["VSD_errors"], dtype=np.float64
            ).reshape(-1)
            for view in ("right", "left")
        }
        vsd_mean_errors = np.asarray([
            _mean_finite([
                vsd_errors_by_view["right"][index],
                vsd_errors_by_view["left"][index],
            ])
            for index in range(len(vsd_taus))
        ], dtype=np.float64)
        vsd_columns = {}
        for index, tau in enumerate(vsd_taus):
            vsd_columns[f"VSD_tau_{tau:.2f}"] = float(
                vsd_mean_errors[index]
            )
            vsd_columns[f"VSD_right_tau_{tau:.2f}"] = float(
                vsd_errors_by_view["right"][index]
            )
            vsd_columns[f"VSD_left_tau_{tau:.2f}"] = float(
                vsd_errors_by_view["left"][index]
            )

        base_residual_mm = (
            fused_pose_base_object[:3, 3] - gt_base_object[:3, 3]
        ) * 1000.0
        fused_object_quaternion = SciPyRotation.from_matrix(
            fused_pose_base_object[:3, :3]
        ).as_quat()
        fused_mesh_quaternion = SciPyRotation.from_matrix(
            fused_pose_base_mesh[:3, :3]
        ).as_quat()
        if args.disable_bop_metrics:
            metric_status = "disabled"
        elif all(
            value == "ok" for value in metric_status_by_view.values()
        ):
            metric_status = "ok"
        else:
            metric_status = "|".join(
                f"{view}:{metric_status_by_view[view]}"
                for view in ("right", "left")
            )
        mssd_m = _mean_finite([
            bop_results["right"]["MSSD_m"],
            bop_results["left"]["MSSD_m"],
        ])
        record.update({
            "estimate_valid": True,
            "metric_status": metric_status,
            "ADD_mm": float(add_m * 1000.0),
            "ADDS_mm": float(adds_m * 1000.0),
            "ADD_0.1d_success": bool(add_m < 0.1 * diameter_m),
            "ADDS_0.1d_success": bool(adds_m < 0.1 * diameter_m),
            "Rotation_error_deg": float(rotation_error),
            "Raw_rotation_error_deg": float(raw_rotation_error),
            "Symmetry_aware_rotation_error_deg": float(
                symmetry_rotation_error
            ),
            "Translation_error_mm": translation_error_object_mm,
            "Translation_error_base_mm": translation_error_object_mm,
            "Translation_error_mesh_origin_mm": translation_error_mesh_mm,
            "Mean_projection_error_px": _mean_finite(
                list(projection_errors.values())
            ),
            "Mean_projection_error_right_px": float(
                projection_errors["right"]
            ),
            "Mean_projection_error_left_px": float(
                projection_errors["left"]
            ),
            "5deg_5cm_success": bool(
                rotation_error < 5.0
                and translation_error_object_mm < 50.0
            ),
            "RPE_translation_mm": float(rpe_t_mm),
            "RPE_rotation_deg": float(rpe_r_deg),
            "VSD_error_mean": _mean_finite([
                bop_results["right"]["VSD_error_mean"],
                bop_results["left"]["VSD_error_mean"],
            ]),
            "VSD_AR_contribution": _mean_finite([
                bop_results["right"]["VSD_AR_contribution"],
                bop_results["left"]["VSD_AR_contribution"],
            ]),
            "VSD_AR_right": float(
                bop_results["right"]["VSD_AR_contribution"]
            ),
            "VSD_AR_left": float(
                bop_results["left"]["VSD_AR_contribution"]
            ),
            "MSSD_mm": float(mssd_m * 1000.0),
            "MSSD_norm_d": float(mssd_m / max(diameter_m, 1e-12)),
            "MSSD_AR_contribution": _mean_finite([
                bop_results["right"]["MSSD_AR_contribution"],
                bop_results["left"]["MSSD_AR_contribution"],
            ]),
            "MSPD_px": _mean_finite([
                bop_results["right"]["MSPD_px"],
                bop_results["left"]["MSPD_px"],
            ]),
            "MSPD_right_px": float(bop_results["right"]["MSPD_px"]),
            "MSPD_left_px": float(bop_results["left"]["MSPD_px"]),
            "MSPD_AR_contribution": _mean_finite([
                bop_results["right"]["MSPD_AR_contribution"],
                bop_results["left"]["MSPD_AR_contribution"],
            ]),
            "MSPD_AR_right": float(
                bop_results["right"]["MSPD_AR_contribution"]
            ),
            "MSPD_AR_left": float(
                bop_results["left"]["MSPD_AR_contribution"]
            ),
            "BOP_AR_contribution": _mean_finite([
                bop_results["right"]["BOP_AR_contribution"],
                bop_results["left"]["BOP_AR_contribution"],
            ]),
            "BOP_AR_right": float(
                bop_results["right"]["BOP_AR_contribution"]
            ),
            "BOP_AR_left": float(
                bop_results["left"]["BOP_AR_contribution"]
            ),
            **vsd_columns,
            "Base_dx_mm": float(base_residual_mm[0]),
            "Base_dy_mm": float(base_residual_mm[1]),
            "Base_dz_mm": float(base_residual_mm[2]),
            "pred_base_mesh_tx_m": float(fused_pose_base_mesh[0, 3]),
            "pred_base_mesh_ty_m": float(fused_pose_base_mesh[1, 3]),
            "pred_base_mesh_tz_m": float(fused_pose_base_mesh[2, 3]),
            "pred_base_mesh_qx": float(fused_mesh_quaternion[0]),
            "pred_base_mesh_qy": float(fused_mesh_quaternion[1]),
            "pred_base_mesh_qz": float(fused_mesh_quaternion[2]),
            "pred_base_mesh_qw": float(fused_mesh_quaternion[3]),
            "pred_base_object_tx_m": float(fused_pose_base_object[0, 3]),
            "pred_base_object_ty_m": float(fused_pose_base_object[1, 3]),
            "pred_base_object_tz_m": float(fused_pose_base_object[2, 3]),
            "pred_base_object_qx": float(fused_object_quaternion[0]),
            "pred_base_object_qy": float(fused_object_quaternion[1]),
            "pred_base_object_qz": float(fused_object_quaternion[2]),
            "pred_base_object_qw": float(fused_object_quaternion[3]),
            "pred_cam_right_tx_m": float(fused_pose_cam["right"][0, 3]),
            "pred_cam_right_ty_m": float(fused_pose_cam["right"][1, 3]),
            "pred_cam_right_tz_m": float(fused_pose_cam["right"][2, 3]),
            "pred_cam_left_tx_m": float(fused_pose_cam["left"][0, 3]),
            "pred_cam_left_ty_m": float(fused_pose_cam["left"][1, 3]),
            "pred_cam_left_tz_m": float(fused_pose_cam["left"][2, 3]),
            "temporal_axis_map": fused_temporal_axis_map,
            "temporal_axis_correction_deg": float(
                fused_temporal_axis_correction_deg
            ),
            "temporal_axis_reference_error_deg": float(
                fused_temporal_axis_reference_error_deg
            ),
            "axis_symmetry_correction_deg": float(
                fused_axis_symmetry_angle_deg
            ),
            "axis_symmetry_flipped": bool(
                fused_axis_symmetry_flipped
            ),
            "axis_symmetry_reference_error_deg": float(
                fused_axis_symmetry_reference_error_deg
            ),
        })

        np.savetxt(
            os.path.join(fused_base_mesh_dir, f"{frame_id:04d}.txt"),
            fused_pose_base_mesh,
        )
        np.savetxt(
            os.path.join(fused_base_object_dir, f"{frame_id:04d}.txt"),
            fused_pose_base_object,
        )
        for view in ("right", "left"):
            np.savetxt(
                os.path.join(
                    fused_camera_dirs[view], f"{frame_id:04d}.txt"
                ),
                fused_pose_cam[view],
            )
            np.savetxt(
                os.path.join(
                    fused_camera_object_dirs[view],
                    f"{frame_id:04d}.txt",
                ),
                fused_pose_cam_object[view],
            )

        if observations is not None:
            visualizations = {}
            for view in ("right", "left"):
                vis = cv2.cvtColor(
                    observations[view]["rgb"], cv2.COLOR_RGB2BGR
                )
                draw_projected_bbox(
                    vis,
                    object_bbox_corners,
                    gt_pose_cam_object[view],
                    readers[view].K,
                    color=(0, 255, 0), thickness=3,
                )
                draw_projected_bbox(
                    vis,
                    object_bbox_corners,
                    fused_pose_cam_object[view],
                    readers[view].K,
                    color=(255, 0, 255), thickness=2,
                )
                if args.draw_pose_axes:
                    axis_length = max(
                        diameter_m * args.axis_ratio,
                        args.min_axis_length_m,
                    )
                    draw_pose_rays_monocolor(
                        vis, gt_pose_cam_object[view], readers[view].K,
                        axis_length, color=(0, 255, 0), thickness=3,
                    )
                    draw_pose_rays_monocolor(
                        vis, fused_pose_cam_object[view], readers[view].K,
                        axis_length, color=(255, 0, 255), thickness=2,
                    )
                if args.draw_origin_gap:
                    draw_origin_gap(
                        vis,
                        gt_pose_cam_object[view],
                        fused_pose_cam_object[view],
                        readers[view].K,
                    )
                if args.show_mask and observations[view]["mask"] is not None:
                    draw_mask_contour(vis, observations[view]["mask"])
                if args.show_overlay:
                    draw_text_panel(
                        vis,
                        [
                            f"{view} | frame {frame_id:04d}",
                            f"green: GT | magenta: fused",
                            f"source: {record['pose_source']}",
                            (
                                "contribution R/L: "
                                f"{record['Right_view_contribution']:.3f}/"
                                f"{record['Left_view_contribution']:.3f}"
                            ),
                        ],
                        origin=(12, 24),
                        line_height=22,
                    )
                visualizations[view] = vis
                if args.save_images:
                    cv2.imwrite(
                        os.path.join(
                            vis_dirs[view], f"{frame_id:04d}.png"
                        ),
                        vis,
                    )

            target_height = min(
                visualizations["right"].shape[0],
                visualizations["left"].shape[0],
            )
            side_by_side = []
            for view in ("right", "left"):
                image = visualizations[view]
                if image.shape[0] != target_height:
                    scale = target_height / image.shape[0]
                    image = cv2.resize(
                        image, None, fx=scale, fy=scale,
                        interpolation=cv2.INTER_LINEAR,
                    )
                side_by_side.append(image)
            combined_vis = np.concatenate(side_by_side, axis=1)

            if args.save_video:
                frame_for_video = combined_vis
                if abs(args.video_scale - 1.0) > 1e-9:
                    frame_for_video = cv2.resize(
                        combined_vis,
                        None,
                        fx=args.video_scale,
                        fy=args.video_scale,
                        interpolation=cv2.INTER_LINEAR,
                    )
                if video_writer is None:
                    height, width = frame_for_video.shape[:2]
                    video_writer = cv2.VideoWriter(
                        video_path,
                        cv2.VideoWriter_fourcc(*"mp4v"),
                        args.fps,
                        (width, height),
                    )
                    if not video_writer.isOpened():
                        raise RuntimeError(
                            f"Cannot open fusion video writer: {video_path}"
                        )
                video_writer.write(frame_for_video)

            if not args.no_display:
                display = combined_vis
                if abs(args.display_scale - 1.0) > 1e-9:
                    display = cv2.resize(
                        combined_vis,
                        None,
                        fx=args.display_scale,
                        fy=args.display_scale,
                        interpolation=cv2.INTER_LINEAR,
                    )
                cv2.imshow(
                    f"MultiView two-view fusion - {args.object_name}",
                    display,
                )
                key = cv2.waitKey(args.wait_ms) & 0xFF
                if key in (ord("q"), 27):
                    stop_requested = True

        records.append(record)
        if (
            args.metrics_flush_interval > 0
            and len(records) % args.metrics_flush_interval == 0
        ):
            xlsx_path = save_fusion_metrics_workbook(
                right_records=view_records_by_view.get("right", []),
                left_records=view_records_by_view.get("left", []),
                fusion_records=records,
                output_dir=fusion_root,
            )
            print(f"[Fusion checkpoint] {xlsx_path}")
        if stop_requested:
            break

    if video_writer is not None:
        video_writer.release()
        print(f"[Saved] fusion video: {video_path}")

    if records:
        xlsx_path = save_fusion_metrics_workbook(
            right_records=view_records_by_view.get("right", []),
            left_records=view_records_by_view.get("left", []),
            fusion_records=records,
            output_dir=fusion_root,
        )
        print(
            "[Saved] three-sheet fusion Excel "
            f"(right / left / fusion): {xlsx_path}"
        )
        print_fusion_decision_summary(records)
    return records


def parse_views(view_arg: str) -> List[str]:
    view_arg = view_arg.lower().strip()
    if view_arg == "both":
        return ["left", "right"]
    if view_arg in ["left", "right"]:
        return [view_arg]
    raise ValueError("--view must be one of: left, right, both")


def resolve_single_object_name(args) -> None:
    requested = str(getattr(args, "object_name", "") or "").strip()
    if requested.lower() in {"", "auto", "infer"}:
        args.object_name = infer_object_name_from_dataset(
            args.dataset_root, args.view
        )
        print(
            f"[Dataset] inferred object_name={args.object_name!r} from "
            f"{os.path.join(os.path.abspath(args.dataset_root), 'evaluation')}"
        )


def build_object_runs(args) -> List[argparse.Namespace]:
    """Build one default object run or expand an optional legacy JSON batch.

    Example:
      {
        "objects": [
          {
            "object_name": "battery",
            "obj_mesh": "models/battery/final_mesh.obj",
            "object_id": 1,
            "model_diameter_m": 0.128,
            "metric_symmetry": "none"
          }
        ]
      }
    """
    if args.objects_config is None:
        if not args.obj_mesh:
            raise ValueError("--obj_mesh is required unless --objects_config is used")
        resolve_single_object_name(args)
        return [args]

    config_path = os.path.abspath(args.objects_config)
    with open(config_path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    objects = payload.get("objects") if isinstance(payload, dict) else payload
    if not isinstance(objects, list) or not objects:
        raise ValueError("--objects_config must contain a non-empty 'objects' list")

    allowed = {
        "object_name", "obj_mesh", "mesh_scale", "model_frame_mode",
        "model_bbox_scale_tolerance", "dataset_root", "view",
        "T_object_mesh", "axis_map", "object_mesh_translation_scale",
        "calibrate_axis_from_first_frame", "axis_calibration_view",
        "axis_calibration_output", "axis_calibration_mode",
        "view_branch_calibration_mode",
        "right_pose_branch_correction", "left_pose_branch_correction",
        "calibrate_proxy_model_from_first_frame",
        "proxy_calibration_view", "proxy_calibration_output",
        "proxy_calibration_iteration", "proxy_calibration_max_points",
        "proxy_calibration_add_weight",
        "proxy_calibration_adds_weight",
        "proxy_calibration_min_improvement_mm",
        "proxy_calibration_max_add_ratio",
        "proxy_calibration_fail_unreliable",
        "model_diameter_m", "models_info", "object_id",
        "models_info_translation_scale", "metric_symmetry",
        "symmetry_axis", "symmetry_samples", "symmetry_allow_axis_flip",
        "symmetry_stabilization", "temporal_axis_stabilization",
        "temporal_axis_min_improvement_deg",
        "temporal_branch_symmetry_mean_ratio",
        "temporal_branch_symmetry_max_ratio",
        "temporal_branch_symmetry_min_mean_mm",
        "temporal_branch_symmetry_min_max_mm",
        "left_first_bbox", "right_first_bbox",
        "use_manual_first_bbox",
        "tracking_profile", "enable_tracking_health_check",
        "register_once_then_track",
        "slender_tool_reinit_interval", "slender_tool_iteration",
        "tracking_health_max_surface_points",
        "tracking_health_min_projected_points",
        "tracking_health_min_compared_points",
        "tracking_health_max_geometry_cost",
        "tracking_health_min_depth_inlier_ratio",
        "tracking_health_min_inside_mask_ratio",
        "tracking_health_min_mask_bbox_iou",
        "tracking_health_axis_min_anisotropy",
        "tracking_health_max_axis_difference_deg",
        "tracking_health_max_translation_jump_mm",
        "tracking_health_max_rotation_jump_deg",
        "tracking_health_max_recovery_frames",
        "tracking_health_temporal_weight",
        "tracking_health_invalid_penalty",
        "enable_fusion", "fusion_failure_policy",
        "fusion_no_pass_fallback", "fusion_max_recovery_frames",
        "fusion_fallback_own_view_weight",
        "fusion_fallback_temporal_weight",
        "fusion_fallback_translation_scale_mm",
        "fusion_fallback_rotation_scale_deg",
        "fusion_fallback_invalid_penalty",
        "fusion_tie_break_view",
        "fusion_gate_mode", "fusion_min_consistency_cues",
        "fusion_view_pass_threshold",
        "fusion_max_candidate_translation_mm",
        "fusion_max_candidate_rotation_deg",
        "fusion_depth_tolerance_m",
        "fusion_occlusion_tolerance_m",
        "fusion_point_tolerance_m",
        "fusion_min_depth_inlier_ratio",
        "fusion_min_mask_bbox_iou",
        "fusion_min_projected_inside_mask_ratio",
        "fusion_min_point_inlier_ratio",
        "fusion_joint_accept_relative_margin",
        "fusion_joint_accept_absolute_margin",
        "fusion_joint_max_deviation_translation_mm",
        "fusion_joint_max_deviation_rotation_deg",
        "fusion_recovery_threshold_factor",
        "fusion_weight_temperature",
    }
    path_fields = {
        "obj_mesh", "dataset_root", "T_object_mesh", "models_info",
        "axis_calibration_output", "proxy_calibration_output",
        "right_pose_branch_correction",
        "left_pose_branch_correction",
    }
    config_dir = os.path.dirname(config_path)
    runs = []
    names = set()
    for index, item in enumerate(objects):
        if not isinstance(item, dict):
            raise TypeError(f"objects[{index}] must be a JSON object")
        item = dict(item)
        if "name" in item and "object_name" not in item:
            item["object_name"] = item.pop("name")
        if "mesh_path" in item and "obj_mesh" not in item:
            item["obj_mesh"] = item.pop("mesh_path")
        if "diameter_m" in item and "model_diameter_m" not in item:
            item["model_diameter_m"] = item.pop("diameter_m")
        symmetry = item.pop("symmetry", None)
        if symmetry is not None:
            if not isinstance(symmetry, dict):
                raise TypeError(f"objects[{index}].symmetry must be an object")
            mapping = {
                "mode": "metric_symmetry",
                "axis": "symmetry_axis",
                "samples": "symmetry_samples",
                "allow_axis_flip": "symmetry_allow_axis_flip",
                "stabilize": "symmetry_stabilization",
            }
            for key, value in symmetry.items():
                target = mapping.get(key, key)
                item.setdefault(target, value)

        unknown = sorted(set(item) - allowed)
        if unknown:
            raise ValueError(f"Unsupported keys in objects[{index}]: {unknown}")

        run_args = copy.deepcopy(args)
        for key, value in item.items():
            if key in path_fields and value is not None and not os.path.isabs(str(value)):
                value = os.path.abspath(os.path.join(config_dir, str(value)))
            setattr(run_args, key, value)
        resolve_single_object_name(run_args)
        if not run_args.object_name or not run_args.obj_mesh:
            raise ValueError(f"objects[{index}] requires object_name and obj_mesh")
        if run_args.object_name in names:
            raise ValueError(f"Duplicate object_name in --objects_config: {run_args.object_name}")
        names.add(run_args.object_name)
        runs.append(run_args)
    return runs


def validate_run_args(args) -> None:
    if str(args.object_name).strip().lower() in {"", "auto", "infer"}:
        raise ValueError(
            "--object_name could not be inferred. Pass it explicitly."
        )
    if args.axis_map not in (None, "", "none"):
        parse_axis_map(args.axis_map)
    if (
        args.axis_calibration_output is not None
        and not args.calibrate_axis_from_first_frame
    ):
        raise ValueError(
            "--axis_calibration_output requires "
            "--calibrate_axis_from_first_frame."
        )
    if (
        args.proxy_calibration_output is not None
        and not args.calibrate_proxy_model_from_first_frame
    ):
        raise ValueError(
            "--proxy_calibration_output requires "
            "--calibrate_proxy_model_from_first_frame."
        )
    if (
        args.calibrate_proxy_model_from_first_frame
        and args.calibrate_axis_from_first_frame
    ):
        raise ValueError(
            "--calibrate_proxy_model_from_first_frame already calibrates one "
            "prediction-only branch for every requested view. Do not combine "
            "it with --calibrate_axis_from_first_frame."
        )
    if (
        args.calibrate_proxy_model_from_first_frame
        and args.proxy_calibration_view not in parse_views(args.view)
    ):
        raise ValueError(
            "--proxy_calibration_view must be included in --view."
        )
    proxy_positive = {
        "proxy_calibration_iteration":
            args.proxy_calibration_iteration,
        "proxy_calibration_max_points":
            args.proxy_calibration_max_points,
        "proxy_calibration_max_add_ratio":
            args.proxy_calibration_max_add_ratio,
    }
    invalid_proxy_positive = [
        name for name, value in proxy_positive.items()
        if float(value) <= 0
    ]
    if invalid_proxy_positive:
        raise ValueError(
            "These proxy calibration parameters must be positive: "
            + ", ".join(invalid_proxy_positive)
        )
    if (
        float(args.proxy_calibration_add_weight) < 0
        or float(args.proxy_calibration_adds_weight) < 0
        or (
            float(args.proxy_calibration_add_weight)
            + float(args.proxy_calibration_adds_weight)
            <= 0
        )
    ):
        raise ValueError(
            "Proxy calibration ADD/ADD-S weights must be non-negative and "
            "at least one must be positive."
        )
    if float(args.proxy_calibration_min_improvement_mm) < 0:
        raise ValueError(
            "--proxy_calibration_min_improvement_mm must be non-negative."
        )
    if abs(float(args.gt_translation_scale) - 1.0) > 1e-12:
        raise ValueError(
            "gt_pose translation is already stored in metres. "
            "Use --gt_translation_scale 1.0. Applying 0.001 would collapse "
            "robot-base poses and invalidate the camera-frame GT."
        )
    if args.running_stride < 1:
        raise ValueError("--running_stride must be >= 1")
    if args.symmetry_samples < 4:
        raise ValueError("--symmetry_samples must be >= 4")
    if args.temporal_axis_min_improvement_deg < 0:
        raise ValueError(
            "--temporal_axis_min_improvement_deg must be >= 0"
        )
    if args.model_bbox_scale_tolerance <= 1.0:
        raise ValueError("--model_bbox_scale_tolerance must be > 1")
    if (
        args.temporal_branch_symmetry_mean_ratio < 0
        or args.temporal_branch_symmetry_max_ratio < 0
        or args.temporal_branch_symmetry_min_mean_mm < 0
        or args.temporal_branch_symmetry_min_max_mm < 0
    ):
        raise ValueError(
            "Temporal branch-symmetry tolerances must be non-negative."
        )
    if args.start_frame < 0 or args.reinit_interval < 0:
        raise ValueError("--start_frame and --reinit_interval must be >= 0")
    tracking_positive = {
        "slender_tool_reinit_interval": args.slender_tool_reinit_interval,
        "slender_tool_iteration": args.slender_tool_iteration,
        "tracking_health_max_surface_points":
            args.tracking_health_max_surface_points,
        "tracking_health_min_projected_points":
            args.tracking_health_min_projected_points,
        "tracking_health_min_compared_points":
            args.tracking_health_min_compared_points,
        "tracking_health_max_geometry_cost":
            args.tracking_health_max_geometry_cost,
        "tracking_health_max_translation_jump_mm":
            args.tracking_health_max_translation_jump_mm,
        "tracking_health_max_rotation_jump_deg":
            args.tracking_health_max_rotation_jump_deg,
        "tracking_health_max_axis_difference_deg":
            args.tracking_health_max_axis_difference_deg,
        "tracking_health_invalid_penalty":
            args.tracking_health_invalid_penalty,
    }
    invalid_tracking_positive = [
        name for name, value in tracking_positive.items()
        if float(value) <= 0
    ]
    if invalid_tracking_positive:
        raise ValueError(
            "These tracking parameters must be positive: "
            + ", ".join(invalid_tracking_positive)
        )
    if int(args.tracking_health_max_recovery_frames) < 0:
        raise ValueError(
            "--tracking_health_max_recovery_frames must be >= 0"
        )
    tracking_ratios = {
        "tracking_health_min_depth_inlier_ratio":
            args.tracking_health_min_depth_inlier_ratio,
        "tracking_health_min_inside_mask_ratio":
            args.tracking_health_min_inside_mask_ratio,
        "tracking_health_min_mask_bbox_iou":
            args.tracking_health_min_mask_bbox_iou,
        "tracking_health_axis_min_anisotropy":
            args.tracking_health_axis_min_anisotropy,
    }
    invalid_tracking_ratios = [
        name for name, value in tracking_ratios.items()
        if not 0.0 <= float(value) <= 1.0
    ]
    if invalid_tracking_ratios:
        raise ValueError(
            "These tracking ratios must be within [0,1]: "
            + ", ".join(invalid_tracking_ratios)
        )
    if float(args.tracking_health_max_axis_difference_deg) > 90.0:
        raise ValueError(
            "--tracking_health_max_axis_difference_deg must be within (0,90]"
        )
    if float(args.tracking_health_temporal_weight) < 0:
        raise ValueError(
            "--tracking_health_temporal_weight must be non-negative"
        )
    if args.max_add_points < 0 or args.max_bop_points < 0 or args.metrics_flush_interval < 0:
        raise ValueError(
            "--max_add_points, --max_bop_points and --metrics_flush_interval must be >= 0"
        )
    if args.max_symmetry_transforms < 1 or args.max_sym_disc_step <= 0:
        raise ValueError("Symmetry discretization parameters must be positive")
    if (
        (args.mesh_scale is not None and args.mesh_scale <= 0)
        or args.gt_translation_scale <= 0
        or args.camera_translation_scale <= 0
        or args.object_mesh_translation_scale <= 0
        or args.models_info_translation_scale <= 0
    ):
        raise ValueError("mesh/GT/camera/object-mesh/models-info scales must be positive")
    if args.display_scale <= 0 or args.video_scale <= 0:
        raise ValueError("display/video scales must be positive")
    if args.metric_symmetry == "models_info" and args.models_info is None:
        raise ValueError("--metric_symmetry models_info requires --models_info")
    positive_fusion_values = {
        "fusion_view_pass_threshold": args.fusion_view_pass_threshold,
        "fusion_max_candidate_translation_mm": args.fusion_max_candidate_translation_mm,
        "fusion_max_candidate_rotation_deg": args.fusion_max_candidate_rotation_deg,
        "fusion_depth_tolerance_m": args.fusion_depth_tolerance_m,
        "fusion_occlusion_tolerance_m": args.fusion_occlusion_tolerance_m,
        "fusion_point_tolerance_m": args.fusion_point_tolerance_m,
        "fusion_weight_temperature": args.fusion_weight_temperature,
        "fusion_max_optimization_translation_mm": args.fusion_max_optimization_translation_mm,
        "fusion_max_optimization_rotation_deg": args.fusion_max_optimization_rotation_deg,
        "fusion_optimization_xtol": args.fusion_optimization_xtol,
        "fusion_optimization_ftol": args.fusion_optimization_ftol,
        "fusion_recovery_threshold_factor": args.fusion_recovery_threshold_factor,
        "fusion_fallback_translation_scale_mm": args.fusion_fallback_translation_scale_mm,
        "fusion_fallback_rotation_scale_deg": args.fusion_fallback_rotation_scale_deg,
        "fusion_fallback_invalid_penalty": args.fusion_fallback_invalid_penalty,
        "fusion_joint_max_deviation_translation_mm": args.fusion_joint_max_deviation_translation_mm,
        "fusion_joint_max_deviation_rotation_deg": args.fusion_joint_max_deviation_rotation_deg,
        "fusion_gt_max_translation_disagreement_mm": args.fusion_gt_max_translation_disagreement_mm,
        "fusion_gt_max_rotation_disagreement_deg": args.fusion_gt_max_rotation_disagreement_deg,
    }
    invalid_positive = [
        name for name, value in positive_fusion_values.items()
        if float(value) <= 0
    ]
    if invalid_positive:
        raise ValueError(
            "These fusion parameters must be positive: "
            + ", ".join(invalid_positive)
        )
    count_values = {
        "fusion_max_surface_points": args.fusion_max_surface_points,
        "fusion_max_observed_points": args.fusion_max_observed_points,
        "fusion_min_observed_points": args.fusion_min_observed_points,
        "fusion_min_projected_points": args.fusion_min_projected_points,
        "fusion_min_compared_points": args.fusion_min_compared_points,
        "fusion_min_consistency_cues": args.fusion_min_consistency_cues,
        "fusion_optimization_max_iterations": args.fusion_optimization_max_iterations,
        "fusion_optimization_max_evaluations": args.fusion_optimization_max_evaluations,
    }
    invalid_counts = [
        name for name, value in count_values.items() if int(value) < 1
    ]
    if invalid_counts:
        raise ValueError(
            "These fusion count parameters must be >= 1: "
            + ", ".join(invalid_counts)
        )
    if int(args.fusion_min_consistency_cues) > 3:
        raise ValueError(
            "--fusion_min_consistency_cues must be within [1,3]"
        )
    if int(args.fusion_max_recovery_frames) < 0:
        raise ValueError("--fusion_max_recovery_frames must be >= 0")
    ratio_values = {
        "fusion_view_pass_threshold": args.fusion_view_pass_threshold,
        "fusion_min_depth_inlier_ratio": args.fusion_min_depth_inlier_ratio,
        "fusion_min_mask_bbox_iou": args.fusion_min_mask_bbox_iou,
        "fusion_min_projected_inside_mask_ratio": args.fusion_min_projected_inside_mask_ratio,
        "fusion_min_point_inlier_ratio": args.fusion_min_point_inlier_ratio,
        "fusion_fallback_own_view_weight": args.fusion_fallback_own_view_weight,
    }
    invalid_ratios = [
        name for name, value in ratio_values.items()
        if not 0.0 <= float(value) <= 1.0
    ]
    if invalid_ratios:
        raise ValueError(
            "These fusion ratios must be within [0,1]: "
            + ", ".join(invalid_ratios)
        )
    fusion_weights = [
        args.fusion_depth_weight,
        args.fusion_inlier_weight,
        args.fusion_mask_weight,
        args.fusion_point_weight,
    ]
    if any(float(value) < 0 for value in fusion_weights) or sum(fusion_weights) <= 0:
        raise ValueError(
            "Fusion geometry weights must be non-negative and at least one "
            "weight must be positive."
        )
    if (
        args.fusion_optimization_regularization < 0
        or args.fusion_joint_accept_relative_margin < 0
        or args.fusion_joint_accept_absolute_margin < 0
        or args.fusion_fallback_temporal_weight < 0
    ):
        raise ValueError(
            "Fusion regularization, fallback temporal weight and "
            "joint-acceptance margins must be non-negative."
        )
    # Parse now so malformed threshold lists fail before a long tracking run.
    parse_float_list(args.vsd_taus, "--vsd_taus")
    parse_float_list(args.vsd_correctness_thresholds, "--vsd_correctness_thresholds")
    parse_float_list(args.mssd_thresholds_d, "--mssd_thresholds_d")
    parse_float_list(args.mspd_thresholds_px, "--mspd_thresholds_px")


def main():
    parser = argparse.ArgumentParser(
        description=(
            "MultiView real-scene left/right tracking, true two-view geometry "
            "fusion, and BOP evaluation with T_base_object GT."
        )
    )
    parser.add_argument("--name", type=str, default="real_multiview_bop_eval")
    parser.add_argument("--dataset_root", type=str, required=True)
    parser.add_argument("--view", choices=["left", "right", "both"], default="right")
    parser.add_argument(
        "--object_name",
        type=str,
        default="auto",
        help=(
            "One object prefix. Default 'auto' infers the unique prefix from "
            "<dataset_root>/evaluation/*_left and *_right."
        ),
    )
    parser.add_argument("--obj_mesh", type=str, default=None)
    parser.add_argument(
        "--objects_config",
        type=str,
        default=None,
        help=(
            "Optional legacy JSON batch. Omit it for the recommended "
            "one-object-per-command workflow."
        ),
    )

    # Model units/frame come from object_pose_setup. --mesh_scale is only a
    # checked manual override; there is no one-size-fits-all 0.001 default.
    parser.add_argument(
        "--mesh_scale",
        type=float,
        default=None,
        help=(
            "Optional manual OBJ-to-metre scale override. By default the "
            "script reads object_model_unit and checks candidate units against "
            "object_bbox_min_m/max_m for this object."
        ),
    )
    parser.add_argument(
        "--model_frame_mode",
        choices=[
            "auto",
            "dataset_gt",
            "raw_bbox_centered",
            "raw_uncentered",
        ],
        default="auto",
        help=(
            "auto: dataset_model_path -> identity; otherwise bbox_center "
            "metadata -> script-centre + rotation only; raw uncentred model -> "
            "full T_gt_model_m."
        ),
    )
    parser.add_argument(
        "--model_bbox_scale_tolerance",
        type=float,
        default=5.0,
        help=(
            "Maximum accepted ratio (and reciprocal) between proxy-model and "
            "metadata bbox diagonals after automatic unit checking."
        ),
    )
    parser.add_argument(
        "--model_diameter_m",
        type=float,
        default=None,
        help="Official model diameter in metres. Preferred for BOP thresholds.",
    )
    parser.add_argument(
        "--models_info",
        type=str,
        default=None,
        help="Optional BOP models_info.json containing diameter and symmetries.",
    )
    parser.add_argument(
        "--object_id",
        type=str,
        default=None,
        help="Object key in --models_info.",
    )
    parser.add_argument(
        "--models_info_translation_scale",
        type=float,
        default=0.001,
        help="Scale for models_info diameter/offset/translation; BOP files normally use mm.",
    )
    parser.add_argument("--gt_translation_scale", type=float, default=1.0,
                        help=("Compatibility option. Real gt_pose translation is already in metres, "
                              "so this must remain 1.0; no scaling is applied."))
    parser.add_argument("--depth_npy_scale", type=float, default=1.0,
                        help="Scale for depth_npy values; use 1.0 when NPY is already metres.")
    parser.add_argument("--depth_png_scale", type=float, default=0.001,
                        help="Scale for uint16 depth PNG values; use 0.001 for millimetres.")


    # Per-frame meta/T_cam_base is authoritative. These legacy files are read
    # only with the explicit compatibility flag below.
    parser.add_argument("--T_base_cam1", type=str, default=None,
                        help="Legacy right-camera camera-to-base fallback.")
    parser.add_argument("--T_base_cam2", type=str, default=None,
                        help="Legacy left-camera camera-to-base fallback.")
    parser.add_argument("--camera_translation_scale", type=float, default=1.0,
                        help="Scale used only by the legacy camera fallback.")
    parser.add_argument(
        "--allow_legacy_camera_extrinsics",
        action="store_true",
        help=(
            "Allow missing meta/T_cam_base and fall back to cam/T_base_cam*.txt. "
            "Do not use for the current captured datasets."
        ),
    )
    parser.add_argument(
        "--allow_legacy_object_setup",
        action="store_true",
        help=(
            "Allow datasets without object_pose_setup. Requires explicit "
            "--mesh_scale and uses legacy model-frame arguments."
        ),
    )
    parser.add_argument("--T_object_mesh", type=str, default=None,
                        help=("Legacy explicit immutable 4x4 ^GT_object T_mesh override. "
                              "By default it is derived from object_pose_setup."))
    parser.add_argument(
        "--axis_map",
        type=str,
        default=None,
        help=(
            "Fixed mesh-axis to GT-object-axis mapping when the origins already "
            "match. Tokens describe mesh +X,+Y,+Z in the object frame, e.g. "
            "'+y,-x,+z'. Must be right-handed. Do not combine with "
            "--T_object_mesh."
        ),
    )
    parser.add_argument(
        "--calibrate_axis_from_first_frame",
        action="store_true",
        help=(
            "Use the first successful frame of EACH view to estimate one "
            "frozen prediction-only local rotation branch. This never modifies "
            "metadata T_object_mesh or GT. Reuse the saved branch TXT files on "
            "a separate sequence for label-independent evaluation."
        ),
    )
    parser.add_argument(
        "--axis_calibration_mode",
        choices=["full", "signed_permutation"],
        default="full",
        help=(
            "Compatibility option retained for older commands. Global model "
            "rotation is no longer learned from prediction; "
            "--view_branch_calibration_mode controls prediction branches."
        ),
    )
    parser.add_argument(
        "--axis_calibration_view",
        choices=["right", "left"],
        default="right",
        help=(
            "Processing-order preference only. Both requested views receive "
            "their own prediction branch; neither view defines GT/model axes."
        ),
    )
    parser.add_argument(
        "--axis_calibration_output",
        type=str,
        default=None,
        help=(
            "Compatibility output path for the immutable metadata-derived "
            "T_object_mesh. Per-view predictions are saved beside it as "
            "<stem>_right_branch.txt and <stem>_left_branch.txt."
        ),
    )
    parser.add_argument(
        "--view_branch_calibration_mode",
        choices=["full", "signed_permutation", "geometry_safe"],
        default="full",
        help=(
            "Frozen per-view pose-branch correction learned on each view's "
            "marked first frame after the global T_object_mesh is known. "
            "full supports an arbitrary proper SO(3) branch; "
            "signed_permutation restricts it to 24 right-handed XYZ branches; "
            "geometry_safe only permits proxy-mesh symmetries and is strongly "
            "recommended for slender asymmetric tools."
        ),
    )
    parser.add_argument(
        "--right_pose_branch_correction",
        type=str,
        default=None,
        help=(
            "Optional saved 4x4 zero-translation local rotation applied on "
            "the right of every direct right-view T_base_mesh prediction."
        ),
    )
    parser.add_argument(
        "--left_pose_branch_correction",
        type=str,
        default=None,
        help=(
            "Optional saved 4x4 zero-translation local rotation applied on "
            "the right of every direct left-view T_base_mesh prediction."
        ),
    )
    parser.add_argument("--object_mesh_translation_scale", type=float, default=1.0,
                        help="Scale applied to the translation of T_object_mesh.")
    parser.add_argument(
        "--calibrate_proxy_model_from_first_frame",
        action="store_true",
        help=(
            "On the first saved-mask registration of EACH requested view, "
            "directly measure one fixed prediction-only rotation that aligns "
            "the predicted object axes to GT. Translation, the common "
            "metadata T_object_mesh and every GT pose remain unchanged."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_view",
        choices=["right", "left"],
        default="right",
        help=(
            "Processing-order preference only. With --view both, both right "
            "and left receive their own prediction-branch calibration."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_output",
        type=str,
        default=None,
        help=(
            "Output filename stem for per-view prediction branches. The "
            "script writes <stem>_right_branch.txt and "
            "<stem>_left_branch.txt plus a report JSON for each view; it "
            "never overwrites T_object_mesh."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_iteration",
        type=int,
        default=10,
        help=(
            "Minimum MultiView refinement iterations for the one initial "
            "registration and the following track_one() calls in this mode."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_max_points",
        type=int,
        default=5000,
        help="Maximum mesh points used only to report first-frame ADD/ADD-S.",
    )
    parser.add_argument(
        "--proxy_calibration_add_weight",
        type=float,
        default=0.8,
        help=(
            "Compatibility weight used only for the reported first-frame "
            "ADD/ADD-S diagnostic score; it does not select the rotation."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_adds_weight",
        type=float,
        default=0.2,
        help="Compatibility ADD-S diagnostic weight.",
    )
    parser.add_argument(
        "--proxy_calibration_min_improvement_mm",
        type=float,
        default=0.5,
        help=(
            "Legacy signed-permutation threshold retained for configuration "
            "compatibility; direct first-frame alignment does not use it."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_max_add_ratio",
        type=float,
        default=0.50,
        help=(
            "Reliability warning threshold for selected first-frame ADD divided "
            "by object diameter."
        ),
    )
    parser.add_argument(
        "--proxy_calibration_fail_unreliable",
        action="store_true",
        help=(
            "Abort instead of warning when the best first-frame axis candidate "
            "still exceeds --proxy_calibration_max_add_ratio."
        ),
    )

    parser.add_argument("--min_mesh_diameter_m", type=float, default=0.01)
    parser.add_argument("--max_mesh_diameter_m", type=float, default=1.0)
    parser.add_argument("--min_valid_gt_z_m", type=float, default=0.05)
    parser.add_argument("--max_valid_gt_z_m", type=float, default=5.0)
    parser.add_argument("--max_valid_gt_norm_m", type=float, default=8.0)
    parser.add_argument("--min_valid_depth_m", type=float, default=0.05)
    parser.add_argument("--max_valid_depth_m", type=float, default=8.0)
    parser.add_argument("--min_axis_length_m", type=float, default=0.02)

    parser.add_argument("--failure_policy", choices=["reregister", "freeze", "skip", "raise"],
                        default="reregister")
    parser.add_argument("--running_stride", type=int, default=1)
    parser.add_argument("--start_frame", type=int, default=0)
    parser.add_argument("--iteration", type=int, default=5)
    parser.add_argument("--reinit_interval", type=int, default=0)
    parser.add_argument(
        "--register_once_then_track",
        action="store_true",
        help=(
            "Use only the saved first-frame mask for one register() call, then "
            "run track_one() for all later frames. Disables periodic/current-"
            "frame SAM re-registration and tracking-health rejection."
        ),
    )
    parser.add_argument(
        "--tracking_profile",
        choices=["standard", "slender_tool"],
        default="standard",
        help=(
            "standard preserves the original tracker. slender_tool enables "
            "periodic mask-based re-registration plus GT-free geometry and "
            "motion health checks for screwdriver-like objects."
        ),
    )
    parser.add_argument(
        "--enable_tracking_health_check",
        action="store_true",
        help=(
            "Enable finite-but-wrong pose detection for the standard profile. "
            "It is enabled automatically by --tracking_profile slender_tool."
        ),
    )
    parser.add_argument(
        "--slender_tool_reinit_interval",
        type=int,
        default=5,
        help=(
            "Default periodic re-registration interval for the slender-tool "
            "profile when --reinit_interval remains zero."
        ),
    )
    parser.add_argument(
        "--slender_tool_iteration",
        type=int,
        default=8,
        help=(
            "Minimum MultiView register/track refinement iterations used by the "
            "slender-tool profile."
        ),
    )
    parser.add_argument(
        "--tracking_health_max_surface_points",
        type=int,
        default=3000,
    )
    parser.add_argument(
        "--tracking_health_min_projected_points",
        type=int,
        default=20,
    )
    parser.add_argument(
        "--tracking_health_min_compared_points",
        type=int,
        default=8,
    )
    parser.add_argument(
        "--tracking_health_max_geometry_cost",
        type=float,
        default=0.72,
    )
    parser.add_argument(
        "--tracking_health_min_depth_inlier_ratio",
        type=float,
        default=0.05,
    )
    parser.add_argument(
        "--tracking_health_min_inside_mask_ratio",
        type=float,
        default=0.15,
    )
    parser.add_argument(
        "--tracking_health_min_mask_bbox_iou",
        type=float,
        default=0.02,
    )
    parser.add_argument(
        "--tracking_health_axis_min_anisotropy",
        type=float,
        default=0.55,
        help=(
            "Apply the GT-free 2-D principal-axis check only when both the "
            "projected model and observed mask are this elongated."
        ),
    )
    parser.add_argument(
        "--tracking_health_max_axis_difference_deg",
        type=float,
        default=40.0,
        help=(
            "Maximum undirected angle between the projected model principal "
            "axis and observed mask principal axis for slender-tool tracking."
        ),
    )
    parser.add_argument(
        "--tracking_health_max_translation_jump_mm",
        type=float,
        default=55.0,
    )
    parser.add_argument(
        "--tracking_health_max_rotation_jump_deg",
        type=float,
        default=35.0,
    )
    parser.add_argument(
        "--tracking_health_max_recovery_frames",
        type=int,
        default=2,
        help=(
            "Maximum consecutive frames allowed to use the GT-free "
            "constant-velocity pose after both tracking and re-registration "
            "fail their observation gates."
        ),
    )
    parser.add_argument(
        "--tracking_health_temporal_weight",
        type=float,
        default=0.20,
    )
    parser.add_argument(
        "--tracking_health_invalid_penalty",
        type=float,
        default=2.0,
    )

    # Initial mask. Saved mask is the default for the real dataset.
    parser.add_argument("--use_manual_first_bbox", action="store_true",
                        help="Generate the first mask from a manually supplied bbox using SAM2.")
    parser.add_argument("--left_first_bbox", type=str, default=None)
    parser.add_argument("--right_first_bbox", type=str, default=None)
    parser.add_argument("--manual_bbox_min_side", type=float, default=20.0)
    parser.add_argument("--manual_bbox_pad", type=float, default=0.05)
    parser.add_argument("--abort_on_bad_first_mask", action="store_true", default=True)
    parser.add_argument("--no_abort_on_bad_first_mask", dest="abort_on_bad_first_mask", action="store_false")
    parser.add_argument("--write_first_mask_to_dataset", action="store_true")
    parser.add_argument("--fallback_first_mask_for_reinit", action="store_true")

    parser.add_argument("--sam2_enable", action="store_true", default=True)
    parser.add_argument("--no_sam2", dest="sam2_enable", action="store_false")
    parser.add_argument("--sam2_checkpoint", type=str,
                        default="./sam2/checkpoints/sam2.1_hiera_large.pt")
    parser.add_argument("--sam2_model_cfg", type=str,
                        default="./sam2/configs/sam2.1/sam2.1_hiera_l.yaml")
    parser.add_argument("--sam2_device", choices=["auto", "cuda", "cpu"], default="auto")
    parser.add_argument("--sam_box_source", choices=["pred", "motion", "saved_mask", "auto"],
                        default="pred")
    parser.add_argument("--sam_box_pad", type=float, default=0.20)
    parser.add_argument("--sam_motion_box_pad", type=float, default=0.35)
    parser.add_argument("--sam_union_box_pad", type=float, default=0.10)
    parser.add_argument("--expected_box_pad", type=float, default=0.10)
    parser.add_argument("--sam_min_box_size", type=float, default=8.0)
    parser.add_argument("--sam_fallback_full_image", action="store_true")
    parser.add_argument("--sam_multimask", action="store_true", default=True)
    parser.add_argument("--no_sam_multimask", dest="sam_multimask", action="store_false")
    parser.add_argument("--sam_use_pose_points", action="store_true", default=True)
    parser.add_argument("--no_sam_pose_points", dest="sam_use_pose_points", action="store_false")
    parser.add_argument("--sam_num_positive_points", type=int, default=10)
    parser.add_argument("--sam_prompt_vertices", type=int, default=300)
    parser.add_argument("--reject_bad_sam_masks", action="store_true", default=True)
    parser.add_argument("--allow_bad_sam_masks", dest="reject_bad_sam_masks", action="store_false")
    parser.add_argument("--mask_min_box_area_ratio", type=float, default=0.20)
    parser.add_argument("--mask_max_box_area_ratio", type=float, default=4.00)
    parser.add_argument("--mask_max_center_dist_norm", type=float, default=1.20)
    parser.add_argument("--mask_min_fill_ratio", type=float, default=0.005)
    parser.add_argument("--mask_max_fill_ratio", type=float, default=0.98)
    parser.add_argument("--pipe_preferred_max_fill_ratio", type=float, default=0.85)
    parser.add_argument("--prefer_saved_mask", action="store_true")
    parser.add_argument("--save_sam_masks", action="store_true")

    # Object symmetries. Metric symmetry is explicit and is independent from
    # optional temporal pose stabilization used only for display/tracking.
    parser.add_argument(
        "--metric_symmetry",
        choices=["none", "axial", "models_info"],
        default="none",
        help="Symmetries used by MSSD/MSPD. Do not mark an object symmetric unless verified.",
    )
    parser.add_argument(
        "--symmetry_stabilization",
        action="store_true",
        default=False,
        help=(
            "Stabilize rotation about the selected local symmetry axis for "
            "tracking/display and reported object-axis rotation. Enable only "
            "when rotation about that axis is physically indistinguishable."
        ),
    )
    parser.add_argument("--no_symmetry_stabilization", dest="symmetry_stabilization", action="store_false")
    parser.add_argument(
        "--temporal_axis_stabilization",
        action="store_true",
        default=True,
        help=(
            "Keep prediction axes on a temporally continuous branch using only "
            "automatically geometry-verified discrete mesh symmetries. The "
            "selected branch is applied to the full prediction used by metrics "
            "and fusion; GT is never consulted after branch initialization."
        ),
    )
    parser.add_argument(
        "--no_temporal_axis_stabilization",
        dest="temporal_axis_stabilization",
        action="store_false",
    )
    parser.add_argument(
        "--temporal_axis_min_improvement_deg",
        type=float,
        default=15.0,
        help=(
            "Apply a non-identity XYZ branch only when it improves temporal "
            "continuity by at least this many degrees; this avoids treating "
            "ordinary object motion as an axis jump."
        ),
    )
    parser.add_argument(
        "--temporal_branch_symmetry_mean_ratio",
        type=float,
        default=0.01,
        help=(
            "Mean bidirectional mesh discrepancy limit, as a fraction of "
            "diameter, for an automatically accepted temporal pose branch."
        ),
    )
    parser.add_argument(
        "--temporal_branch_symmetry_max_ratio",
        type=float,
        default=0.03,
        help=(
            "Maximum bidirectional mesh discrepancy limit, as a fraction of "
            "diameter, for an automatically accepted temporal pose branch."
        ),
    )
    parser.add_argument(
        "--temporal_branch_symmetry_min_mean_mm",
        type=float,
        default=1.0,
        help="Minimum absolute mean discrepancy tolerance in millimetres.",
    )
    parser.add_argument(
        "--temporal_branch_symmetry_min_max_mm",
        type=float,
        default=3.0,
        help="Minimum absolute maximum discrepancy tolerance in millimetres.",
    )
    parser.add_argument("--symmetry_axis", choices=["auto", "x", "y", "z"], default="auto",
                        help="Cylinder axis in bbox-centred OBJ coordinates. auto selects the largest bbox extent.")
    parser.add_argument("--symmetry_samples", type=int, default=180,
                        help="Number of axial rotations tested over 360 degrees; 180 gives 2-degree resolution.")
    parser.add_argument("--symmetry_allow_axis_flip", action="store_true", default=False,
                        help="Also treat reversal of the cylinder axis as symmetry (appropriate when both ends are equivalent).")
    parser.add_argument("--no_symmetry_axis_flip", dest="symmetry_allow_axis_flip", action="store_false")
    parser.add_argument("--symmetry_initial_reference", choices=["raw"], default="raw",
                        help="Kept for configuration compatibility; GT is never used to alter a prediction.")
    parser.add_argument(
        "--max_sym_disc_step",
        type=float,
        default=0.01,
        help="Maximum continuous-symmetry discretization step relative to object diameter.",
    )
    parser.add_argument("--max_symmetry_transforms", type=int, default=720)

    # True two-view fusion. Enabled automatically for --view both.
    parser.add_argument(
        "--enable_fusion",
        action="store_true",
        default=True,
        help=(
            "For --view both, validate right/left hypotheses against both "
            "RGB-D observations and produce a common base-frame fused pose."
        ),
    )
    parser.add_argument(
        "--no_fusion",
        dest="enable_fusion",
        action="store_false",
        help="Run both cameras independently without the fusion stage.",
    )
    parser.add_argument(
        "--fusion_failure_policy",
        choices=["previous", "fail"],
        default="previous",
        help=(
            "When no current right/left candidate exists, validate the previous "
            "fused pose on the current observations or mark the frame failed. "
            "Current candidates are handled by --fusion_no_pass_fallback."
        ),
    )
    parser.add_argument(
        "--fusion_no_pass_fallback",
        choices=["best_single", "fail"],
        default="best_single",
        help=(
            "When neither current candidate satisfies every two-view hard "
            "gate but at least one current pose exists, select the lower "
            "GT-free geometry/temporal score by default instead of writing an "
            "empty fusion row. Use fail for strict protocol diagnostics."
        ),
    )
    parser.add_argument(
        "--fusion_max_recovery_frames",
        type=int,
        default=2,
        help=(
            "Maximum consecutive frames allowed to reuse the previous fused "
            "pose when no current right/left pose exists. Current candidates "
            "always take precedence."
        ),
    )
    parser.add_argument(
        "--fusion_fallback_own_view_weight",
        type=float,
        default=0.25,
        help=(
            "Weight of a candidate's originating-view geometry score when "
            "ranking current single-view fallbacks; the remaining weight uses "
            "its two-view reporting energy."
        ),
    )
    parser.add_argument(
        "--fusion_fallback_temporal_weight",
        type=float,
        default=0.10,
        help=(
            "Small GT-free temporal penalty used only to break ambiguous "
            "current-candidate fallback decisions."
        ),
    )
    parser.add_argument(
        "--fusion_fallback_translation_scale_mm",
        type=float,
        default=50.0,
    )
    parser.add_argument(
        "--fusion_fallback_rotation_scale_deg",
        type=float,
        default=30.0,
    )
    parser.add_argument(
        "--fusion_fallback_invalid_penalty",
        type=float,
        default=1.0,
    )
    parser.add_argument(
        "--fusion_tie_break_view",
        choices=["right", "left"],
        default="right",
        help="Deterministic tie break used only for numerically equal scores.",
    )
    parser.add_argument("--fusion_max_surface_points", type=int, default=6000)
    parser.add_argument("--fusion_max_observed_points", type=int, default=4000)
    parser.add_argument("--fusion_min_observed_points", type=int, default=30)
    parser.add_argument("--fusion_min_projected_points", type=int, default=30)
    parser.add_argument("--fusion_min_compared_points", type=int, default=15)
    parser.add_argument(
        "--fusion_depth_tolerance_m",
        type=float,
        default=0.030,
        help=(
            "Depth inlier and truncation tolerance in metres. The balanced "
            "default is 30 mm for real RGB-D data and proxy meshes."
        ),
    )
    parser.add_argument(
        "--fusion_occlusion_tolerance_m",
        type=float,
        default=0.025,
        help=(
            "A model surface farther than the measured depth by this amount is "
            "treated as occluded rather than a free-space violation."
        ),
    )
    parser.add_argument(
        "--fusion_point_tolerance_m",
        type=float,
        default=0.035,
        help=(
            "Observed-to-model point-cloud inlier tolerance in metres. "
            "The balanced default allows moderate proxy-mesh mismatch."
        ),
    )
    parser.add_argument("--fusion_depth_weight", type=float, default=0.35)
    parser.add_argument("--fusion_inlier_weight", type=float, default=0.20)
    parser.add_argument("--fusion_mask_weight", type=float, default=0.20)
    parser.add_argument("--fusion_point_weight", type=float, default=0.25)
    parser.add_argument(
        "--fusion_view_pass_threshold",
        type=float,
        default=0.65,
        help="Maximum normalized geometry cost for each camera view.",
    )
    parser.add_argument(
        "--fusion_gate_mode",
        choices=["balanced", "strict"],
        default="balanced",
        help=(
            "balanced requires the total cost and at least N available cue "
            "groups; strict reproduces the original all-individual-gates rule."
        ),
    )
    parser.add_argument(
        "--fusion_min_consistency_cues",
        type=int,
        default=2,
        help=(
            "Minimum passing cue groups per camera in balanced mode. Available "
            "groups are depth, silhouette mask and point cloud; if fewer groups "
            "exist, all available groups are required."
        ),
    )
    parser.add_argument(
        "--fusion_min_depth_inlier_ratio",
        type=float,
        default=0.15,
    )
    parser.add_argument(
        "--fusion_min_mask_bbox_iou",
        type=float,
        default=0.03,
    )
    parser.add_argument(
        "--fusion_min_projected_inside_mask_ratio",
        type=float,
        default=0.12,
    )
    parser.add_argument(
        "--fusion_min_point_inlier_ratio",
        type=float,
        default=0.05,
    )
    parser.add_argument(
        "--fusion_max_candidate_translation_mm",
        type=float,
        default=50.0,
        help="Maximum right-left translation disagreement for joint optimization.",
    )
    parser.add_argument(
        "--fusion_max_candidate_rotation_deg",
        type=float,
        default=30.0,
        help=(
            "Maximum symmetry-aware right-left rotation disagreement for "
            "joint optimization."
        ),
    )
    parser.add_argument(
        "--fusion_weight_temperature",
        "--fusion_contribution_tau",
        dest="fusion_weight_temperature",
        type=float,
        default=0.10,
        help=(
            "Paper-formula tau in exp(-(E_v-E_min)/tau). The resulting "
            "right/left weights are exported and initialize joint SE(3) "
            "optimization."
        ),
    )
    parser.add_argument(
        "--fusion_max_optimization_translation_mm",
        type=float,
        default=20.0,
    )
    parser.add_argument(
        "--fusion_max_optimization_rotation_deg",
        type=float,
        default=10.0,
    )
    parser.add_argument(
        "--fusion_optimization_max_iterations",
        type=int,
        default=15,
    )
    parser.add_argument(
        "--fusion_optimization_max_evaluations",
        type=int,
        default=120,
    )
    parser.add_argument(
        "--fusion_optimization_regularization",
        type=float,
        default=0.01,
    )
    parser.add_argument("--fusion_optimization_xtol", type=float, default=1e-3)
    parser.add_argument("--fusion_optimization_ftol", type=float, default=1e-4)
    parser.add_argument(
        "--fusion_joint_accept_relative_margin",
        type=float,
        default=0.03,
        help=(
            "Accept joint pose only when its geometry cost is no more than this "
            "fraction worse than the best passing single-view hypothesis."
        ),
    )
    parser.add_argument(
        "--fusion_joint_accept_absolute_margin",
        type=float,
        default=0.005,
    )
    parser.add_argument(
        "--fusion_joint_max_deviation_translation_mm",
        type=float,
        default=20.0,
        help=(
            "Reject a jointly optimized pose if it moves farther than this "
            "from the best passing single-view pose."
        ),
    )
    parser.add_argument(
        "--fusion_joint_max_deviation_rotation_deg",
        type=float,
        default=15.0,
        help=(
            "Reject a jointly optimized pose if its symmetry-aware rotation "
            "deviation from the best passing single-view pose exceeds this."
        ),
    )
    parser.add_argument(
        "--fusion_recovery_threshold_factor",
        type=float,
        default=1.30,
        help="Relaxation factor when validating the previous fused pose.",
    )
    parser.add_argument(
        "--fusion_gt_max_translation_disagreement_mm",
        type=float,
        default=1.0,
        help=(
            "Dataset audit only: maximum allowed disagreement between the "
            "right/left copies of T_base_object GT."
        ),
    )
    parser.add_argument(
        "--fusion_gt_max_rotation_disagreement_deg",
        type=float,
        default=0.2,
        help=(
            "Dataset audit only: maximum allowed rotation disagreement between "
            "the right/left copies of T_base_object GT."
        ),
    )

    # BOP 2019/2020 evaluation protocol.
    parser.add_argument("--disable_bop_metrics", action="store_true")
    parser.add_argument(
        "--metric_failure_policy",
        choices=["raise", "nan"],
        default="raise",
        help="Raise on an invalid BOP evaluation by default; 'nan' is diagnostic only.",
    )
    parser.add_argument(
        "--vsd_delta_mm",
        type=float,
        default=15.0,
        help=(
            "Depth tolerance used to decide whether a rendered surface is "
            "visible in the measured depth image."
        ),
    )
    parser.add_argument(
        "--vsd_taus",
        type=str,
        default="0.05,0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45,0.50",
        help=(
            "VSD surface-misalignment tolerances. Because normalized_by_diameter "
            "is enabled, 0.05 means 5 percent of the object diameter."
        ),
    )
    parser.add_argument(
        "--vsd_correctness_thresholds",
        type=str,
        default="0.05,0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45,0.50",
        help=(
            "Thresholds applied to the resulting VSD error when computing "
            "per-frame VSD average-recall contribution."
        ),
    )
    parser.add_argument(
        "--mssd_thresholds_d",
        type=str,
        default="0.05,0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45,0.50",
    )
    parser.add_argument(
        "--mspd_thresholds_px",
        type=str,
        default="5,10,15,20,25,30,35,40,45,50",
    )

    parser.add_argument("--axis_ratio", type=float, default=0.5)
    parser.add_argument(
        "--max_add_points",
        type=int,
        default=0,
        help="0 uses all mesh vertices for ADD/ADD-S; positive values subsample deterministically.",
    )
    parser.add_argument(
        "--max_bop_points",
        type=int,
        default=0,
        help="0 uses all mesh vertices; positive values enable deterministic subsampling.",
    )
    parser.add_argument(
        "--metrics_flush_interval",
        type=int,
        default=50,
        help="Rewrite the per-frame Excel checkpoint every N rows; 0 disables checkpoints.",
    )
    parser.add_argument("--min_mask_pixels", type=int, default=100)
    parser.add_argument("--show_mask", action="store_true")
    parser.add_argument("--show_overlay", action="store_true",
                        help="Show only a compact frame/source/symmetry label. Metrics remain hidden.")
    parser.add_argument("--draw_pose_axes", action="store_true", default=True)
    parser.add_argument("--no_pose_axes", dest="draw_pose_axes", action="store_false")
    parser.add_argument("--draw_origin_gap", action="store_true",
                        help="Draw the white line between GT and prediction origins; disabled by default.")
    parser.add_argument("--save_video", action="store_true")
    parser.add_argument("--save_images", action="store_true")
    parser.add_argument("--fps", type=float, default=20.0)
    parser.add_argument("--display_scale", type=float, default=1.5)
    parser.add_argument("--video_scale", type=float, default=1.0)
    parser.add_argument("--wait_ms", type=int, default=1)
    parser.add_argument("--no_display", action="store_true")
    parser.add_argument("--debug", type=int, default=0)
    parser.add_argument("--output_root", type=str,
                        default="./results/real_scene_multiview_tracking")
    args = parser.parse_args()
    print(f"[Build] multiview_bop_eval.py {SCRIPT_BUILD_ID}")
    print(f"[Script path] {os.path.abspath(__file__)}")
    print(
        "[Excel fusion columns] "
        + ", ".join(output_name for _, output_name in FUSION_METRIC_COLUMN_MAP)
    )

    object_runs = build_object_runs(args)
    for run_args in object_runs:
        validate_run_args(run_args)

    seed_everything(0)
    date_str = f"{datetime.now():%Y-%m-%d_%H-%M-%S}"
    save_root = os.path.join(args.output_root, args.name, date_str)
    os.makedirs(save_root, exist_ok=True)

    try:
        for run_index, run_args in enumerate(object_runs, start=1):
            print(
                f"\n[Object {run_index}/{len(object_runs)}] "
                f"{run_args.object_name} | mesh={run_args.obj_mesh}"
            )
            requested_views = parse_views(run_args.view)
            setattr(run_args, "_view_branch_corrections", {})
            setattr(run_args, "_view_branch_diagnostics", {})
            setattr(run_args, "_view_branch_paths", {})
            setattr(run_args, "_view_branch_calibration_frames", {})
            setattr(run_args, "_fixed_model_definition", None)
            setattr(run_args, "_fixed_model_transform_saved", False)
            if run_args.calibrate_proxy_model_from_first_frame:
                proxy_view = str(run_args.proxy_calibration_view)
                if proxy_view not in requested_views:
                    raise ValueError(
                        "--proxy_calibration_view must be included in --view. "
                        f"Requested view={run_args.view!r}, calibration "
                        f"view={proxy_view!r}."
                    )
                requested_views = [
                    proxy_view,
                    *[
                        view for view in requested_views
                        if view != proxy_view
                    ],
                ]
                print(
                    "[Proxy calibration] global T_object_mesh remains fixed. "
                    "Each requested view will use its own first saved-mask "
                    "registration to measure one fixed prediction-only "
                    "rotation offset directly from pred/GT orientation. "
                    f"{proxy_view} is processed first only for deterministic "
                    "ordering."
                )
            if run_args.calibrate_axis_from_first_frame:
                calibration_view = (
                    run_args.axis_calibration_view
                    if run_args.axis_calibration_view in requested_views
                    else requested_views[0]
                )
                setattr(
                    run_args,
                    "_axis_calibration_view_resolved",
                    calibration_view,
                )
                requested_views = [
                    calibration_view,
                    *[
                        view for view in requested_views
                        if view != calibration_view
                    ],
                ]
                print(
                    f"[Prediction branch calibration] {calibration_view} is "
                    "processed first only for deterministic ordering. "
                    "T_object_mesh comes exclusively from object_pose_setup "
                    "and stays immutable; each requested view receives its "
                    "own prediction-only branch correction."
                )
            else:
                setattr(
                    run_args, "_axis_calibration_view_resolved", None
                )
            setattr(
                run_args,
                "_defer_view_metric_export",
                bool(
                    run_args.enable_fusion
                    and set(requested_views) == {"right", "left"}
                ),
            )
            view_records_by_view: Dict[str, List[Dict]] = {}
            for view in requested_views:
                view_records = process_one_view(
                    run_args, view=view, save_root=save_root
                )
                view_records_by_view[view] = view_records
            if (
                run_args.enable_fusion
                and set(requested_views) == {"right", "left"}
            ):
                fusion_records = process_two_view_fusion(
                    run_args,
                    save_root=save_root,
                    view_records_by_view=view_records_by_view,
                )
    except KeyboardInterrupt:
        print("\n[Stopped] Interrupted by user.")
    finally:
        if not args.no_display:
            cv2.destroyAllWindows()

    print(f"Done. Results folder: {save_root}")


if __name__ == "__main__":
    main()
