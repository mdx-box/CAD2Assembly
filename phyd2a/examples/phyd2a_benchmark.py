
from __future__ import annotations

import argparse
import csv
import colorsys
import gc
import hashlib
import heapq
import json
import math
import os
import shutil
import sys
import time
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Deque, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import trimesh
import genesis as gs


CACHE_VERSION = 14
PLANNER_VERSION = "v19-contact-calibrated-virtual-screw"
DEFAULT_LLM_PRIOR_NUT_IDS = [
    "009_COMPOUND101",
    "010_COMPOUND102",
    "013_COMPOUND121",
    "014_COMPOUND122",
    "015_COMPOUND123",
    "016_COMPOUND124",
    "017_COMPOUND125",
    "018_COMPOUND126",
    "019_COMPOUND127",
]


@dataclass(frozen=True)
class ApproachConfig:
    """Method-specific switches used by the unified benchmark entry point.

    All approaches share preprocessing, Genesis dynamics, exact swept-mesh
    validation, disassembly success criteria, progressive no-progress timeout,
    statistics, and replay.  Only the semantic prior, sequence policy and path
    search algorithm are changed.
    """

    canonical_name: str
    display_name: str
    use_llm_prior: bool
    strict_prior: bool
    search_strategy: str
    candidate_order: str
    enable_direct_release: bool
    force_rotation: bool
    auto_rotation_depth: int
    enable_diagonal_actions: bool
    coupled_action_scale: float
    enable_blocker_guidance: bool
    enable_structural_guidance: bool


APPROACH_CONFIGS: Dict[str, ApproachConfig] = {
    "full_phyd2a": ApproachConfig(
        "full_phyd2a", "Full PhyD2A", True, True,
        "adaptive_best_first", "top-down-left-right", True, False, 2,
        True, 0.45, True, True,
    ),
    "phyd2a_bfs": ApproachConfig(
        "phyd2a_bfs", "PhyD2A with BFS (no LLM prior)", False, False,
        "bfs", "exterior", True, False, 2,
        True, 0.45, True, False,
    ),
    "phyd2a_dfs": ApproachConfig(
        "phyd2a_dfs", "PhyD2A with DFS (no LLM prior)", False, False,
        "dfs", "exterior", True, False, 2,
        True, 0.45, True, False,
    ),
    "phyd2a_no_llm": ApproachConfig(
        "phyd2a_no_llm", "PhyD2A without LLM prior", False, False,
        "adaptive_best_first", "exterior", True, False, 2,
        True, 0.45, True, True,
    ),
    "rrt": ApproachConfig(
        "rrt", "RRT", False, False,
        "rrt", "random", False, True, 1,
        False, 0.0, False, False,
    ),
    "rrt_star": ApproachConfig(
        "rrt_star", "RRT*", False, False,
        "rrt_star", "random", False, True, 1,
        False, 0.0, False, False,
    ),
    "bi_rrt": ApproachConfig(
        "bi_rrt", "Bi-RRT", False, False,
        "bi_rrt", "random", False, True, 1,
        False, 0.0, False, False,
    ),
    "bk_rrt": ApproachConfig(
        "bk_rrt", "Behavioral Kinodynamic RRT", False, False,
        "bk_rrt", "exterior", False, True, 1,
        True, 0.45, True, False,
    ),
}

APPROACH_ALIASES = {
    "full": "full_phyd2a",
    "phyd2a": "full_phyd2a",
    "full-phyd2a": "full_phyd2a",
    "full_phyd2a": "full_phyd2a",
    "rrt": "rrt",
    "rrt*": "rrt_star",
    "rrt-star": "rrt_star",
    "rrt_star": "rrt_star",
    "rrtstart": "rrt_star",
    "rrt-start": "rrt_star",
    "birrt": "bi_rrt",
    "bi-rrt": "bi_rrt",
    "bi_rrt": "bi_rrt",
    "bkrrt": "bk_rrt",
    "bk-rrt": "bk_rrt",
    "bk_rrt": "bk_rrt",
    "behavioral-kinodynamic-rrt": "bk_rrt",
    "phyd2a-bfs": "phyd2a_bfs",
    "phyd2a-with-bfs": "phyd2a_bfs",
    "phyd2a_bfs": "phyd2a_bfs",
    "bfs": "phyd2a_bfs",
    "phyd2a-dfs": "phyd2a_dfs",
    "phyd2a-with-dfs": "phyd2a_dfs",
    "phyd2a_dfs": "phyd2a_dfs",
    "dfs": "phyd2a_dfs",
    "phyd2a-no-llm": "phyd2a_no_llm",
    "phyd2a_without_llm": "phyd2a_no_llm",
    "phyd2a-without-llm-prior": "phyd2a_no_llm",
    "phyd2a_no_llm": "phyd2a_no_llm",
    "no-llm": "phyd2a_no_llm",
}


def normalize_approach_name(value: str) -> str:
    key = str(value).strip().lower().replace(" ", "-")
    canonical = APPROACH_ALIASES.get(key, key.replace("-", "_"))
    if canonical not in APPROACH_CONFIGS:
        valid = ", ".join(APPROACH_CONFIGS)
        raise ValueError(f"Unknown --approach {value!r}. Valid approaches: {valid}")
    return canonical


_FREE_DOF_ORDER_CACHE: Optional[str] = None


@dataclass
class LLMPriorOperation:
    """Operation parameters proposed by the LLM for one part.

    ``mating_parts`` identifies the parts that form the known screw/thread
    connection.  Their ordinary rigid-body collision response is replaced by a
    virtual screw joint during the unscrew primitive.  Collisions with every
    other part remain enabled and are checked independently with python-fcl.
    """

    operation: str = "physics_bfs"
    axis: Optional[np.ndarray] = None
    turns: float = 3.0
    distance: Optional[float] = None
    frames: int = 120
    handedness: float = 0.0
    axial_sign: float = 0.0
    max_turns: float = 12.0
    axis_mode: str = "auto"
    pitch: Optional[float] = None
    clearance: Optional[float] = None
    mating_parts: Tuple[str, ...] = ()


@dataclass
class LLMPriorPlan:
    """Structured LLM prior consumed by the sequence/path planner."""

    enabled: bool
    source: str
    priority_order: List[str]
    operations: Dict[str, LLMPriorOperation]
    base_part_id: Optional[str] = None
    planning_policy: Dict[str, object] = field(default_factory=dict)

    @classmethod
    def disabled(cls) -> "LLMPriorPlan":
        return cls(
            enabled=False,
            source="disabled",
            priority_order=[],
            operations={},
            base_part_id=None,
            planning_policy={},
        )


# ============================================================================
# Compatibility helpers
# ============================================================================


def _model_field_names(cls) -> Optional[set]:
    fields = getattr(cls, "model_fields", None)
    if fields is None:
        fields = getattr(cls, "__fields__", None)
    if fields is None:
        return None
    return set(fields.keys())


def construct_compatible(cls, **kwargs):
    """Construct a Genesis/Pydantic options object while dropping unknown fields."""
    field_names = _model_field_names(cls)
    if field_names is not None:
        kwargs = {k: v for k, v in kwargs.items() if k in field_names}
    return cls(**kwargs)


def scene_profiling_kwargs(show_fps: bool = False) -> dict:
    if hasattr(gs.options, "ProfilingOptions"):
        return {
            "profiling_options": construct_compatible(
                gs.options.ProfilingOptions,
                show_FPS=show_fps,
            )
        }
    # Compatibility with older Genesis versions.
    return {"show_FPS": show_fps}


def make_surface(color_rgb: Sequence[float]):
    return gs.surfaces.Default(
        color=tuple(float(x) for x in color_rgb[:3]),
        roughness=0.62,
        metallic=0.05,
    )


def require_python_fcl():
    """Fail early when exact mesh collision validation is unavailable."""
    try:
        trimesh.collision.CollisionManager()
    except Exception as exc:
        raise RuntimeError(
            "This corrected planner requires python-fcl to prevent tunnelling. "
            "Install it with `conda install -c conda-forge python-fcl` or "
            "`pip install python-fcl`."
        ) from exc


# ============================================================================
# Basic geometry utilities
# ============================================================================


def as_mesh(scene_or_mesh) -> Optional[trimesh.Trimesh]:
    if isinstance(scene_or_mesh, trimesh.Scene):
        meshes = [
            geom
            for geom in scene_or_mesh.geometry.values()
            if isinstance(geom, trimesh.Trimesh) and len(geom.faces) > 0
        ]
        if not meshes:
            return None
        return trimesh.util.concatenate(meshes)
    if isinstance(scene_or_mesh, trimesh.Trimesh):
        return scene_or_mesh
    return None


def to_numpy(x) -> np.ndarray:
    if hasattr(x, "detach"):
        return x.detach().cpu().numpy()
    return np.asarray(x)


def normalize_vector(v, fallback=(0.0, 0.0, 1.0)) -> np.ndarray:
    v = np.asarray(v, dtype=np.float64)
    n = float(np.linalg.norm(v))
    if n < 1e-12:
        return np.asarray(fallback, dtype=np.float64)
    return v / n


def normalize_quat(q) -> np.ndarray:
    q = np.asarray(q, dtype=np.float64)
    n = float(np.linalg.norm(q))
    if n < 1e-12:
        return np.array([1.0, 0.0, 0.0, 0.0], dtype=np.float64)
    q = q / n
    # Canonicalize sign so q and -q hash consistently.
    if q[0] < 0:
        q = -q
    return q


def quat_distance(q1, q2) -> float:
    q1 = normalize_quat(q1)
    q2 = normalize_quat(q2)
    dot = abs(float(np.dot(q1, q2)))
    dot = float(np.clip(dot, -1.0, 1.0))
    return 2.0 * math.acos(dot)


def quat_to_matrix(q) -> np.ndarray:
    w, x, y, z = normalize_quat(q)
    return np.array(
        [
            [1 - 2 * (y * y + z * z), 2 * (x * y - z * w), 2 * (x * z + y * w)],
            [2 * (x * y + z * w), 1 - 2 * (x * x + z * z), 2 * (y * z - x * w)],
            [2 * (x * z - y * w), 2 * (y * z + x * w), 1 - 2 * (x * x + y * y)],
        ],
        dtype=np.float64,
    )


def axis_angle_to_quat(axis, angle: float) -> np.ndarray:
    axis = normalize_vector(axis)
    half = 0.5 * float(angle)
    s = math.sin(half)
    return normalize_quat(
        np.array(
            [math.cos(half), axis[0] * s, axis[1] * s, axis[2] * s],
            dtype=np.float64,
        )
    )


def quat_multiply(q1, q2) -> np.ndarray:
    w1, x1, y1, z1 = normalize_quat(q1)
    w2, x2, y2, z2 = normalize_quat(q2)
    return normalize_quat(
        np.array(
            [
                w1 * w2 - x1 * x2 - y1 * y2 - z1 * z2,
                w1 * x2 + x1 * w2 + y1 * z2 - z1 * y2,
                w1 * y2 - x1 * z2 + y1 * w2 + z1 * x2,
                w1 * z2 + x1 * y2 - y1 * x2 + z1 * w2,
            ],
            dtype=np.float64,
        )
    )


def quat_conjugate(q) -> np.ndarray:
    q = normalize_quat(q)
    return np.array([q[0], -q[1], -q[2], -q[3]], dtype=np.float64)


def signed_rotation_about_axis(q0, q1, axis) -> float:
    """Approximate signed relative rotation around a world-space axis."""
    axis = normalize_vector(axis)
    q_rel = quat_multiply(normalize_quat(q1), quat_conjugate(normalize_quat(q0)))
    if q_rel[0] < 0.0:
        q_rel = -q_rel
    vector = q_rel[1:4]
    vector_norm = float(np.linalg.norm(vector))
    if vector_norm < 1e-12:
        return 0.0
    angle = 2.0 * math.atan2(vector_norm, max(float(q_rel[0]), 1e-12))
    rel_axis = vector / vector_norm
    return float(angle * np.sign(np.dot(rel_axis, axis)))


def transform_mesh_by_qpos(mesh: trimesh.Trimesh, qpos: np.ndarray) -> trimesh.Trimesh:
    qpos = np.asarray(qpos, dtype=np.float64)
    transform = np.eye(4, dtype=np.float64)
    transform[:3, :3] = quat_to_matrix(qpos[3:7])
    transform[:3, 3] = qpos[:3]
    result = mesh.copy()
    result.apply_transform(transform)
    return result


def qpos_to_transform(qpos: np.ndarray) -> np.ndarray:
    """Convert [x, y, z, qw, qx, qy, qz] to a homogeneous transform."""
    qpos = np.asarray(qpos, dtype=np.float64)
    transform = np.eye(4, dtype=np.float64)
    transform[:3, :3] = quat_to_matrix(qpos[3:7])
    transform[:3, 3] = qpos[:3]
    return transform


def slerp_quat(q0, q1, t: float) -> np.ndarray:
    """Shortest-path quaternion interpolation in Genesis w-x-y-z order."""
    q0 = normalize_quat(q0)
    q1 = normalize_quat(q1)
    dot = float(np.dot(q0, q1))
    if dot < 0.0:
        q1 = -q1
        dot = -dot
    dot = float(np.clip(dot, -1.0, 1.0))
    if dot > 0.9995:
        return normalize_quat((1.0 - t) * q0 + t * q1)
    theta = math.acos(dot)
    sin_theta = math.sin(theta)
    return normalize_quat(
        math.sin((1.0 - t) * theta) / sin_theta * q0
        + math.sin(t * theta) / sin_theta * q1
    )


def interpolate_pose_segment(
    q0: np.ndarray,
    q1: np.ndarray,
    translation_step: float,
    rotation_step: float,
) -> List[np.ndarray]:
    """Densely sample a rigid-body segment for swept collision validation."""
    q0 = np.asarray(q0, dtype=np.float64)
    q1 = np.asarray(q1, dtype=np.float64)
    trans = float(np.linalg.norm(q1[:3] - q0[:3]))
    rot = float(quat_distance(q0[3:7], q1[3:7]))
    n_trans = int(math.ceil(trans / max(float(translation_step), 1e-6)))
    n_rot = int(math.ceil(rot / max(float(rotation_step), 1e-4)))
    n = max(n_trans, n_rot, 1)
    samples = []
    for i in range(1, n + 1):
        t = i / float(n)
        q = np.empty(7, dtype=np.float64)
        q[:3] = (1.0 - t) * q0[:3] + t * q1[:3]
        q[3:7] = slerp_quat(q0[3:7], q1[3:7], t)
        samples.append(q)
    return samples


def aabb_overlap(min_a, max_a, min_b, max_b) -> bool:
    return bool(np.all(max_a >= min_b) and np.all(max_b >= min_a))


def aabb_contains(min_a, max_a, min_b, max_b) -> bool:
    return bool(np.all(min_a <= min_b) and np.all(max_a >= max_b))


def is_pose_disassembled_from_assets(
    assets: Dict[str, "PartAsset"],
    move_id: str,
    still_ids: Sequence[str],
    qpos: np.ndarray,
) -> bool:
    """
    Apply the same convex-hull completion test used by Assemble-Them-All.

    This validates an LLM operation primitive before it is accepted as a
    successful removal. It is deliberately not used as the physics contact
    model; Genesis still handles contact during BFS.
    """
    if not still_ids:
        return True

    still_hulls = [
        transform_mesh_by_qpos(assets[pid].local_hull, make_initial_qpos(assets[pid]))
        for pid in still_ids
    ]
    combined = trimesh.util.concatenate(still_hulls)
    try:
        still_hull = combined.convex_hull
    except Exception:
        still_hull = combined

    move_hull = transform_mesh_by_qpos(assets[move_id].local_hull, qpos)
    has_collision = None
    try:
        manager = trimesh.collision.CollisionManager()
        manager.add_object("still", still_hull)
        has_collision = bool(manager.in_collision_single(move_hull))
    except Exception:
        pass

    if has_collision is None:
        min_m, max_m = move_hull.bounds
        min_s, max_s = still_hull.bounds
        has_collision = aabb_overlap(min_m, max_m, min_s, max_s)
    if has_collision:
        return False

    min_m, max_m = move_hull.bounds
    min_s, max_s = still_hull.bounds
    return not (
        aabb_contains(min_m, max_m, min_s, max_s)
        or aabb_contains(min_s, max_s, min_m, max_m)
    )


def mesh_bounds_center(mesh: trimesh.Trimesh) -> np.ndarray:
    return 0.5 * (mesh.vertices.min(axis=0) + mesh.vertices.max(axis=0))


def clean_mesh(mesh: trimesh.Trimesh, fill_holes: bool = True) -> trimesh.Trimesh:
    mesh = mesh.copy()
    try:
        mesh.remove_unreferenced_vertices()
    except Exception:
        pass
    try:
        mesh.merge_vertices()
    except Exception:
        pass
    try:
        trimesh.repair.fix_normals(mesh)
    except Exception:
        pass
    if fill_holes and not mesh.is_watertight:
        try:
            trimesh.repair.fill_holes(mesh)
        except Exception:
            pass
    try:
        mesh.remove_unreferenced_vertices()
    except Exception:
        pass
    return mesh


# ============================================================================
# Mesh simplification and cache preprocessing
# ============================================================================


def _quadric_simplify(mesh: trimesh.Trimesh, target_faces: int, aggression: int):
    """Try all common trimesh simplification API variants."""
    method_names = ["simplify_quadric_decimation", "simplify_quadratic_decimation"]
    errors = []
    for name in method_names:
        method = getattr(mesh, name, None)
        if method is None:
            continue
        attempts = [
            lambda: method(face_count=int(target_faces), aggression=int(aggression)),
            lambda: method(face_count=int(target_faces)),
            lambda: method(int(target_faces)),
        ]
        for attempt in attempts:
            try:
                result = attempt()
                result = as_mesh(result)
                if result is not None and len(result.faces) > 0:
                    return clean_mesh(result, fill_holes=False)
            except Exception as exc:
                errors.append(f"{name}: {exc}")
    if errors:
        raise RuntimeError("; ".join(errors[-3:]))
    raise RuntimeError("No trimesh quadric decimator is available.")


def _cluster_mesh_once(mesh: trimesh.Trimesh, cell_size: float) -> Optional[trimesh.Trimesh]:
    if cell_size <= 0:
        return mesh.copy()

    vertices = np.asarray(mesh.vertices, dtype=np.float64)
    faces = np.asarray(mesh.faces, dtype=np.int64)
    origin = vertices.min(axis=0)
    keys = np.floor((vertices - origin) / cell_size).astype(np.int64)

    unique_keys, inverse = np.unique(keys, axis=0, return_inverse=True)
    n_new = len(unique_keys)
    if n_new < 4:
        return None

    sums = np.zeros((n_new, 3), dtype=np.float64)
    counts = np.bincount(inverse, minlength=n_new).astype(np.float64)
    np.add.at(sums, inverse, vertices)
    new_vertices = sums / np.maximum(counts[:, None], 1.0)

    new_faces = inverse[faces]
    valid = (
        (new_faces[:, 0] != new_faces[:, 1])
        & (new_faces[:, 1] != new_faces[:, 2])
        & (new_faces[:, 0] != new_faces[:, 2])
    )
    new_faces = new_faces[valid]
    if len(new_faces) < 4:
        return None

    # Remove duplicate triangles independently of winding.
    canonical = np.sort(new_faces, axis=1)
    _, unique_idx = np.unique(canonical, axis=0, return_index=True)
    new_faces = new_faces[np.sort(unique_idx)]

    result = trimesh.Trimesh(vertices=new_vertices, faces=new_faces, process=True)
    result = as_mesh(result)
    if result is None or len(result.faces) < 4:
        return None
    return clean_mesh(result, fill_holes=False)


def voxel_cluster_simplify(mesh: trimesh.Trimesh, target_faces: int) -> trimesh.Trimesh:
    """Dependency-free fallback simplifier based on adaptive vertex clustering."""
    if len(mesh.faces) <= target_faces:
        return mesh.copy()

    extents = np.maximum(mesh.extents, 1e-12)
    max_extent = float(np.max(extents))
    lo = max_extent * 1e-7
    hi = max_extent
    best = None

    # Find the smallest cell size producing <= target_faces.
    for _ in range(22):
        mid = 0.5 * (lo + hi)
        candidate = _cluster_mesh_once(mesh, mid)
        if candidate is None:
            hi = mid
            continue
        if len(candidate.faces) <= target_faces:
            best = candidate
            hi = mid
        else:
            lo = mid

    if best is None:
        try:
            best = mesh.convex_hull
        except Exception:
            best = mesh.copy()
    return clean_mesh(best, fill_holes=False)


def simplify_collision_mesh(
    mesh: trimesh.Trimesh,
    target_faces: int,
    aggression: int,
) -> Tuple[trimesh.Trimesh, str]:
    if len(mesh.faces) <= target_faces:
        return mesh.copy(), "unchanged"

    try:
        simplified = _quadric_simplify(mesh, target_faces, aggression)
        if 4 <= len(simplified.faces) <= len(mesh.faces):
            return simplified, "quadric"
    except Exception as exc:
        print(f"[Preprocess] Quadric simplification unavailable: {exc}")

    simplified = voxel_cluster_simplify(mesh, target_faces)
    return simplified, "voxel-cluster"


def generate_part_colors(part_ids: Sequence[str], seed: int) -> Dict[str, Tuple[float, float, float]]:
    """Generate bright, separated colors using a shuffled golden-ratio hue sequence."""
    rng = np.random.default_rng(seed)
    ids = list(sorted(part_ids))
    rng.shuffle(ids)
    hue0 = float(rng.random())
    golden = 0.6180339887498949
    result = {}
    for idx, pid in enumerate(ids):
        hue = (hue0 + idx * golden) % 1.0
        saturation = 0.58 + 0.18 * float(rng.random())
        value = 0.78 + 0.16 * float(rng.random())
        result[pid] = tuple(float(v) for v in colorsys.hsv_to_rgb(hue, saturation, value))
    return result


@dataclass
class NormalizationInfo:
    original_center: np.ndarray
    scale: float
    target_size: float
    original_min: np.ndarray
    original_max: np.ndarray

    def to_normalized_position(self, position) -> np.ndarray:
        return (np.asarray(position, dtype=np.float64) - self.original_center) * self.scale

    def to_original_position(self, position) -> np.ndarray:
        return np.asarray(position, dtype=np.float64) / self.scale + self.original_center

    def qpos_to_original(self, qpos) -> np.ndarray:
        qpos = np.asarray(qpos, dtype=np.float64).copy()
        qpos[:3] = self.to_original_position(qpos[:3])
        qpos[3:7] = normalize_quat(qpos[3:7])
        return qpos

    def to_json(self) -> dict:
        return {
            "original_center": self.original_center.tolist(),
            "scale": float(self.scale),
            "target_size": float(self.target_size),
            "original_min": self.original_min.tolist(),
            "original_max": self.original_max.tolist(),
        }


@dataclass
class PartAsset:
    part_id: str
    original_path: str
    visual_path: str
    collision_path: str
    center: np.ndarray
    center_original: np.ndarray
    local_visual_mesh: trimesh.Trimesh
    local_collision_mesh: trimesh.Trimesh
    local_hull: trimesh.Trimesh
    color: Tuple[float, float, float]
    source_faces: int
    collision_faces: int
    collision_watertight: bool


def _source_signatures(assembly_dir: str, obj_files: Sequence[str]) -> dict:
    result = {}
    for filename in obj_files:
        path = os.path.join(assembly_dir, filename)
        stat = os.stat(path)
        result[filename] = {
            "size": int(stat.st_size),
            "mtime_ns": int(stat.st_mtime_ns),
        }
    return result


def _cache_key(params: dict, sources: dict) -> str:
    text = json.dumps({"params": params, "sources": sources}, sort_keys=True)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _read_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return None


def preprocess_assembly_assets(
    assembly_dir: str,
    cache_dir: str,
    normalize_size: float,
    collision_face_num: int,
    simplify_aggressiveness: int,
    color_seed: int,
    rebuild_cache: bool = False,
    force_hull_for_nonwatertight: bool = False,
) -> Tuple[Dict[str, PartAsset], NormalizationInfo]:
    """
    Normalize all parts with one global transform and build persistent visual/collision caches.

    Original repository normalization uses a common assembly bounding box. Here the same
    idea is retained, but target_size is expressed in Genesis meters.
    """
    assembly_dir = os.path.abspath(assembly_dir)
    cache_dir = os.path.abspath(cache_dir)
    obj_files = sorted(
        f for f in os.listdir(assembly_dir)
        if f.lower().endswith(".obj") and f != "assembly.obj"
    )
    if len(obj_files) <= 1:
        raise RuntimeError(f"Need at least two OBJ parts in {assembly_dir}; found {len(obj_files)}.")

    params = {
        "cache_version": CACHE_VERSION,
        "normalize_size": float(normalize_size),
        "collision_face_num": int(collision_face_num),
        "simplify_aggressiveness": int(simplify_aggressiveness),
        "force_hull_for_nonwatertight": bool(force_hull_for_nonwatertight),
    }
    signatures = _source_signatures(assembly_dir, obj_files)
    key = _cache_key(params, signatures)
    manifest_path = os.path.join(cache_dir, "manifest.json")
    manifest = _read_json(manifest_path)

    cache_valid = (
        not rebuild_cache
        and manifest is not None
        and manifest.get("cache_key") == key
        and all(
            os.path.isfile(os.path.join(cache_dir, "visual", f))
            and os.path.isfile(os.path.join(cache_dir, "collision", f))
            for f in obj_files
        )
    )

    colors = generate_part_colors([Path(f).stem for f in obj_files], color_seed)

    if cache_valid:
        print(f"[Preprocess] Reusing mesh cache: {cache_dir}")
        norm_json = manifest["normalization"]
        normalization = NormalizationInfo(
            original_center=np.asarray(norm_json["original_center"], dtype=np.float64),
            scale=float(norm_json["scale"]),
            target_size=float(norm_json["target_size"]),
            original_min=np.asarray(norm_json["original_min"], dtype=np.float64),
            original_max=np.asarray(norm_json["original_max"], dtype=np.float64),
        )
    else:
        print(f"[Preprocess] Building mesh cache: {cache_dir}")
        if os.path.isdir(cache_dir):
            shutil.rmtree(cache_dir)
        os.makedirs(os.path.join(cache_dir, "visual"), exist_ok=True)
        os.makedirs(os.path.join(cache_dir, "collision"), exist_ok=True)

        raw_meshes: Dict[str, trimesh.Trimesh] = {}
        all_vertices = []
        for filename in obj_files:
            path = os.path.join(assembly_dir, filename)
            mesh = as_mesh(trimesh.load(path, process=True, maintain_order=True))
            if mesh is None or len(mesh.faces) == 0:
                raise RuntimeError(f"Empty or invalid mesh: {path}")
            mesh = clean_mesh(mesh, fill_holes=False)
            raw_meshes[filename] = mesh
            all_vertices.append(np.asarray(mesh.vertices, dtype=np.float64))

        stacked = np.concatenate(all_vertices, axis=0)
        original_min = stacked.min(axis=0)
        original_max = stacked.max(axis=0)
        original_center = 0.5 * (original_min + original_max)
        max_side = float(np.max(original_max - original_min))
        if max_side < 1e-12:
            raise RuntimeError("Assembly bounding box is degenerate.")
        scale = float(normalize_size) / max_side
        normalization = NormalizationInfo(
            original_center=original_center,
            scale=scale,
            target_size=float(normalize_size),
            original_min=original_min,
            original_max=original_max,
        )

        part_meta = {}
        for filename in obj_files:
            pid = Path(filename).stem
            raw = raw_meshes[filename]
            normalized = raw.copy()
            normalized.vertices = (normalized.vertices - original_center.reshape(1, 3)) * scale

            center_norm = mesh_bounds_center(normalized)
            center_orig = normalization.to_original_position(center_norm)
            local_visual = normalized.copy()
            local_visual.vertices -= center_norm.reshape(1, 3)
            local_visual = clean_mesh(local_visual, fill_holes=False)

            collision, method = simplify_collision_mesh(
                local_visual,
                target_faces=collision_face_num,
                aggression=simplify_aggressiveness,
            )
            collision = clean_mesh(collision, fill_holes=False)
            if not collision.is_watertight and force_hull_for_nonwatertight:
                try:
                    collision = collision.convex_hull
                    method += "+convex-hull"
                except Exception:
                    pass

            visual_path = os.path.join(cache_dir, "visual", filename)
            collision_path = os.path.join(cache_dir, "collision", filename)
            local_visual.export(visual_path)
            collision.export(collision_path)

            part_meta[pid] = {
                "filename": filename,
                "center": center_norm.tolist(),
                "center_original": center_orig.tolist(),
                "source_faces": int(len(raw.faces)),
                "visual_faces": int(len(local_visual.faces)),
                "collision_faces": int(len(collision.faces)),
                "collision_watertight": bool(collision.is_watertight),
                "simplification": method,
            }
            print(
                f"[Preprocess] {pid}: source={len(raw.faces)} faces -> "
                f"collision={len(collision.faces)} faces ({method}), "
                f"watertight={collision.is_watertight}"
            )

        manifest = {
            "cache_key": key,
            "params": params,
            "sources": signatures,
            "normalization": normalization.to_json(),
            "parts": part_meta,
        }
        with open(manifest_path, "w", encoding="utf-8") as file:
            json.dump(manifest, file, indent=2, ensure_ascii=False)

    # Load cached assets into memory.
    assets: Dict[str, PartAsset] = {}
    part_meta = manifest["parts"]
    for filename in obj_files:
        pid = Path(filename).stem
        meta = part_meta[pid]
        visual_path = os.path.join(cache_dir, "visual", filename)
        collision_path = os.path.join(cache_dir, "collision", filename)
        visual_mesh = as_mesh(trimesh.load(visual_path, process=True, maintain_order=True))
        collision_mesh = as_mesh(trimesh.load(collision_path, process=True, maintain_order=True))
        if visual_mesh is None or collision_mesh is None:
            raise RuntimeError(f"Failed to reload cached meshes for {pid}.")
        visual_mesh = clean_mesh(visual_mesh, fill_holes=False)
        collision_mesh = clean_mesh(collision_mesh, fill_holes=False)
        try:
            local_hull = visual_mesh.convex_hull
        except Exception:
            local_hull = collision_mesh.convex_hull

        assets[pid] = PartAsset(
            part_id=pid,
            original_path=os.path.join(assembly_dir, filename),
            visual_path=visual_path,
            collision_path=collision_path,
            center=np.asarray(meta["center"], dtype=np.float64),
            center_original=np.asarray(meta["center_original"], dtype=np.float64),
            local_visual_mesh=visual_mesh,
            local_collision_mesh=collision_mesh,
            local_hull=local_hull,
            color=colors[pid],
            source_faces=int(meta["source_faces"]),
            collision_faces=int(meta["collision_faces"]),
            collision_watertight=bool(meta["collision_watertight"]),
        )

    print("\n========== Normalization ==========")
    print(f"Original bbox min: {normalization.original_min.tolist()}")
    print(f"Original bbox max: {normalization.original_max.tolist()}")
    print(f"Original center:   {normalization.original_center.tolist()}")
    print(f"Scale:             {normalization.scale:.10g}")
    print(f"Normalized max side: {normalization.target_size:.6f} m")

    print("\n========== Part Colors ==========")
    for pid in sorted(assets):
        rgb = assets[pid].color
        print(f"{pid}: RGB=({rgb[0]:.3f}, {rgb[1]:.3f}, {rgb[2]:.3f})")

    return assets, normalization


# ============================================================================
# Assembly geometry and operation primitives
# ============================================================================


def make_initial_qpos(asset: PartAsset) -> np.ndarray:
    return np.concatenate(
        [asset.center.astype(np.float64), np.array([1.0, 0.0, 0.0, 0.0])]
    )


def compute_assembly_bounds(assets: Dict[str, PartAsset], part_ids: Optional[Iterable[str]] = None):
    ids = list(assets.keys()) if part_ids is None else list(part_ids)
    vertices = [
        assets[pid].local_visual_mesh.vertices + assets[pid].center.reshape(1, 3)
        for pid in ids
    ]
    stacked = np.concatenate(vertices, axis=0)
    xyz_min = stacked.min(axis=0)
    xyz_max = stacked.max(axis=0)
    center = 0.5 * (xyz_min + xyz_max)
    extent = float(np.linalg.norm(xyz_max - xyz_min))
    return xyz_min, xyz_max, center, max(extent, 1e-6)


class ExactCollisionValidator:
    """Exact swept-mesh validator for ordinary extraction and screw motion.

    Ordinary extraction is intentionally strict:

    * contacts absent in the assembled CAD state are treated as new contacts;
    * an initially intersecting pair may retain only a small numerical growth
      margin, never the former centimetre-scale release envelope;
    * once an initial contact has genuinely cleared, any later contact with the
      same part is classified as re-contact and rejected;
    * the final pose must be collision free.

    The previous release envelope could allow a cover to penetrate completely
    through a bolt head: the cover/bolt pair was already touching initially, so
    every later collision with that bolt was mistakenly treated as harmless
    release noise.  Dense sampling alone cannot prevent tunnelling when those
    samples are explicitly accepted.  This implementation validates the full
    visual meshes and tracks contact clearance/re-contact along the whole path.
    """

    def __init__(
        self,
        assets: Dict[str, PartAsset],
        move_id: str,
        still_ids: Sequence[str],
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
        mesh_mode: str = "visual",
        log_initial_contacts: bool = True,
    ):
        self.assets = assets
        self.move_id = move_id
        self.still_ids = list(still_ids)
        self.penetration_tolerance = max(float(penetration_tolerance), 0.0)
        self.sweep_translation_step = max(float(sweep_translation_step), 1e-6)
        self.sweep_rotation_step = max(float(sweep_rotation_step), 1e-4)
        self.mesh_mode = str(mesh_mode).lower().strip()
        self.log_initial_contacts = bool(log_initial_contacts)
        if self.mesh_mode not in {"visual", "collision"}:
            raise ValueError(f"Unsupported validation mesh mode: {mesh_mode!r}")

        # Ordinary extraction keeps the high-resolution visual mesh. Semantic
        # screw sweeps may explicitly use the cached simplified collision mesh
        # because hundreds of rotations around the screw axis otherwise dominate
        # the complete sequence budget. A high-resolution final-pose check is
        # still performed after the fast sweep.
        self.move_validation_mesh = (
            self.assets[self.move_id].local_collision_mesh
            if self.mesh_mode == "collision"
            else self.assets[self.move_id].local_visual_mesh
        )
        self.move_part_diagonal = max(
            float(np.linalg.norm(self.move_validation_mesh.extents)),
            1e-6,
        )

        self.existing_contact_ratio = 0.25
        self.existing_contact_hard_ratio = 0.50
        self.existing_contact_soft_abs = max(
            self.penetration_tolerance,
            0.5 * self.sweep_translation_step,
            2.5e-4,
        )
        self.existing_contact_hard_abs = max(
            1.5 * self.penetration_tolerance,
            self.sweep_translation_step,
            5.0e-4,
        )
        # A contact that is absent at the exact CAD pose may appear after the
        # first sub-millimetre motion solely because the simplified collision
        # mesh and Genesis SDF do not share the visual mesh's zero level set.
        # The previous hard cap of 0.5 mm was smaller than that representation
        # error (the reported failures were only 0.377--0.400 mm), so normal
        # screw motion was rejected as a newly created collision.  Calibrate a
        # bounded numerical skin from the user-selected penetration tolerance
        # and sweep resolution.  Macroscopic collisions remain invalid.
        numerical_skin = max(
            0.50 * self.penetration_tolerance,
            0.25 * self.sweep_translation_step,
            1.0e-4,
        )
        numerical_skin_cap = max(
            2.5e-3,
            0.01 * self.move_part_diagonal,
        )
        self.new_contact_tolerance = min(numerical_skin, numerical_skin_cap)
        self.contact_clearance_confirmation_distance = max(
            self.sweep_translation_step,
            0.003 * self.move_part_diagonal,
            0.5 * self.penetration_tolerance,
        )

        self.manager = None
        self.available = False
        self.initial_qpos = make_initial_qpos(self.assets[self.move_id])
        self.initial_depths: Dict[str, float] = {}
        self.last_blockers: set[str] = set()
        self.last_soft_contacts: Dict[str, float] = {}
        self.blocker_counts: Dict[str, int] = {}

        if not self.still_ids:
            self.available = True
            return

        try:
            manager = trimesh.collision.CollisionManager()
            for pid in self.still_ids:
                manager.add_object(
                    pid,
                    (
                        self.assets[pid].local_collision_mesh
                        if self.mesh_mode == "collision"
                        else self.assets[pid].local_visual_mesh
                    ),
                    transform=qpos_to_transform(make_initial_qpos(self.assets[pid])),
                )
            self.manager = manager
            self.available = True
            self.initial_depths = self.penetration_profile(self.initial_qpos)

            entries = []
            for pid, depth in sorted(self.initial_depths.items()):
                if depth <= 0.0:
                    continue
                entries.append(
                    f"{pid}:{depth:.6g}->soft:{self._allowed_depth(pid):.6g}"
                    f"->hard:{self._hard_allowed_depth(pid):.6g}"
                )
            if entries and self.log_initial_contacts:
                print(
                    f"[Collision] Exact initial-contact limits for {move_id}: "
                    + ", ".join(entries)
                )
                print(
                    f"[Collision] Re-contact confirmation distance for {move_id}: "
                    f"{self.contact_clearance_confirmation_distance:.6g} m."
                )
        except Exception as exc:
            self.manager = None
            self.available = False
            print(
                "[Warning] python-fcl exact swept collision validation is "
                f"unavailable: {exc}. Install python-fcl for reliable "
                "anti-tunnelling checks."
            )

    def _allowed_depth(self, pid: str) -> float:
        initial = float(self.initial_depths.get(pid, 0.0))
        if initial > 0.0:
            growth = max(
                self.existing_contact_soft_abs,
                self.existing_contact_ratio * initial,
            )
            return initial + growth
        return self.new_contact_tolerance

    def _hard_allowed_depth(self, pid: str) -> float:
        initial = float(self.initial_depths.get(pid, 0.0))
        if initial > 0.0:
            growth = max(
                self.existing_contact_hard_abs,
                self.existing_contact_hard_ratio * initial,
            )
            return initial + growth
        return self.new_contact_tolerance

    def _record_blocker(self, pid: str):
        pid = str(pid)
        self.last_blockers.add(pid)
        self.blocker_counts[pid] = self.blocker_counts.get(pid, 0) + 1

    def clear_last_blockers(self):
        self.last_blockers.clear()
        self.last_soft_contacts.clear()

    def penetration_profile(self, qpos: np.ndarray) -> Dict[str, float]:
        if not self.available or self.manager is None:
            return {}
        collided, names, contacts = self.manager.in_collision_single(
            self.move_validation_mesh,
            transform=qpos_to_transform(qpos),
            return_names=True,
            return_data=True,
        )
        if not collided:
            return {}

        depths: Dict[str, float] = {str(name): 0.0 for name in names}
        for contact in contacts:
            pid = next(
                (name for name in contact.names if name != "__external"),
                None,
            )
            if pid is not None:
                depths[str(pid)] = max(
                    depths.get(str(pid), 0.0),
                    max(float(contact.depth), 0.0),
                )
        return depths

    def pose_is_valid(
        self,
        qpos: np.ndarray,
        release_direction: Optional[np.ndarray] = None,
        allow_initial_release: bool = False,
    ) -> Tuple[bool, str]:
        """Validate one pose using strict exact-mesh contact growth limits.

        ``release_direction`` and ``allow_initial_release`` are retained for API
        compatibility, but they no longer enlarge the collision threshold.
        Direction alone is not evidence that an initial contact is separating.
        """
        del release_direction, allow_initial_release
        if not self.available:
            return True, "fcl-unavailable"

        depths = self.penetration_profile(qpos)
        for pid, depth in depths.items():
            if pid not in self.initial_depths:
                allowed = self.new_contact_tolerance
                if depth > allowed:
                    self._record_blocker(pid)
                    return (
                        False,
                        f"new exact contact with {pid}: depth={depth:.6g}, "
                        f"allowed={allowed:.6g}",
                    )
                continue

            soft = self._allowed_depth(pid)
            hard = self._hard_allowed_depth(pid)
            if depth > hard:
                self._record_blocker(pid)
                return (
                    False,
                    f"existing exact-contact growth with {pid}: "
                    f"depth={depth:.6g}, soft={soft:.6g}, hard={hard:.6g}",
                )
            if depth > soft:
                self.last_soft_contacts[pid] = max(
                    self.last_soft_contacts.get(pid, 0.0),
                    float(depth - soft),
                )
        return True, "ok"

    def initial_contact_depth_sum(self, qpos: np.ndarray) -> float:
        depths = self.penetration_profile(qpos)
        return float(sum(depths.get(pid, 0.0) for pid in self.initial_depths))

    def unscrew_pose_is_valid(
        self,
        qpos: np.ndarray,
        axis: np.ndarray,
        origin: np.ndarray,
        lateral_limit: float,
    ) -> Tuple[bool, str]:
        """Validator for an explicitly known screw-removal operation.

        Only this semantic operation receives a broader transient allowance for
        initially mating/thread contacts. New contacts remain strict.
        """
        if not self.available:
            return True, "fcl-unavailable"

        qpos = np.asarray(qpos, dtype=np.float64)
        axis = normalize_vector(axis)
        delta = qpos[:3] - np.asarray(origin, dtype=np.float64)
        axial = float(np.dot(delta, axis))
        lateral = float(np.linalg.norm(delta - axial * axis))
        if lateral > max(float(lateral_limit), 1e-6):
            return (
                False,
                f"left screw corridor: lateral={lateral:.6g}, "
                f"limit={lateral_limit:.6g}",
            )

        depths = self.penetration_profile(qpos)
        for pid, depth in depths.items():
            if pid not in self.initial_depths:
                if depth > self.new_contact_tolerance:
                    self._record_blocker(pid)
                    return False, f"new exact contact with {pid}: depth={depth:.6g}"
                continue

            initial = float(self.initial_depths.get(pid, 0.0))
            cap = initial + max(
                4.0 * self.penetration_tolerance,
                2.0 * initial,
                0.01,
            )
            if depth > cap:
                self._record_blocker(pid)
                return (
                    False,
                    f"excessive mating-contact growth with {pid}: "
                    f"depth={depth:.6g}, cap={cap:.6g}",
                )
        return True, "ok"

    def unscrew_segment_is_valid(
        self,
        q0: np.ndarray,
        q1: np.ndarray,
        axis: np.ndarray,
        origin: np.ndarray,
        lateral_limit: float,
    ) -> Tuple[bool, str]:
        if not self.available:
            return True, "fcl-unavailable"
        self.clear_last_blockers()
        for sample in interpolate_pose_segment(
            q0,
            q1,
            translation_step=self.sweep_translation_step,
            rotation_step=self.sweep_rotation_step,
        ):
            valid, reason = self.unscrew_pose_is_valid(
                sample,
                axis=axis,
                origin=origin,
                lateral_limit=lateral_limit,
            )
            if not valid:
                return False, reason
        return True, "ok"

    def unscrew_path_is_valid(
        self,
        path: Sequence[np.ndarray],
        axis: np.ndarray,
        origin: np.ndarray,
        lateral_limit: float,
        min_removal_distance: float,
    ) -> Tuple[bool, str]:
        if path is None or len(path) < 2:
            return False, "path contains fewer than two poses"
        previous = np.asarray(path[0], dtype=np.float64)
        for index, qpos in enumerate(path[1:], start=1):
            qpos = np.asarray(qpos, dtype=np.float64)
            valid, reason = self.unscrew_segment_is_valid(
                previous,
                qpos,
                axis=axis,
                origin=origin,
                lateral_limit=lateral_limit,
            )
            if not valid:
                return False, f"segment {index - 1}->{index}: {reason}"
            previous = qpos

        displacement = float(np.linalg.norm(previous[:3] - np.asarray(path[0])[:3]))
        if displacement < min_removal_distance:
            return False, f"insufficient displacement: {displacement:.6g}"
        return self.pose_is_collision_free(previous)

    def pose_is_collision_free(self, qpos: np.ndarray) -> Tuple[bool, str]:
        """The final extracted pose must not intersect any remaining part."""
        if not self.available:
            return True, "fcl-unavailable"
        depths = self.penetration_profile(qpos)
        blocking = {
            pid: depth
            for pid, depth in depths.items()
            if depth > max(0.20 * self.new_contact_tolerance, 1e-7)
        }
        if blocking:
            pid, depth = max(blocking.items(), key=lambda item: item[1])
            self._record_blocker(pid)
            return False, f"final exact collision with {pid}: depth={depth:.6g}"
        return True, "ok"

    def segment_is_valid(
        self,
        q0: np.ndarray,
        q1: np.ndarray,
        release_direction: Optional[np.ndarray] = None,
        allow_initial_release: bool = False,
    ) -> Tuple[bool, str]:
        if not self.available:
            return True, "fcl-unavailable"
        self.clear_last_blockers()
        for sample in interpolate_pose_segment(
            q0,
            q1,
            translation_step=self.sweep_translation_step,
            rotation_step=self.sweep_rotation_step,
        ):
            valid, reason = self.pose_is_valid(
                sample,
                release_direction=release_direction,
                allow_initial_release=allow_initial_release,
            )
            if not valid:
                return False, reason
        return True, "ok"

    def path_is_valid(
        self,
        path: Sequence[np.ndarray],
        min_removal_distance: float,
        allow_initial_release: bool = True,
        deadline: Optional[float] = None,
    ) -> Tuple[bool, str]:
        """Validate a complete path and reject clear-then-recontact tunnelling."""
        del allow_initial_release
        self.clear_last_blockers()
        if path is None or len(path) < 2:
            return False, "path contains fewer than two poses"

        q0 = np.asarray(path[0], dtype=np.float64)
        previous = q0
        cleared = {pid: False for pid in self.initial_depths}
        absence_start: Dict[str, Optional[float]] = {
            pid: None for pid in self.initial_depths
        }
        contact_eps = max(0.20 * self.new_contact_tolerance, 1e-7)

        for segment_index, endpoint in enumerate(path[1:], start=1):
            endpoint = np.asarray(endpoint, dtype=np.float64)
            samples = interpolate_pose_segment(
                previous,
                endpoint,
                translation_step=self.sweep_translation_step,
                rotation_step=self.sweep_rotation_step,
            )
            for sample in samples:
                if deadline is not None and time.monotonic() >= float(deadline):
                    return False, "exact swept validation timeout"
                displacement = float(np.linalg.norm(sample[:3] - q0[:3]))
                depths = self.penetration_profile(sample)

                for pid, depth in depths.items():
                    if pid not in self.initial_depths:
                        if depth > self.new_contact_tolerance:
                            self._record_blocker(pid)
                            return (
                                False,
                                f"segment {segment_index - 1}->{segment_index}: "
                                f"new exact contact with {pid}: depth={depth:.6g}, "
                                f"allowed={self.new_contact_tolerance:.6g}",
                            )

                for pid, initial_depth in self.initial_depths.items():
                    depth = float(depths.get(pid, 0.0))
                    if depth <= contact_eps:
                        if absence_start[pid] is None:
                            absence_start[pid] = displacement
                        elif (
                            displacement - float(absence_start[pid])
                            >= self.contact_clearance_confirmation_distance
                        ):
                            cleared[pid] = True
                        continue

                    if cleared[pid]:
                        self._record_blocker(pid)
                        return (
                            False,
                            f"segment {segment_index - 1}->{segment_index}: "
                            f"re-contact with {pid} after the initial contact "
                            f"had cleared; depth={depth:.6g}",
                        )

                    absence_start[pid] = None
                    hard = self._hard_allowed_depth(pid)
                    if depth > hard:
                        self._record_blocker(pid)
                        return (
                            False,
                            f"segment {segment_index - 1}->{segment_index}: "
                            f"existing exact-contact growth with {pid}: "
                            f"initial={initial_depth:.6g}, depth={depth:.6g}, "
                            f"hard={hard:.6g}",
                        )
            previous = endpoint

        displacement = float(np.linalg.norm(previous[:3] - q0[:3]))
        if displacement < float(min_removal_distance):
            return (
                False,
                f"removal displacement {displacement:.6g} < "
                f"{min_removal_distance:.6g}",
            )

        valid, reason = self.pose_is_collision_free(previous)
        if not valid:
            return False, reason
        return True, "ok"

def removal_distance_threshold(
    assets: Dict[str, PartAsset],
    move_id: str,
    absolute_distance: float,
    part_diagonal_factor: float,
) -> float:
    part_diag = float(np.linalg.norm(assets[move_id].local_collision_mesh.extents))
    return max(float(absolute_distance), float(part_diagonal_factor) * part_diag)


def validate_kinematic_extraction_path(
    assets: Dict[str, PartAsset],
    move_id: str,
    still_ids: Sequence[str],
    path: Sequence[np.ndarray],
    penetration_tolerance: float,
    sweep_translation_step: float,
    sweep_rotation_step: float,
    min_removal_distance: float,
    min_removal_distance_factor: float,
) -> Tuple[bool, str]:
    validator = ExactCollisionValidator(
        assets=assets,
        move_id=move_id,
        still_ids=still_ids,
        penetration_tolerance=penetration_tolerance,
        sweep_translation_step=sweep_translation_step,
        sweep_rotation_step=sweep_rotation_step,
    )
    threshold = removal_distance_threshold(
        assets,
        move_id,
        absolute_distance=min_removal_distance,
        part_diagonal_factor=min_removal_distance_factor,
    )
    valid, reason = validator.path_is_valid(path, threshold)
    if not valid:
        return False, reason
    if not is_pose_disassembled_from_assets(assets, move_id, still_ids, path[-1]):
        return False, "final pose has not cleared the remaining-assembly envelope"
    return True, "ok"


def compute_parking_position(
    assets: Dict[str, PartAsset],
    step_index: int,
    total_steps: int,
    distance_scale: float = 1.35,
    spacing_scale: float = 0.30,
) -> np.ndarray:
    _, _, center, extent = compute_assembly_bounds(assets)
    total_steps = max(int(total_steps), 1)
    cols = int(math.ceil(math.sqrt(total_steps)))
    row, col = divmod(int(step_index), cols)
    base = center + np.array([distance_scale * extent, 0.0, 0.25 * extent])
    offset_y = (col - 0.5 * (cols - 1)) * spacing_scale * extent
    offset_z = row * spacing_scale * extent
    return base + np.array([0.0, offset_y, offset_z])


def interpolate_qpos(q0, q1, n_frames: int) -> List[np.ndarray]:
    q0 = np.asarray(q0, dtype=np.float64)
    q1 = np.asarray(q1, dtype=np.float64)
    result = []
    for i in range(1, max(int(n_frames), 1) + 1):
        t = i / float(max(int(n_frames), 1))
        q = (1.0 - t) * q0 + t * q1
        q[3:7] = normalize_quat(q[3:7])
        result.append(q)
    return result


def downsample_pose_path(
    path: Optional[Sequence[np.ndarray]],
    max_frames: int,
) -> List[np.ndarray]:
    """Uniformly downsample a pose path while preserving both endpoints."""
    if not path:
        return []
    frames = [np.asarray(q, dtype=np.float64).copy() for q in path]
    max_frames = max(int(max_frames), 2)
    if len(frames) <= max_frames:
        return frames
    indices = np.linspace(0, len(frames) - 1, max_frames)
    indices = np.unique(np.rint(indices).astype(np.int64))
    if indices[0] != 0:
        indices = np.insert(indices, 0, 0)
    if indices[-1] != len(frames) - 1:
        indices = np.append(indices, len(frames) - 1)
    return [frames[int(i)].copy() for i in indices]


def append_parking_motion(
    path: List[np.ndarray],
    assets: Dict[str, PartAsset],
    step_index: int,
    total_steps: int,
    distance_scale: float,
    spacing_scale: float,
    parking_frames: int,
) -> List[np.ndarray]:
    """Route a removed part to parking through an exterior clearance corridor.

    The previous implementation directly interpolated from the extraction pose
    to the parking pose. That straight segment could cross the remaining
    assembly even when the physics-generated extraction itself was valid.
    """
    if not path:
        return path

    result = [np.asarray(q, dtype=np.float64).copy() for q in path]
    xyz_min, xyz_max, center, extent = compute_assembly_bounds(assets)
    target = result[-1].copy()
    target[:3] = compute_parking_position(
        assets,
        step_index=step_index,
        total_steps=total_steps,
        distance_scale=distance_scale,
        spacing_scale=spacing_scale,
    )

    current = result[-1].copy()
    outward = normalize_vector(current[:3] - center, fallback=(0.0, 0.0, 1.0))
    clearance_radius = 1.75 * extent
    current_radius = float(np.linalg.norm(current[:3] - center))

    waypoint_radial = current.copy()
    waypoint_radial[:3] = center + outward * max(clearance_radius, current_radius)

    transit_z = float(xyz_max[2] + 0.75 * extent)
    waypoint_up = waypoint_radial.copy()
    waypoint_up[2] = max(transit_z, waypoint_radial[2])

    waypoint_over = waypoint_up.copy()
    waypoint_over[0] = target[0]
    waypoint_over[1] = target[1]

    total_frames = max(int(parking_frames), 4)
    counts = [
        max(total_frames // 4, 1),
        max(total_frames // 4, 1),
        max(total_frames // 4, 1),
    ]
    counts.append(max(total_frames - sum(counts), 1))
    for waypoint, frames in zip(
        [waypoint_radial, waypoint_up, waypoint_over, target], counts
    ):
        result.extend(interpolate_qpos(result[-1], waypoint, frames))
    return result


def load_axis_map(path: Optional[str]) -> Dict[str, np.ndarray]:
    if path is None:
        return {}
    with open(path, "r", encoding="utf-8") as file:
        data = json.load(file)
    result = {}
    for pid, axis in data.items():
        axis = normalize_vector(axis)
        result[str(pid)] = axis
    return result


def _parse_prior_operation(
    value,
    default_turns: float,
    default_distance: Optional[float],
    default_frames: int,
    default_handedness: float,
    default_axis_mode: str,
) -> LLMPriorOperation:
    if isinstance(value, str):
        value = {"operation": value}
    if value is None:
        value = {}
    if not isinstance(value, dict):
        raise ValueError(f"Invalid LLM prior operation: {value!r}")

    operation = str(value.get("operation", value.get("type", "physics_bfs"))).lower()
    aliases = {
        "axis_locked_unscrew": "unscrew",
        "llm_prior_unscrew": "unscrew",
        "bfs": "physics_bfs",
        "generic": "physics_bfs",
    }
    operation = aliases.get(operation, operation)
    if operation not in {"unscrew", "physics_bfs"}:
        raise ValueError(f"Unsupported LLM prior operation: {operation}")

    axis = value.get("axis")
    axis = None if axis is None else normalize_vector(axis)
    distance = value.get("distance", default_distance)
    if distance is not None:
        distance = float(distance)

    pitch = value.get("pitch")
    if pitch is not None:
        pitch = float(pitch)
        if pitch <= 0.0:
            pitch = None
    clearance = value.get("clearance")
    if clearance is not None:
        clearance = max(float(clearance), 0.0)
    mating_parts_raw = value.get(
        "mating_parts",
        value.get("thread_parts", value.get("paired_parts", [])),
    )
    if mating_parts_raw is None:
        mating_parts_raw = []
    if isinstance(mating_parts_raw, str):
        mating_parts_raw = [mating_parts_raw]
    if not isinstance(mating_parts_raw, (list, tuple)):
        raise ValueError("mating_parts must be a list of part IDs")

    return LLMPriorOperation(
        operation=operation,
        axis=axis,
        turns=float(value.get("turns", default_turns)),
        distance=distance,
        frames=max(int(value.get("frames", default_frames)), 2),
        handedness=float(value.get("handedness", default_handedness)),
        axial_sign=float(value.get("axial_sign", 0.0)),
        max_turns=max(float(value.get("max_turns", 12.0)), 0.5),
        axis_mode=str(value.get("axis_mode", default_axis_mode)),
        pitch=pitch,
        clearance=clearance,
        mating_parts=tuple(str(x) for x in mating_parts_raw),
    )


def load_llm_prior_plan(
    assets: Dict[str, PartAsset],
    enabled: bool,
    prior_file: Optional[str],
    explicit_part_ids: Optional[Sequence[str]],
    default_part_ids: Sequence[str],
    axis_map: Dict[str, np.ndarray],
    enable_unscrew: bool,
    default_turns: float,
    default_distance: Optional[float],
    default_frames: int,
    default_handedness: float,
    default_axis_mode: str,
    explicit_base_part_id: Optional[str] = None,
) -> LLMPriorPlan:
    """
    Load a structured prior generated by an LLM or use the known fastener list.

    Supported JSON example::

        {
          "priority_order": ["009_COMPOUND101", "010_COMPOUND102"],
          "default_operation": {
            "operation": "unscrew",
            "axis": [0, 0, 1],
            "turns": 3,
            "handedness": -1,
            "axial_sign": 1
          },
          "operations": {}
        }
    """
    if not enabled:
        # Strict baseline isolation: a non-Full approach must not open or parse
        # the LLM prior file at all, including for base-part metadata.  The
        # stationary base is supplied independently through --base-part-id.
        plan = LLMPriorPlan.disabled()
        if explicit_base_part_id not in (None, ""):
            plan.base_part_id = str(explicit_base_part_id)
        plan.planning_policy = {}
        print("\n========== LLM Disassembly Prior ==========")
        print("Enabled: False")
        print(
            "[Prior] This approach does not read an LLM prior file and does not "
            "use semantic ordering, semantic fastener labels, unscrew primitives, "
            "or LLM planning policies."
        )
        return plan

    source = "built-in fastener knowledge"
    file_data = {}
    if prior_file is not None:
        with open(prior_file, "r", encoding="utf-8") as file:
            file_data = json.load(file)
        if not isinstance(file_data, dict):
            raise ValueError("--llm-prior-file must contain a JSON object.")
        source = os.path.abspath(prior_file)

    if explicit_part_ids is not None:
        requested_order = [str(pid) for pid in explicit_part_ids]
        source = "--prior-part-ids"
    elif prior_file is not None:
        requested_order = file_data.get(
            "priority_order",
            file_data.get("prior_part_ids", file_data.get("parts", [])),
        )
        requested_order = [str(pid) for pid in requested_order]
    else:
        requested_order = [str(pid) for pid in default_part_ids]

    # Preserve LLM order while removing duplicate IDs.
    deduplicated = []
    seen = set()
    for pid in requested_order:
        if pid not in seen:
            deduplicated.append(pid)
            seen.add(pid)

    missing = [pid for pid in deduplicated if pid not in assets]
    priority_order = [pid for pid in deduplicated if pid in assets]

    raw_operations = file_data.get("operations", file_data.get("part_operations", {}))
    if raw_operations is None:
        raw_operations = {}
    if not isinstance(raw_operations, dict):
        raise ValueError("LLM prior field 'operations' must be a JSON object.")

    # A compact JSON may define one operation template shared by all prior
    # fasteners. Per-part entries in ``operations`` override only the fields
    # that differ. This keeps the LLM prior semantic rather than encoding the
    # complete structural sequence by hand.
    raw_default_operation = file_data.get("default_operation", {})
    if raw_default_operation is None:
        raw_default_operation = {}
    if not isinstance(raw_default_operation, (dict, str)):
        raise ValueError(
            "LLM prior field 'default_operation' must be a JSON object or string."
        )

    # The base/reference part is a hard planning constraint.  A CLI value
    # has highest priority; otherwise a reviewed LLM/CAD policy JSON may define
    # it.  Unlike v8, the JSON value is intentionally honored because the user
    # has explicitly identified 000_COMPOUND004 as the physical fixture/base.
    suggested_base = file_data.get(
        "base_part_id",
        file_data.get("anchor_part_id", file_data.get("reference_part_id")),
    )
    raw_base_part_id = (
        explicit_base_part_id
        if explicit_base_part_id not in (None, "")
        else suggested_base
    )
    base_part_id = None if raw_base_part_id in (None, "") else str(raw_base_part_id)
    if base_part_id is not None and base_part_id not in assets:
        raise ValueError(
            f"Configured base part {base_part_id!r} is not present in the assembly."
        )
    if base_part_id in priority_order:
        print(
            f"[Prior] Base part {base_part_id} was present in priority_order; "
            "it will remain stationary and is removed from the candidate list."
        )
        priority_order = [pid for pid in priority_order if pid != base_part_id]

    operations: Dict[str, LLMPriorOperation] = {}
    for pid in priority_order:
        default_operation = "unscrew" if enable_unscrew else "physics_bfs"
        per_part_value = raw_operations.get(pid)
        if per_part_value is None:
            if isinstance(raw_default_operation, dict):
                raw_value = dict(raw_default_operation)
                raw_value.setdefault("operation", default_operation)
            elif isinstance(raw_default_operation, str) and raw_default_operation.strip():
                raw_value = raw_default_operation
            else:
                raw_value = {"operation": default_operation}
        elif isinstance(per_part_value, dict) and isinstance(raw_default_operation, dict):
            raw_value = dict(raw_default_operation)
            raw_value.update(per_part_value)
            raw_value.setdefault("operation", default_operation)
        else:
            raw_value = per_part_value
        operation = _parse_prior_operation(
            raw_value,
            default_turns=default_turns,
            default_distance=default_distance,
            default_frames=default_frames,
            default_handedness=default_handedness,
            default_axis_mode=default_axis_mode,
        )
        if pid in axis_map:
            operation.axis = axis_map[pid]
        if not enable_unscrew and operation.operation == "unscrew":
            operation.operation = "physics_bfs"
        operations[pid] = operation

    raw_policy = file_data.get("planning_policy", {})
    if raw_policy is None:
        raw_policy = {}
    if not isinstance(raw_policy, dict):
        raise ValueError("LLM prior field 'planning_policy' must be a JSON object.")

    # Defaults used only when the semantic prior is enabled. Collision-derived
    # blockers are never deferred merely because their geometric centre lies
    # below the currently tested part.
    planning_policy = {
        "structural_order": "top_down_left_right",
        "top_axis": "z",
        "top_direction": "descending",
        "left_axis": "x",
        "left_direction": "ascending",
        "height_reference": "upper_surface",
        "layer_tolerance_ratio": 0.06,
        "defer_lower_blockers": False,
        "strict_top_layer": True,
        "translation_priority": ["+z", "-z", "+x", "-x", "+y", "-y"],
        "rotation_search_start_depth": 3,
        "random_rotation_actions": 12,
        "random_rotation_scale": 0.35,
    }
    planning_policy.update(raw_policy)

    print("\n========== LLM Disassembly Prior ==========")
    print("Enabled: True")
    print("Source:", source)
    print("Priority order present in assembly:", priority_order)
    print("Base/reference part:", base_part_id if base_part_id is not None else "auto")
    if missing:
        print("[Prior] IDs not found in this assembly:", missing)
    if not priority_order:
        print("[Prior] No known prior parts were found; geometric exterior ordering will be used.")
    else:
        for index, pid in enumerate(priority_order, start=1):
            operation = operations[pid]
            print(
                f"  {index:02d}. {pid}: operation={operation.operation}, "
                f"axis={None if operation.axis is None else operation.axis.round(4).tolist()}, "
                f"mating_parts={list(operation.mating_parts)}"
            )

    return LLMPriorPlan(
        enabled=True,
        source=source,
        priority_order=priority_order,
        operations=operations,
        base_part_id=base_part_id,
        planning_policy=planning_policy,
    )


def orient_axis_outward(axis, part_center, assembly_center) -> np.ndarray:
    axis = normalize_vector(axis)
    outward = np.asarray(part_center) - np.asarray(assembly_center)
    if np.linalg.norm(outward) > 1e-12 and np.dot(axis, outward) < 0:
        axis = -axis
    return axis


def estimate_unscrew_axis(
    asset: PartAsset,
    assembly_center: np.ndarray,
    mode: str = "auto",
) -> np.ndarray:
    vertices = np.asarray(asset.local_visual_mesh.vertices, dtype=np.float64)
    centered = vertices - vertices.mean(axis=0, keepdims=True)
    try:
        covariance = np.cov(centered.T)
        values, vectors = np.linalg.eigh(covariance)
        order = np.argsort(values)
        values, vectors = values[order], vectors[:, order]
        if mode == "min":
            axis = vectors[:, 0]
        elif mode == "max":
            axis = vectors[:, -1]
        elif values[-1] > 1.8 * max(values[-2], 1e-12):
            axis = vectors[:, -1]
        else:
            axis = vectors[:, 0]
    except Exception:
        extents = asset.local_visual_mesh.extents
        if mode == "max":
            axis = np.eye(3)[int(np.argmax(extents))]
        elif mode == "min":
            axis = np.eye(3)[int(np.argmin(extents))]
        else:
            axis = np.eye(3)[int(np.argmin(extents))]
    return orient_axis_outward(axis, asset.center, assembly_center)


def make_unscrew_path(
    asset: PartAsset,
    assets: Dict[str, PartAsset],
    step_index: int,
    total_steps: int,
    turns: float,
    distance: Optional[float],
    frames: int,
    axis: Optional[np.ndarray],
    axis_mode: str,
    handedness: float,
    parking_enabled: bool,
    parking_distance_scale: float,
    parking_spacing_scale: float,
    parking_frames: int,
) -> List[np.ndarray]:
    """Generate a two-stage screw-removal primitive.

    The previous implementation coupled the *entire* extraction distance to
    the screw rotation.  For the current assembly this produced an unrealistically
    large pitch (about 0.16 m/rev after normalization), forcing the nut through
    neighbouring geometry.  Here the helical phase uses a pitch estimated from
    fastener diameter, followed by a pure axial withdrawal phase.
    """
    _, _, assembly_center, assembly_extent = compute_assembly_bounds(assets)
    axis = (
        estimate_unscrew_axis(asset, assembly_center, axis_mode)
        if axis is None
        else orient_axis_outward(axis, asset.center, assembly_center)
    )

    vertices = np.asarray(asset.local_visual_mesh.vertices, dtype=np.float64)
    axial_coord = vertices @ axis
    axial_length = max(float(np.ptp(axial_coord)), 1e-6)
    radial_vec = vertices - np.outer(axial_coord, axis)
    radial_radius = float(np.percentile(np.linalg.norm(radial_vec, axis=1), 95))
    diameter = max(2.0 * radial_radius, 1e-6)

    # Generic metric-thread prior: pitch is roughly 8-15% of nominal diameter.
    pitch = float(np.clip(
        0.12 * diameter,
        0.0015 * assembly_extent,
        0.02 * assembly_extent,
    ))
    helical_distance = min(
        abs(float(turns)) * pitch,
        max(0.75 * axial_length, 0.04 * assembly_extent),
    )
    if distance is None or distance <= 0:
        extraction_distance = max(
            1.25 * axial_length,
            0.08 * assembly_extent,
        )
    else:
        extraction_distance = float(distance)

    total_frames = max(int(frames), 12)
    helical_frames = max(int(round(0.60 * total_frames)), 6)
    extract_frames = max(total_frames - helical_frames, 6)
    q0 = make_initial_qpos(asset)
    result = [q0.copy()]

    # Stage A: screw motion with a physically plausible pitch.
    for i in range(1, helical_frames + 1):
        t = i / float(helical_frames)
        angle = handedness * 2.0 * math.pi * turns * t
        q = q0.copy()
        q[:3] = q0[:3] + axis * helical_distance * t
        q[3:7] = quat_multiply(axis_angle_to_quat(axis, angle), q0[3:7])
        result.append(q)

    # Stage B: after thread release, withdraw without additional rotation.
    q_release = result[-1].copy()
    for i in range(1, extract_frames + 1):
        t = i / float(extract_frames)
        q = q_release.copy()
        q[:3] = q_release[:3] + axis * extraction_distance * t
        result.append(q)

    if parking_enabled:
        result = append_parking_motion(
            result,
            assets,
            step_index,
            total_steps,
            parking_distance_scale,
            parking_spacing_scale,
            parking_frames,
        )
    print(
        f"[Unscrew] {asset.part_id}: axis={axis.round(4).tolist()}, "
        f"turns={turns}, pitch={pitch:.6f}, helical={helical_distance:.6f}, "
        f"extract={extraction_distance:.6f}, handedness={handedness:+.0f}"
    )
    return result


# ============================================================================
# Genesis path-planning world
# ============================================================================


def _collision_mesh_kwargs(
    asset: PartAsset,
    fixed: bool,
    visualization: bool,
    collision_face_num: int,
    collision_mode: str,
    watertighten: int,
) -> dict:
    # Preserve holes, shafts and fastener geometry in SDF mode.  Automatically
    # convexifying every non-watertight proxy turns screws and thin plates into
    # oversized solid hulls and creates false blocking contacts.  Genesis
    # watertightening is used instead; convexification is now explicit only.
    convexify = collision_mode == "convex"
    return {
        "file": asset.collision_path,
        "pos": tuple(asset.center.tolist()),
        "quat": (1.0, 0.0, 0.0, 0.0),
        "fixed": fixed,
        "collision": True,
        "visualization": visualization,
        "decimate": True,
        "decimate_face_num": int(collision_face_num),
        "decimate_aggressiveness": 2,
        "convexify": convexify,
        "decompose_object_error_threshold": 0.08,
        "watertighten": 0 if asset.collision_watertight else int(watertighten),
        "recompute_inertia": not fixed,
        "align": False,
        "file_meshes_are_zup": True,
    }


class GenesisDisassemblyWorld:
    def __init__(
        self,
        assets: Dict[str, PartAsset],
        move_id: str,
        still_ids: Sequence[str],
        show_viewer: bool,
        dt: float,
        substeps: int,
        force_mag: float,
        density: float,
        friction: float,
        sdf_cell_size: float,
        sdf_min_res: int,
        sdf_max_res: int,
        collision_face_num: int,
        collision_mode: str,
        watertighten: int,
        free_dof_order: str,
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
        min_removal_distance: float,
        min_removal_distance_factor: float,
    ):
        self.assets = assets
        self.move_id = move_id
        self.still_ids = list(still_ids)
        self.show_viewer = show_viewer
        self.dt = float(dt)
        self.substeps = int(substeps)
        self.force_mag = float(force_mag)
        self.density = float(density)
        self.friction = float(friction)
        self.sdf_cell_size = float(sdf_cell_size)
        self.sdf_min_res = int(sdf_min_res)
        self.sdf_max_res = int(sdf_max_res)
        self.collision_face_num = int(collision_face_num)
        self.collision_mode = collision_mode
        self.watertighten = int(watertighten)
        self.free_dof_order_request = free_dof_order
        self.penetration_tolerance = float(penetration_tolerance)
        self.sweep_translation_step = float(sweep_translation_step)
        self.sweep_rotation_step = float(sweep_rotation_step)
        self.min_removal_distance = removal_distance_threshold(
            assets,
            move_id,
            absolute_distance=min_removal_distance,
            part_diagonal_factor=min_removal_distance_factor,
        )

        self.scene = None
        self.move_entity = None
        self.still_entities = {}
        self._collision_manager = None
        self._still_hull_world = None
        self._free_dof_order = "angular-linear"
        self._initial_qpos = make_initial_qpos(self.assets[self.move_id])
        self.last_invalid_qpos: Optional[np.ndarray] = None
        self._path_validator = ExactCollisionValidator(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=self.still_ids,
            penetration_tolerance=self.penetration_tolerance,
            sweep_translation_step=self.sweep_translation_step,
            sweep_rotation_step=self.sweep_rotation_step,
        )

        self._build_scene()
        self._build_disassembly_checker()

    def _build_scene(self):
        _, _, center, extent = compute_assembly_bounds(self.assets)
        scene_kwargs = dict(
            sim_options=construct_compatible(
                gs.options.SimOptions,
                dt=self.dt,
                substeps=self.substeps,
            ),
            rigid_options=construct_compatible(
                gs.options.RigidOptions,
                gravity=(0.0, 0.0, 0.0),
                enable_collision=True,
                constraint_timeconst=0.01,
                max_contacts=8192,
                max_collision_pairs=4096,
                enable_multi_contact=True,
            ),
            viewer_options=construct_compatible(
                gs.options.ViewerOptions,
                camera_pos=(
                    float(center[0] + 1.2 * extent),
                    float(center[1] - 1.6 * extent),
                    float(center[2] + 0.9 * extent),
                ),
                camera_lookat=tuple(center.tolist()),
                camera_fov=45,
            ),
            show_viewer=self.show_viewer,
        )
        scene_kwargs.update(scene_profiling_kwargs(False))
        self.scene = gs.Scene(**scene_kwargs)

        for pid in self.still_ids:
            asset = self.assets[pid]
            add_kwargs = dict(
                material=construct_compatible(
                    gs.materials.Rigid,
                    rho=self.density,
                    friction=self.friction,
                    sdf_cell_size=self.sdf_cell_size,
                    sdf_min_res=self.sdf_min_res,
                    sdf_max_res=self.sdf_max_res,
                ),
                morph=construct_compatible(
                    gs.morphs.Mesh,
                    **_collision_mesh_kwargs(
                        asset,
                        fixed=True,
                        visualization=self.show_viewer,
                        collision_face_num=self.collision_face_num,
                        collision_mode=self.collision_mode,
                        watertighten=self.watertighten,
                    ),
                ),
            )
            if self.show_viewer:
                add_kwargs["surface"] = make_surface(asset.color)
            self.still_entities[pid] = self.scene.add_entity(**add_kwargs)

        move_asset = self.assets[self.move_id]
        add_kwargs = dict(
            material=construct_compatible(
                    gs.materials.Rigid,
                    rho=self.density,
                    friction=self.friction,
                    sdf_cell_size=self.sdf_cell_size,
                    sdf_min_res=self.sdf_min_res,
                    sdf_max_res=self.sdf_max_res,
                ),
            morph=construct_compatible(
                gs.morphs.Mesh,
                **_collision_mesh_kwargs(
                    move_asset,
                    fixed=False,
                    visualization=self.show_viewer,
                    collision_face_num=self.collision_face_num,
                    collision_mode=self.collision_mode,
                    watertighten=self.watertighten,
                ),
            ),
        )
        if self.show_viewer:
            add_kwargs["surface"] = make_surface(move_asset.color)
        self.move_entity = self.scene.add_entity(**add_kwargs)

        self.scene.build()
        self.set_qpos(self.initial_qpos(), zero_velocity=True)
        self.scene.step()
        self._free_dof_order = self._resolve_free_dof_order()

    def initial_qpos(self) -> np.ndarray:
        return self._initial_qpos.copy()

    def get_qpos(self) -> np.ndarray:
        return to_numpy(self.move_entity.get_qpos()).astype(np.float64)

    def set_qpos(self, qpos: np.ndarray, zero_velocity: bool = True):
        qpos = np.asarray(qpos, dtype=np.float64).copy()
        qpos[3:7] = normalize_quat(qpos[3:7])
        self.move_entity.set_qpos(qpos, zero_velocity=zero_velocity)
        if zero_velocity:
            self.move_entity.set_dofs_velocity(None)

    def _resolve_free_dof_order(self) -> str:
        global _FREE_DOF_ORDER_CACHE
        request = self.free_dof_order_request
        if request in ("linear-angular", "angular-linear"):
            return request
        if _FREE_DOF_ORDER_CACHE is not None:
            return _FREE_DOF_ORDER_CACHE

        # Auto-detect in collision-free space by exciting DOF 0.
        q_saved = self.get_qpos()
        _, _, center, extent = compute_assembly_bounds(self.assets)
        q_test = q_saved.copy()
        q_test[:3] = center + np.array([4.0 * extent, 4.0 * extent, 4.0 * extent])
        self.set_qpos(q_test, zero_velocity=True)
        self.scene.step()
        before = self.get_qpos()
        control = np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        control[0] = max(self.force_mag, 1.0)
        for _ in range(6):
            self.move_entity.control_dofs_force(control, np.arange(self.move_entity.n_dofs))
            self.scene.step()
        after = self.get_qpos()
        self.move_entity.control_dofs_force(
            np.zeros(self.move_entity.n_dofs), np.arange(self.move_entity.n_dofs)
        )
        position_change = float(np.linalg.norm(after[:3] - before[:3]))
        rotation_change = quat_distance(after[3:7], before[3:7])
        self.set_qpos(q_saved, zero_velocity=True)
        self.scene.step()

        if position_change > max(1e-8, 0.1 * rotation_change):
            detected = "linear-angular"
        else:
            detected = "angular-linear"
        _FREE_DOF_ORDER_CACHE = detected
        print(
            f"[Genesis] Auto-detected free DOF order: {detected} "
            f"(position_delta={position_change:.3e}, rotation_delta={rotation_change:.3e})"
        )
        return detected

    def _action_to_control(self, action: np.ndarray) -> np.ndarray:
        action = np.asarray(action, dtype=np.float64)
        if action.shape != (6,):
            raise ValueError("Action must be [Fx, Fy, Fz, Tx, Ty, Tz].")
        norm = float(np.linalg.norm(action))
        if norm < 1e-12:
            return np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        action = action / norm * self.force_mag
        if self._free_dof_order == "linear-angular":
            ordered = action
        else:
            ordered = np.concatenate([action[3:6], action[0:3]])
        control = np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        control[: min(len(control), 6)] = ordered[: min(len(control), 6)]
        return control

    def _wrench_to_control(self, force: np.ndarray, torque: np.ndarray) -> np.ndarray:
        force = np.asarray(force, dtype=np.float64).reshape(3)
        torque = np.asarray(torque, dtype=np.float64).reshape(3)
        if self._free_dof_order == "linear-angular":
            ordered = np.concatenate([force, torque])
        else:
            ordered = np.concatenate([torque, force])
        control = np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        control[: min(len(control), 6)] = ordered[: min(len(control), 6)]
        return control

    def step_unscrew_wrench(
        self,
        force: np.ndarray,
        torque: np.ndarray,
        n_steps: int,
        axis: np.ndarray,
        origin: np.ndarray,
        lateral_limit: float,
    ) -> Tuple[List[np.ndarray], bool, str]:
        """Apply a physical force/torque and validate only true unscrew errors."""
        control = self._wrench_to_control(force, torque)
        dofs = np.arange(self.move_entity.n_dofs)
        states: List[np.ndarray] = []
        previous = self.get_qpos()
        self.last_invalid_qpos = None
        valid = True
        reason = "ok"
        for _ in range(max(int(n_steps), 1)):
            self.move_entity.control_dofs_force(control, dofs)
            self.scene.step()
            current = self.get_qpos()
            current[3:7] = normalize_quat(current[3:7])
            segment_valid, segment_reason = self._path_validator.unscrew_segment_is_valid(
                previous,
                current,
                axis=axis,
                origin=origin,
                lateral_limit=lateral_limit,
            )
            if not segment_valid:
                valid = False
                reason = segment_reason
                self.last_invalid_qpos = current.copy()
                break
            states.append(current.copy())
            previous = current
        self.move_entity.control_dofs_force(np.zeros_like(control), dofs)
        if not valid:
            self.set_qpos(previous, zero_velocity=True)
        return states, valid, reason

    def step_action(
        self, action: np.ndarray, n_steps: int
    ) -> Tuple[List[np.ndarray], bool, str]:
        """Advance dynamics while validating every discrete swept segment.

        Pure or mixed translational actions receive the bounded initial-contact
        release envelope.  Rotation-only actions do not.  This prevents the
        first Genesis depenetration impulse from invalidating an otherwise
        correct pull-out direction while keeping new contacts strict.
        """
        action = np.asarray(action, dtype=np.float64)
        control = self._action_to_control(action)
        translation = action[:3]
        release_direction = (
            normalize_vector(translation)
            if float(np.linalg.norm(translation)) > 1e-12
            else None
        )

        dofs = np.arange(self.move_entity.n_dofs)
        states: List[np.ndarray] = []
        previous = self.get_qpos()
        self.last_invalid_qpos = None
        valid = True
        reason = "ok"
        for _ in range(max(int(n_steps), 1)):
            self.move_entity.control_dofs_force(control, dofs)
            self.scene.step()
            current = self.get_qpos()
            current[3:7] = normalize_quat(current[3:7])
            segment_valid, segment_reason = self._path_validator.segment_is_valid(
                previous,
                current,
                release_direction=release_direction,
                allow_initial_release=release_direction is not None,
            )
            if not segment_valid:
                valid = False
                reason = segment_reason
                self.last_invalid_qpos = current.copy()
                break
            states.append(current.copy())
            previous = current
        self.move_entity.control_dofs_force(np.zeros_like(control), dofs)
        if not valid:
            self.set_qpos(previous, zero_velocity=True)
        return states, valid, reason

    def _build_disassembly_checker(self):
        if not self.still_ids:
            return
        still_hulls = []
        for pid in self.still_ids:
            still_hulls.append(
                transform_mesh_by_qpos(self.assets[pid].local_hull, make_initial_qpos(self.assets[pid]))
            )
        combined = trimesh.util.concatenate(still_hulls)
        try:
            self._still_hull_world = combined.convex_hull
        except Exception:
            self._still_hull_world = combined
        try:
            manager = trimesh.collision.CollisionManager()
            manager.add_object("still", self._still_hull_world)
            self._collision_manager = manager
        except Exception as exc:
            print(f"[Warning] python-fcl unavailable; AABB fallback will be used: {exc}")
            self._collision_manager = None

    def is_disassembled(self, qpos: Optional[np.ndarray] = None) -> bool:
        if not self.still_ids or self._still_hull_world is None:
            return True
        qpos = self.get_qpos() if qpos is None else np.asarray(qpos, dtype=np.float64)

        # Never accept the initial assembled pose as "Start with goal".
        # A real removal must first move a meaningful distance.
        displacement = float(np.linalg.norm(qpos[:3] - self._initial_qpos[:3]))
        if displacement < self.min_removal_distance:
            return False

        # The final pose must be free from every remaining collision mesh, not
        # merely separated from one coarse convex hull.
        collision_free, _ = self._path_validator.pose_is_collision_free(qpos)
        if not collision_free:
            return False

        move_hull = transform_mesh_by_qpos(self.assets[self.move_id].local_hull, qpos)
        has_collision = None
        if self._collision_manager is not None:
            try:
                has_collision = bool(self._collision_manager.in_collision_single(move_hull))
            except Exception:
                pass
        if has_collision is None:
            min_m, max_m = move_hull.bounds
            min_s, max_s = self._still_hull_world.bounds
            has_collision = aabb_overlap(min_m, max_m, min_s, max_s)
        if has_collision:
            return False

        min_m, max_m = move_hull.bounds
        min_s, max_s = self._still_hull_world.bounds
        return not (
            aabb_contains(min_m, max_m, min_s, max_s)
            or aabb_contains(min_s, max_s, min_m, max_m)
        )

    def close(self):
        try:
            self.move_entity = None
            self.still_entities.clear()
            self._collision_manager = None
            self._still_hull_world = None
            self._path_validator = None
            if self.scene is not None:
                scene = self.scene
                self.scene = None
                if hasattr(scene, "destroy"):
                    scene.destroy()
                del scene
            gc.collect()
        except Exception as exc:
            print(f"[Warning] Scene cleanup failed: {exc}")


# ============================================================================
# Adaptive physics unscrew planner
# ============================================================================


class SemanticScrewWorld:
    """Genesis world for one semantically known screw operation.

    A detailed CAD screw and its threaded hole often overlap after mesh
    simplification and SDF conversion.  Letting the generic contact solver
    handle that pair makes the free screw explode sideways at the first step.
    Here the known mating pair is therefore removed from *dynamic* collision
    response and replaced by a one-DOF virtual screw joint.  Genesis still
    supplies the force/torque response, while the virtual joint projects the
    resulting pose onto the screw manifold.  All non-mating collisions remain
    active in Genesis and are checked independently with python-fcl.
    """

    def __init__(
        self,
        assets: Dict[str, PartAsset],
        move_id: str,
        still_ids: Sequence[str],
        mating_parts: Sequence[str],
        show_viewer: bool,
        dt: float,
        substeps: int,
        density: float,
        friction: float,
        sdf_cell_size: float,
        sdf_min_res: int,
        sdf_max_res: int,
        collision_face_num: int,
        collision_mode: str,
        watertighten: int,
        free_dof_order: str,
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
    ):
        self.assets = assets
        self.move_id = move_id
        self.all_still_ids = list(still_ids)
        mating_set = {str(pid) for pid in mating_parts if pid in still_ids}
        self.mating_parts = sorted(mating_set)
        self.collision_still_ids = [
            pid for pid in self.all_still_ids if pid not in mating_set
        ]
        self.show_viewer = bool(show_viewer)
        self.dt = float(dt)
        self.substeps = int(substeps)
        self.density = float(density)
        self.friction = float(friction)
        self.sdf_cell_size = float(sdf_cell_size)
        self.sdf_min_res = int(sdf_min_res)
        self.sdf_max_res = int(sdf_max_res)
        self.collision_face_num = int(collision_face_num)
        self.collision_mode = str(collision_mode)
        self.watertighten = int(watertighten)
        self.free_dof_order_request = str(free_dof_order)
        self.initial = make_initial_qpos(self.assets[self.move_id])
        self.scene = None
        self.move_entity = None
        self.fixed_entities: Dict[str, object] = {}
        self._free_dof_order = "angular-linear"

        self.nonmating_validator = ExactCollisionValidator(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=self.collision_still_ids,
            penetration_tolerance=penetration_tolerance,
            sweep_translation_step=sweep_translation_step,
            sweep_rotation_step=sweep_rotation_step,
        )
        self.final_validator = ExactCollisionValidator(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=self.all_still_ids,
            penetration_tolerance=penetration_tolerance,
            sweep_translation_step=sweep_translation_step,
            sweep_rotation_step=sweep_rotation_step,
        )
        self._build_scene()

    def _build_scene(self):
        _, _, center, extent = compute_assembly_bounds(self.assets)
        scene_kwargs = dict(
            sim_options=construct_compatible(
                gs.options.SimOptions,
                dt=self.dt,
                substeps=self.substeps,
            ),
            rigid_options=construct_compatible(
                gs.options.RigidOptions,
                gravity=(0.0, 0.0, 0.0),
                enable_collision=True,
                constraint_timeconst=0.02,
                max_contacts=8192,
                max_collision_pairs=4096,
                enable_multi_contact=True,
            ),
            viewer_options=construct_compatible(
                gs.options.ViewerOptions,
                camera_pos=(
                    float(center[0] + 1.2 * extent),
                    float(center[1] - 1.6 * extent),
                    float(center[2] + 0.9 * extent),
                ),
                camera_lookat=tuple(center.tolist()),
                camera_fov=45,
                run_in_thread=False,
            ),
            show_viewer=self.show_viewer,
        )
        scene_kwargs.update(scene_profiling_kwargs(False))
        self.scene = gs.Scene(**scene_kwargs)

        # Non-mating parts keep full rigid collision response.
        for pid in self.collision_still_ids:
            asset = self.assets[pid]
            kwargs = dict(
                material=construct_compatible(
                    gs.materials.Rigid,
                    rho=self.density,
                    friction=self.friction,
                    sdf_cell_size=self.sdf_cell_size,
                    sdf_min_res=self.sdf_min_res,
                    sdf_max_res=self.sdf_max_res,
                ),
                morph=construct_compatible(
                    gs.morphs.Mesh,
                    **_collision_mesh_kwargs(
                        asset,
                        fixed=True,
                        visualization=self.show_viewer,
                        collision_face_num=self.collision_face_num,
                        collision_mode=self.collision_mode,
                        watertighten=self.watertighten,
                    ),
                ),
            )
            if self.show_viewer:
                kwargs["surface"] = make_surface(asset.color)
            self.fixed_entities[pid] = self.scene.add_entity(**kwargs)

        # Mating parts are visual-only in this special operation.  Their
        # thread/hole interaction is represented by the virtual screw joint.
        if self.show_viewer:
            for pid in self.mating_parts:
                asset = self.assets[pid]
                self.fixed_entities[pid] = self.scene.add_entity(
                    material=construct_compatible(gs.materials.Rigid, rho=self.density),
                    morph=construct_compatible(
                        gs.morphs.Mesh,
                        file=asset.visual_path,
                        pos=tuple(asset.center.tolist()),
                        quat=(1.0, 0.0, 0.0, 0.0),
                        fixed=True,
                        collision=False,
                        visualization=True,
                        decimate=False,
                        convexify=False,
                        watertighten=0,
                        recompute_inertia=False,
                        align=False,
                        file_meshes_are_zup=True,
                    ),
                    surface=make_surface(asset.color),
                )

        asset = self.assets[self.move_id]
        kwargs = dict(
            material=construct_compatible(
                gs.materials.Rigid,
                rho=self.density,
                friction=self.friction,
                sdf_cell_size=self.sdf_cell_size,
                sdf_min_res=self.sdf_min_res,
                sdf_max_res=self.sdf_max_res,
            ),
            morph=construct_compatible(
                gs.morphs.Mesh,
                **_collision_mesh_kwargs(
                    asset,
                    fixed=False,
                    visualization=self.show_viewer,
                    collision_face_num=self.collision_face_num,
                    collision_mode=self.collision_mode,
                    watertighten=self.watertighten,
                ),
            ),
        )
        if self.show_viewer:
            kwargs["surface"] = make_surface(asset.color)
        self.move_entity = self.scene.add_entity(**kwargs)
        self.scene.build()
        self.set_qpos(self.initial, zero_velocity=True)
        self._free_dof_order = self._resolve_free_dof_order()

    def get_qpos(self) -> np.ndarray:
        q = to_numpy(self.move_entity.get_qpos()).astype(np.float64)
        q[3:7] = normalize_quat(q[3:7])
        return q

    def set_qpos(self, qpos: np.ndarray, zero_velocity: bool = True):
        q = np.asarray(qpos, dtype=np.float64).copy()
        q[3:7] = normalize_quat(q[3:7])
        self.move_entity.set_qpos(q, zero_velocity=zero_velocity)
        if zero_velocity:
            self.move_entity.set_dofs_velocity(None)

    def _resolve_free_dof_order(self) -> str:
        global _FREE_DOF_ORDER_CACHE
        if self.free_dof_order_request in {"linear-angular", "angular-linear"}:
            return self.free_dof_order_request
        if _FREE_DOF_ORDER_CACHE is not None:
            return _FREE_DOF_ORDER_CACHE

        q_saved = self.get_qpos()
        _, _, center, extent = compute_assembly_bounds(self.assets)
        q_test = q_saved.copy()
        q_test[:3] = center + np.array([4.0 * extent, 4.0 * extent, 4.0 * extent])
        self.set_qpos(q_test, zero_velocity=True)
        before = self.get_qpos()
        control = np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        control[0] = 0.5
        dofs = np.arange(self.move_entity.n_dofs)
        for _ in range(5):
            self.move_entity.control_dofs_force(control, dofs)
            self.scene.step()
        after = self.get_qpos()
        self.move_entity.control_dofs_force(np.zeros_like(control), dofs)
        position_change = float(np.linalg.norm(after[:3] - before[:3]))
        rotation_change = quat_distance(after[3:7], before[3:7])
        self.set_qpos(q_saved, zero_velocity=True)
        detected = (
            "linear-angular"
            if position_change > max(1e-8, 0.1 * rotation_change)
            else "angular-linear"
        )
        _FREE_DOF_ORDER_CACHE = detected
        print(
            f"[Genesis] Auto-detected free DOF order: {detected} "
            f"(position_delta={position_change:.3e}, rotation_delta={rotation_change:.3e})"
        )
        return detected

    def _wrench_to_control(self, force: np.ndarray, torque: np.ndarray) -> np.ndarray:
        force = np.asarray(force, dtype=np.float64).reshape(3)
        torque = np.asarray(torque, dtype=np.float64).reshape(3)
        ordered = (
            np.concatenate([force, torque])
            if self._free_dof_order == "linear-angular"
            else np.concatenate([torque, force])
        )
        result = np.zeros(self.move_entity.n_dofs, dtype=np.float64)
        result[: min(6, len(result))] = ordered[: min(6, len(result))]
        return result

    def apply_wrench_once(
        self,
        force: np.ndarray,
        torque: np.ndarray,
    ) -> np.ndarray:
        control = self._wrench_to_control(force, torque)
        dofs = np.arange(self.move_entity.n_dofs)
        self.move_entity.control_dofs_force(control, dofs)
        self.scene.step()
        raw = self.get_qpos()
        self.move_entity.control_dofs_force(np.zeros_like(control), dofs)
        return raw

    def project_thread_step(
        self,
        previous: np.ndarray,
        origin: np.ndarray,
        initial_quat: np.ndarray,
        axis: np.ndarray,
        torque_sign: float,
        pitch: float,
        theta_progress: float,
        axial_force: float,
        torque: float,
        max_angle_step: float,
    ) -> Tuple[np.ndarray, float, bool, str, float, float]:
        """Apply a wrench, then project the response onto z=p*theta/(2*pi)."""
        raw = self.apply_wrench_once(
            force=axis * axial_force,
            torque=axis * (torque_sign * torque),
        )
        raw_dtheta = torque_sign * signed_rotation_about_axis(
            previous[3:7], raw[3:7], axis
        )
        raw_dz = float(np.dot(raw[:3] - previous[:3], axis))
        theta_from_z = max(raw_dz, 0.0) * (2.0 * math.pi / max(pitch, 1e-9))
        # Both generalized-force channels contribute to the same screw DOF.
        dtheta = 0.75 * max(raw_dtheta, 0.0) + 0.25 * theta_from_z
        # A virtual screw joint converts the applied wrench into its single
        # generalized coordinate.  The small commanded term prevents a free-
        # body solver with per-step velocity reset from reporting exactly zero
        # motion even though a non-zero wrench is applied.  Collision checks
        # against non-mating parts still gate every projected increment.
        generalized_effort = abs(float(torque)) + abs(float(axial_force)) * pitch / (2.0 * math.pi)
        commanded_dtheta = 50.0 * generalized_effort * self.dt
        dtheta = max(dtheta, commanded_dtheta)
        dtheta = float(np.clip(dtheta, 0.0, max_angle_step))
        new_theta = theta_progress + dtheta
        z = pitch * new_theta / (2.0 * math.pi)
        projected = self.initial.copy()
        projected[:3] = np.asarray(origin, dtype=np.float64) + axis * z
        projected[3:7] = quat_multiply(
            axis_angle_to_quat(axis, torque_sign * new_theta),
            initial_quat,
        )
        valid, reason = self.nonmating_validator.segment_is_valid(previous, projected)
        if not valid:
            self.set_qpos(previous, zero_velocity=True)
            return previous.copy(), theta_progress, False, reason, raw_dtheta, raw_dz
        self.set_qpos(projected, zero_velocity=True)
        return projected, new_theta, True, "ok", raw_dtheta, raw_dz

    def project_extract_step(
        self,
        previous: np.ndarray,
        origin: np.ndarray,
        fixed_quat: np.ndarray,
        axis: np.ndarray,
        z_progress: float,
        axial_force: float,
        max_translation_step: float,
    ) -> Tuple[np.ndarray, float, bool, str, float]:
        raw = self.apply_wrench_once(force=axis * axial_force, torque=np.zeros(3))
        raw_dz = float(np.dot(raw[:3] - previous[:3], axis))
        commanded_dz = 0.10 * abs(float(axial_force)) * self.dt
        dz = max(max(raw_dz, 0.0), commanded_dz)
        dz = float(np.clip(dz, 0.0, max_translation_step))
        new_z = z_progress + dz
        projected = self.initial.copy()
        projected[:3] = np.asarray(origin, dtype=np.float64) + axis * new_z
        projected[3:7] = normalize_quat(fixed_quat)
        valid, reason = self.nonmating_validator.segment_is_valid(previous, projected)
        if not valid:
            self.set_qpos(previous, zero_velocity=True)
            return previous.copy(), z_progress, False, reason, raw_dz
        self.set_qpos(projected, zero_velocity=True)
        return projected, new_z, True, "ok", raw_dz

    def final_is_clear(self, qpos: np.ndarray) -> Tuple[bool, str]:
        return self.final_validator.pose_is_collision_free(qpos)

    def close(self):
        try:
            if self.show_viewer and self.scene is not None and getattr(self.scene, "viewer", None) is not None:
                try:
                    self.scene.viewer.stop()
                except Exception:
                    pass
            self.move_entity = None
            self.fixed_entities.clear()
            self.scene = None
            self.nonmating_validator = None
            self.final_validator = None
        finally:
            gc.collect()


class SemanticScrewPlanner:
    """Force-driven unscrew primitive with a virtual screw-joint projection."""

    def __init__(
        self,
        assets: Dict[str, PartAsset],
        move_id: str,
        still_ids: Sequence[str],
        operation: LLMPriorOperation,
        show_viewer: bool,
        dt: float,
        substeps: int,
        force_mag: float,
        unscrew_axial_force: float,
        unscrew_torque: float,
        unscrew_probe_chunks: int,
        unscrew_max_chunks: int,
        unscrew_stall_chunks: int,
        unscrew_lateral_factor: float,
        unscrew_force_growth: float,
        unscrew_max_force_scale: float,
        density: float,
        friction: float,
        sdf_cell_size: float,
        sdf_min_res: int,
        sdf_max_res: int,
        frame_skip: int,
        collision_face_num: int,
        collision_mode: str,
        watertighten: int,
        free_dof_order: str,
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
        min_removal_distance: float,
        min_removal_distance_factor: float,
    ):
        del force_mag
        self.assets = assets
        self.move_id = move_id
        self.still_ids = list(still_ids)
        self.operation = operation
        self.show_viewer = bool(show_viewer)
        self.dt = float(dt)
        self.substeps = int(substeps)
        self.density = float(density)
        self.friction = float(friction)
        self.sdf_cell_size = float(sdf_cell_size)
        self.sdf_min_res = int(sdf_min_res)
        self.sdf_max_res = int(sdf_max_res)
        self.frame_skip = max(int(frame_skip), 1)
        self.collision_face_num = int(collision_face_num)
        self.collision_mode = collision_mode
        self.watertighten = int(watertighten)
        self.free_dof_order = free_dof_order
        self.penetration_tolerance = float(penetration_tolerance)
        self.sweep_translation_step = float(sweep_translation_step)
        self.sweep_rotation_step = float(sweep_rotation_step)
        self.unscrew_lateral_factor = max(float(unscrew_lateral_factor), 0.05)
        self.min_removal_distance = max(float(min_removal_distance), 0.0)
        self.min_removal_distance_factor = max(float(min_removal_distance_factor), 0.0)
        self.max_chunks = max(int(unscrew_max_chunks), 20)
        self.stall_chunks = max(int(unscrew_stall_chunks), 3)
        self.force_growth = max(float(unscrew_force_growth), 1.0)
        self.max_force_scale = max(float(unscrew_max_force_scale), 1.0)
        self.probe_chunks = max(int(unscrew_probe_chunks), 1)
        # Protect the normalized small fasteners from impulsive wrenches.
        requested_force = max(float(unscrew_axial_force), 1e-4)
        requested_torque = max(float(unscrew_torque), 1e-5)
        self.axial_force = min(requested_force, 3.0)
        self.torque = min(requested_torque, 0.30)
        if self.axial_force < requested_force or self.torque < requested_torque:
            print(
                f"[Semantic Unscrew] Clamped requested wrench for numerical stability: "
                f"force {requested_force:g}->{self.axial_force:g} N, "
                f"torque {requested_torque:g}->{self.torque:g} N*m"
            )
        self.initial = make_initial_qpos(self.assets[self.move_id])
        self.axis = self._resolve_axis()
        self.pitch = self._resolve_pitch()
        self.clearance = self._resolve_clearance()
        self._mating_inference_evidence: Dict[str, dict] = {}
        self.mating_parts = self._resolve_mating_parts()
        self.max_turns = max(abs(float(self.operation.max_turns)), 0.25)
        self.target_turns = min(
            max(abs(float(self.operation.turns)), 0.25),
            self.max_turns,
        )
        self.target_theta = 2.0 * math.pi * self.target_turns
        self.thread_release_distance = self.pitch * self.target_turns
        self.target_extraction_distance = self._resolve_extraction_distance()
        self.max_angle_step = 0.12
        _, _, _, assembly_extent = compute_assembly_bounds(self.assets)
        self.max_translation_step = max(0.0015 * assembly_extent, 5e-4)
        self.blockers: set[str] = set()
        self.world: Optional[SemanticScrewWorld] = None
        self.last_failure_reason = "not started"
        self.search_diagnostics: Dict[str, object] = {}

    def _resolve_axis(self) -> np.ndarray:
        if self.operation.axis is not None:
            axis = normalize_vector(self.operation.axis)
        else:
            _, _, assembly_center, _ = compute_assembly_bounds(self.assets)
            axis = estimate_unscrew_axis(
                self.assets[self.move_id],
                assembly_center,
                mode=self.operation.axis_mode,
            )
        if abs(self.operation.axial_sign) > 1e-9:
            axis = axis * np.sign(self.operation.axial_sign)
        else:
            _, _, assembly_center, _ = compute_assembly_bounds(self.assets)
            axis = orient_axis_outward(
                axis,
                self.assets[self.move_id].center,
                assembly_center,
            )
        return normalize_vector(axis)

    def _mesh_world_vertices(self, pid: str, mesh_mode: str = "visual") -> np.ndarray:
        asset = self.assets[pid]
        mesh = (
            asset.local_collision_mesh
            if str(mesh_mode).lower() == "collision"
            else asset.local_visual_mesh
        )
        return (
            np.asarray(mesh.vertices, dtype=np.float64)
            + asset.center.reshape(1, 3)
        )

    def _axis_corridor_metrics(self, pid: str) -> dict:
        """Measure whether a part belongs to the screw's coaxial stack.

        A clearance hole normally has *no* exact mesh intersection with the
        screw in the nominal CAD pose.  Initial-intersection-only inference
        therefore cannot identify the plate, washer, insert, or threaded boss
        that the screw actually passes through.  This test uses axial interval
        overlap and the closest sampled radius around the known screw axis.
        """
        moving = self._mesh_world_vertices(self.move_id, "visual")
        origin = self.initial[:3]
        rel_moving = moving - origin.reshape(1, 3)
        moving_axial = rel_moving @ self.axis
        moving_radial = rel_moving - np.outer(moving_axial, self.axis)
        moving_radius = np.linalg.norm(moving_radial, axis=1)
        screw_min = float(moving_axial.min())
        screw_max = float(moving_axial.max())
        screw_length = max(screw_max - screw_min, 1e-6)
        head_radius = max(float(np.percentile(moving_radius, 95)), 1e-6)

        part = self._mesh_world_vertices(pid, "visual")
        rel_part = part - origin.reshape(1, 3)
        part_axial = rel_part @ self.axis
        part_radial = rel_part - np.outer(part_axial, self.axis)
        part_radius = np.linalg.norm(part_radial, axis=1)
        part_min = float(part_axial.min())
        part_max = float(part_axial.max())

        axial_margin = max(
            self.clearance,
            self.penetration_tolerance,
            0.03 * screw_length,
        )
        in_axial_slab = (
            (part_axial >= screw_min - axial_margin)
            & (part_axial <= screw_max + axial_margin)
        )
        minimum_radius = (
            float(part_radius[in_axial_slab].min())
            if np.any(in_axial_slab)
            else float("inf")
        )
        axial_gap = max(screw_min - part_max, part_min - screw_max, 0.0)
        corridor_radius = head_radius + max(
            self.penetration_tolerance,
            0.15 * head_radius,
        )
        in_corridor = bool(
            axial_gap <= axial_margin
            and minimum_radius <= corridor_radius
        )
        return {
            "in_corridor": in_corridor,
            "axial_gap": float(axial_gap),
            "axial_margin": float(axial_margin),
            "minimum_radius": (
                float(minimum_radius)
                if np.isfinite(minimum_radius)
                else None
            ),
            "corridor_radius": float(corridor_radius),
            "screw_length": float(screw_length),
            "head_radius": float(head_radius),
        }

    def _pair_distance_at_initial(self, pid: str, mesh_mode: str) -> float:
        """Return the exact FCL surface gap for one pair, or infinity."""
        try:
            manager = trimesh.collision.CollisionManager()
            fixed_asset = self.assets[pid]
            fixed_mesh = (
                fixed_asset.local_collision_mesh
                if mesh_mode == "collision"
                else fixed_asset.local_visual_mesh
            )
            move_asset = self.assets[self.move_id]
            move_mesh = (
                move_asset.local_collision_mesh
                if mesh_mode == "collision"
                else move_asset.local_visual_mesh
            )
            manager.add_object(
                pid,
                fixed_mesh,
                transform=qpos_to_transform(make_initial_qpos(fixed_asset)),
            )
            result = manager.min_distance_single(
                move_mesh,
                transform=qpos_to_transform(self.initial),
            )
            if isinstance(result, tuple):
                result = result[0]
            return max(float(result), 0.0)
        except Exception:
            return float("inf")

    def _mating_gap_limit(self) -> float:
        diameter = self._fastener_diameter()
        return max(
            0.75 * self.penetration_tolerance,
            0.03 * diameter,
            2.5e-4,
        )

    def _mating_probe_qposes(self) -> List[np.ndarray]:
        """Sample one outward turn of the virtual screw at fine resolution."""
        screw_min, screw_max = self._projection_interval(self.move_id)
        screw_length = max(screw_max - screw_min, 1e-6)
        max_probe_axial = min(
            self.pitch,
            max(2.0 * self.penetration_tolerance, 0.10 * screw_length),
        )
        samples: List[np.ndarray] = []
        sample_count = 16
        for torque_sign in (-1.0, 1.0):
            for index in range(1, sample_count + 1):
                ratio = index / float(sample_count)
                theta = ratio * 2.0 * math.pi
                axial = min(self.pitch * ratio, max_probe_axial)
                qpos = self.initial.copy()
                qpos[:3] = self.initial[:3] + self.axis * axial
                qpos[3:7] = quat_multiply(
                    axis_angle_to_quat(self.axis, torque_sign * theta),
                    self.initial[3:7],
                )
                samples.append(qpos)
        return samples

    def _resolve_mating_parts(self) -> List[str]:
        explicit = [
            pid for pid in self.operation.mating_parts
            if pid in self.still_ids and pid in self.assets
        ]
        if explicit:
            for pid in explicit:
                self._mating_inference_evidence[pid] = {
                    "selected": True,
                    "evidence": ["explicit-prior"],
                }
            print(
                f"[Semantic Unscrew] Using explicit mating parts for "
                f"{self.move_id}: {explicit}"
            )
            return explicit

        # Calibrate both representations.  Visual meshes preserve the CAD
        # geometry, while collision meshes reproduce the simplified geometry
        # used by Genesis.  A one-turn micro-probe catches thread/hole pairs
        # that are separated at q0 but touch after the first helical increment.
        evidence: Dict[str, set[str]] = {
            pid: set() for pid in self.still_ids
        }
        peak_depth: Dict[str, float] = {
            pid: 0.0 for pid in self.still_ids
        }
        probe_qposes = self._mating_probe_qposes()
        for mesh_mode in ("visual", "collision"):
            validator = ExactCollisionValidator(
                assets=self.assets,
                move_id=self.move_id,
                still_ids=self.still_ids,
                penetration_tolerance=self.penetration_tolerance,
                sweep_translation_step=self.sweep_translation_step,
                sweep_rotation_step=self.sweep_rotation_step,
                mesh_mode=mesh_mode,
                log_initial_contacts=False,
            )
            for pid, depth in validator.initial_depths.items():
                if depth > 0.0:
                    evidence[pid].add(f"initial-{mesh_mode}")
                    peak_depth[pid] = max(peak_depth[pid], float(depth))
            if validator.available:
                for qpos in probe_qposes:
                    for pid, depth in validator.penetration_profile(qpos).items():
                        if depth > 0.0:
                            evidence[pid].add(f"thread-probe-{mesh_mode}")
                            peak_depth[pid] = max(peak_depth[pid], float(depth))

        gap_limit = self._mating_gap_limit()
        inferred: List[str] = []
        for pid in self.still_ids:
            metrics = self._axis_corridor_metrics(pid)
            visual_gap = self._pair_distance_at_initial(pid, "visual")
            collision_gap = self._pair_distance_at_initial(pid, "collision")
            minimum_gap = min(visual_gap, collision_gap)
            initial_contact = any(
                item.startswith("initial-") for item in evidence[pid]
            )
            probe_contact = any(
                item.startswith("thread-probe-") for item in evidence[pid]
            )
            selected = bool(
                initial_contact
                or (
                    metrics["in_corridor"]
                    and (probe_contact or minimum_gap <= gap_limit)
                )
            )
            if selected:
                inferred.append(pid)
            self._mating_inference_evidence[pid] = {
                "selected": selected,
                "evidence": sorted(evidence[pid]),
                "peak_probe_depth": float(peak_depth[pid]),
                "minimum_surface_gap": (
                    float(minimum_gap)
                    if np.isfinite(minimum_gap)
                    else None
                ),
                **metrics,
            }

        inferred = sorted(set(inferred))
        if not inferred:
            print(
                f"[Semantic Unscrew] Warning: no mating parts inferred for "
                f"{self.move_id}. Add per-part mating_parts to the prior if the "
                "CAD model does not preserve the hole/thread neighbourhood."
            )
        else:
            summaries = []
            for pid in inferred:
                item = self._mating_inference_evidence[pid]
                gap = item["minimum_surface_gap"]
                gap_text = "n/a" if gap is None else f"{gap:.4g}"
                reasons = ",".join(item["evidence"]) or "near-axis-gap"
                summaries.append(f"{pid}({reasons}; gap={gap_text})")
            print(
                f"[Semantic Unscrew] Inferred mating stack for {self.move_id}: "
                + ", ".join(summaries)
            )
        return inferred

    def _promote_plausible_mating_blockers(
        self, blockers: Iterable[str]
    ) -> List[str]:
        """Promote only near-axis, near-surface blockers to the virtual joint."""
        promoted: List[str] = []
        promotion_gap = max(
            2.0 * self._mating_gap_limit(),
            self.penetration_tolerance,
        )
        for pid in blockers:
            if (
                pid in self.mating_parts
                or pid not in self.still_ids
                or pid not in self.assets
            ):
                continue
            item = self._mating_inference_evidence.get(pid)
            if item is None:
                metrics = self._axis_corridor_metrics(pid)
                minimum_gap = min(
                    self._pair_distance_at_initial(pid, "visual"),
                    self._pair_distance_at_initial(pid, "collision"),
                )
                item = {
                    **metrics,
                    "minimum_surface_gap": (
                        float(minimum_gap)
                        if np.isfinite(minimum_gap)
                        else None
                    ),
                    "evidence": [],
                }
                self._mating_inference_evidence[pid] = item
            gap = item.get("minimum_surface_gap")
            probe_contact = any(
                str(value).startswith("thread-probe-")
                for value in item.get("evidence", [])
            )
            if (
                bool(item.get("in_corridor"))
                and (
                    probe_contact
                    or (gap is not None and float(gap) <= promotion_gap)
                )
            ):
                promoted.append(pid)

        if promoted:
            self.mating_parts = sorted(
                set(self.mating_parts).union(promoted)
            )
            print(
                f"[Semantic Unscrew] Reclassified near-axis numerical blockers "
                f"as mating parts for {self.move_id}: {sorted(promoted)}. "
                "All other collisions remain active."
            )
        return sorted(promoted)

    def _fastener_diameter(self) -> float:
        mesh = self.assets[self.move_id].local_visual_mesh
        vertices = np.asarray(mesh.vertices, dtype=np.float64)
        axial = vertices @ self.axis
        radial = vertices - np.outer(axial, self.axis)
        return max(2.0 * float(np.percentile(np.linalg.norm(radial, axis=1), 90)), 1e-6)

    def _resolve_pitch(self) -> float:
        if self.operation.pitch is not None and self.operation.pitch > 0:
            return float(self.operation.pitch)
        _, _, _, extent = compute_assembly_bounds(self.assets)
        diameter = self._fastener_diameter()
        return float(np.clip(0.12 * diameter, 0.001 * extent, 0.01 * extent))

    def _resolve_clearance(self) -> float:
        if self.operation.clearance is not None:
            return float(self.operation.clearance)
        _, _, _, extent = compute_assembly_bounds(self.assets)
        return max(0.5 * self.pitch, 0.003 * extent)

    def _projection_interval(self, pid: str) -> Tuple[float, float]:
        asset = self.assets[pid]
        world_vertices = (
            np.asarray(asset.local_visual_mesh.vertices, dtype=np.float64)
            + asset.center.reshape(1, 3)
        )
        projected = world_vertices @ self.axis
        return float(projected.min()), float(projected.max())

    def _resolve_extraction_distance(self) -> float:
        """Return the total axial displacement measured from the CAD pose.

        The former implementation used the complete projection interval of
        every mating compound.  For the satellite model, one mating compound
        spans most of the product height, so an ordinary screw was incorrectly
        assigned an extraction distance of about 0.263 m.  The screw only needs
        to clear its own shank/thread engagement plus a small safety margin.

        When ``operation.distance`` is provided, it is treated as the requested
        total axial displacement rather than merely a lower bound beneath a
        compound-sized estimate.
        """
        screw_min, screw_max = self._projection_interval(self.move_id)
        screw_axial_length = max(screw_max - screw_min, 1e-6)
        _, _, _, extent = compute_assembly_bounds(self.assets)

        minimum = max(
            self.thread_release_distance + self.clearance,
            0.5 * screw_axial_length + self.clearance,
        )
        if self.operation.distance is not None and self.operation.distance > 0:
            requested = max(float(self.operation.distance), minimum)
            return float(np.clip(requested, minimum, 0.50 * extent))

        required = max(
            minimum,
            1.10 * screw_axial_length + self.clearance,
            0.025 * extent,
        )
        # A semantic screw primitive should never require travelling through a
        # large fraction of the complete product merely because a mating part
        # is represented by one large compound mesh.
        return float(np.clip(required, minimum, 0.35 * extent))

    def _make_validated_kinematic_path(
        self,
        torque_sign: float,
        extraction_distance: Optional[float] = None,
    ) -> List[np.ndarray]:
        """Build the deterministic virtual-screw path used by the old planner.

        The path follows the same screw manifold as the force-driven planner:
        first z = pitch * theta / (2*pi), then pure axial withdrawal.  It is not
        blindly accepted; ``_validate_kinematic_path`` performs dense exact-mesh
        swept collision validation against every non-mating part and requires a
        collision-free final pose against the complete remaining assembly.
        """
        total_frames = max(int(self.operation.frames), 36)
        helical_frames = max(int(round(0.60 * total_frames)), 18)
        extract_frames = max(total_frames - helical_frames, 18)
        q0 = self.initial.copy()
        path = [q0.copy()]

        for index in range(1, helical_frames + 1):
            ratio = index / float(helical_frames)
            theta = ratio * self.target_theta
            axial = ratio * self.thread_release_distance
            qpos = q0.copy()
            qpos[:3] = q0[:3] + self.axis * axial
            qpos[3:7] = quat_multiply(
                axis_angle_to_quat(self.axis, torque_sign * theta),
                q0[3:7],
            )
            path.append(qpos)

        release = path[-1].copy()
        start_axial = self.thread_release_distance
        end_axial = max(
            self.target_extraction_distance
            if extraction_distance is None
            else float(extraction_distance),
            start_axial,
        )
        for index in range(1, extract_frames + 1):
            ratio = index / float(extract_frames)
            axial = (1.0 - ratio) * start_axial + ratio * end_axial
            qpos = release.copy()
            qpos[:3] = q0[:3] + self.axis * axial
            path.append(qpos)
        return path

    def _validate_kinematic_path(
        self,
        path: Sequence[np.ndarray],
        deadline: Optional[float] = None,
    ) -> Tuple[bool, str, set[str]]:
        mating_set = set(self.mating_parts)
        nonmating_ids = [pid for pid in self.still_ids if pid not in mating_set]
        # Fast semantic sweep: use the cached collision meshes and a bounded
        # angular sampling interval. Rotation about a known screw axis should not
        # spend several minutes per fastener. The final pose is checked again
        # below with the full visual meshes.
        nonmating_validator = ExactCollisionValidator(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=nonmating_ids,
            penetration_tolerance=self.penetration_tolerance,
            sweep_translation_step=max(self.sweep_translation_step, 0.0025),
            sweep_rotation_step=max(self.sweep_rotation_step, 0.12),
            mesh_mode="collision",
        )
        threshold = removal_distance_threshold(
            self.assets,
            self.move_id,
            absolute_distance=self.min_removal_distance,
            part_diagonal_factor=self.min_removal_distance_factor,
        )
        valid, reason = nonmating_validator.path_is_valid(
            path, threshold, deadline=deadline
        )
        blockers = set(nonmating_validator.last_blockers)
        if not valid:
            return False, f"non-mating swept validation failed: {reason}", blockers

        final_validator = ExactCollisionValidator(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=self.still_ids,
            penetration_tolerance=self.penetration_tolerance,
            sweep_translation_step=self.sweep_translation_step,
            sweep_rotation_step=self.sweep_rotation_step,
        )
        clear, clear_reason = final_validator.pose_is_collision_free(path[-1])
        blockers.update(final_validator.last_blockers)
        if not clear:
            return False, f"final screw pose is not clear: {clear_reason}", blockers
        if not is_pose_disassembled_from_assets(
            self.assets,
            self.move_id,
            self.still_ids,
            path[-1],
        ):
            return (
                False,
                "final screw pose is not outside the convex hull of the "
                "remaining assembly",
                blockers,
            )

        axial = float(np.dot(path[-1][:3] - self.initial[:3], self.axis))
        if axial + 1e-9 < self.target_extraction_distance:
            return (
                False,
                f"axial displacement {axial:.6g} < target "
                f"{self.target_extraction_distance:.6g}",
                blockers,
            )
        return True, "ok", blockers

    def _kinematic_extraction_candidates(self) -> List[float]:
        """Generate bounded axial targets without using compound thickness."""
        base = float(self.target_extraction_distance)
        screw_min, screw_max = self._projection_interval(self.move_id)
        screw_length = max(screw_max - screw_min, 1e-6)
        _, _, _, extent = compute_assembly_bounds(self.assets)
        upper = min(
            0.75 * extent,
            max(
                base,
                3.0 * screw_length + self.clearance,
                base + 0.10 * extent,
            ),
        )
        if upper <= base + 1e-9:
            return [base]
        return [float(x) for x in np.linspace(base, upper, 4)]

    def _try_validated_kinematic_paths(
        self, deadline: Optional[float] = None
    ):
        attempted_signs: List[float] = []
        best_path = [self.initial.copy()]
        reasons: List[str] = []
        extraction_candidates = self._kinematic_extraction_candidates()
        for torque_sign in self._torque_candidates():
            if deadline is not None and time.monotonic() >= float(deadline):
                reasons.append("deterministic exact validation timeout")
                break
            attempted_signs.append(float(torque_sign))
            for extraction_distance in extraction_candidates:
                if deadline is not None and time.monotonic() >= float(deadline):
                    reasons.append("deterministic exact validation timeout")
                    break
                path = self._make_validated_kinematic_path(
                    torque_sign,
                    extraction_distance=extraction_distance,
                )
                if len(path) > len(best_path):
                    best_path = path
                valid, reason, blockers = self._validate_kinematic_path(
                    path, deadline=deadline
                )
                self.blockers.update(blockers)
                if valid:
                    self.target_extraction_distance = float(extraction_distance)
                    print(
                        f"[Semantic Unscrew] Validated deterministic screw path for "
                        f"{self.move_id}: torque_sign={torque_sign:+.0f}, "
                        f"turns={self.target_turns:.3g}, pitch={self.pitch:.6g}, "
                        f"axial={self.target_extraction_distance:.6g}"
                    )
                    return True, path, attempted_signs, reasons
                reasons.append(
                    f"sign {torque_sign:+.0f}, axial {extraction_distance:.6g}: "
                    f"{reason}"
                )
                print(
                    f"[Semantic Unscrew] Deterministic candidate sign "
                    f"{torque_sign:+.0f}, axial={extraction_distance:.6g} "
                    f"rejected for {self.move_id}: {reason}"
                )
                # A longer axial target can solve a final mating collision, but
                # it cannot repair a swept collision with a non-mating blocker.
                if not reason.startswith("final screw pose is not"):
                    break
        return False, best_path, attempted_signs, reasons

    def _make_world(self) -> SemanticScrewWorld:
        return SemanticScrewWorld(
            assets=self.assets,
            move_id=self.move_id,
            still_ids=self.still_ids,
            mating_parts=self.mating_parts,
            show_viewer=self.show_viewer,
            dt=self.dt,
            substeps=self.substeps,
            density=self.density,
            friction=self.friction,
            sdf_cell_size=self.sdf_cell_size,
            sdf_min_res=self.sdf_min_res,
            sdf_max_res=self.sdf_max_res,
            collision_face_num=self.collision_face_num,
            collision_mode=self.collision_mode,
            watertighten=self.watertighten,
            free_dof_order=self.free_dof_order,
            penetration_tolerance=self.penetration_tolerance,
            sweep_translation_step=self.sweep_translation_step,
            sweep_rotation_step=self.sweep_rotation_step,
        )

    def _torque_candidates(self) -> List[float]:
        if abs(self.operation.handedness) > 1e-9:
            sign = float(np.sign(self.operation.handedness))
            return [sign, -sign]
        # Counter-clockwise viewed along the outward axis is the conventional
        # first attempt; the opposite sign remains an automatic fallback.
        return [-1.0, 1.0]

    def _run_candidate(self, torque_sign: float, max_time: float):
        world = self._make_world()
        self.world = world
        path = [self.initial.copy()]
        theta = 0.0
        z = 0.0
        scale = 1.0
        stall = 0
        started = time.time()
        previous = self.initial.copy()
        try:
            print(
                f"[Semantic Unscrew] {self.move_id}: axis={self.axis.round(5).tolist()}, "
                f"mating_parts={self.mating_parts}, torque_sign={torque_sign:+.0f}, "
                f"pitch={self.pitch:.6g}, turns={self.target_turns:.3g}, "
                f"target_extract={self.target_extraction_distance:.6g}"
            )

            # Stage A: force/torque-driven virtual screw joint.
            for chunk in range(self.max_chunks):
                if time.time() - started > max_time:
                    return "Timeout", path, "unscrew timeout"
                progressed = False
                for _ in range(self.frame_skip):
                    state, new_theta, valid, reason, raw_dtheta, raw_dz = world.project_thread_step(
                        previous=previous,
                        origin=self.initial[:3],
                        initial_quat=self.initial[3:7],
                        axis=self.axis,
                        torque_sign=torque_sign,
                        pitch=self.pitch,
                        theta_progress=theta,
                        axial_force=self.axial_force * scale,
                        torque=self.torque * scale,
                        max_angle_step=self.max_angle_step,
                    )
                    if not valid:
                        self.blockers.update(world.nonmating_validator.last_blockers)
                        path.append(np.asarray(state, dtype=np.float64).copy())
                        return "Failure", path, reason
                    if new_theta > theta + 1e-8:
                        progressed = True
                    theta = new_theta
                    z = self.pitch * theta / (2.0 * math.pi)
                    previous = state
                    path.append(state.copy())
                    if theta >= self.target_theta - 1e-6:
                        break
                if theta >= self.target_theta - 1e-6:
                    break
                if progressed:
                    stall = max(stall - 1, 0)
                else:
                    stall += 1
                if stall >= self.stall_chunks:
                    if scale < self.max_force_scale - 1e-9:
                        scale = min(scale * self.force_growth, self.max_force_scale)
                        stall = 0
                        print(
                            f"[Semantic Unscrew] {self.move_id}: increasing wrench scale "
                            f"to {scale:.3g} during thread release."
                        )
                    else:
                        return "Failure", path, "no positive screw-coordinate response"
                if chunk == 0 or (chunk + 1) % 10 == 0:
                    print(
                        f"[Semantic Unscrew] {self.move_id}: thread chunk={chunk + 1}, "
                        f"theta={theta:.4g}/{self.target_theta:.4g}, z={z:.6g}, scale={scale:.3g}"
                    )
            if theta < self.target_theta - 1e-6:
                return "Failure", path, "maximum chunks reached before thread release"

            # Stage B: after the virtual thread is released, pull axially while
            # keeping the screw centreline and orientation constrained.
            release_quat = path[-1][3:7].copy()
            stall = 0
            screw_min, screw_max = self._projection_interval(self.move_id)
            completion_check_spacing = max(
                self.clearance,
                0.25 * max(screw_max - screw_min, 1e-6),
            )
            next_completion_check = self.target_extraction_distance
            last_clear = False
            last_clear_reason = "target axial displacement not reached"
            for chunk in range(self.max_chunks):
                if time.time() - started > max_time:
                    return "Timeout", path, "extraction timeout"
                progressed = False
                for _ in range(self.frame_skip):
                    state, new_z, valid, reason, raw_dz = world.project_extract_step(
                        previous=previous,
                        origin=self.initial[:3],
                        fixed_quat=release_quat,
                        axis=self.axis,
                        z_progress=z,
                        axial_force=self.axial_force * scale,
                        max_translation_step=self.max_translation_step,
                    )
                    if not valid:
                        self.blockers.update(world.nonmating_validator.last_blockers)
                        path.append(np.asarray(state, dtype=np.float64).copy())
                        return "Failure", path, reason
                    if new_z > z + 1e-9:
                        progressed = True
                    z = new_z
                    previous = state
                    path.append(state.copy())
                    if z >= next_completion_check:
                        last_clear, last_clear_reason = world.final_is_clear(state)
                        if last_clear:
                            hull_clear = is_pose_disassembled_from_assets(
                                self.assets,
                                self.move_id,
                                self.still_ids,
                                state,
                            )
                            if hull_clear:
                                print(
                                    f"[Semantic Unscrew] Removed {self.move_id}: "
                                    f"theta={theta:.4g} rad, axial={z:.6g}, "
                                    f"torque_sign={torque_sign:+.0f}"
                                )
                                return "Success", path, "ok"
                            last_clear = False
                            last_clear_reason = (
                                "outside exact contacts but still inside the "
                                "remaining-assembly convex hull"
                            )
                            next_completion_check = (
                                z + completion_check_spacing
                            )
                if progressed:
                    stall = max(stall - 1, 0)
                else:
                    stall += 1
                if stall >= self.stall_chunks:
                    if scale < self.max_force_scale - 1e-9:
                        scale = min(scale * self.force_growth, self.max_force_scale)
                        stall = 0
                        print(
                            f"[Semantic Unscrew] {self.move_id}: increasing axial-force scale "
                            f"to {scale:.3g}."
                        )
                    else:
                        return "Failure", path, "axial extraction stalled"
                if chunk == 0 or (chunk + 1) % 10 == 0:
                    print(
                        f"[Semantic Unscrew] {self.move_id}: extract chunk={chunk + 1}, "
                        f"z={z:.6g}/{self.target_extraction_distance:.6g}, "
                        f"final_clear={last_clear} ({last_clear_reason})"
                    )
            return "Failure", path, "maximum chunks reached before full extraction"
        finally:
            world.close()
            self.world = None

    def plan(self, max_time: float):
        """Plan one semantic fastener removal without wall-clock lockup.

        A deterministic virtual-screw path is generated and exactly validated
        first.  This restores the reliable behaviour of the earlier satellite
        planner while preserving strict non-mating collision checks.  Genesis
        force-driven rollout remains a secondary fallback for geometry that
        cannot be represented by the deterministic screw manifold.
        """
        started = time.monotonic()
        deadline = started + max(float(max_time), 0.0)
        best_path = [self.initial.copy()]
        reasons: List[str] = []

        valid, path, attempted_signs, deterministic_reasons = (
            self._try_validated_kinematic_paths(deadline=deadline)
        )
        reasons.extend(deterministic_reasons)
        if len(path) > len(best_path):
            best_path = path
        adaptive_promotions: List[str] = []
        if not valid:
            adaptive_promotions = self._promote_plausible_mating_blockers(
                self.blockers
            )
            if (
                adaptive_promotions
                and time.monotonic() < deadline
            ):
                reasons.append(
                    "expanded virtual screw joint with near-axis contact "
                    f"parts {adaptive_promotions}"
                )
                (
                    valid,
                    retry_path,
                    retry_signs,
                    retry_reasons,
                ) = self._try_validated_kinematic_paths(deadline=deadline)
                attempted_signs.extend(retry_signs)
                reasons.extend(retry_reasons)
                if len(retry_path) > len(best_path):
                    best_path = retry_path
                if valid:
                    path = retry_path
        if valid:
            self.last_failure_reason = ""
            self.search_diagnostics = {
                "strategy": (
                    "validated_kinematic_virtual_screw_adaptive_mating"
                    if adaptive_promotions
                    else "validated_kinematic_virtual_screw"
                ),
                "attempted_torque_signs": attempted_signs,
                "failure_reason": "",
                "blockers": sorted(self.blockers),
                "mating_parts": list(self.mating_parts),
                "adaptive_mating_promotions": adaptive_promotions,
                "mating_inference": self._mating_inference_evidence,
                "pitch": float(self.pitch),
                "target_turns": float(self.target_turns),
                "max_turns": float(self.max_turns),
                "target_extraction_distance": float(self.target_extraction_distance),
            }
            return "Success", time.monotonic() - started, path

        print(
            f"[Semantic Unscrew] Exact deterministic path was unavailable for "
            f"{self.move_id}; trying the force-driven virtual screw fallback."
        )
        physical_signs: List[float] = []
        for torque_sign in self._torque_candidates():
            physical_signs.append(float(torque_sign))
            elapsed = time.monotonic() - started
            remaining = float(max_time) - elapsed
            if remaining <= 0.0:
                reasons.append("force-driven fallback: semantic unscrew timeout")
                break
            status, candidate_path, reason = self._run_candidate(
                torque_sign,
                remaining,
            )
            if len(candidate_path) > len(best_path):
                best_path = candidate_path
            if status == "Success":
                self.last_failure_reason = ""
                self.search_diagnostics = {
                    "strategy": "force_driven_virtual_screw_fallback",
                    "attempted_torque_signs": physical_signs,
                    "deterministic_attempted_torque_signs": attempted_signs,
                    "failure_reason": "",
                    "blockers": sorted(self.blockers),
                    "mating_parts": list(self.mating_parts),
                    "adaptive_mating_promotions": adaptive_promotions,
                    "mating_inference": self._mating_inference_evidence,
                    "pitch": float(self.pitch),
                    "target_turns": float(self.target_turns),
                    "max_turns": float(self.max_turns),
                    "target_extraction_distance": float(self.target_extraction_distance),
                }
                return "Success", time.monotonic() - started, candidate_path
            reasons.append(f"force sign {torque_sign:+.0f}: {reason}")
            print(
                f"[Semantic Unscrew] Force-driven candidate torque sign "
                f"{torque_sign:+.0f} failed for {self.move_id}: {reason}"
            )

        self.last_failure_reason = "; ".join(reasons) or "all unscrew candidates failed"
        self.search_diagnostics = {
            "strategy": "validated_kinematic_then_force_driven",
            "attempted_torque_signs": physical_signs,
            "deterministic_attempted_torque_signs": attempted_signs,
            "failure_reason": self.last_failure_reason,
            "blockers": sorted(self.blockers),
            "mating_parts": list(self.mating_parts),
            "adaptive_mating_promotions": adaptive_promotions,
            "mating_inference": self._mating_inference_evidence,
            "pitch": float(self.pitch),
            "target_turns": float(self.target_turns),
            "max_turns": float(self.max_turns),
            "target_extraction_distance": float(self.target_extraction_distance),
        }
        print(
            f"[Semantic Unscrew] All semantic unscrew candidates failed for "
            f"{self.move_id}: {self.last_failure_reason}"
        )
        status = (
            "Timeout"
            if "timeout" in self.last_failure_reason.lower()
            else "Failure"
        )
        return status, time.monotonic() - started, best_path

    def close(self):
        if self.world is not None:
            self.world.close()
            self.world = None
        gc.collect()


# ============================================================================
# BFS path planner
# ============================================================================


@dataclass
class SearchNode:
    qpos: np.ndarray
    path: List[np.ndarray]
    depth: int
    score: float = 0.0
    last_action: Optional[np.ndarray] = None
    action_trace: List[List[float]] = field(default_factory=list)


@dataclass
class RRTTreeNode:
    qpos: np.ndarray
    parent: Optional[int]
    edge_path: List[np.ndarray]
    depth: int
    cost: float
    last_action: Optional[np.ndarray] = None


class GenesisBFSPathPlanner:
    """Physics rollout planner with FIFO-BFS and adaptive best-first modes.

    The adaptive mode is a contact-guided receding-horizon beam search.  It
    retains the reliable translation-first prior, but also evaluates diagonal
    planar forces and simultaneous force/torque actions.  Physics rollouts are
    scored by outward progress, AABB clearance, reduction of initial contact
    penetration, and path cost.  Only the highest-scoring successors are kept,
    which avoids exhaustive enumeration of every 6-DoF action sequence.
    """

    def __init__(
        self,
        assets: Dict[str, PartAsset],
        move_id: str,
        still_ids: Sequence[str],
        rotation: bool,
        show_viewer: bool,
        dt: float,
        substeps: int,
        force_mag: float,
        density: float,
        friction: float,
        sdf_cell_size: float,
        sdf_min_res: int,
        sdf_max_res: int,
        frame_skip: int,
        trans_dist_th: float,
        quat_dist_th: float,
        collision_face_num: int,
        collision_mode: str,
        watertighten: int,
        free_dof_order: str,
        max_action_chunks: int,
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
        min_removal_distance: float,
        min_removal_distance_factor: float,
        enable_direct_release: bool = True,
        translation_priority: Optional[Sequence[str]] = None,
        random_rotation_actions: int = 12,
        random_rotation_scale: float = 0.35,
        search_strategy: str = "adaptive_best_first",
        beam_width: int = 8,
        action_top_k: int = 24,
        enable_diagonal_actions: bool = True,
        coupled_action_scale: float = 0.45,
        rrt_max_nodes: int = 400,
        rrt_goal_bias: float = 0.20,
        rrt_neighbor_count: int = 8,
        rrt_connect_distance_factor: float = 0.35,
        rrt_goal_root_count: int = 8,
    ):
        self.assets = assets
        self.move_id = move_id
        self.still_ids = list(still_ids)
        self.rotation = bool(rotation)
        self.frame_skip = int(frame_skip)
        self.trans_dist_th = float(trans_dist_th)
        self.quat_dist_th = float(quat_dist_th)
        self.max_action_chunks = int(max_action_chunks)
        self.enable_direct_release = bool(enable_direct_release)
        self.translation_priority = tuple(
            translation_priority
            or ("+z", "-z", "+x", "-x", "+y", "-y")
        )
        self.random_rotation_actions = max(int(random_rotation_actions), 0)
        self.random_rotation_scale = max(float(random_rotation_scale), 0.0)
        self.search_strategy = str(search_strategy).strip().lower().replace("-", "_")
        supported_strategies = {
            "bfs", "dfs", "adaptive_best_first", "rrt", "rrt_star",
            "bi_rrt", "bk_rrt",
        }
        if self.search_strategy not in supported_strategies:
            self.search_strategy = "adaptive_best_first"
        self.beam_width = max(int(beam_width), 1)
        self.action_top_k = max(int(action_top_k), 6)
        self.enable_diagonal_actions = bool(enable_diagonal_actions)
        self.coupled_action_scale = max(float(coupled_action_scale), 0.0)
        self.rrt_max_nodes = max(int(rrt_max_nodes), 2)
        self.rrt_goal_bias = float(np.clip(rrt_goal_bias, 0.0, 1.0))
        self.rrt_neighbor_count = max(int(rrt_neighbor_count), 1)
        self.rrt_connect_distance_factor = max(
            float(rrt_connect_distance_factor), 0.02
        )
        self.rrt_goal_root_count = max(int(rrt_goal_root_count), 1)
        digest = hashlib.sha256(
            f"{self.move_id}|{self.search_strategy}".encode("utf-8")
        ).digest()
        self._rrt_rng = np.random.default_rng(
            int.from_bytes(digest[:8], byteorder="little", signed=False)
        )

        self.world = GenesisDisassemblyWorld(
            assets=assets,
            move_id=move_id,
            still_ids=still_ids,
            show_viewer=show_viewer,
            dt=dt,
            substeps=substeps,
            force_mag=force_mag,
            density=density,
            friction=friction,
            sdf_cell_size=sdf_cell_size,
            sdf_min_res=sdf_min_res,
            sdf_max_res=sdf_max_res,
            collision_face_num=collision_face_num,
            collision_mode=collision_mode,
            watertighten=watertighten,
            free_dof_order=free_dof_order,
            penetration_tolerance=penetration_tolerance,
            sweep_translation_step=sweep_translation_step,
            sweep_rotation_step=sweep_rotation_step,
            min_removal_distance=min_removal_distance,
            min_removal_distance_factor=min_removal_distance_factor,
        )
        (
            self.translation_actions,
            self.rotation_actions,
            self.coupled_actions,
        ) = self._make_action_groups()
        self.blockers: set[str] = set()
        self.solution_mode = "physics_bfs"
        self.last_failure_reason = "not started"
        self.last_expand_reason = "not started"
        self.search_diagnostics: Dict[str, object] = {}
        self.failure_reason_counts: Dict[str, int] = defaultdict(int)
        self._expanded_nodes = 0
        self._evaluated_actions = 0
        self._generated_successors = 0
        self._collision_rejections = 0
        self._best_score = -float("inf")
        self._best_failure_score = -float("inf")
        self._best_failure_path: List[np.ndarray] = []
        self._rrt_nodes = 0
        self._rrt_rewires = 0
        self._birrt_start_nodes = 0
        self._birrt_goal_nodes = 0

        self._initial_qpos = self.world.initial_qpos().astype(np.float64)
        assembly_min, assembly_max, assembly_center, assembly_extent = compute_assembly_bounds(
            self.assets, [self.move_id, *self.still_ids]
        )
        self._assembly_min = np.asarray(assembly_min, dtype=np.float64)
        self._assembly_max = np.asarray(assembly_max, dtype=np.float64)
        self._assembly_center = np.asarray(assembly_center, dtype=np.float64)
        self._assembly_extent = max(float(assembly_extent), 1e-6)
        self._outward = normalize_vector(
            self._initial_qpos[:3] - self._assembly_center,
            fallback=(0.0, 0.0, 1.0),
        )
        self._part_diagonal = max(
            float(np.linalg.norm(self.assets[self.move_id].local_hull.extents)),
            1e-6,
        )
        self._initial_contact_sum = self.world._path_validator.initial_contact_depth_sum(
            self._initial_qpos
        )
        self._still_bounds: Optional[Tuple[np.ndarray, np.ndarray]] = None
        if self.still_ids:
            vertices = []
            for pid in self.still_ids:
                asset = self.assets[pid]
                vertices.append(
                    np.asarray(asset.local_visual_mesh.vertices, dtype=np.float64)
                    + asset.center.reshape(1, 3)
                )
            if vertices:
                stacked = np.concatenate(vertices, axis=0)
                self._still_bounds = (stacked.min(axis=0), stacked.max(axis=0))

    @staticmethod
    def _label_to_direction(label: str) -> Optional[np.ndarray]:
        table = {
            "+x": np.array([1.0, 0.0, 0.0]),
            "-x": np.array([-1.0, 0.0, 0.0]),
            "+y": np.array([0.0, 1.0, 0.0]),
            "-y": np.array([0.0, -1.0, 0.0]),
            "+z": np.array([0.0, 0.0, 1.0]),
            "-z": np.array([0.0, 0.0, -1.0]),
        }
        return table.get(str(label).strip().lower())

    @staticmethod
    def _append_unique_direction(
        target: List[np.ndarray], direction: np.ndarray, cosine_threshold: float = 0.998
    ) -> None:
        direction = normalize_vector(direction)
        if any(float(np.dot(direction, old)) > cosine_threshold for old in target):
            return
        target.append(direction)

    @staticmethod
    def _append_unique_action(
        target: List[np.ndarray], action: np.ndarray, cosine_threshold: float = 0.998
    ) -> None:
        action = np.asarray(action, dtype=np.float64)
        norm = float(np.linalg.norm(action))
        if norm < 1e-12:
            return
        unit = action / norm
        for old in target:
            old_norm = float(np.linalg.norm(old))
            if old_norm > 1e-12 and float(np.dot(unit, old / old_norm)) > cosine_threshold:
                return
        target.append(action)

    def _principal_axes(self) -> List[np.ndarray]:
        vertices = np.asarray(
            self.assets[self.move_id].local_collision_mesh.vertices,
            dtype=np.float64,
        )
        try:
            centered = vertices - vertices.mean(axis=0, keepdims=True)
            _, eigvec = np.linalg.eigh(np.cov(centered.T))
            return [normalize_vector(eigvec[:, i]) for i in range(3)]
        except Exception:
            return [axis.copy() for axis in np.eye(3)]

    def _make_action_groups(
        self,
    ) -> Tuple[List[np.ndarray], List[np.ndarray], List[np.ndarray]]:
        """Construct cardinal, diagonal, rotational, and coupled 6-DoF actions."""
        outward = self._outward if hasattr(self, "_outward") else normalize_vector(
            self.assets[self.move_id].center - compute_assembly_bounds(
                self.assets, [self.move_id, *self.still_ids]
            )[2],
            fallback=(0.0, 0.0, 1.0),
        )

        primary_dirs: List[np.ndarray] = []
        for label in self.translation_priority:
            direction = self._label_to_direction(label)
            if direction is not None:
                self._append_unique_direction(primary_dirs, direction)
        for label in ("+z", "-z", "+x", "-x", "+y", "-y"):
            direction = self._label_to_direction(label)
            assert direction is not None
            self._append_unique_direction(primary_dirs, direction)

        geometry_dirs: List[np.ndarray] = []
        self._append_unique_direction(geometry_dirs, outward)
        move_center = self.assets[self.move_id].center
        for pid in self.still_ids:
            separation = move_center - self.assets[pid].center
            if np.linalg.norm(separation) > 1e-9:
                self._append_unique_direction(geometry_dirs, separation)
                self._append_unique_direction(geometry_dirs, -separation)

        principal_axes = self._principal_axes()
        for axis_i in principal_axes:
            self._append_unique_direction(geometry_dirs, axis_i)
            self._append_unique_direction(geometry_dirs, -axis_i)
            self._append_unique_direction(geometry_dirs, outward + 0.5 * axis_i)
            self._append_unique_direction(geometry_dirs, outward - 0.5 * axis_i)

        diagonal_dirs: List[np.ndarray] = []
        if self.enable_diagonal_actions:
            basis = np.eye(3)
            for i, j in ((0, 1), (0, 2), (1, 2)):
                for si in (-1.0, 1.0):
                    for sj in (-1.0, 1.0):
                        self._append_unique_direction(
                            diagonal_dirs, si * basis[i] + sj * basis[j]
                        )
            for sx in (-1.0, 1.0):
                for sy in (-1.0, 1.0):
                    for sz in (-1.0, 1.0):
                        self._append_unique_direction(
                            diagonal_dirs,
                            np.array([sx, sy, sz], dtype=np.float64),
                        )

        geometry_dirs = [
            d for d in geometry_dirs
            if not any(float(np.dot(d, p)) > 0.998 for p in primary_dirs)
        ]
        up = np.array([0.0, 0.0, 1.0], dtype=np.float64)
        geometry_dirs.sort(
            key=lambda d: (float(np.dot(d, up)), float(np.dot(d, outward))),
            reverse=True,
        )
        diagonal_dirs.sort(
            key=lambda d: (float(np.dot(d, outward)), float(np.dot(d, up))),
            reverse=True,
        )

        translation_actions: List[np.ndarray] = []
        for direction in [*primary_dirs, *diagonal_dirs, *geometry_dirs]:
            action = np.zeros(6, dtype=np.float64)
            action[:3] = direction
            self._append_unique_action(translation_actions, action)

        rotation_actions: List[np.ndarray] = []
        random_axes: List[np.ndarray] = []
        if self.rotation:
            for axis in np.eye(3):
                for sign in (1.0, -1.0):
                    action = np.zeros(6, dtype=np.float64)
                    action[3:6] = sign * axis
                    self._append_unique_action(rotation_actions, action)

            digest = hashlib.sha256(self.move_id.encode("utf-8")).digest()
            seed = int.from_bytes(digest[:8], byteorder="little", signed=False)
            rng = np.random.default_rng(seed)
            for _ in range(self.random_rotation_actions):
                axis = rng.normal(size=3)
                if np.linalg.norm(axis) < 1e-9:
                    continue
                axis = normalize_vector(axis)
                random_axes.append(axis)
                action = np.zeros(6, dtype=np.float64)
                action[3:6] = axis
                self._append_unique_action(rotation_actions, action)

        coupled_actions: List[np.ndarray] = []
        if self.rotation and self.coupled_action_scale > 0.0:
            translation_dirs = [a[:3] for a in translation_actions[:24]]
            torque_pool = [*principal_axes, *np.eye(3), *random_axes[:6]]
            for direction in translation_dirs:
                direction = normalize_vector(direction)
                local_torque_axes: List[np.ndarray] = []
                for candidate in torque_pool:
                    candidate = normalize_vector(candidate)
                    if abs(float(np.dot(candidate, direction))) < 0.94:
                        self._append_unique_direction(local_torque_axes, candidate, 0.985)
                tangent = np.cross(direction, outward)
                if np.linalg.norm(tangent) > 1e-8:
                    self._append_unique_direction(local_torque_axes, tangent, 0.985)
                # Axial spin is useful for cylindrical or keyed parts.
                self._append_unique_direction(local_torque_axes, direction, 0.985)

                for torque_axis in local_torque_axes[:3]:
                    for sign in (1.0, -1.0):
                        action = np.zeros(6, dtype=np.float64)
                        action[:3] = direction
                        action[3:6] = (
                            sign * self.coupled_action_scale * torque_axis
                        )
                        self._append_unique_action(coupled_actions, action, 0.995)

        return translation_actions, rotation_actions, coupled_actions

    def state_key(self, qpos: np.ndarray) -> Tuple[int, ...]:
        qpos = np.asarray(qpos, dtype=np.float64)
        pos_res = max(self.trans_dist_th, 1e-6)
        key = list(np.rint(qpos[:3] / pos_res).astype(np.int64))
        if self.rotation:
            q = normalize_quat(qpos[3:7])
            quat_res = max(self.quat_dist_th / 4.0, 0.02)
            key.extend(np.rint(q / quat_res).astype(np.int64).tolist())
        return tuple(int(v) for v in key)

    @staticmethod
    def _direction_text(direction: np.ndarray) -> str:
        direction = normalize_vector(direction)
        axes = [
            ("+x", np.array([1.0, 0.0, 0.0])),
            ("-x", np.array([-1.0, 0.0, 0.0])),
            ("+y", np.array([0.0, 1.0, 0.0])),
            ("-y", np.array([0.0, -1.0, 0.0])),
            ("+z", np.array([0.0, 0.0, 1.0])),
            ("-z", np.array([0.0, 0.0, -1.0])),
        ]
        label, score = max(
            ((label, float(np.dot(direction, axis))) for label, axis in axes),
            key=lambda item: item[1],
        )
        if score > 0.999:
            return label
        return np.array2string(direction, precision=3, suppress_small=True)

    def _required_direct_release_distance(
        self,
        q0: np.ndarray,
        direction: np.ndarray,
    ) -> float:
        direction = normalize_vector(direction)
        move_hull = transform_mesh_by_qpos(
            self.assets[self.move_id].local_hull,
            q0,
        )
        move_projection = np.asarray(move_hull.vertices) @ direction

        still_projection = []
        for pid in self.still_ids:
            still_hull = transform_mesh_by_qpos(
                self.assets[pid].local_hull,
                make_initial_qpos(self.assets[pid]),
            )
            still_projection.append(np.asarray(still_hull.vertices) @ direction)

        if not still_projection:
            return self.world.min_removal_distance

        still_projection = np.concatenate(still_projection, axis=0)
        margin = max(
            2.0 * self.world._path_validator.sweep_translation_step,
            0.02 * self._part_diagonal,
            1.25 * self.world._path_validator.new_contact_tolerance,
        )
        required = (
            float(np.max(still_projection))
            - float(np.min(move_projection))
            + margin
        )
        return max(required, 1.20 * self.world.min_removal_distance, margin)

    def _aabb_clearance_metric(self, qpos: np.ndarray) -> float:
        if self._still_bounds is None:
            return self._part_diagonal
        move_hull = transform_mesh_by_qpos(
            self.assets[self.move_id].local_hull,
            qpos,
        )
        move_min, move_max = move_hull.bounds
        still_min, still_max = self._still_bounds
        positive_gap = np.maximum(
            np.maximum(still_min - move_max, move_min - still_max),
            0.0,
        )
        if np.any(positive_gap > 0.0):
            return float(np.linalg.norm(positive_gap))
        overlap = np.minimum(move_max, still_max) - np.maximum(move_min, still_min)
        return -float(max(np.min(overlap), 0.0))

    def _state_score(self, qpos: np.ndarray, depth: int) -> float:
        qpos = np.asarray(qpos, dtype=np.float64)
        delta = qpos[:3] - self._initial_qpos[:3]
        displacement = float(np.linalg.norm(delta)) / self._part_diagonal
        outward_progress = float(np.dot(delta, self._outward)) / self._part_diagonal
        clearance = self._aabb_clearance_metric(qpos) / self._part_diagonal
        contact_now = self.world._path_validator.initial_contact_depth_sum(qpos)
        contact_scale = max(
            self._initial_contact_sum,
            self.world._path_validator.new_contact_tolerance,
            1e-6,
        )
        contact_release = (self._initial_contact_sum - contact_now) / contact_scale
        rotation = quat_distance(qpos[3:7], self._initial_qpos[3:7])
        score = (
            2.2 * outward_progress
            + 1.0 * displacement
            + 2.8 * clearance
            + 1.5 * contact_release
            - 0.08 * float(depth)
            - 0.025 * rotation
        )
        return float(score)

    def _blocker_escape_direction(self, qpos: np.ndarray) -> np.ndarray:
        current = np.asarray(qpos[:3], dtype=np.float64)
        target = 1.5 * self._outward
        for pid in sorted(self.blockers):
            if pid not in self.assets:
                continue
            away = current - self.assets[pid].center
            if np.linalg.norm(away) > 1e-9:
                target += normalize_vector(away)
        return normalize_vector(target, fallback=self._outward)

    def _action_prior(self, node: SearchNode, action: np.ndarray) -> float:
        action = np.asarray(action, dtype=np.float64)
        trans_norm = float(np.linalg.norm(action[:3]))
        rot_norm = float(np.linalg.norm(action[3:6]))
        escape = self._blocker_escape_direction(node.qpos)
        score = 0.0
        if trans_norm > 1e-12:
            direction = action[:3] / trans_norm
            score += 1.8 * float(np.dot(direction, escape))
            score += 0.7 * float(np.dot(direction, self._outward))
            nonzero = int(np.count_nonzero(np.abs(direction) > 0.20))
            if nonzero >= 2:
                score += 0.12
        else:
            score -= 0.75
        if rot_norm > 1e-12 and trans_norm > 1e-12:
            score += 0.30 + 0.08 * min(node.depth, 3)
        elif rot_norm > 1e-12:
            score -= 0.10
        if node.last_action is not None:
            previous = np.asarray(node.last_action, dtype=np.float64)
            if np.linalg.norm(previous) > 1e-12 and np.linalg.norm(action) > 1e-12:
                score += 0.18 * float(
                    np.dot(previous / np.linalg.norm(previous), action / np.linalg.norm(action))
                )
        return float(score)

    @staticmethod
    def _action_signature(action: np.ndarray) -> Tuple[int, ...]:
        action = np.asarray(action, dtype=np.float64)
        norm = max(float(np.linalg.norm(action)), 1e-12)
        return tuple(np.rint(action / norm * 1000.0).astype(int).tolist())

    def _rank_actions_for_node(self, node: SearchNode) -> List[np.ndarray]:
        if self.search_strategy == "bfs":
            return [*self.translation_actions, *self.coupled_actions, *self.rotation_actions]

        candidates = [*self.translation_actions]
        if self.rotation:
            candidates.extend(self.coupled_actions)
            candidates.extend(self.rotation_actions)
        ranked = sorted(
            candidates,
            key=lambda action: self._action_prior(node, action),
            reverse=True,
        )

        # Preserve all six cardinal translations as a reliable base set, then
        # fill the remaining budget with score-ranked diagonal/coupled actions.
        selected: List[np.ndarray] = []
        signatures = set()
        for action in self.translation_actions[:6]:
            sig = self._action_signature(action)
            if sig not in signatures:
                selected.append(action)
                signatures.add(sig)
        for action in ranked:
            sig = self._action_signature(action)
            if sig in signatures:
                continue
            selected.append(action)
            signatures.add(sig)
            if len(selected) >= self.action_top_k:
                break
        return selected

    def _try_direct_kinematic_release(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        verbose: bool,
    ) -> Tuple[str, List[np.ndarray], str]:
        validator = self.world._path_validator
        pure_directions: List[np.ndarray] = []
        for action in self.translation_actions:
            action = np.asarray(action, dtype=np.float64)
            if float(np.linalg.norm(action[3:6])) > 1e-12:
                continue
            direction = action[:3]
            if float(np.linalg.norm(direction)) < 1e-12:
                continue
            direction = normalize_vector(direction)
            if any(float(np.dot(direction, old)) > 0.9995 for old in pure_directions):
                continue
            pure_directions.append(direction)

        nominal_step = max(
            validator.sweep_translation_step,
            min(0.01, 0.04 * self._part_diagonal),
        )
        print(
            f"[Direct release] {self.move_id}: probing "
            f"{len(pure_directions)} straight swept-mesh directions before "
            "force-driven search."
        )

        best_path = [initial.copy()]
        best_reason = "no straight release direction"
        best_score = self._state_score(initial, 0)
        first_failures = 0
        for direction in pure_directions:
            if time.time() - start_time > max_time:
                return "Timeout", best_path, "direct probe timeout"

            label = self._direction_text(direction)
            distance = self._required_direct_release_distance(initial, direction)
            n_steps = max(int(math.ceil(distance / nominal_step)), 1)
            n_steps = min(n_steps, 400)
            step_distance = distance / float(n_steps)

            path = [initial.copy()]
            previous = initial.copy()
            failed_reason = "not disassembled"
            for index in range(1, n_steps + 1):
                qpos = initial.copy()
                qpos[:3] = initial[:3] + direction * min(
                    index * step_distance, distance
                )
                valid, reason = validator.segment_is_valid(
                    previous,
                    qpos,
                    release_direction=direction,
                    allow_initial_release=True,
                )
                if not valid:
                    self.blockers.update(validator.last_blockers)
                    failed_reason = reason
                    break
                path.append(qpos.copy())
                previous = qpos

            score = self._state_score(path[-1], 1)
            if score > best_score or len(path) > len(best_path):
                best_score = score
                best_path = [q.copy() for q in path]
                best_reason = f"direction={label}: {failed_reason}"
            self._consider_failure_path(path, qpos=path[-1], penalty=0.05)

            if len(path) == n_steps + 1:
                final_qpos = path[-1]
                if self.world.is_disassembled(final_qpos):
                    path_valid, reason = validator.path_is_valid(
                        path,
                        self.world.min_removal_distance,
                        allow_initial_release=True,
                    )
                    if path_valid:
                        print(
                            f"[Direct release] Success for {self.move_id}: "
                            f"direction={label}, distance={distance:.6g} m, "
                            f"frames={len(path)}."
                        )
                        return "Success", path, "ok"
                    failed_reason = reason
                else:
                    failed_reason = "final pose did not clear remaining-assembly envelope"

            if verbose or first_failures < 3:
                print(
                    f"    [Direct release reject] {self.move_id} "
                    f"direction={label}: {failed_reason}"
                )
                first_failures += 1

        return "Failure", best_path, best_reason

    def _consider_failure_path(
        self,
        path: Sequence[np.ndarray],
        qpos: Optional[np.ndarray] = None,
        penalty: float = 0.0,
    ) -> None:
        if not path:
            return
        terminal = np.asarray(qpos if qpos is not None else path[-1], dtype=np.float64)
        # Score diagnostic candidates without requiring them to be collision-free.
        delta = terminal[:3] - self._initial_qpos[:3]
        displacement = float(np.linalg.norm(delta)) / self._part_diagonal
        outward = float(np.dot(delta, self._outward)) / self._part_diagonal
        score = 1.5 * outward + displacement - float(penalty)
        if score > self._best_failure_score or (
            abs(score - self._best_failure_score) < 1e-9
            and len(path) > len(self._best_failure_path)
        ):
            self._best_failure_score = score
            self._best_failure_path = [
                np.asarray(item, dtype=np.float64).copy() for item in path
            ]

    def _record_failure(self, reason: str) -> None:
        reason = str(reason or "unknown failure")
        self.last_expand_reason = reason
        self.failure_reason_counts[reason] += 1

    def _expand_action(
        self,
        node: SearchNode,
        action: np.ndarray,
        visited: set,
        start_time: float,
        max_time: float,
        verbose: bool,
        collision_rejections: int,
    ):
        self._evaluated_actions += 1
        self.world.set_qpos(node.qpos, zero_velocity=True)
        branch_path = [np.asarray(q, dtype=np.float64).copy() for q in node.path]
        terminal = node.qpos.copy()
        no_progress_count = 0
        branch_invalid = False
        branch_seen = {self.state_key(node.qpos)}
        self.last_expand_reason = "no measurable progress"

        for _ in range(self.max_action_chunks):
            if time.time() - start_time > max_time:
                self._record_failure("path search timeout")
                return "Timeout", terminal, branch_path, branch_invalid, collision_rejections
            step_states, motion_valid, invalid_reason = self.world.step_action(
                action, self.frame_skip
            )
            if not motion_valid:
                collision_rejections += 1
                self._collision_rejections += 1
                self.blockers.update(self.world._path_validator.last_blockers)
                branch_invalid = True
                self._record_failure(invalid_reason)
                diagnostic_path = [q.copy() for q in branch_path]
                invalid_qpos = getattr(self.world, "last_invalid_qpos", None)
                if invalid_qpos is not None:
                    diagnostic_path.append(np.asarray(invalid_qpos).copy())
                self._consider_failure_path(
                    diagnostic_path,
                    qpos=(invalid_qpos if invalid_qpos is not None else terminal),
                    penalty=0.15,
                )
                if verbose or collision_rejections <= 3:
                    print(
                        f"    [Collision reject] {self.move_id} "
                        f"action={np.asarray(action).round(4).tolist()}: {invalid_reason}"
                    )
                break
            if not step_states:
                self._record_failure("physics rollout returned no state")
                self._consider_failure_path(branch_path, qpos=terminal, penalty=0.25)
                break

            branch_path.extend(q.copy() for q in step_states)
            new_qpos = step_states[-1].copy()
            terminal = new_qpos.copy()

            if self.world.is_disassembled(new_qpos):
                valid_path, reason = self.world._path_validator.path_is_valid(
                    branch_path,
                    self.world.min_removal_distance,
                    allow_initial_release=True,
                )
                if valid_path:
                    return "Success", terminal, branch_path, branch_invalid, collision_rejections
                branch_invalid = True
                collision_rejections += 1
                self._collision_rejections += 1
                self.blockers.update(self.world._path_validator.last_blockers)
                self._record_failure(reason)
                self._consider_failure_path(branch_path, qpos=new_qpos, penalty=0.10)
                if verbose or collision_rejections <= 3:
                    print(f"    [Path reject] {self.move_id}: {reason}")
                break

            new_key = self.state_key(new_qpos)
            if new_key in visited or new_key in branch_seen:
                no_progress_count += 1
            else:
                no_progress_count = 0
                branch_seen.add(new_key)

            if no_progress_count >= 2:
                self._record_failure("repeated discretized state")
                self._consider_failure_path(branch_path, qpos=new_qpos, penalty=0.05)
                break

        return "Continue", terminal, branch_path, branch_invalid, collision_rejections

    def _finalize_search(
        self,
        status: str,
        started: float,
        path: Sequence[np.ndarray],
        reason: str,
    ):
        output_path = [np.asarray(q).copy() for q in path]
        if status != "Success":
            self.last_failure_reason = str(reason)
            output_terminal = output_path[-1] if output_path else self._initial_qpos
            output_delta = output_terminal[:3] - self._initial_qpos[:3]
            output_progress_score = (
                1.5 * float(np.dot(output_delta, self._outward)) / self._part_diagonal
                + float(np.linalg.norm(output_delta)) / self._part_diagonal
            )
            if (
                self._best_failure_path
                and self._best_failure_score > output_progress_score
            ):
                output_path = [q.copy() for q in self._best_failure_path]
        else:
            self.last_failure_reason = ""
        blocker_counts = {}
        if self.world is not None and self.world._path_validator is not None:
            blocker_counts = dict(self.world._path_validator.blocker_counts)
        self.search_diagnostics = {
            "strategy": self.search_strategy,
            "expanded_nodes": int(self._expanded_nodes),
            "evaluated_actions": int(self._evaluated_actions),
            "generated_successors": int(self._generated_successors),
            "collision_rejections": int(self._collision_rejections),
            "best_heuristic_score": (
                None if not math.isfinite(self._best_score) else float(self._best_score)
            ),
            "failure_reason": self.last_failure_reason,
            "failure_reason_counts": dict(self.failure_reason_counts),
            "blockers": sorted(self.blockers),
            "blocker_counts": blocker_counts,
            "initial_contact_depths": {
                pid: float(depth)
                for pid, depth in self.world._path_validator.initial_depths.items()
            },
            "translation_action_count": int(len(self.translation_actions)),
            "rotation_action_count": int(len(self.rotation_actions)),
            "coupled_action_count": int(len(self.coupled_actions)),
            "beam_width": int(self.beam_width),
            "action_top_k": int(self.action_top_k),
            "rrt_max_nodes": int(self.rrt_max_nodes),
            "rrt_goal_bias": float(self.rrt_goal_bias),
            "rrt_neighbor_count": int(self.rrt_neighbor_count),
            "rrt_connect_distance_factor": float(self.rrt_connect_distance_factor),
            "rrt_goal_root_count": int(self.rrt_goal_root_count),
            "rrt_nodes": int(self._rrt_nodes),
            "rrt_rewires": int(self._rrt_rewires),
            "birrt_start_nodes": int(self._birrt_start_nodes),
            "birrt_goal_nodes": int(self._birrt_goal_nodes),
        }
        return status, time.time() - started, output_path

    def _plan_fifo(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        max_depth: int,
        verbose: bool,
        seed_path: List[np.ndarray],
    ):
        queue: Deque[SearchNode] = deque(
            [SearchNode(initial, [initial], 0, score=self._state_score(initial, 0))]
        )
        visited = {self.state_key(initial)}
        best_path = [q.copy() for q in seed_path] if seed_path else [initial.copy()]
        best_score = self._state_score(best_path[-1], 0)
        self._best_score = best_score
        collision_rejections = 0
        rotation_activation_printed = False

        while queue:
            if time.time() - start_time > max_time:
                return self._finalize_search(
                    "Timeout", start_time, best_path, "path search timeout"
                )
            node = queue.popleft()
            self._expanded_nodes += 1
            if node.depth >= max_depth:
                continue

            translation_successor = False
            for action in self.translation_actions:
                (
                    result,
                    terminal,
                    branch_path,
                    branch_invalid,
                    collision_rejections,
                ) = self._expand_action(
                    node,
                    action,
                    visited,
                    start_time,
                    max_time,
                    verbose,
                    collision_rejections,
                )
                if result == "Timeout":
                    return self._finalize_search(
                        "Timeout", start_time, best_path, "path search timeout"
                    )
                if result == "Success":
                    return self._finalize_search("Success", start_time, branch_path, "ok")
                if branch_invalid or self.state_key(terminal) == self.state_key(node.qpos):
                    continue

                score = self._state_score(terminal, node.depth + 1)
                child = SearchNode(
                    terminal,
                    branch_path,
                    node.depth + 1,
                    score=score,
                    last_action=np.asarray(action).copy(),
                    action_trace=[*node.action_trace, np.asarray(action).tolist()],
                )
                visited.add(self.state_key(terminal))
                queue.append(child)
                self._generated_successors += 1
                translation_successor = True
                if score > best_score:
                    best_score = score
                    self._best_score = score
                    best_path = [q.copy() for q in branch_path]

            if self.rotation and not translation_successor:
                if not rotation_activation_printed:
                    print(
                        f"[Search] No valid translational successor for {self.move_id}; "
                        "enabling pure and coupled 6-DoF actions."
                    )
                    rotation_activation_printed = True
                for action in [*self.coupled_actions, *self.rotation_actions]:
                    (
                        result,
                        terminal,
                        branch_path,
                        branch_invalid,
                        collision_rejections,
                    ) = self._expand_action(
                        node,
                        action,
                        visited,
                        start_time,
                        max_time,
                        verbose,
                        collision_rejections,
                    )
                    if result == "Timeout":
                        return self._finalize_search(
                            "Timeout", start_time, best_path, "path search timeout"
                        )
                    if result == "Success":
                        return self._finalize_search("Success", start_time, branch_path, "ok")
                    if branch_invalid or self.state_key(terminal) == self.state_key(node.qpos):
                        continue
                    score = self._state_score(terminal, node.depth + 1)
                    visited.add(self.state_key(terminal))
                    queue.append(
                        SearchNode(
                            terminal,
                            branch_path,
                            node.depth + 1,
                            score=score,
                            last_action=np.asarray(action).copy(),
                            action_trace=[*node.action_trace, np.asarray(action).tolist()],
                        )
                    )
                    self._generated_successors += 1
                    if score > best_score:
                        best_score = score
                        self._best_score = score
                        best_path = [q.copy() for q in branch_path]

        common_reason = (
            max(self.failure_reason_counts, key=self.failure_reason_counts.get)
            if self.failure_reason_counts
            else "search queue exhausted"
        )
        return self._finalize_search(
            "Failure", start_time, best_path, f"search queue exhausted: {common_reason}"
        )

    def _plan_adaptive_best_first(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        max_depth: int,
        verbose: bool,
        seed_path: List[np.ndarray],
    ):
        initial_score = self._state_score(initial, 0)
        root = SearchNode(initial, [initial], 0, score=initial_score)
        counter = 0
        frontier = [(-initial_score, 0, counter, root)]
        visited = {self.state_key(initial)}
        best_path = [q.copy() for q in seed_path] if seed_path else [initial.copy()]
        best_score = max(initial_score, self._state_score(best_path[-1], 0))
        self._best_score = best_score
        collision_rejections = 0

        print(
            f"[Adaptive search] {self.move_id}: beam_width={self.beam_width}, "
            f"action_top_k={self.action_top_k}, translations={len(self.translation_actions)}, "
            f"coupled={len(self.coupled_actions)}, rotations={len(self.rotation_actions)}"
        )

        while frontier:
            if time.time() - start_time > max_time:
                return self._finalize_search(
                    "Timeout", start_time, best_path, "path search timeout"
                )
            _, _, _, node = heapq.heappop(frontier)
            self._expanded_nodes += 1
            if node.depth >= max_depth:
                continue

            actions = self._rank_actions_for_node(node)
            successors: List[SearchNode] = []
            for action in actions:
                (
                    result,
                    terminal,
                    branch_path,
                    branch_invalid,
                    collision_rejections,
                ) = self._expand_action(
                    node,
                    action,
                    visited,
                    start_time,
                    max_time,
                    verbose,
                    collision_rejections,
                )
                if result == "Timeout":
                    return self._finalize_search(
                        "Timeout", start_time, best_path, "path search timeout"
                    )
                if result == "Success":
                    self.solution_mode = "adaptive_contact_guided_best_first"
                    return self._finalize_search("Success", start_time, branch_path, "ok")
                if branch_invalid or self.state_key(terminal) == self.state_key(node.qpos):
                    continue

                score = self._state_score(terminal, node.depth + 1)
                child = SearchNode(
                    qpos=terminal,
                    path=branch_path,
                    depth=node.depth + 1,
                    score=score,
                    last_action=np.asarray(action, dtype=np.float64).copy(),
                    action_trace=[*node.action_trace, np.asarray(action).tolist()],
                )
                successors.append(child)
                self._generated_successors += 1
                if score > best_score:
                    best_score = score
                    self._best_score = score
                    best_path = [q.copy() for q in branch_path]

            successors.sort(key=lambda item: item.score, reverse=True)
            for child in successors[: self.beam_width]:
                child_key = self.state_key(child.qpos)
                if child_key in visited:
                    continue
                visited.add(child_key)
                counter += 1
                heapq.heappush(
                    frontier,
                    (-child.score, child.depth, counter, child),
                )

            if verbose and successors:
                print(
                    f"    [Adaptive frontier] depth={node.depth + 1}, "
                    f"generated={len(successors)}, kept={min(len(successors), self.beam_width)}, "
                    f"best_score={best_score:.4f}, frontier={len(frontier)}"
                )

        common_reason = (
            max(self.failure_reason_counts, key=self.failure_reason_counts.get)
            if self.failure_reason_counts
            else "adaptive frontier exhausted"
        )
        return self._finalize_search(
            "Failure",
            start_time,
            best_path,
            f"adaptive frontier exhausted: {common_reason}",
        )


    def _plan_dfs(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        max_depth: int,
        verbose: bool,
        seed_path: List[np.ndarray],
    ):
        """Depth-first counterpart of the FIFO physics search.

        The action model and exact path validation are identical to BFS; only
        frontier discipline changes.  Children are pushed in reverse priority
        order so the first configured action is explored first.
        """
        stack: List[SearchNode] = [
            SearchNode(initial, [initial], 0, score=self._state_score(initial, 0))
        ]
        visited = {self.state_key(initial)}
        best_path = [q.copy() for q in seed_path] if seed_path else [initial.copy()]
        best_score = self._state_score(best_path[-1], 0)
        self._best_score = best_score
        collision_rejections = 0

        while stack:
            if time.time() - start_time > max_time:
                return self._finalize_search(
                    "Timeout", start_time, best_path, "path search timeout"
                )
            node = stack.pop()
            self._expanded_nodes += 1
            if node.depth >= max_depth:
                continue

            actions = [*self.translation_actions]
            if self.rotation:
                actions.extend(self.coupled_actions)
                actions.extend(self.rotation_actions)
            children: List[SearchNode] = []
            for action in actions:
                result, terminal, branch_path, branch_invalid, collision_rejections = (
                    self._expand_action(
                        node, action, visited, start_time, max_time, verbose,
                        collision_rejections,
                    )
                )
                if result == "Timeout":
                    return self._finalize_search(
                        "Timeout", start_time, best_path, "path search timeout"
                    )
                if result == "Success":
                    self.solution_mode = "physics_dfs"
                    return self._finalize_search(
                        "Success", start_time, branch_path, "ok"
                    )
                key = self.state_key(terminal)
                if branch_invalid or key == self.state_key(node.qpos) or key in visited:
                    continue
                score = self._state_score(terminal, node.depth + 1)
                visited.add(key)
                children.append(
                    SearchNode(
                        terminal,
                        branch_path,
                        node.depth + 1,
                        score=score,
                        last_action=np.asarray(action).copy(),
                        action_trace=[*node.action_trace, np.asarray(action).tolist()],
                    )
                )
                self._generated_successors += 1
                if score > best_score:
                    best_score = score
                    self._best_score = score
                    best_path = [q.copy() for q in branch_path]

            for child in reversed(children):
                stack.append(child)

        common_reason = (
            max(self.failure_reason_counts, key=self.failure_reason_counts.get)
            if self.failure_reason_counts else "search stack exhausted"
        )
        return self._finalize_search(
            "Failure", start_time, best_path,
            f"search stack exhausted: {common_reason}",
        )

    def _tree_state_distance(self, q0: np.ndarray, q1: np.ndarray) -> float:
        q0 = np.asarray(q0, dtype=np.float64)
        q1 = np.asarray(q1, dtype=np.float64)
        translation = float(np.linalg.norm(q0[:3] - q1[:3])) / self._part_diagonal
        rotation = quat_distance(q0[3:7], q1[3:7]) / math.pi
        return translation + 0.20 * rotation

    def _edge_cost(self, edge_path: Sequence[np.ndarray]) -> float:
        if edge_path is None or len(edge_path) < 2:
            return 0.0
        total = 0.0
        previous = np.asarray(edge_path[0], dtype=np.float64)
        for qpos in edge_path[1:]:
            qpos = np.asarray(qpos, dtype=np.float64)
            total += float(np.linalg.norm(qpos[:3] - previous[:3])) / self._part_diagonal
            total += 0.05 * quat_distance(previous[3:7], qpos[3:7]) / math.pi
            previous = qpos
        return float(total)

    @staticmethod
    def _tree_is_ancestor(
        nodes: Sequence[RRTTreeNode], ancestor: int, descendant: int
    ) -> bool:
        current: Optional[int] = descendant
        while current is not None:
            if current == ancestor:
                return True
            current = nodes[current].parent
        return False

    def _tree_path(
        self, nodes: Sequence[RRTTreeNode], node_index: int
    ) -> List[np.ndarray]:
        chain: List[int] = []
        current: Optional[int] = int(node_index)
        while current is not None:
            chain.append(current)
            current = nodes[current].parent
        chain.reverse()
        path = [np.asarray(nodes[chain[0]].qpos, dtype=np.float64).copy()]
        for index in chain[1:]:
            edge = nodes[index].edge_path
            if not edge:
                path.append(np.asarray(nodes[index].qpos, dtype=np.float64).copy())
                continue
            start = 1 if np.allclose(path[-1], edge[0], atol=1e-9) else 0
            path.extend(np.asarray(q, dtype=np.float64).copy() for q in edge[start:])
        return path

    def _propagate_tree_costs(
        self, nodes: List[RRTTreeNode], root_index: int
    ) -> None:
        queue = deque([int(root_index)])
        while queue:
            parent_index = queue.popleft()
            parent = nodes[parent_index]
            for child_index, child in enumerate(nodes):
                if child.parent != parent_index:
                    continue
                child.depth = parent.depth + 1
                child.cost = parent.cost + self._edge_cost(child.edge_path)
                queue.append(child_index)

    def _sample_rrt_target(self, behavior: Optional[str] = None) -> np.ndarray:
        rng = self._rrt_rng
        target = self._initial_qpos.copy()
        behavior = str(behavior or "uniform")
        if behavior == "outward":
            direction = normalize_vector(
                self._outward + 0.30 * rng.normal(size=3), fallback=self._outward
            )
            radius = rng.uniform(0.9, 1.8) * self._assembly_extent
            target[:3] = self._assembly_center + direction * radius
        elif behavior == "blocker_escape":
            direction = normalize_vector(
                self._blocker_escape_direction(self._initial_qpos)
                + 0.20 * rng.normal(size=3),
                fallback=self._outward,
            )
            radius = rng.uniform(0.7, 1.6) * self._assembly_extent
            target[:3] = self._initial_qpos[:3] + direction * radius
        elif rng.random() < self.rrt_goal_bias:
            direction = normalize_vector(rng.normal(size=3), fallback=self._outward)
            if float(np.dot(direction, self._outward)) < 0.0:
                direction = -direction
            radius = rng.uniform(0.9, 1.8) * self._assembly_extent
            target[:3] = self._assembly_center + direction * radius
        else:
            margin = 0.65 * self._assembly_extent
            target[:3] = rng.uniform(
                self._assembly_min - margin,
                self._assembly_max + margin,
            )

        if self.rotation and rng.random() < 0.55:
            axis = normalize_vector(rng.normal(size=3))
            angle = rng.uniform(-math.pi, math.pi)
            target[3:7] = quat_multiply(
                axis_angle_to_quat(axis, angle), self._initial_qpos[3:7]
            )
        return target

    def _tree_action_candidates(
        self,
        node_qpos: np.ndarray,
        target_qpos: np.ndarray,
        last_action: Optional[np.ndarray],
        behavioral: bool,
    ) -> List[np.ndarray]:
        node = SearchNode(
            np.asarray(node_qpos, dtype=np.float64),
            [np.asarray(node_qpos, dtype=np.float64)],
            0,
            last_action=last_action,
        )
        actions = [*self.translation_actions]
        if self.rotation:
            actions.extend(self.coupled_actions)
            actions.extend(self.rotation_actions)
        delta = np.asarray(target_qpos[:3] - node_qpos[:3], dtype=np.float64)
        target_direction = normalize_vector(delta, fallback=self._outward)

        def score(action: np.ndarray) -> float:
            action = np.asarray(action, dtype=np.float64)
            trans_norm = float(np.linalg.norm(action[:3]))
            rot_norm = float(np.linalg.norm(action[3:6]))
            value = 0.0
            if trans_norm > 1e-12:
                value += 2.0 * float(np.dot(action[:3] / trans_norm, target_direction))
            elif rot_norm > 1e-12:
                value -= 0.15
            if behavioral:
                value += 1.25 * self._action_prior(node, action)
                if trans_norm > 1e-12 and rot_norm > 1e-12:
                    value += 0.25
            else:
                # Random tie-breaking prevents a hidden deterministic PhyD2A
                # action ordering from leaking into classical RRT baselines.
                value += float(self._rrt_rng.uniform(-0.08, 0.08))
            return value

        actions.sort(key=score, reverse=True)
        if not behavioral and self._rrt_rng.random() < 0.30:
            # Exploration branch: permit pure rotations and non-greedy forces,
            # otherwise a target-directed RRT can degenerate into repeated
            # straight translations on tightly constrained parts.
            self._rrt_rng.shuffle(actions)
        budget = min(len(actions), max(6, min(self.action_top_k, 18)))
        return actions[:budget]

    def _rollout_tree_edge(
        self,
        nodes: Sequence[RRTTreeNode],
        parent_index: int,
        action: np.ndarray,
        visited: set,
        start_time: float,
        max_time: float,
        verbose: bool,
        collision_rejections: int,
    ):
        parent_path = self._tree_path(nodes, parent_index)
        parent = nodes[parent_index]
        search_node = SearchNode(
            qpos=parent.qpos.copy(),
            path=parent_path,
            depth=parent.depth,
            score=self._state_score(parent.qpos, parent.depth),
            last_action=parent.last_action,
        )
        result, terminal, branch_path, invalid, collision_rejections = self._expand_action(
            search_node, action, visited, start_time, max_time, verbose,
            collision_rejections,
        )
        edge_start = max(len(parent_path) - 1, 0)
        edge_path = [np.asarray(q).copy() for q in branch_path[edge_start:]]
        return result, terminal, branch_path, edge_path, invalid, collision_rejections

    def _rewire_rrt_star(
        self,
        nodes: List[RRTTreeNode],
        new_index: int,
    ) -> int:
        new_node = nodes[new_index]
        radius = max(self.rrt_connect_distance_factor, 0.10)
        candidates = sorted(
            (
                (self._tree_state_distance(new_node.qpos, node.qpos), index)
                for index, node in enumerate(nodes[:-1])
            ),
            key=lambda item: item[0],
        )[: self.rrt_neighbor_count]
        rewired = 0
        validator = self.world._path_validator
        for distance, index in candidates:
            if distance > radius or index == 0:
                continue
            if self._tree_is_ancestor(nodes, index, new_index):
                continue
            neighbor = nodes[index]
            segment = [new_node.qpos.copy(), *interpolate_pose_segment(
                new_node.qpos,
                neighbor.qpos,
                translation_step=validator.sweep_translation_step,
                rotation_step=validator.sweep_rotation_step,
            )]
            candidate_cost = new_node.cost + self._edge_cost(segment)
            if candidate_cost + 1e-6 >= neighbor.cost:
                continue
            valid, _ = validator.segment_is_valid(new_node.qpos, neighbor.qpos)
            if not valid:
                continue
            neighbor.parent = new_index
            neighbor.edge_path = segment
            neighbor.depth = new_node.depth + 1
            neighbor.cost = candidate_cost
            self._propagate_tree_costs(nodes, index)
            rewired += 1
        return rewired

    def _plan_rrt_family(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        max_depth: int,
        verbose: bool,
        seed_path: List[np.ndarray],
        variant: str,
    ):
        behavioral = variant == "bk_rrt"
        star = variant == "rrt_star"
        nodes: List[RRTTreeNode] = [
            RRTTreeNode(initial.copy(), None, [initial.copy()], 0, 0.0, None)
        ]
        visited = {self.state_key(initial)}
        best_path = [q.copy() for q in seed_path] if seed_path else [initial.copy()]
        best_score = self._state_score(best_path[-1], 0)
        self._best_score = best_score
        collision_rejections = 0
        rewires = 0
        behavior_cycle = ("outward", "blocker_escape", "uniform", "uniform")

        print(
            f"[{variant.upper()}] {self.move_id}: max_nodes={self.rrt_max_nodes}, "
            f"goal_bias={self.rrt_goal_bias:.3f}, max_depth={max_depth}"
        )
        while len(nodes) < self.rrt_max_nodes:
            if time.time() - start_time > max_time:
                self._rrt_nodes = len(nodes)
                self._rrt_rewires = rewires
                return self._finalize_search(
                    "Timeout", start_time, best_path, "path search timeout"
                )

            behavior = None
            if behavioral:
                behavior = behavior_cycle[(len(nodes) - 1) % len(behavior_cycle)]
            target = self._sample_rrt_target(behavior)
            nearest = sorted(
                range(len(nodes)),
                key=lambda index: self._tree_state_distance(nodes[index].qpos, target),
            )
            parent_candidates = nearest[: (min(self.rrt_neighbor_count, 3) if star else 1)]
            best_candidate = None
            for parent_index in parent_candidates:
                parent = nodes[parent_index]
                if parent.depth >= max_depth:
                    continue
                actions = self._tree_action_candidates(
                    parent.qpos, target, parent.last_action, behavioral
                )
                # Classical RRT extends once toward the random sample. RRT*
                # may evaluate two local controls for parent optimization.
                action_budget = 2 if star else 1
                for action in actions[:action_budget]:
                    result, terminal, branch_path, edge_path, invalid, collision_rejections = (
                        self._rollout_tree_edge(
                            nodes, parent_index, action, visited, start_time,
                            max_time, verbose, collision_rejections,
                        )
                    )
                    if result == "Timeout":
                        return self._finalize_search(
                            "Timeout", start_time, best_path, "path search timeout"
                        )
                    if result == "Success":
                        self.solution_mode = {
                            "rrt": "physics_rrt",
                            "rrt_star": "physics_rrt_star",
                            "bk_rrt": "behavioral_kinodynamic_rrt",
                        }[variant]
                        self._rrt_nodes = len(nodes)
                        self._rrt_rewires = rewires
                        return self._finalize_search(
                            "Success", start_time, branch_path, "ok"
                        )
                    key = self.state_key(terminal)
                    if invalid or key in visited or key == self.state_key(parent.qpos):
                        continue
                    edge_cost = self._edge_cost(edge_path)
                    objective = parent.cost + edge_cost
                    objective -= 0.08 * self._state_score(terminal, parent.depth + 1)
                    candidate = (
                        objective, parent_index, np.asarray(action).copy(),
                        terminal.copy(), edge_path, branch_path,
                    )
                    if best_candidate is None or objective < best_candidate[0]:
                        best_candidate = candidate

            if best_candidate is None:
                continue
            _, parent_index, action, terminal, edge_path, branch_path = best_candidate
            parent = nodes[parent_index]
            child = RRTTreeNode(
                qpos=terminal.copy(),
                parent=parent_index,
                edge_path=[q.copy() for q in edge_path],
                depth=parent.depth + 1,
                cost=parent.cost + self._edge_cost(edge_path),
                last_action=action.copy(),
            )
            nodes.append(child)
            child_index = len(nodes) - 1
            visited.add(self.state_key(terminal))
            self._generated_successors += 1
            if star:
                rewires += self._rewire_rrt_star(nodes, child_index)

            score = self._state_score(terminal, child.depth)
            if score > best_score:
                best_score = score
                self._best_score = score
                best_path = self._tree_path(nodes, child_index)

        self._rrt_nodes = len(nodes)
        self._rrt_rewires = rewires
        return self._finalize_search(
            "Failure", start_time, best_path,
            f"{variant} exhausted {self.rrt_max_nodes} nodes",
        )

    def _make_birrt_goal_roots(self, initial: np.ndarray) -> List[np.ndarray]:
        validator = self.world._path_validator
        directions: List[np.ndarray] = [
            self._outward.copy(),
            np.array([1.0, 0.0, 0.0]), np.array([-1.0, 0.0, 0.0]),
            np.array([0.0, 1.0, 0.0]), np.array([0.0, -1.0, 0.0]),
            np.array([0.0, 0.0, 1.0]), np.array([0.0, 0.0, -1.0]),
        ]
        while len(directions) < self.rrt_goal_root_count * 2:
            directions.append(normalize_vector(self._rrt_rng.normal(size=3)))
        roots: List[np.ndarray] = []
        for direction in directions:
            if len(roots) >= self.rrt_goal_root_count:
                break
            direction = normalize_vector(direction)
            distance = max(
                1.10 * self._assembly_extent,
                self._required_direct_release_distance(initial, direction),
            )
            qpos = initial.copy()
            qpos[:3] = initial[:3] + direction * distance
            free, _ = validator.pose_is_collision_free(qpos)
            if free and self.world.is_disassembled(qpos):
                roots.append(qpos)
        return roots

    def _kinematic_tree_extend(
        self,
        nodes: List[RRTTreeNode],
        target: np.ndarray,
        max_depth: int,
    ) -> Optional[int]:
        candidates = sorted(
            range(len(nodes)),
            key=lambda index: self._tree_state_distance(nodes[index].qpos, target),
        )
        validator = self.world._path_validator
        max_translation = max(0.18 * self._part_diagonal, 2.0 * validator.sweep_translation_step)
        max_rotation = max(0.20, 2.0 * validator.sweep_rotation_step)
        for parent_index in candidates[: self.rrt_neighbor_count]:
            parent = nodes[parent_index]
            if parent.depth >= max_depth:
                continue
            delta = target[:3] - parent.qpos[:3]
            distance = float(np.linalg.norm(delta))
            alpha_t = 1.0 if distance <= max_translation else max_translation / max(distance, 1e-12)
            rotation_distance = quat_distance(parent.qpos[3:7], target[3:7])
            alpha_r = 1.0 if rotation_distance <= max_rotation else max_rotation / max(rotation_distance, 1e-12)
            alpha = min(alpha_t, alpha_r)
            qpos = parent.qpos.copy()
            qpos[:3] = parent.qpos[:3] + alpha * delta
            qpos[3:7] = slerp_quat(parent.qpos[3:7], target[3:7], alpha)
            valid, _ = validator.segment_is_valid(parent.qpos, qpos)
            if not valid:
                continue
            edge = [parent.qpos.copy(), *interpolate_pose_segment(
                parent.qpos, qpos,
                validator.sweep_translation_step,
                validator.sweep_rotation_step,
            )]
            nodes.append(
                RRTTreeNode(
                    qpos=qpos,
                    parent=parent_index,
                    edge_path=edge,
                    depth=parent.depth + 1,
                    cost=parent.cost + self._edge_cost(edge),
                )
            )
            return len(nodes) - 1
        return None

    def _plan_bi_rrt(
        self,
        initial: np.ndarray,
        start_time: float,
        max_time: float,
        max_depth: int,
        verbose: bool,
        seed_path: List[np.ndarray],
    ):
        goal_roots = self._make_birrt_goal_roots(initial)
        if not goal_roots:
            return self._finalize_search(
                "Failure", start_time, seed_path,
                "Bi-RRT could not construct a collision-free exterior goal set",
            )
        start_tree = [RRTTreeNode(initial.copy(), None, [initial.copy()], 0, 0.0)]
        goal_tree = [
            RRTTreeNode(root.copy(), None, [root.copy()], 0, 0.0)
            for root in goal_roots
        ]
        visited = {self.state_key(initial)}
        best_path = [q.copy() for q in seed_path] if seed_path else [initial.copy()]
        best_score = self._state_score(best_path[-1], 0)
        collision_rejections = 0
        connect_limit = self.rrt_connect_distance_factor
        validator = self.world._path_validator

        print(
            f"[BI_RRT] {self.move_id}: goal_roots={len(goal_roots)}, "
            f"max_nodes={self.rrt_max_nodes}, connect_factor={connect_limit:.3f}"
        )
        while len(start_tree) + len(goal_tree) < 2 * self.rrt_max_nodes:
            if time.time() - start_time > max_time:
                self._birrt_start_nodes = len(start_tree)
                self._birrt_goal_nodes = len(goal_tree)
                return self._finalize_search(
                    "Timeout", start_time, best_path, "path search timeout"
                )

            target = self._sample_rrt_target()
            parent_index = min(
                range(len(start_tree)),
                key=lambda index: self._tree_state_distance(start_tree[index].qpos, target),
            )
            parent = start_tree[parent_index]
            new_start_index = None
            if parent.depth < max_depth:
                actions = self._tree_action_candidates(
                    parent.qpos, target, parent.last_action, False
                )
                for action in actions[:2]:
                    result, terminal, branch_path, edge_path, invalid, collision_rejections = (
                        self._rollout_tree_edge(
                            start_tree, parent_index, action, visited, start_time,
                            max_time, verbose, collision_rejections,
                        )
                    )
                    if result == "Timeout":
                        return self._finalize_search(
                            "Timeout", start_time, best_path, "path search timeout"
                        )
                    if result == "Success":
                        self.solution_mode = "bidirectional_rrt"
                        return self._finalize_search(
                            "Success", start_time, branch_path, "ok"
                        )
                    key = self.state_key(terminal)
                    if invalid or key in visited or key == self.state_key(parent.qpos):
                        continue
                    start_tree.append(
                        RRTTreeNode(
                            terminal.copy(), parent_index, edge_path,
                            parent.depth + 1,
                            parent.cost + self._edge_cost(edge_path),
                            np.asarray(action).copy(),
                        )
                    )
                    new_start_index = len(start_tree) - 1
                    visited.add(key)
                    break

            # Grow the exterior tree backwards through exact kinematic local
            # connections. Reversing these collision-free segments is valid.
            goal_target = (
                start_tree[new_start_index].qpos.copy()
                if new_start_index is not None
                else self._sample_rrt_target()
            )
            self._kinematic_tree_extend(
                goal_tree, goal_target, max_depth=max_depth * 2
            )

            if new_start_index is None:
                continue
            start_node = start_tree[new_start_index]
            nearest_goal_index = min(
                range(len(goal_tree)),
                key=lambda index: self._tree_state_distance(
                    start_node.qpos, goal_tree[index].qpos
                ),
            )
            goal_node = goal_tree[nearest_goal_index]
            distance = self._tree_state_distance(start_node.qpos, goal_node.qpos)
            score = self._state_score(start_node.qpos, start_node.depth)
            if score > best_score:
                best_score = score
                self._best_score = score
                best_path = self._tree_path(start_tree, new_start_index)
            if distance > connect_limit:
                continue
            valid, _ = validator.segment_is_valid(start_node.qpos, goal_node.qpos)
            if not valid:
                continue

            start_path = self._tree_path(start_tree, new_start_index)
            connector = interpolate_pose_segment(
                start_node.qpos, goal_node.qpos,
                validator.sweep_translation_step,
                validator.sweep_rotation_step,
            )
            root_to_goal = self._tree_path(goal_tree, nearest_goal_index)
            goal_to_root = [q.copy() for q in reversed(root_to_goal)]
            full_path = [*start_path, *connector]
            if goal_to_root:
                offset = 1 if np.allclose(full_path[-1], goal_to_root[0], atol=1e-9) else 0
                full_path.extend(goal_to_root[offset:])
            valid_path, reason = validator.path_is_valid(
                full_path, self.world.min_removal_distance,
                allow_initial_release=True,
            )
            if valid_path and self.world.is_disassembled(full_path[-1]):
                self.solution_mode = "bidirectional_rrt"
                self._birrt_start_nodes = len(start_tree)
                self._birrt_goal_nodes = len(goal_tree)
                return self._finalize_search(
                    "Success", start_time, full_path, "ok"
                )
            self._record_failure(f"Bi-RRT connection rejected: {reason}")

        self._birrt_start_nodes = len(start_tree)
        self._birrt_goal_nodes = len(goal_tree)
        return self._finalize_search(
            "Failure", start_time, best_path,
            "Bi-RRT exhausted its node budget",
        )

    def plan(self, max_time: float, max_depth: int, verbose: bool = False):
        start_time = time.time()
        initial = self._initial_qpos.copy()
        direct_best_path = [initial.copy()]
        direct_reason = "direct release not attempted"

        if verbose:
            print(
                f"[Search] {self.move_id} translation order: "
                f"{list(self.translation_priority)}"
            )

        if self.enable_direct_release:
            direct_status, direct_best_path, direct_reason = (
                self._try_direct_kinematic_release(
                    initial=initial,
                    start_time=start_time,
                    max_time=max_time,
                    verbose=verbose,
                )
            )
            if direct_status == "Success":
                self.solution_mode = "direct_kinematic_release"
                self._best_score = self._state_score(direct_best_path[-1], 1)
                return self._finalize_search(
                    "Success", start_time, direct_best_path, "ok"
                )
            if direct_status == "Timeout":
                return self._finalize_search(
                    "Timeout", start_time, direct_best_path, direct_reason
                )
            if verbose:
                print(
                    f"[Direct release] {self.move_id} failed: {direct_reason}. "
                    f"Falling back to {self.search_strategy}."
                )
        elif verbose:
            print(
                f"[Direct release] Disabled for {self.move_id}; using "
                f"{self.search_strategy}."
            )

        if self.search_strategy == "bfs":
            self.solution_mode = "physics_bfs"
            return self._plan_fifo(
                initial, start_time, max_time, max_depth, verbose, direct_best_path
            )
        if self.search_strategy == "dfs":
            self.solution_mode = "physics_dfs"
            return self._plan_dfs(
                initial, start_time, max_time, max_depth, verbose, direct_best_path
            )
        if self.search_strategy in {"rrt", "rrt_star", "bk_rrt"}:
            self.solution_mode = {
                "rrt": "physics_rrt",
                "rrt_star": "physics_rrt_star",
                "bk_rrt": "behavioral_kinodynamic_rrt",
            }[self.search_strategy]
            return self._plan_rrt_family(
                initial, start_time, max_time, max_depth, verbose,
                direct_best_path, self.search_strategy,
            )
        if self.search_strategy == "bi_rrt":
            self.solution_mode = "bidirectional_rrt"
            return self._plan_bi_rrt(
                initial, start_time, max_time, max_depth, verbose, direct_best_path
            )

        self.solution_mode = "adaptive_contact_guided_best_first"
        result = self._plan_adaptive_best_first(
            initial,
            start_time,
            max_time,
            max_depth,
            verbose,
            direct_best_path,
        )
        if result[0] != "Success" and direct_reason not in {
            "direct release not attempted",
            "no straight release direction",
        }:
            self.last_failure_reason = (
                f"{self.last_failure_reason}; best direct probe: {direct_reason}"
            )
            self.search_diagnostics["failure_reason"] = self.last_failure_reason
        return result

    def close(self):
        if self.world is not None:
            self.world.close()
            self.world = None
        gc.collect()


# ============================================================================
# Progressive sequence planner
# ============================================================================


class GenesisProgressiveSequencePlanner:
    def __init__(
        self,
        assets: Dict[str, PartAsset],
        rotation: bool,
        show_viewer: bool,
        dt: float,
        substeps: int,
        force_mag: float,
        density: float,
        friction: float,
        sdf_cell_size: float,
        sdf_min_res: int,
        sdf_max_res: int,
        frame_skip: int,
        trans_dist_th: float,
        quat_dist_th: float,
        collision_face_num: int,
        collision_mode: str,
        watertighten: int,
        free_dof_order: str,
        max_action_chunks: int,
        candidate_order: str,
        enable_direct_release: bool,
        penetration_tolerance: float,
        sweep_translation_step: float,
        sweep_rotation_step: float,
        min_removal_distance: float,
        min_removal_distance_factor: float,
        unscrew_axial_force: float,
        unscrew_torque: float,
        unscrew_probe_chunks: int,
        unscrew_max_chunks: int,
        unscrew_stall_chunks: int,
        unscrew_lateral_factor: float,
        unscrew_force_growth: float,
        unscrew_max_force_scale: float,
        prior_unscrew_max_retries: int,
        max_progressive_depth: int,
        max_sequence_trials: int,
        auto_rotation_depth: int,
        search_strategy: str,
        adaptive_beam_width: int,
        adaptive_action_top_k: int,
        enable_diagonal_actions: bool,
        coupled_action_scale: float,
        diagnostic_max_frames_per_trial: int,
        rrt_max_nodes: int,
        rrt_goal_bias: float,
        rrt_neighbor_count: int,
        rrt_connect_distance_factor: float,
        rrt_goal_root_count: int,
        enable_blocker_guidance: bool,
        enable_structural_guidance: bool,
        approach_name: str,
    ):
        self.assets = assets
        self.part_ids = sorted(assets.keys())
        self.rotation = rotation
        self.show_viewer = show_viewer
        self.dt = dt
        self.substeps = substeps
        self.force_mag = force_mag
        self.density = density
        self.friction = friction
        self.sdf_cell_size = sdf_cell_size
        self.sdf_min_res = sdf_min_res
        self.sdf_max_res = sdf_max_res
        self.frame_skip = frame_skip
        self.trans_dist_th = trans_dist_th
        self.quat_dist_th = quat_dist_th
        self.collision_face_num = collision_face_num
        self.collision_mode = collision_mode
        self.watertighten = watertighten
        self.free_dof_order = free_dof_order
        self.max_action_chunks = max_action_chunks
        self.candidate_order = candidate_order
        self.enable_direct_release = bool(enable_direct_release)
        self.penetration_tolerance = penetration_tolerance
        self.sweep_translation_step = sweep_translation_step
        self.sweep_rotation_step = sweep_rotation_step
        self.min_removal_distance = min_removal_distance
        self.min_removal_distance_factor = min_removal_distance_factor
        self.unscrew_axial_force = unscrew_axial_force
        self.unscrew_torque = unscrew_torque
        self.unscrew_probe_chunks = unscrew_probe_chunks
        self.unscrew_max_chunks = unscrew_max_chunks
        self.unscrew_stall_chunks = unscrew_stall_chunks
        self.unscrew_lateral_factor = unscrew_lateral_factor
        self.unscrew_force_growth = unscrew_force_growth
        self.unscrew_max_force_scale = unscrew_max_force_scale
        self.prior_unscrew_max_retries = max(int(prior_unscrew_max_retries), 1)
        self.max_progressive_depth = max(int(max_progressive_depth), 1)
        self.max_sequence_trials = max(int(max_sequence_trials), 1)
        self.auto_rotation_depth = max(int(auto_rotation_depth), 1)
        self.search_strategy = str(search_strategy).strip().lower().replace("-", "_")
        self.adaptive_beam_width = max(int(adaptive_beam_width), 1)
        self.adaptive_action_top_k = max(int(adaptive_action_top_k), 6)
        self.enable_diagonal_actions = bool(enable_diagonal_actions)
        self.coupled_action_scale = max(float(coupled_action_scale), 0.0)
        self.diagnostic_max_frames_per_trial = max(
            int(diagnostic_max_frames_per_trial), 2
        )
        self.rrt_max_nodes = max(int(rrt_max_nodes), 2)
        self.rrt_goal_bias = float(np.clip(rrt_goal_bias, 0.0, 1.0))
        self.rrt_neighbor_count = max(int(rrt_neighbor_count), 1)
        self.rrt_connect_distance_factor = max(
            float(rrt_connect_distance_factor), 0.02
        )
        self.rrt_goal_root_count = max(int(rrt_goal_root_count), 1)
        self.enable_blocker_guidance = bool(enable_blocker_guidance)
        self.enable_structural_guidance = bool(enable_structural_guidance)
        self.approach_name = str(approach_name)

        # Translation-first action policy. Cardinal translations are always
        # attempted before any rotational action. Rotational exploration is
        # activated only after the current state has no valid translational
        # successor, which matches the requested human-like strategy.
        self.translation_priority_labels: Tuple[str, ...] = (
            "+z", "-z", "+x", "-x", "+y", "-y"
        )
        self.random_rotation_actions = 12
        self.random_rotation_scale = 0.35

        # Optional explicit structural groups. Lower groups are not searched
        # while an upper group remains unresolved. This prevents the lower
        # chassis from being removed before upper plates and inserts.
        self.strict_top_layer = False
        self.structural_priority_groups: List[List[str]] = []
        self.structural_group_by_part: Dict[str, int] = {}

        # Configured when plan_sequence receives the reviewed LLM/CAD policy.
        self.spatial_policy: Dict[str, object] = {}
        self.spatial_mode = "top_down_left_right"
        self.top_axis = 2
        self.left_axis = 0
        self.front_axis = 1
        self.top_descending = True
        self.left_ascending = True
        self.front_ascending = True
        self.layer_tolerance = 0.05
        self.defer_lower_blockers = False
        self.height_reference = "upper_surface"
        self._spatial_reference_top = 0.0

    @staticmethod
    def _axis_to_index(value, default: int) -> int:
        if isinstance(value, int) and value in (0, 1, 2):
            return value
        key = str(value).strip().lower()
        return {"x": 0, "y": 1, "z": 2}.get(key, default)

    def _configure_spatial_policy(
        self,
        policy: Optional[Dict[str, object]],
        candidate_ids: Sequence[str],
    ) -> None:
        policy = dict(policy or {})
        if not self.enable_structural_guidance:
            # Classical sampling baselines do not inherit PhyD2A's top-layer
            # ordering or strict structural gating.
            policy = {
                "structural_order": self.candidate_order,
                "strict_top_layer": False,
                "defer_lower_blockers": False,
                "translation_priority": ["+x", "-x", "+y", "-y", "+z", "-z"],
                "rotation_search_start_depth": 1,
                "random_rotation_actions": 12,
                "random_rotation_scale": 0.35,
            }
        self.spatial_policy = policy
        mode = str(
            policy.get(
                "structural_order",
                policy.get("candidate_order", self.candidate_order),
            )
        ).strip().lower().replace("-", "_")
        if mode in {"top_down_left_right", "topdown_leftright", "human_spatial"}:
            self.spatial_mode = "top_down_left_right"
        elif mode in {"random", "exterior"}:
            self.spatial_mode = mode
        else:
            self.spatial_mode = "top_down_left_right"

        self.top_axis = self._axis_to_index(policy.get("top_axis", "z"), 2)
        self.left_axis = self._axis_to_index(policy.get("left_axis", "x"), 0)
        self.front_axis = self._axis_to_index(policy.get("front_axis", "y"), 1)
        self.top_descending = str(policy.get("top_direction", "descending")).lower() != "ascending"
        self.left_ascending = str(policy.get("left_direction", "ascending")).lower() != "descending"
        self.front_ascending = str(policy.get("front_direction", "ascending")).lower() != "descending"
        self.defer_lower_blockers = bool(policy.get("defer_lower_blockers", False))
        self.strict_top_layer = bool(policy.get("strict_top_layer", False))

        labels = policy.get(
            "translation_priority",
            ["+z", "-z", "+x", "-x", "+y", "-y"],
        )
        if isinstance(labels, str):
            labels = [item.strip() for item in labels.split(",") if item.strip()]
        cleaned_labels: List[str] = []
        for label in list(labels or []):
            key = str(label).strip().lower().replace(" ", "")
            if key in {"+x", "-x", "+y", "-y", "+z", "-z"} and key not in cleaned_labels:
                cleaned_labels.append(key)
        for key in ("+z", "-z", "+x", "-x", "+y", "-y"):
            if key not in cleaned_labels:
                cleaned_labels.append(key)
        self.translation_priority_labels = tuple(cleaned_labels)
        self.random_rotation_actions = max(
            int(policy.get("random_rotation_actions", 12)), 0
        )
        self.random_rotation_scale = max(
            float(policy.get("random_rotation_scale", 0.35)), 0.0
        )
        self.auto_rotation_depth = max(
            int(policy.get("rotation_search_start_depth", self.auto_rotation_depth)), 1
        )

        # v12 deliberately ignores any hand-authored structural part groups.
        # Structural layers are inferred online from the geometry of the
        # currently remaining assembly, then validated by physical search.
        raw_groups = policy.get("structural_priority_groups", [])
        if raw_groups:
            print(
                "[Policy] Ignoring 'structural_priority_groups'. v12 infers "
                "structural layers automatically from current CAD geometry."
            )
        self.structural_priority_groups = []
        self.structural_group_by_part = {}

        self.height_reference = str(policy.get("height_reference", "upper_surface")).strip().lower()
        if self.height_reference not in {"center", "upper_surface", "lower_surface"}:
            self.height_reference = "center"

        ids = [pid for pid in candidate_ids if pid in self.assets]
        if ids:
            xyz_min, xyz_max, _, _ = compute_assembly_bounds(self.assets, ids)
            axis_extent = max(float(xyz_max[self.top_axis] - xyz_min[self.top_axis]), 1e-6)
        else:
            axis_extent = 1.0
        explicit_tol = policy.get("layer_tolerance")
        if explicit_tol is None:
            ratio = max(float(policy.get("layer_tolerance_ratio", 0.06)), 1e-4)
            self.layer_tolerance = max(axis_extent * ratio, 1e-5)
        else:
            self.layer_tolerance = max(float(explicit_tol), 1e-5)

        top_values = [self._part_top_value(pid) for pid in ids]
        if top_values:
            self._spatial_reference_top = (
                max(top_values) if self.top_descending else min(top_values)
            )
        else:
            self._spatial_reference_top = 0.0

        print("\n========== Structural Disassembly Policy ==========")
        print("Mode:", self.spatial_mode)
        print("Fixed base:", policy.get("base_part_id", "provided separately"))
        print(
            "Spatial order: top axis=",
            "xyz"[self.top_axis],
            "descending" if self.top_descending else "ascending",
            "; left axis=",
            "xyz"[self.left_axis],
            "ascending" if self.left_ascending else "descending",
        )
        print("Height reference:", self.height_reference)
        print("Layer tolerance:", self.layer_tolerance)
        print("Defer lower blockers:", self.defer_lower_blockers)
        print("Strict highest inferred-layer gating:", self.strict_top_layer)
        print("Translation-first order:", list(self.translation_priority_labels))
        print("Rotation search starts at depth:", self.auto_rotation_depth)
        print("Random rotational actions:", self.random_rotation_actions)
        print("Structural layers: inferred automatically from remaining CAD geometry")

    def _part_world_bounds(self, pid: str) -> Tuple[np.ndarray, np.ndarray]:
        asset = self.assets[pid]
        bounds = np.asarray(asset.local_visual_mesh.bounds, dtype=float)
        return bounds[0] + asset.center, bounds[1] + asset.center

    def _part_top_value(self, pid: str) -> float:
        asset = self.assets[pid]
        if self.height_reference == "center":
            return float(asset.center[self.top_axis])
        bmin, bmax = self._part_world_bounds(pid)
        if self.height_reference == "lower_surface":
            return float(bmin[self.top_axis])
        return float(bmax[self.top_axis])

    def _spatial_key(self, pid: str):
        asset = self.assets[pid]
        top_value = self._part_top_value(pid)
        if self.top_descending:
            layer_float = (self._spatial_reference_top - top_value) / self.layer_tolerance
        else:
            layer_float = (top_value - self._spatial_reference_top) / self.layer_tolerance
        layer = max(int(math.floor(layer_float + 1e-9)), 0)

        left = float(asset.center[self.left_axis])
        front = float(asset.center[self.front_axis])
        if not self.left_ascending:
            left = -left
        if not self.front_ascending:
            front = -front
        # Primary: upper layers first. Secondary: left-to-right. Tertiary:
        # front-to-back and stable part id for deterministic planning.
        return (layer, left, front, pid)

    def _structural_level(self, pid: str) -> int:
        """Return an automatically inferred top-down geometric layer."""
        return int(self._spatial_key(pid)[0])

    def _candidate_order_key(self, pid: str):
        return (self._structural_level(pid), *self._spatial_key(pid))

    def _restrict_to_highest_unresolved_level(
        self,
        active_queue: Deque[Tuple[str, int]],
        inactive_queue: Deque[Tuple[str, int]],
        remaining: Sequence[str],
        base_part_id: Optional[str],
    ) -> Tuple[Deque[Tuple[str, int]], Deque[Tuple[str, int]], Optional[int], bool]:
        """Activate only the highest unresolved inferred layer.

        ``exhausted`` becomes true when parts remain in the highest group but
        every queued search state for that group has already been consumed.
        Descending to lower groups in that situation would violate the requested
        top-down physical rule, so the sequence planner stops explicitly.
        """
        candidates = [
            pid for pid in remaining
            if pid != base_part_id and pid in self.assets
        ]
        if not candidates:
            return active_queue, inactive_queue, None, False

        level = min(self._structural_level(pid) for pid in candidates)
        level_ids = {pid for pid in candidates if self._structural_level(pid) == level}

        active_level = deque((pid, d) for pid, d in active_queue if pid in level_ids)
        active_lower = deque((pid, d) for pid, d in active_queue if pid not in level_ids)
        inactive_level = deque((pid, d) for pid, d in inactive_queue if pid in level_ids)
        inactive_lower = deque((pid, d) for pid, d in inactive_queue if pid not in level_ids)

        lower_waiting = self._sort_queue_spatial(
            deque(list(active_lower) + list(inactive_lower))
        )
        if active_level:
            return (
                self._sort_queue_spatial(active_level),
                self._sort_queue_spatial(deque(list(inactive_level) + list(lower_waiting))),
                level,
                False,
            )
        if inactive_level:
            min_top_depth = min(int(depth) for _, depth in inactive_level)
            min_lower_depth = (
                min(int(depth) for _, depth in lower_waiting)
                if lower_waiting
                else math.inf
            )
            # Do not exhaust progressively deeper searches in the highest layer
            # before every lower candidate has received its cheaper trial. The
            # geometric layer is a priority, not proof of a hard precedence.
            if min_lower_depth < min_top_depth:
                waiting = self._sort_queue_spatial(
                    deque(list(inactive_level) + list(lower_waiting))
                )
                return deque(), waiting, level, True
            return (
                self._sort_queue_spatial(inactive_level),
                lower_waiting,
                level,
                False,
            )
        return deque(), lower_waiting, level, True

    def _sort_queue_spatial(
        self,
        queue: Deque[Tuple[str, int]],
    ) -> Deque[Tuple[str, int]]:
        # Deduplicate while preserving the greatest progressive depth already
        # reached for each candidate.
        depth_by_pid: Dict[str, int] = {}
        for pid, depth in queue:
            depth_by_pid[pid] = max(depth_by_pid.get(pid, 0), int(depth))
        items = [(pid, depth) for pid, depth in depth_by_pid.items()]
        if self.spatial_mode == "top_down_left_right":
            # Progressive depth is the primary key. Otherwise a difficult upper
            # part can consume depth 1..N before an easy lower part receives even
            # one trial. Spatial order remains the tie-breaker within each depth.
            items.sort(key=lambda item: (item[1], self._candidate_order_key(item[0])))
        return deque(items)

    def _print_spatial_order(self, ids: Sequence[str]) -> None:
        if self.spatial_mode != "top_down_left_right":
            return
        print("[Sequence] Fixed-base structural order (top-to-bottom, left-to-right):")
        for index, pid in enumerate(sorted(ids, key=self._candidate_order_key), start=1):
            key = self._spatial_key(pid)
            inferred_level = self._structural_level(pid)
            center = self.assets[pid].center
            print(
                f"  {index:02d}. {pid}: inferred_level={inferred_level}, layer={key[0]}, "
                f"height={center[self.top_axis]:.6g}, "
                f"left_coord={center[self.left_axis]:.6g}"
            )

    def _ordered_candidates(self, ids: Sequence[str], rng) -> List[str]:
        ids = list(ids)
        mode = self.spatial_mode
        if mode == "random":
            rng.shuffle(ids)
            return ids
        if mode == "top_down_left_right":
            return sorted(ids, key=self._candidate_order_key)

        _, _, center, extent = compute_assembly_bounds(self.assets, ids)

        def score(pid: str):
            asset = self.assets[pid]
            radial = float(np.linalg.norm(asset.center - center) / max(extent, 1e-9))
            size = float(np.linalg.norm(asset.local_visual_mesh.extents) / max(extent, 1e-9))
            return radial - 0.15 * size

        return sorted(ids, key=score, reverse=True)


    def _build_queue_for_current_state(
        self,
        remaining: Sequence[str],
        base_part_id: Optional[str],
        attempted_depth: Dict[Tuple[Tuple[str, ...], str], int],
        rng,
    ) -> Deque[Tuple[str, int]]:
        """Build the next informative candidate queue for the exact state.

        Search depth is state-dependent.  Once any part is removed, every
        previously unresolved part becomes a new planning problem and must be
        reinserted at depth 1.  In the same state, only the next untried depth is
        queued.  This fixes the old behaviour where a part exhausted once was
        silently lost forever even after another part had been removed.
        """
        signature = tuple(sorted(remaining))
        candidates = [
            pid
            for pid in remaining
            if pid != base_part_id and pid in self.assets
        ]
        ordered = self._ordered_candidates(candidates, rng)
        items: List[Tuple[str, int]] = []
        for pid in ordered:
            previous = int(attempted_depth.get((signature, pid), 0))
            next_depth = previous + 1
            if next_depth <= self.max_progressive_depth:
                items.append((pid, max(next_depth, 1)))
        return self._sort_queue_spatial(deque(items))

    @staticmethod
    def _remove_ids_from_queue(queue: Deque[Tuple[str, int]], ids: set[str]):
        return deque((pid, depth) for pid, depth in queue if pid not in ids)

    def _promote_blockers(
        self,
        blockers: Sequence[str],
        move_id: str,
        remaining: Sequence[str],
        active_queue: Deque[Tuple[str, int]],
        inactive_queue: Deque[Tuple[str, int]],
        base_part_id: Optional[str],
        minimum_depth: int = 1,
    ) -> Tuple[Deque[Tuple[str, int]], Deque[Tuple[str, int]]]:
        """Move blockers forward without resetting their progressive depth.

        The previous implementation always inserted a promoted blocker at
        depth 1. Mutual blocker pairs therefore ping-ponged forever and never
        reached the deeper multi-action searches used by the original
        progressive queue. This version preserves the largest queued depth.
        """
        valid = []
        move_layer = self._structural_level(move_id) if move_id in self.assets else 0
        lower_physical_blockers = []
        for pid in blockers:
            if pid == move_id or pid == base_part_id or pid not in remaining:
                continue
            if (
                pid in self.assets
                and self._structural_level(pid) > move_layer
            ):
                lower_physical_blockers.append(pid)
            if pid not in valid:
                valid.append(pid)
        if lower_physical_blockers:
            print(
                f"[Sequence] Physical blockers override the geometric layer "
                f"heuristic for {move_id}: {sorted(lower_physical_blockers)}"
            )
        if not valid:
            return active_queue, inactive_queue

        depth_by_pid: Dict[str, int] = {}
        for pid, depth in list(active_queue) + list(inactive_queue):
            depth_by_pid[pid] = max(depth_by_pid.get(pid, 0), int(depth))

        blocked_set = set(valid)
        active_queue = self._remove_ids_from_queue(active_queue, blocked_set)
        inactive_queue = self._remove_ids_from_queue(inactive_queue, blocked_set)
        promoted = []
        for pid in reversed(valid):
            depth = max(depth_by_pid.get(pid, 1), int(minimum_depth), 1)
            depth = min(depth, self.max_progressive_depth)
            active_queue.appendleft((pid, depth))
            promoted.append((pid, depth))
        promoted.reverse()
        print(
            f"[Sequence] Promoting blockers ahead of {move_id} "
            f"without depth reset: {promoted}"
        )
        return active_queue, inactive_queue

    def plan_sequence(
        self,
        seq_max_time: float,
        path_max_time: float,
        seed: int,
        save_dir: Optional[str],
        llm_prior: LLMPriorPlan,
        base_part_id: Optional[str],
        strict_prior: bool,
        prior_initial_depth: int,
        parking_enabled: bool,
        parking_distance_scale: float,
        parking_spacing_scale: float,
        parking_frames: int,
        verbose: bool = True,
    ):
        rng = np.random.default_rng(seed)
        remaining = list(self.part_ids)
        if base_part_id is not None and base_part_id not in remaining:
            raise ValueError(f"Base part {base_part_id!r} is not present in loaded assets.")
        if base_part_id is None:
            raise ValueError(
                "v12 requires an explicit fixed base part. Set base_part_id in the "
                "LLM prior JSON or pass --base-part-id."
            )

        structural_candidates = [pid for pid in remaining if pid != base_part_id]
        self._configure_spatial_policy(llm_prior.planning_policy, structural_candidates)
        print(f"[Base] {base_part_id} is fixed and permanently excluded from removal candidates.")
        self._print_spatial_order(structural_candidates)

        prior_order = [
            pid for pid in llm_prior.priority_order
            if pid in remaining and pid != base_part_id
        ]
        prior_operations = llm_prior.operations

        initial_candidates = [pid for pid in remaining if pid != base_part_id]
        prior_phase = bool(strict_prior and prior_order)
        prior_retry_counts: Dict[str, int] = {pid: 0 for pid in prior_order}
        if prior_phase:
            # LLM-confirmed fasteners are a mandatory first phase.  The
            # adaptive unscrew planner resolves force/torque direction; general
            # sequence search begins only after every prior fastener is removed.
            active_queue = deque((pid, prior_initial_depth) for pid in prior_order)
        else:
            ordered = self._ordered_candidates(initial_candidates, rng)
            prior_rank = {pid: i for i, pid in enumerate(prior_order)}
            ordered.sort(key=lambda pid: (0, prior_rank[pid]) if pid in prior_rank else (1, 0))
            active_queue = deque((pid, 1) for pid in ordered)
        inactive_queue: Deque[Tuple[str, int]] = deque()

        sequence: List[str] = []
        paths: Dict[str, List[np.ndarray]] = {}
        operations: Dict[str, str] = {}
        trial_log = []
        diagnostic_events = []
        # Attempts are keyed by the exact remaining-part set. Repeating the
        # same part at the same or shallower depth cannot add information.
        attempted_depth: Dict[Tuple[Tuple[str, ...], str], int] = {}
        blocker_graph: Dict[str, set[str]] = {}
        t0 = time.monotonic()
        last_progress_time = t0
        last_progress_part: Optional[str] = None
        longest_no_progress_interval = 0.0
        total_path_time = 0.0
        trials = 0
        final_status = "Failure"
        # The inferred top layer is tried first, but it is a geometric
        # heuristic rather than a proven precedence constraint.  If that layer
        # is exhausted in the current state, the gate is relaxed so that a
        # lower removable part may change the state and unblock it.
        layer_gate_relaxed_for_state = False

        if save_dir:
            os.makedirs(save_dir, exist_ok=True)

        while True:
            removable_remaining = [pid for pid in remaining if pid != base_part_id]
            # Physics-based disassembly requires N-1 motions. The final part is
            # the stationary base/reference part and remains visible in replay.
            if not removable_remaining or (base_part_id is None and len(remaining) <= 1):
                final_status = "Success"
                break
            now = time.monotonic()
            no_progress_elapsed = now - last_progress_time
            longest_no_progress_interval = max(
                longest_no_progress_interval, no_progress_elapsed
            )
            if seq_max_time > 0.0 and no_progress_elapsed >= seq_max_time:
                final_status = "Timeout"
                print(
                    f"[Sequence] No part has been removed for "
                    f"{no_progress_elapsed:.3f}s (limit={seq_max_time:.3f}s). "
                    "Stopping this stalled remaining-part state."
                )
                break
            if trials >= self.max_sequence_trials:
                final_status = "TrialLimit"
                print(
                    f"[Sequence] Reached --max-sequence-trials="
                    f"{self.max_sequence_trials}."
                )
                break

            if (
                not prior_phase
                and self.strict_top_layer
                and not layer_gate_relaxed_for_state
            ):
                (
                    active_queue,
                    inactive_queue,
                    current_level,
                    level_exhausted,
                ) = self._restrict_to_highest_unresolved_level(
                    active_queue,
                    inactive_queue,
                    remaining,
                    base_part_id,
                )
                if level_exhausted:
                    unresolved_level = [
                        pid for pid in remaining
                        if pid != base_part_id
                        and self._structural_level(pid) == current_level
                    ]
                    print(
                        "[Sequence] Highest inferred layer "
                        f"{current_level} exhausted in the current state: "
                        f"{unresolved_level}. Because inferred height is only a "
                        "priority heuristic, the hard gate is now relaxed. "
                        "Lower candidates will be tested; after any successful "
                        "removal, all unresolved parts are reinserted and the "
                        "top-layer preference is restored."
                    )
                    layer_gate_relaxed_for_state = True
                    active_queue = self._build_queue_for_current_state(
                        remaining=remaining,
                        base_part_id=base_part_id,
                        attempted_depth=attempted_depth,
                        rng=rng,
                    )
                    inactive_queue = deque()
                    if not active_queue:
                        final_status = "Failure"
                        break
                    continue

            if not active_queue:
                if not inactive_queue:
                    final_status = "Failure"
                    break
                active_queue, inactive_queue = self._sort_queue_spatial(inactive_queue), deque()

            if not prior_phase:
                active_queue = self._sort_queue_spatial(active_queue)
            move_id, max_depth = active_queue.popleft()
            if move_id not in remaining or move_id == base_part_id:
                continue
            max_depth = int(max_depth)
            if max_depth > self.max_progressive_depth:
                print(
                    f"[Sequence] Skipping {move_id}: requested depth {max_depth} "
                    f"exceeds max {self.max_progressive_depth}."
                )
                continue

            state_signature = tuple(sorted(remaining))
            attempt_key = (state_signature, move_id)
            previous_depth = attempted_depth.get(attempt_key, 0)
            if max_depth <= previous_depth:
                max_depth = previous_depth + 1
            if max_depth > self.max_progressive_depth:
                continue
            attempted_depth[attempt_key] = max_depth

            # Cap each path attempt by the remaining no-progress budget. This
            # prevents a final 300-second attempt from starting at 3599 seconds
            # and making a nominal 3600-second timeout finish near 3900 seconds.
            if seq_max_time > 0.0:
                no_progress_remaining = max(
                    seq_max_time - (time.monotonic() - last_progress_time), 0.0
                )
                trial_time_budget = min(float(path_max_time), no_progress_remaining)
            else:
                trial_time_budget = float(path_max_time)
            if trial_time_budget <= 1e-3:
                final_status = "Timeout"
                print(
                    "[Sequence] No-progress budget was exhausted before the next "
                    f"trial for {move_id}."
                )
                break
            if trial_time_budget + 1e-9 < float(path_max_time):
                print(
                    f"[Sequence] Capping {move_id} path budget to "
                    f"{trial_time_budget:.3f}s to respect the no-progress timeout."
                )

            still_ids = [pid for pid in remaining if pid != move_id]
            trials += 1
            print(
                f"\n# Trial {trials} | Move={move_id} | depth={max_depth} | "
                f"remaining={len(remaining)}"
            )

            operation = "physics_bfs"
            status = "Failure"
            path_time = 0.0
            path = None
            trial_blockers: set[str] = set()
            failure_reason = ""
            planner_diagnostics: Dict[str, object] = {}

            # ------------------------------------------------------------
            # Stage 1: LLM-confirmed fasteners use a deterministic virtual
            # screw manifold with dense exact-mesh validation. Genesis
            # force-driven rollout is retained only as a secondary fallback.
            # A failed screw primitive is never replaced by random translation.
            # ------------------------------------------------------------
            prior_operation = prior_operations.get(move_id)
            if prior_operation is not None and prior_operation.operation == "unscrew":
                operation = "llm_force_driven_virtual_screw"
                unscrew_planner = None
                try:
                    unscrew_planner = SemanticScrewPlanner(
                        assets=self.assets,
                        move_id=move_id,
                        still_ids=still_ids,
                        operation=prior_operation,
                        show_viewer=self.show_viewer,
                        dt=self.dt,
                        substeps=self.substeps,
                        force_mag=self.force_mag,
                        unscrew_axial_force=self.unscrew_axial_force,
                        unscrew_torque=self.unscrew_torque,
                        unscrew_probe_chunks=self.unscrew_probe_chunks,
                        unscrew_max_chunks=self.unscrew_max_chunks,
                        unscrew_stall_chunks=self.unscrew_stall_chunks,
                        unscrew_lateral_factor=self.unscrew_lateral_factor,
                        unscrew_force_growth=self.unscrew_force_growth,
                        unscrew_max_force_scale=self.unscrew_max_force_scale,
                        density=self.density,
                        friction=self.friction,
                        sdf_cell_size=self.sdf_cell_size,
                        sdf_min_res=self.sdf_min_res,
                        sdf_max_res=self.sdf_max_res,
                        frame_skip=self.frame_skip,
                        collision_face_num=self.collision_face_num,
                        collision_mode=self.collision_mode,
                        watertighten=self.watertighten,
                        free_dof_order=self.free_dof_order,
                        penetration_tolerance=self.penetration_tolerance,
                        sweep_translation_step=self.sweep_translation_step,
                        sweep_rotation_step=self.sweep_rotation_step,
                        min_removal_distance=self.min_removal_distance,
                        min_removal_distance_factor=self.min_removal_distance_factor,
                    )
                    status, path_time, path = unscrew_planner.plan(trial_time_budget)
                    trial_blockers.update(unscrew_planner.blockers)
                    failure_reason = unscrew_planner.last_failure_reason
                    planner_diagnostics = dict(unscrew_planner.search_diagnostics)
                    if status == "Success":
                        strategy = str(planner_diagnostics.get("strategy", ""))
                        if strategy.startswith("validated_kinematic_virtual_screw"):
                            operation = "llm_validated_virtual_screw"
                        elif strategy == "force_driven_virtual_screw_fallback":
                            operation = "llm_force_driven_virtual_screw_fallback"
                finally:
                    if unscrew_planner is not None:
                        unscrew_planner.close()
                        del unscrew_planner
                        gc.collect()

                if status == "Success" and parking_enabled and path:
                    path = append_parking_motion(
                        path,
                        self.assets,
                        step_index=len(sequence),
                        total_steps=len(self.part_ids) - 1,
                        distance_scale=parking_distance_scale,
                        spacing_scale=parking_spacing_scale,
                        parking_frames=parking_frames,
                    )
                elif status != "Success":
                    print(
                        f"[LLM Prior] Semantic screw operation failed for {move_id} "
                        "after deterministic exact validation and force-driven fallback. "
                        "It will not be replaced by a random translational BFS."
                    )

            # ------------------------------------------------------------
            # Stage 2: ordinary Genesis physics search. This is used for all
            # non-prior parts and as a correctness fallback for failed priors.
            # ------------------------------------------------------------
            if status != "Success" and not (
                prior_operation is not None and prior_operation.operation == "unscrew"
            ):
                planner = None
                planner_solution_mode = "physics_bfs"
                try:
                    use_rotation = self.rotation or max_depth >= self.auto_rotation_depth
                    if use_rotation and not self.rotation:
                        print(
                            f"[Search] Escalating {move_id} to 6-DoF search at "
                            f"depth={max_depth}."
                        )
                    planner = GenesisBFSPathPlanner(
                        assets=self.assets,
                        move_id=move_id,
                        still_ids=still_ids,
                        rotation=use_rotation,
                        show_viewer=self.show_viewer,
                        dt=self.dt,
                        substeps=self.substeps,
                        force_mag=self.force_mag,
                        density=self.density,
                        friction=self.friction,
                        sdf_cell_size=self.sdf_cell_size,
                        sdf_min_res=self.sdf_min_res,
                        sdf_max_res=self.sdf_max_res,
                        frame_skip=self.frame_skip,
                        trans_dist_th=self.trans_dist_th,
                        quat_dist_th=self.quat_dist_th,
                        collision_face_num=self.collision_face_num,
                        collision_mode=self.collision_mode,
                        watertighten=self.watertighten,
                        free_dof_order=self.free_dof_order,
                        max_action_chunks=self.max_action_chunks,
                        penetration_tolerance=self.penetration_tolerance,
                        sweep_translation_step=self.sweep_translation_step,
                        sweep_rotation_step=self.sweep_rotation_step,
                        min_removal_distance=self.min_removal_distance,
                        min_removal_distance_factor=self.min_removal_distance_factor,
                        enable_direct_release=self.enable_direct_release,
                        translation_priority=self.translation_priority_labels,
                        random_rotation_actions=self.random_rotation_actions,
                        random_rotation_scale=self.random_rotation_scale,
                        search_strategy=self.search_strategy,
                        beam_width=self.adaptive_beam_width,
                        action_top_k=self.adaptive_action_top_k,
                        enable_diagonal_actions=self.enable_diagonal_actions,
                        coupled_action_scale=self.coupled_action_scale,
                        rrt_max_nodes=self.rrt_max_nodes,
                        rrt_goal_bias=self.rrt_goal_bias,
                        rrt_neighbor_count=self.rrt_neighbor_count,
                        rrt_connect_distance_factor=self.rrt_connect_distance_factor,
                        rrt_goal_root_count=self.rrt_goal_root_count,
                    )
                    status, path_time, path = planner.plan(
                        max_time=trial_time_budget,
                        max_depth=max_depth,
                        verbose=False,
                    )
                    planner_solution_mode = planner.solution_mode
                    trial_blockers.update(planner.blockers)
                    failure_reason = planner.last_failure_reason
                    planner_diagnostics = dict(planner.search_diagnostics)
                finally:
                    if planner is not None:
                        planner.close()
                        del planner
                        gc.collect()

                if operation == "physics_bfs":
                    operation = (
                        f"llm_prior_{planner_solution_mode}"
                        if prior_operation is not None
                        else planner_solution_mode
                    )

                if parking_enabled and status == "Success" and path:
                    path = append_parking_motion(
                        path,
                        self.assets,
                        step_index=len(sequence),
                        total_steps=len(self.part_ids) - 1,
                        distance_scale=parking_distance_scale,
                        spacing_scale=parking_spacing_scale,
                        parking_frames=parking_frames,
                    )

            total_path_time += path_time
            original_path_frames = int(len(path)) if path is not None else 0
            diagnostic_path = downsample_pose_path(
                path, self.diagnostic_max_frames_per_trial
            )
            trial_entry = {
                "trial": trials,
                "move_id": move_id,
                "max_depth": max_depth,
                "status": status,
                "path_time": float(path_time),
                "operation": operation,
                "blockers": sorted(trial_blockers),
                "failure_reason": failure_reason if status != "Success" else "",
                "path_frames": original_path_frames,
                "diagnostic_path_frames": int(len(diagnostic_path)),
                "remaining_count_before": int(len(remaining)),
                "successful_step": int(len(sequence) + 1) if status == "Success" else None,
                "planner_diagnostics": planner_diagnostics,
            }
            trial_log.append(trial_entry)
            diagnostic_events.append(
                {
                    **trial_entry,
                    "removed_before": list(sequence),
                    "path": diagnostic_path,
                }
            )
            print(
                f"Status={status} | path_time={path_time:.3f}s | "
                f"total_path_time={total_path_time:.3f}s"
            )
            if status != "Success":
                print(
                    f"[Failure diagnostic] part={move_id} | reason={failure_reason or status} "
                    f"| blockers={sorted(trial_blockers)} | "
                    f"saved_frames={len(diagnostic_path)}/{original_path_frames}"
                )

            if status == "Success":
                sequence.append(move_id)
                paths[move_id] = path
                operations[move_id] = operation
                remaining.remove(move_id)
                progress_now = time.monotonic()
                interval = progress_now - last_progress_time
                longest_no_progress_interval = max(
                    longest_no_progress_interval, interval
                )
                last_progress_time = progress_now
                last_progress_part = move_id
                print(
                    f"[Progress] Removed {move_id}; reset the no-progress timer "
                    f"after {interval:.3f}s."
                )
                if not prior_phase:
                    # The remaining-part set has changed.  Rebuild the queue
                    # from scratch so every unresolved part is reconsidered in
                    # the new geometry, beginning at depth 1 for this new state.
                    layer_gate_relaxed_for_state = False
                    active_queue = self._build_queue_for_current_state(
                        remaining=remaining,
                        base_part_id=base_part_id,
                        attempted_depth=attempted_depth,
                        rng=rng,
                    )
                    inactive_queue = deque()
                    print(
                        f"[Sequence] State changed after removing {move_id}; "
                        "reinserted all unresolved candidates for the new "
                        "remaining-part set."
                    )
            elif prior_phase and move_id in prior_order:
                prior_retry_counts[move_id] = prior_retry_counts.get(move_id, 0) + 1
                if prior_retry_counts[move_id] < self.prior_unscrew_max_retries:
                    active_queue.append((move_id, prior_initial_depth))
                    print(
                        f"[Prior] Re-queueing mandatory fastener {move_id}: "
                        f"retry {prior_retry_counts[move_id] + 1}/"
                        f"{self.prior_unscrew_max_retries}."
                    )
                else:
                    print(
                        f"[Prior] Mandatory unscrew failed for {move_id} after "
                        f"{self.prior_unscrew_max_retries} attempts. The semantic "
                        "unscrew primitive remains mandatory for this fastener, "
                        "but the hard prior phase is relaxed so collision-derived "
                        "blockers can be tested first."
                    )
                    # An LLM order is a priority, not proof of physical
                    # feasibility.  Humans likewise remove an actually blocking
                    # cover/retainer before retrying the screw.  Keep the failed
                    # fastener tagged as ``unscrew`` (never replace it by random
                    # translational BFS), but continue the physics-guided
                    # sequence instead of terminating the whole assembly.
                    prior_phase = False
                    layer_gate_relaxed_for_state = True
                    active_queue = self._build_queue_for_current_state(
                        remaining=remaining,
                        base_part_id=base_part_id,
                        attempted_depth=attempted_depth,
                        rng=rng,
                    )
                    inactive_queue = deque()
            else:
                next_depth = max_depth + 1
                if next_depth <= self.max_progressive_depth:
                    inactive_queue.append((move_id, next_depth))
                else:
                    print(
                        f"[Sequence] {move_id} exhausted progressive depth "
                        f"{self.max_progressive_depth}."
                    )

            if status != "Success" and seq_max_time > 0.0:
                stalled_for = time.monotonic() - last_progress_time
                longest_no_progress_interval = max(
                    longest_no_progress_interval, stalled_for
                )
                if stalled_for >= seq_max_time:
                    final_status = "Timeout"
                    print(
                        f"[Sequence] Trial ended without progress and the remaining "
                        f"state has now stalled for {stalled_for:.3f}s "
                        f"(limit={seq_max_time:.3f}s)."
                    )
                    break

            if prior_phase and not any(pid in remaining for pid in prior_order):
                prior_phase = False
                layer_gate_relaxed_for_state = False
                active_queue.clear()
                inactive_queue.clear()
                structural_remaining = [
                    pid for pid in remaining if pid != base_part_id
                ]
                # Recompute layer tolerance and the top reference after the
                # fasteners have been removed. No structural part IDs or layer
                # membership are supplied by JSON.
                self._configure_spatial_policy(
                    llm_prior.planning_policy, structural_remaining
                )
                self._print_spatial_order(structural_remaining)
                ordered = self._ordered_candidates(
                    structural_remaining, rng
                )
                active_queue.extend((pid, 1) for pid in ordered)
                active_queue = self._sort_queue_spatial(active_queue)
                print(
                    "[Prior] All LLM-confirmed fasteners were physically unscrewed; "
                    "releasing the remaining parts to fixed-base, top-down, "
                    "left-to-right progressive collision search."
                )
                print("[Sequence] Structural candidate order:", [pid for pid, _ in active_queue])

            if (
                status != "Success"
                and trial_blockers
                and not prior_phase
                and self.enable_blocker_guidance
            ):
                blocker_graph.setdefault(move_id, set()).update(trial_blockers)
                # Mutual blocker pairs are cycles, not an ordering relation.
                # Immediate promotion would only ping-pong the same two parts.
                promotable = []
                cyclic = []
                for pid in trial_blockers:
                    if move_id in blocker_graph.get(pid, set()):
                        cyclic.append(pid)
                    else:
                        promotable.append(pid)
                if cyclic:
                    print(
                        f"[Sequence] Mutual blocker cycle detected for {move_id}: "
                        f"{sorted(cyclic)}. Keeping progressive depth growth "
                        "instead of immediate ping-pong promotion."
                    )
                ranked_blockers = sorted(
                    promotable,
                    key=lambda pid: (
                        0 if pid in remaining else 1,
                        -float(np.linalg.norm(self.assets[pid].center)) if pid in self.assets else 0.0,
                    ),
                )
                active_queue, inactive_queue = self._promote_blockers(
                    blockers=ranked_blockers,
                    move_id=move_id,
                    remaining=remaining,
                    active_queue=active_queue,
                    inactive_queue=inactive_queue,
                    base_part_id=base_part_id,
                    minimum_depth=max_depth,
                )
                if ranked_blockers:
                    # A collision-derived precedence relation is stronger than
                    # a height heuristic. Keep promoted blockers at the front
                    # and relax the inferred-layer gate for this exact state.
                    layer_gate_relaxed_for_state = True
                    inactive_queue = self._sort_queue_spatial(inactive_queue)
                    print(
                        f"[Sequence] Relaxing geometric layer gating so physical "
                        f"blockers of {move_id} are tested first: "
                        f"{ranked_blockers}"
                    )
                elif self.spatial_mode == "top_down_left_right":
                    active_queue = self._sort_queue_spatial(active_queue)
                    inactive_queue = self._sort_queue_spatial(inactive_queue)

            if verbose:
                print("Sequence:", sequence)
                print("Active queue:", list(active_queue))
                print("Inactive queue:", list(inactive_queue))

        planning_wall_time = time.monotonic() - t0
        planning_complete = final_status == "Success"

        # A complete order exists only after N-1 valid motions have been found.
        # On timeout/failure, appending unresolved parts is not a solution; it
        # merely hides the fact that their relative order is unknown.
        resolved_base_part_id = base_part_id
        if planning_complete and resolved_base_part_id is None and len(remaining) == 1:
            resolved_base_part_id = remaining[0]

        if planning_complete:
            residual_parts = list(remaining)
            full_disassembly_sequence = list(sequence) + [
                pid for pid in residual_parts if pid not in sequence
            ]
            full_assembly_sequence = list(reversed(full_disassembly_sequence))
        else:
            full_disassembly_sequence = list(sequence)
            full_assembly_sequence = []

        print("\n========== Final Planning Result ==========")
        print(f"Status: {final_status}")
        movable_total = max(len(self.part_ids) - 1, 0)
        success_rate_percent = (
            100.0 * len(sequence) / movable_total if movable_total > 0 else 100.0
        )
        print(f"Disassembly motions: {len(sequence)}/{movable_total}")
        print(
            f"Disassembly success rate: {len(sequence)}/{movable_total} "
            f"({success_rate_percent:.2f}%)"
        )
        print(f"Trials: {trials}")
        print(f"Total path planning time: {total_path_time:.3f}s")
        print(f"Planning wall time: {planning_wall_time:.3f}s")
        print(f"No-progress timeout: {seq_max_time:.3f}s")
        print(
            f"Longest interval without a successful removal: "
            f"{longest_no_progress_interval:.3f}s"
        )
        print("Base/reference part:", resolved_base_part_id)
        if planning_complete:
            print("Disassembly motion sequence:", sequence)
            print("Full disassembly order (base last):", full_disassembly_sequence)
            print("Full assembly sequence (base first):", full_assembly_sequence)
        else:
            print("Partial validated disassembly sequence:", sequence)
            print("Unresolved parts (order unknown):", list(remaining))
            print(
                "[Result] No complete assembly/disassembly order is emitted "
                "because planning did not succeed."
            )

        result_payload = {
            "approach": self.approach_name,
            "status": final_status,
            "complete_sequence_available": planning_complete,
            # `sequence` is retained as the motion sequence for backward
            # compatibility with path saving and replay.
            "sequence": sequence,
            "motion_sequence": sequence,
            "disassembly_sequence": full_disassembly_sequence,
            "assembly_sequence": full_assembly_sequence,
            "base_part_id": resolved_base_part_id,
            "paths": paths,
            "operations": operations,
            "remaining": remaining,
            "trials": trials,
            "total_path_time": total_path_time,
            "planning_wall_time": planning_wall_time,
            "no_progress_timeout_s": float(seq_max_time),
            "longest_no_progress_interval_s": float(longest_no_progress_interval),
            "last_progress_part": last_progress_part,
            "simulation_dt": float(self.dt),
            "trial_log": trial_log,
            "diagnostic_events": diagnostic_events,
            "part_count_total": int(len(self.part_ids)),
            "part_count_movable": int(max(len(self.part_ids) - 1, 0)),
            "part_count_removed": int(len(sequence)),
            "part_count_unresolved": int(len([pid for pid in remaining if pid != resolved_base_part_id])),
            "blocker_graph": {
                pid: sorted(values) for pid, values in blocker_graph.items()
            },
            "llm_prior": {
                "enabled": llm_prior.enabled,
                "source": llm_prior.source,
                "priority_order": list(prior_order),
                "base_part_id": resolved_base_part_id,
                "planning_policy": dict(llm_prior.planning_policy),
                "operations": {
                    pid: {
                        "operation": op.operation,
                        "axis": None if op.axis is None else op.axis.tolist(),
                        "turns": op.turns,
                        "distance": op.distance,
                        "frames": op.frames,
                        "handedness": op.handedness,
                        "axis_mode": op.axis_mode,
                    }
                    for pid, op in prior_operations.items()
                },
            },
        }
        movable_count = int(result_payload["part_count_movable"])
        removed_count = int(result_payload["part_count_removed"])
        result_payload["success_fraction"] = f"{removed_count}/{movable_count}"
        result_payload["success_rate"] = (
            float(removed_count) / float(movable_count) if movable_count > 0 else 1.0
        )
        result_payload["part_statistics"] = build_part_statistics(
            self.part_ids,
            result_payload,
        )
        print_part_statistics(result_payload["part_statistics"])
        return result_payload


# ============================================================================
# Statistics and result persistence
# ============================================================================


def utc_now_iso() -> str:
    """Return an ISO-8601 UTC timestamp suitable for experiment logs."""
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def build_part_statistics(
    part_ids: Sequence[str],
    result: dict,
) -> Dict[str, dict]:
    """Aggregate per-part search statistics from the complete trial log.

    ``search_attempts`` is the number of top-level sequence-planning trials for
    the part.  One such attempt may internally contain many Genesis simulation
    steps or BFS nodes; ``cumulative_search_time_s`` records the corresponding
    wall-clock path-planning time returned by that attempt.
    """
    sequence = list(result.get("motion_sequence", result.get("sequence", [])))
    sequence_index = {pid: i + 1 for i, pid in enumerate(sequence)}
    base_part_id = result.get("base_part_id")
    remaining = set(result.get("remaining", []))
    successful_operations = result.get("operations", {})
    paths = result.get("paths", {})

    stats: Dict[str, dict] = {}
    for pid in sorted(part_ids):
        stats[pid] = {
            "part_id": pid,
            "is_base_part": pid == base_part_id,
            "resolved": pid == base_part_id or pid in sequence_index,
            "final_state": (
                "fixed_base"
                if pid == base_part_id
                else "removed"
                if pid in sequence_index
                else "unresolved"
                if pid in remaining
                else "not_attempted"
            ),
            "disassembly_step": sequence_index.get(pid),
            "search_attempts": 0,
            "successful_attempts": 0,
            "failed_attempts": 0,
            "timeout_attempts": 0,
            "cumulative_search_time_s": 0.0,
            "average_search_time_s": 0.0,
            "minimum_search_time_s": None,
            "maximum_search_time_s": None,
            "maximum_search_depth": 0,
            "tried_depths": [],
            "first_trial": None,
            "last_trial": None,
            "successful_trial": None,
            "successful_operation": successful_operations.get(pid),
            "successful_path_frames": int(len(paths.get(pid, []))),
            "successful_search_time_s": None,
            "successful_motion_duration_s": (
                float(len(paths.get(pid, []))) * float(result.get("simulation_dt", 0.0))
                if pid in sequence_index
                else None
            ),
            "time_to_success_s": None,
            "last_failure_reason": None,
            "encountered_blockers": [],
            "operation_attempt_counts": {},
        }

    for entry in result.get("trial_log", []):
        pid = entry.get("move_id")
        if pid not in stats:
            continue
        item = stats[pid]
        trial = int(entry.get("trial", 0))
        depth = int(entry.get("max_depth", 0))
        status = str(entry.get("status", "Failure"))
        elapsed = float(entry.get("path_time", 0.0))
        operation = str(entry.get("operation", "unknown"))

        item["search_attempts"] += 1
        item["cumulative_search_time_s"] += elapsed
        item["maximum_search_depth"] = max(item["maximum_search_depth"], depth)
        if depth not in item["tried_depths"]:
            item["tried_depths"].append(depth)
        item["first_trial"] = trial if item["first_trial"] is None else min(item["first_trial"], trial)
        item["last_trial"] = trial if item["last_trial"] is None else max(item["last_trial"], trial)
        item["minimum_search_time_s"] = (
            elapsed
            if item["minimum_search_time_s"] is None
            else min(item["minimum_search_time_s"], elapsed)
        )
        item["maximum_search_time_s"] = (
            elapsed
            if item["maximum_search_time_s"] is None
            else max(item["maximum_search_time_s"], elapsed)
        )
        item["operation_attempt_counts"][operation] = (
            item["operation_attempt_counts"].get(operation, 0) + 1
        )

        if status == "Success":
            item["successful_attempts"] += 1
            item["successful_trial"] = trial
            item["successful_search_time_s"] = elapsed
            item["time_to_success_s"] = item["cumulative_search_time_s"]
        elif status == "Timeout":
            item["timeout_attempts"] += 1
            item["last_failure_reason"] = entry.get("failure_reason") or "Timeout"
        else:
            item["failed_attempts"] += 1
            item["last_failure_reason"] = entry.get("failure_reason") or status

        blockers = set(item["encountered_blockers"])
        blockers.update(entry.get("blockers", []))
        item["encountered_blockers"] = sorted(blockers)

    for item in stats.values():
        attempts = int(item["search_attempts"])
        item["tried_depths"] = sorted(item["tried_depths"])
        if attempts > 0:
            item["average_search_time_s"] = item["cumulative_search_time_s"] / attempts
        # Round only exported scalar timings; retain enough precision for analysis.
        for key in (
            "cumulative_search_time_s",
            "average_search_time_s",
            "minimum_search_time_s",
            "maximum_search_time_s",
            "successful_search_time_s",
            "successful_motion_duration_s",
            "time_to_success_s",
        ):
            if item[key] is not None:
                item[key] = round(float(item[key]), 6)
    return stats


def print_part_statistics(part_statistics: Dict[str, dict]):
    """Print a compact experiment table after sequence planning."""
    print("\n========== Per-Part Search Statistics ==========")
    print(
        f"{'Part':<22} {'State':<11} {'Attempts':>8} "
        f"{'Depth':>6} {'Time(s)':>10} {'Step':>6}"
    )
    print("-" * 72)
    for pid in sorted(part_statistics):
        item = part_statistics[pid]
        step = "-" if item["disassembly_step"] is None else str(item["disassembly_step"])
        print(
            f"{pid:<22} {item['final_state']:<11} "
            f"{item['search_attempts']:>8d} "
            f"{item['maximum_search_depth']:>6d} "
            f"{item['cumulative_search_time_s']:>10.3f} "
            f"{step:>6}"
        )


def write_part_statistics_csv(path: str, part_statistics: Dict[str, dict]):
    fieldnames = [
        "part_id",
        "is_base_part",
        "resolved",
        "final_state",
        "disassembly_step",
        "search_attempts",
        "successful_attempts",
        "failed_attempts",
        "timeout_attempts",
        "cumulative_search_time_s",
        "average_search_time_s",
        "minimum_search_time_s",
        "maximum_search_time_s",
        "maximum_search_depth",
        "tried_depths",
        "first_trial",
        "last_trial",
        "successful_trial",
        "successful_operation",
        "successful_path_frames",
        "successful_search_time_s",
        "successful_motion_duration_s",
        "time_to_success_s",
        "last_failure_reason",
        "encountered_blockers",
        "operation_attempt_counts",
    ]
    with open(path, "w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for pid in sorted(part_statistics):
            row = dict(part_statistics[pid])
            row["tried_depths"] = json.dumps(row["tried_depths"], ensure_ascii=False)
            row["encountered_blockers"] = json.dumps(
                row["encountered_blockers"], ensure_ascii=False
            )
            row["operation_attempt_counts"] = json.dumps(
                row["operation_attempt_counts"], ensure_ascii=False
            )
            writer.writerow(row)


def write_experiment_excel(
    path: str,
    result: dict,
    part_statistics: Dict[str, dict],
) -> bool:
    """Write a formatted multi-sheet XLSX experiment report.

    The workbook contains an experiment summary, per-part statistics, the full
    trial log, and the online blocker graph.  ``openpyxl`` is imported lazily so
    planning can still run in minimal environments; install it with
    ``pip install openpyxl`` when XLSX export is required.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "[Excel] openpyxl is not installed; XLSX export was skipped. "
            "Install it with: pip install openpyxl"
        )
        return False

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    part_sheet = workbook.create_sheet("Part Statistics")
    trial_sheet = workbook.create_sheet("Trial Log")
    blocker_sheet = workbook.create_sheet("Blocker Graph")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    success_fill = PatternFill("solid", fgColor="E2F0D9")
    failure_fill = PatternFill("solid", fgColor="FCE4D6")
    timeout_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(sheet, row: int, start_col: int, end_col: int):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def autofit(sheet, minimum: int = 10, maximum: int = 42):
        for column_cells in sheet.columns:
            width = minimum
            column_index = column_cells[0].column
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, maximum))
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    movable = int(result.get("part_count_movable", 0))
    removed = int(result.get("part_count_removed", 0))
    success_rate = float(result.get("success_rate", removed / movable if movable else 1.0))
    summary_rows = [
        ("Metric", "Value"),
        ("Planning status", result.get("status")),
        ("Complete sequence available", bool(result.get("complete_sequence_available", False))),
        ("Total CAD parts", int(result.get("part_count_total", len(part_statistics)))),
        ("Movable parts", movable),
        ("Successfully removed parts", removed),
        ("Unresolved movable parts", int(result.get("part_count_unresolved", 0))),
        ("Success fraction", result.get("success_fraction", f"{removed}/{movable}")),
        ("Success rate", success_rate),
        ("Sequence trials", int(result.get("trials", 0))),
        ("Total path-planning time (s)", float(result.get("total_path_time", 0.0))),
        ("Planning wall time (s)", float(result.get("planning_wall_time", 0.0))),
        ("Base/reference part", result.get("base_part_id")),
        ("Successful disassembly order", ", ".join(result.get("motion_sequence", result.get("sequence", [])))),
        ("Unresolved parts", ", ".join(result.get("remaining", []))),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    style_header(summary_sheet, 1, 1, 2)
    summary_sheet.freeze_panes = "A2"
    summary_sheet["A8"].font = bold_font
    summary_sheet["B9"].number_format = "0.00%"
    for row in summary_sheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary_sheet["A17"] = "Successful parts and successful-search time"
    summary_sheet["A17"].fill = subheader_fill
    summary_sheet["A17"].font = bold_font
    summary_sheet["A18"] = "Part ID"
    summary_sheet["B18"] = "Successful search time (s)"
    summary_sheet["C18"] = "Simulated motion duration (s)"
    style_header(summary_sheet, 18, 1, 3)
    row_index = 19
    for pid in result.get("motion_sequence", result.get("sequence", [])):
        item = part_statistics.get(pid, {})
        summary_sheet.cell(row=row_index, column=1, value=pid)
        summary_sheet.cell(
            row=row_index,
            column=2,
            value=item.get("successful_search_time_s"),
        )
        summary_sheet.cell(row=row_index, column=2).number_format = "0.000000"
        summary_sheet.cell(
            row=row_index,
            column=3,
            value=item.get("successful_motion_duration_s"),
        )
        summary_sheet.cell(row=row_index, column=3).number_format = "0.000000"
        row_index += 1

    part_headers = [
        "part_id", "is_base_part", "resolved", "final_state",
        "disassembly_step", "search_attempts", "successful_attempts",
        "failed_attempts", "timeout_attempts", "cumulative_search_time_s",
        "average_search_time_s", "minimum_search_time_s",
        "maximum_search_time_s", "successful_search_time_s",
        "successful_motion_duration_s", "time_to_success_s", "maximum_search_depth", "tried_depths",
        "first_trial", "last_trial", "successful_trial",
        "successful_operation", "successful_path_frames",
        "last_failure_reason", "encountered_blockers",
        "operation_attempt_counts",
    ]
    part_sheet.append(part_headers)
    for pid in sorted(part_statistics):
        item = part_statistics[pid]
        row = []
        for key in part_headers:
            value = item.get(key)
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        part_sheet.append(row)
    style_header(part_sheet, 1, 1, len(part_headers))
    part_sheet.freeze_panes = "A2"
    part_sheet.auto_filter.ref = part_sheet.dimensions
    for row in part_sheet.iter_rows(min_row=2):
        state = str(row[3].value)
        fill = success_fill if state in {"removed", "fixed_base"} else failure_fill
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[3].fill = fill
    for col in range(10, 16):
        for cell in part_sheet.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                item.number_format = "0.000000"

    trial_headers = [
        "trial", "move_id", "max_depth", "status", "path_time",
        "operation", "blockers", "failure_reason", "path_frames",
        "diagnostic_path_frames", "remaining_count_before",
        "successful_step", "planner_diagnostics",
    ]
    trial_sheet.append(trial_headers)
    for entry in result.get("trial_log", []):
        row = []
        for key in trial_headers:
            value = entry.get(key)
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        trial_sheet.append(row)
    style_header(trial_sheet, 1, 1, len(trial_headers))
    trial_sheet.freeze_panes = "A2"
    trial_sheet.auto_filter.ref = trial_sheet.dimensions
    for row in trial_sheet.iter_rows(min_row=2):
        status = str(row[3].value)
        if status == "Success":
            row[3].fill = success_fill
        elif status == "Timeout":
            row[3].fill = timeout_fill
        else:
            row[3].fill = failure_fill
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if trial_sheet.max_row >= 2:
        trial_sheet.conditional_formatting.add(
            f"E2:E{trial_sheet.max_row}",
            CellIsRule(operator="greaterThan", formula=["60"], fill=timeout_fill),
        )

    blocker_sheet.append(["Blocked part", "Blocking part", "Relation"])
    for blocked, blockers in sorted(result.get("blocker_graph", {}).items()):
        if not blockers:
            blocker_sheet.append([blocked, None, "No recorded blocker"])
        for blocker in blockers:
            blocker_sheet.append([blocked, blocker, f"{blocker} -> {blocked}"])
    style_header(blocker_sheet, 1, 1, 3)
    blocker_sheet.freeze_panes = "A2"
    blocker_sheet.auto_filter.ref = blocker_sheet.dimensions
    for row in blocker_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for sheet in workbook.worksheets:
        autofit(sheet)
        sheet.sheet_view.showGridLines = False
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 70
    part_sheet.column_dimensions["W"].width = 45
    trial_sheet.column_dimensions["H"].width = 55
    trial_sheet.column_dimensions["M"].width = 70

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    workbook.save(path)
    print(f"[Excel] Experiment workbook saved: {path}")
    return True


def update_saved_summary(output_dir: str, updates: dict):
    """Merge replay/video timing data into saved JSON artifacts."""
    summary_path = os.path.join(output_dir, "plan_summary.json")
    if not os.path.isfile(summary_path):
        return
    with open(summary_path, "r", encoding="utf-8") as file:
        summary = json.load(file)
    summary.update(updates)
    with open(summary_path, "w", encoding="utf-8") as file:
        json.dump(summary, file, indent=2, ensure_ascii=False)

    process_path = os.path.join(output_dir, "process_statistics.json")
    process_data = {}
    if os.path.isfile(process_path):
        with open(process_path, "r", encoding="utf-8") as file:
            process_data = json.load(file)
    process_data.update(updates)
    with open(process_path, "w", encoding="utf-8") as file:
        json.dump(process_data, file, indent=2, ensure_ascii=False)


def save_results(
    output_dir: str,
    result: dict,
    assets: Dict[str, PartAsset],
    normalization: NormalizationInfo,
    args,
):
    save_started = time.perf_counter()
    os.makedirs(output_dir, exist_ok=True)
    path_dir = os.path.join(output_dir, "paths")
    diagnostic_path_dir = os.path.join(output_dir, "diagnostic_paths")
    # Do not leave stale paths or sequence files from an earlier run.
    if os.path.isdir(path_dir):
        shutil.rmtree(path_dir)
    if os.path.isdir(diagnostic_path_dir):
        shutil.rmtree(diagnostic_path_dir)
    os.makedirs(path_dir, exist_ok=True)
    os.makedirs(diagnostic_path_dir, exist_ok=True)
    for stale_name in (
        "disassembly_sequence.txt",
        "assembly_sequence.txt",
        "partial_disassembly_sequence.txt",
        "unresolved_parts.txt",
        "base_part.txt",
        "part_statistics.json",
        "part_statistics.csv",
        "experiment_statistics.xlsx",
        "process_statistics.json",
    ):
        stale_path = os.path.join(output_dir, stale_name)
        if os.path.isfile(stale_path):
            os.remove(stale_path)

    if getattr(args, "save_video", False):
        video_name = getattr(args, "video_filename", "disassembly_demo.mp4")
        stale_video = video_name if os.path.isabs(video_name) else os.path.join(output_dir, video_name)
        if os.path.isfile(stale_video):
            os.remove(stale_video)
    if getattr(args, "save_diagnostic_video", False):
        diagnostic_video_name = getattr(
            args, "diagnostic_video_filename", "planning_diagnostics.mp4"
        )
        stale_diagnostic_video = (
            diagnostic_video_name
            if os.path.isabs(diagnostic_video_name)
            else os.path.join(output_dir, diagnostic_video_name)
        )
        if os.path.isfile(stale_diagnostic_video):
            os.remove(stale_diagnostic_video)

    path_files = {}
    for step, pid in enumerate(result["sequence"]):
        path = np.asarray(result["paths"][pid], dtype=np.float64)
        original_path = np.asarray([normalization.qpos_to_original(q) for q in path])
        normalized_file = os.path.join(path_dir, f"{step:03d}_{pid}_normalized.npy")
        original_file = os.path.join(path_dir, f"{step:03d}_{pid}_original.npy")
        np.save(normalized_file, path)
        np.save(original_file, original_path)
        path_files[pid] = {
            "normalized": os.path.relpath(normalized_file, output_dir),
            "original": os.path.relpath(original_file, output_dir),
        }

    diagnostic_event_records = []
    for event in result.get("diagnostic_events", []):
        trial = int(event.get("trial", len(diagnostic_event_records) + 1))
        pid = str(event.get("move_id", "unknown"))
        event_path = np.asarray(event.get("path", []), dtype=np.float64)
        normalized_rel = None
        original_rel = None
        if event_path.ndim == 2 and event_path.shape[1] == 7 and len(event_path) > 0:
            original_event_path = np.asarray(
                [normalization.qpos_to_original(q) for q in event_path],
                dtype=np.float64,
            )
            normalized_file = os.path.join(
                diagnostic_path_dir, f"trial_{trial:04d}_{pid}_normalized.npy"
            )
            original_file = os.path.join(
                diagnostic_path_dir, f"trial_{trial:04d}_{pid}_original.npy"
            )
            np.save(normalized_file, event_path)
            np.save(original_file, original_event_path)
            normalized_rel = os.path.relpath(normalized_file, output_dir)
            original_rel = os.path.relpath(original_file, output_dir)
        record = {key: value for key, value in event.items() if key != "path"}
        record["path_file_normalized"] = normalized_rel
        record["path_file_original"] = original_rel
        diagnostic_event_records.append(record)

    summary = {
        "planner_version": PLANNER_VERSION,
        "approach": result.get("approach", getattr(args, "approach", None)),
        "status": result["status"],
        "complete_sequence_available": result.get("complete_sequence_available", False),
        "motion_sequence": result.get("motion_sequence", result["sequence"]),
        "disassembly_sequence": result.get("disassembly_sequence", result["sequence"]),
        "assembly_sequence": result["assembly_sequence"],
        "base_part_id": result.get("base_part_id"),
        "remaining": result["remaining"],
        "success_fraction": result.get("success_fraction"),
        "success_rate": result.get("success_rate"),
        "operations": result["operations"],
        "trials": result["trials"],
        "total_path_time": result["total_path_time"],
        "planning_wall_time": result.get("planning_wall_time"),
        "simulation_dt": result.get("simulation_dt"),
        "trial_log": result["trial_log"],
        "diagnostic_events": diagnostic_event_records,
        "blocker_graph": result.get("blocker_graph", {}),
        "llm_prior": result.get("llm_prior", {}),
        "normalization": normalization.to_json(),
        "colors": {pid: list(asset.color) for pid, asset in assets.items()},
        "mesh_statistics": {
            pid: {
                "source_faces": asset.source_faces,
                "collision_faces": asset.collision_faces,
                "collision_watertight": asset.collision_watertight,
            }
            for pid, asset in assets.items()
        },
        "path_files": path_files,
        "part_counts": {
            "total": int(result.get("part_count_total", len(assets))),
            "movable": int(result.get("part_count_movable", max(len(assets) - 1, 0))),
            "removed": int(result.get("part_count_removed", len(result.get("sequence", [])))),
            "unresolved": int(result.get("part_count_unresolved", len(result.get("remaining", [])))),
        },
        "part_statistics": result.get("part_statistics", {}),
        "timing": {
            "run_started_at_utc": result.get("run_started_at_utc"),
            "run_start_epoch": result.get("run_start_epoch"),
            "genesis_initialization_time_s": result.get("genesis_initialization_time"),
            "preprocessing_time_s": result.get("preprocessing_time"),
            "planning_path_time_s": result.get("total_path_time"),
            "planning_wall_time_s": result.get("planning_wall_time"),
            "elapsed_before_save_s": result.get("elapsed_before_save"),
            "result_save_time_s": None,
            "total_process_time_before_replay_s": None,
            "end_to_end_time_including_video_s": None,
        },
        "video": {
            "requested": bool(getattr(args, "save_video", False)),
            "status": (
                "pending"
                if result.get("status") == "Success" and getattr(args, "save_video", False)
                else "not_requested_or_unsuccessful"
            ),
            "file": None,
            "mode": "validated_final_plan",
        },
        "diagnostic_video": {
            "requested": bool(getattr(args, "save_diagnostic_video", False)),
            "status": (
                "pending"
                if result.get("status") != "Success"
                and diagnostic_event_records
                and getattr(args, "save_diagnostic_video", False)
                else "not_requested_or_not_needed"
            ),
            "file": None,
            "mode": "all_planning_trials",
        },
        "arguments": vars(args),
    }

    with open(os.path.join(output_dir, "llm_prior_used.json"), "w", encoding="utf-8") as file:
        json.dump(result.get("llm_prior", {}), file, indent=2, ensure_ascii=False)

    with open(os.path.join(output_dir, "disassembly_motion_sequence.txt"), "w", encoding="utf-8") as file:
        file.write("\n".join(result.get("motion_sequence", result["sequence"])) + "\n")

    if result.get("complete_sequence_available", False):
        with open(os.path.join(output_dir, "disassembly_sequence.txt"), "w", encoding="utf-8") as file:
            file.write("\n".join(result.get("disassembly_sequence", result["sequence"])) + "\n")
        with open(os.path.join(output_dir, "assembly_sequence.txt"), "w", encoding="utf-8") as file:
            file.write("\n".join(result["assembly_sequence"]) + "\n")
        with open(os.path.join(output_dir, "base_part.txt"), "w", encoding="utf-8") as file:
            file.write(str(result.get("base_part_id") or "") + "\n")
    else:
        with open(os.path.join(output_dir, "partial_disassembly_sequence.txt"), "w", encoding="utf-8") as file:
            file.write("\n".join(result.get("motion_sequence", result["sequence"])) + "\n")
        with open(os.path.join(output_dir, "unresolved_parts.txt"), "w", encoding="utf-8") as file:
            file.write("\n".join(result.get("remaining", [])) + "\n")

    part_statistics = result.get("part_statistics", {})
    with open(os.path.join(output_dir, "part_statistics.json"), "w", encoding="utf-8") as file:
        json.dump(part_statistics, file, indent=2, ensure_ascii=False)
    write_part_statistics_csv(
        os.path.join(output_dir, "part_statistics.csv"),
        part_statistics,
    )
    write_experiment_excel(
        os.path.join(output_dir, "experiment_statistics.xlsx"),
        result,
        part_statistics,
    )

    save_time = time.perf_counter() - save_started
    elapsed_before_save = float(result.get("elapsed_before_save", 0.0) or 0.0)
    total_before_replay = elapsed_before_save + save_time
    summary["timing"]["result_save_time_s"] = round(save_time, 6)
    summary["timing"]["total_process_time_before_replay_s"] = round(total_before_replay, 6)

    process_statistics = {
        "approach": result.get("approach", getattr(args, "approach", None)),
        "status": result.get("status"),
        "run_started_at_utc": result.get("run_started_at_utc"),
        "part_counts": summary["part_counts"],
        "total_sequence_trials": int(result.get("trials", 0)),
        "success_fraction": result.get("success_fraction"),
        "success_rate": result.get("success_rate"),
        "timing": summary["timing"],
        "per_part_search_attempts": {
            pid: int(item.get("search_attempts", 0))
            for pid, item in part_statistics.items()
        },
        "per_part_search_time_s": {
            pid: float(item.get("cumulative_search_time_s", 0.0))
            for pid, item in part_statistics.items()
        },
        "video": summary["video"],
        "diagnostic_video": summary["diagnostic_video"],
    }
    with open(os.path.join(output_dir, "process_statistics.json"), "w", encoding="utf-8") as file:
        json.dump(process_statistics, file, indent=2, ensure_ascii=False)

    # Write the master summary last so it includes file-saving overhead.
    with open(os.path.join(output_dir, "plan_summary.json"), "w", encoding="utf-8") as file:
        json.dump(summary, file, indent=2, ensure_ascii=False)

    print(f"[Save] Results written to: {output_dir}")
    print(f"[Save] Per-part statistics: {os.path.join(output_dir, 'part_statistics.csv')}")
    print(f"[Save] Excel statistics: {os.path.join(output_dir, 'experiment_statistics.xlsx')}")
    print(f"[Save] Diagnostic trial paths: {diagnostic_path_dir}")
    print(f"[Save] Process statistics: {os.path.join(output_dir, 'process_statistics.json')}")
    print(f"[Timing] Total process time before replay: {total_before_replay:.3f}s")
    return summary


# ============================================================================
# Final high-resolution replay
# ============================================================================


def compute_scene_camera(assets: Dict[str, PartAsset]):
    _, _, center, extent = compute_assembly_bounds(assets)
    camera_pos = (
        float(center[0] + 1.2 * extent),
        float(center[1] - 1.6 * extent),
        float(center[2] + 0.9 * extent),
    )
    return camera_pos, tuple(center.tolist())


def load_saved_replay(output_dir: str):
    """Load validated paths and diagnostic trial paths from saved results."""
    summary_path = os.path.join(output_dir, "plan_summary.json")
    if not os.path.isfile(summary_path):
        raise FileNotFoundError(f"Replay summary not found: {summary_path}")
    with open(summary_path, "r", encoding="utf-8") as file:
        summary = json.load(file)

    sequence = summary.get(
        "motion_sequence",
        summary.get("disassembly_sequence", []),
    )
    path_files = summary.get("path_files", {})
    paths: Dict[str, List[np.ndarray]] = {}
    for pid in sequence:
        info = path_files.get(pid)
        if info is None:
            print(f"[Replay] No saved path metadata for {pid}; skipping its motion.")
            continue
        path_file = os.path.join(output_dir, info["normalized"])
        if not os.path.isfile(path_file):
            print(f"[Replay] Missing saved path for {pid}: {path_file}")
            continue
        array = np.load(path_file)
        paths[pid] = [np.asarray(q, dtype=np.float64) for q in array]

    diagnostic_events = []
    for record in summary.get("diagnostic_events", []):
        event = dict(record)
        path_file = record.get("path_file_normalized")
        event_path = []
        if path_file:
            absolute = os.path.join(output_dir, path_file)
            if os.path.isfile(absolute):
                array = np.load(absolute)
                event_path = [np.asarray(q, dtype=np.float64) for q in array]
        event["path"] = event_path
        diagnostic_events.append(event)
    return summary, list(sequence), paths, diagnostic_events


def relaunch_replay_in_fresh_process():
    """Replace the planning process with a fresh Genesis replay process.

    Genesis keeps native renderer/solver resources globally. Rebuilding many
    planning scenes and then constructing an interactive viewer in the same
    interpreter can leave stale native state and trigger a bare SIGSEGV.
    ``execv`` replaces the entire process image, guaranteeing a clean Genesis
    initialization without requiring unreliable in-process scene destruction.
    """
    script_path = os.path.abspath(__file__)
    filtered_args = []
    for arg in sys.argv[1:]:
        if arg in {"--replay-only", "--rebuild-cache", "--vis-plan"}:
            continue
        filtered_args.append(arg)
    filtered_args.append("--replay-only")
    command = [sys.executable, script_path, *filtered_args]
    print("[Replay] Restarting Genesis in a fresh process for stable visualization.")
    print("[Replay] Command:", " ".join(command))
    sys.stdout.flush()
    sys.stderr.flush()
    os.execv(sys.executable, command)


def resolve_video_path(output_dir: str, filename: str) -> str:
    path = filename if os.path.isabs(filename) else os.path.join(output_dir, filename)
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    return os.path.abspath(path)


def visualize_disassembly_replay(
    assets: Dict[str, PartAsset],
    sequence: Sequence[str],
    paths: Dict[str, List[np.ndarray]],
    dt: float,
    replay_fps: float,
    repeat: bool,
    intro_seconds: float,
    output_dir: str,
    save_video: bool,
    video_filename: str,
    video_width: int,
    video_height: int,
    video_fps: float,
    video_step_hold_seconds: float,
    video_final_hold_seconds: float,
    video_frame_stride: int,
    show_viewer: bool,
    replay_mode: str = "final",
    diagnostic_events: Optional[Sequence[dict]] = None,
    diagnostic_reset_hold_seconds: float = 0.25,
):
    """Replay either the validated final plan or all diagnostic search trials.

    ``replay_mode='diagnostic'`` reproduces the actual top-level trial order.
    Successful parts remain at their final/parking pose.  A failed trial follows
    its best observed trajectory, pauses at the failure state, and then resets
    only the attempted part before the next trial.  This makes incomplete runs
    observable instead of silently skipping video generation.
    """
    replay_mode = str(replay_mode).strip().lower()
    diagnostic_events = list(diagnostic_events or [])
    if replay_mode == "diagnostic":
        if not diagnostic_events:
            print("[Replay] No diagnostic trial path to visualize.")
            return
    elif not sequence:
        print("[Replay] No successful path to visualize.")
        return

    camera_pos, camera_lookat = compute_scene_camera(assets)
    scene_kwargs = dict(
        sim_options=construct_compatible(gs.options.SimOptions, dt=dt, substeps=1),
        rigid_options=construct_compatible(
            gs.options.RigidOptions,
            gravity=(0.0, 0.0, 0.0),
            enable_collision=False,
        ),
        viewer_options=construct_compatible(
            gs.options.ViewerOptions,
            camera_pos=camera_pos,
            camera_lookat=camera_lookat,
            camera_fov=45,
            refresh_rate=max(int(replay_fps), 1),
            run_in_thread=False,
        ),
        show_viewer=bool(show_viewer),
    )
    scene_kwargs.update(scene_profiling_kwargs(False))
    scene = gs.Scene(**scene_kwargs)

    entities = {}
    for pid in sorted(assets):
        asset = assets[pid]
        morph = construct_compatible(
            gs.morphs.Mesh,
            file=asset.visual_path,
            pos=tuple(asset.center.tolist()),
            quat=(1.0, 0.0, 0.0, 0.0),
            fixed=False,
            collision=False,
            visualization=True,
            decimate=False,
            convexify=False,
            recompute_inertia=False,
            align=False,
            file_meshes_are_zup=True,
        )
        entities[pid] = scene.add_entity(
            material=construct_compatible(gs.materials.Rigid, rho=100.0),
            morph=morph,
            surface=make_surface(asset.color),
        )

    record_camera = None
    if save_video:
        camera_kwargs = dict(
            res=(int(video_width), int(video_height)),
            pos=camera_pos,
            lookat=camera_lookat,
            fov=45,
        )
        try:
            record_camera = scene.add_camera(**camera_kwargs, GUI=False)
        except TypeError:
            record_camera = scene.add_camera(**camera_kwargs)

    scene.build()
    initial_qpos = {pid: make_initial_qpos(asset) for pid, asset in assets.items()}
    frame_sleep = 1.0 / max(float(replay_fps), 1.0)
    recording_fps = max(float(video_fps), 1.0)
    video_stride = max(int(video_frame_stride), 1)
    recorded_frames = 0
    replay_frames = 0
    recording_active = False
    recording_started = None
    video_path = resolve_video_path(output_dir, video_filename)

    def set_pose(pid: str, qpos: np.ndarray):
        if pid not in entities:
            return
        q = np.asarray(qpos, dtype=np.float64).copy()
        q[3:7] = normalize_quat(q[3:7])
        entities[pid].set_qpos(q, zero_velocity=True)
        entities[pid].set_dofs_velocity(None)

    def advance_frame(record: bool):
        nonlocal recorded_frames, replay_frames
        scene.step()
        replay_frames += 1
        if record_camera is not None and record and (replay_frames - 1) % video_stride == 0:
            record_camera.render()
            recorded_frames += 1
        time.sleep(frame_sleep)

    def hold(seconds: float, record: bool):
        for _ in range(max(int(max(seconds, 0.0) * replay_fps), 1)):
            advance_frame(record)

    def reset_all(record: bool):
        for pid, qpos in initial_qpos.items():
            set_pose(pid, qpos)
        hold(intro_seconds, record)

    def replay_final(record: bool):
        reset_all(record)
        for step, pid in enumerate(sequence):
            path = paths.get(pid)
            if not path:
                continue
            print(f"[Replay] {step + 1}/{len(sequence)}: {pid}, frames={len(path)}")
            for qpos in path:
                set_pose(pid, qpos)
                advance_frame(record)
            hold(video_step_hold_seconds, record)
        hold(video_final_hold_seconds, record)

    def replay_diagnostics(record: bool):
        reset_all(record)
        successful_final_pose: Dict[str, np.ndarray] = {}
        for event_index, event in enumerate(diagnostic_events, start=1):
            pid = str(event.get("move_id", ""))
            status = str(event.get("status", "Failure"))
            trial = int(event.get("trial", event_index))
            path = event.get("path") or []
            blockers = event.get("blockers", [])
            failure_reason = event.get("failure_reason", "")

            # Reassert already removed parts in case a viewer/backend changed a
            # pose while resetting the current attempted part.
            for removed_pid, final_pose in successful_final_pose.items():
                set_pose(removed_pid, final_pose)
            if pid in initial_qpos and pid not in successful_final_pose:
                set_pose(pid, initial_qpos[pid])
                advance_frame(record)

            print(
                f"[Diagnostic replay] trial={trial} part={pid} status={status} "
                f"frames={len(path)} blockers={blockers} reason={failure_reason}"
            )
            for qpos in path:
                set_pose(pid, qpos)
                advance_frame(record)
            hold(video_step_hold_seconds, record)

            if status == "Success" and path:
                successful_final_pose[pid] = np.asarray(path[-1], dtype=np.float64).copy()
            else:
                if pid in initial_qpos:
                    set_pose(pid, initial_qpos[pid])
                hold(diagnostic_reset_hold_seconds, record)

        hold(video_final_hold_seconds, record)

    print("\n========== Genesis High-Resolution Replay ==========")
    print(f"Replay mode: {replay_mode}")
    if save_video:
        print(f"[Video] Recording to: {video_path}")
    print("Close the viewer or press Ctrl+C to stop after the video is saved.")

    video_metadata = {
        "requested": bool(save_video),
        "status": "not_requested",
        "mode": replay_mode,
        "file": None,
        "resolution": [int(video_width), int(video_height)],
        "fps": float(recording_fps),
        "frame_stride": int(video_stride),
        "recorded_frames": 0,
        "duration_s": 0.0,
        "recording_wall_time_s": 0.0,
        "saved_at_utc": None,
    }
    summary_field = "diagnostic_video" if replay_mode == "diagnostic" else "video"

    def replay_once(record: bool):
        if replay_mode == "diagnostic":
            replay_diagnostics(record)
        else:
            replay_final(record)

    try:
        if record_camera is not None:
            if os.path.isfile(video_path):
                os.remove(video_path)
            record_camera.render()
            record_camera.start_recording()
            recording_active = True
            recording_started = time.perf_counter()

        replay_once(record=record_camera is not None)

        if record_camera is not None:
            record_camera.stop_recording(
                save_to_filename=video_path,
                fps=max(int(round(recording_fps)), 1),
            )
            recording_active = False
            recording_wall = time.perf_counter() - recording_started
            video_metadata.update(
                {
                    "status": "saved" if os.path.isfile(video_path) else "save_requested",
                    "file": os.path.relpath(video_path, output_dir),
                    "recorded_frames": int(recorded_frames),
                    "duration_s": round(recorded_frames / recording_fps, 6),
                    "recording_wall_time_s": round(recording_wall, 6),
                    "saved_at_utc": utc_now_iso(),
                }
            )
            print(
                f"[Video] Saved: {video_path} | frames={recorded_frames} | "
                f"duration={recorded_frames / recording_fps:.2f}s"
            )

        summary_path = os.path.join(output_dir, "plan_summary.json")
        run_start_epoch = None
        summary_timing = {}
        if os.path.isfile(summary_path):
            with open(summary_path, "r", encoding="utf-8") as file:
                existing_summary = json.load(file)
            summary_timing = dict(existing_summary.get("timing", {}))
            run_start_epoch = summary_timing.get("run_start_epoch")
        if run_start_epoch is not None:
            summary_timing["end_to_end_time_including_video_s"] = round(
                max(time.time() - float(run_start_epoch), 0.0), 6
            )
        summary_timing[
            "diagnostic_replay_frames" if replay_mode == "diagnostic" else "replay_frames"
        ] = int(replay_frames)
        summary_timing[
            "diagnostic_video_recording_wall_time_s"
            if replay_mode == "diagnostic"
            else "video_recording_wall_time_s"
        ] = video_metadata.get("recording_wall_time_s", 0.0)
        update_saved_summary(
            output_dir,
            {
                summary_field: video_metadata,
                "timing": summary_timing,
            },
        )

        if not show_viewer:
            print("[Replay] Headless replay finished.")
            return
        if repeat:
            print("[Replay] Video saved. Continuing repeated viewer replay.")
            while True:
                replay_once(record=False)
        else:
            print("[Replay] Finished. Viewer remains open; press Ctrl+C to exit.")
            while True:
                advance_frame(record=False)

    except KeyboardInterrupt:
        print("[Replay] Stopped by user.")
    except Exception as exc:
        video_metadata.update(
            {
                "status": "failed",
                "error": f"{type(exc).__name__}: {exc}",
            }
        )
        update_saved_summary(output_dir, {summary_field: video_metadata})
        raise
    finally:
        if recording_active and record_camera is not None:
            try:
                record_camera.stop_recording(
                    save_to_filename=video_path,
                    fps=max(int(round(recording_fps)), 1),
                )
            except Exception:
                pass


# ============================================================================
# CLI
# ============================================================================


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="Efficient Genesis assembly-by-disassembly sequence and path planner.",
    )
    parser.add_argument("--dir", required=True, help="Dataset directory under --assets-root, e.g. satellite")
    parser.add_argument("--id", required=True, help="Assembly id, e.g. 00005")
    parser.add_argument("--assets-root", default="assets")
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--cache-dir", default=None)
    parser.add_argument("--rebuild-cache", action="store_true")

    # Preprocessing.
    parser.add_argument("--normalize-size", type=float, default=1.0, help="Maximum assembly bbox side in Genesis meters")
    parser.add_argument("--collision-face-num", type=int, default=5000)
    parser.add_argument("--simplify-aggressiveness", type=int, default=5, choices=range(0, 9))
    parser.add_argument("--force-hull-for-nonwatertight", action="store_true")
    parser.add_argument("--color-seed", type=int, default=17)

    # Genesis physics.
    parser.add_argument("--cpu", action="store_true")
    parser.add_argument("--dt", type=float, default=0.0025)
    parser.add_argument("--substeps", type=int, default=4)
    parser.add_argument("--density", type=float, default=200.0)
    parser.add_argument("--friction", type=float, default=0.01, help="Near-frictionless contact, matching the original planner")
    parser.add_argument("--sdf-cell-size", type=float, default=0.005, help="Genesis SDF cell size in normalized meters")
    parser.add_argument("--sdf-min-res", type=int, default=32)
    parser.add_argument("--sdf-max-res", type=int, default=128)
    parser.add_argument("--force-mag", type=float, default=30.0)
    parser.add_argument("--frame-skip", type=int, default=8)
    parser.add_argument("--collision-mode", choices=["sdf", "convex"], default="sdf")
    parser.add_argument("--watertighten", type=int, default=1, help="Used only by Genesis versions exposing this field")
    parser.add_argument(
        "--free-dof-order",
        choices=["auto", "linear-angular", "angular-linear"],
        default="auto",
        help="Ordering of free-body force DOFs; auto performs a short calibration",
    )
    parser.add_argument(
        "--penetration-tolerance",
        type=float,
        default=0.0025,
        help=(
            "Contact calibration scale: full growth margin for CAD contacts "
            "and twice the bounded numerical skin for newly appearing "
            "collision-mesh/SDF contacts"
        ),
    )
    parser.add_argument(
        "--sweep-translation-step",
        type=float,
        default=0.002,
        help="Translation sampling interval for anti-tunnelling validation",
    )
    parser.add_argument(
        "--sweep-rotation-step",
        type=float,
        default=0.05,
        help="Rotation sampling interval in radians for anti-tunnelling validation",
    )
    parser.add_argument(
        "--min-removal-distance",
        type=float,
        default=0.01,
        help="Absolute minimum displacement before a part may be declared removed",
    )
    parser.add_argument(
        "--min-removal-distance-factor",
        type=float,
        default=0.15,
        help="Additional minimum displacement as a fraction of the moving part diagonal",
    )

    # Unified comparison method. Method-specific switches are applied after
    # parsing so all approaches share the same physics and evaluation settings.
    parser.add_argument(
        "--approach",
        default="full_phyd2a",
        help=(
            "Comparison approach: full_phyd2a, rrt, rrt_star (or RRT*), "
            "bi_rrt, bk_rrt, phyd2a_bfs, phyd2a_dfs, phyd2a_no_llm."
        ),
    )

    # Search.
    parser.add_argument("--rotation", action="store_true")
    parser.add_argument(
        "--seq-max-time",
        "--no-progress-timeout",
        dest="seq_max_time",
        type=float,
        default=1200.0,
        help=(
            "No-progress timeout in seconds. The timer resets after every "
            "successfully removed part; it is no longer a fixed total runtime cap. "
            "Default: 1200 s. Use <=0 to disable this guard."
        ),
    )
    parser.add_argument("--path-max-time", type=float, default=300.0, help="Wall-clock guard for one path; deterministic semantic screw validation normally finishes before this limit")
    parser.add_argument("--seed", type=int, default=1)
    parser.add_argument(
        "--candidate-order",
        choices=["top-down-left-right", "exterior", "random"],
        default=None,
        help=(
            "Candidate ordering. Default is top-down-left-right with the LLM "
            "prior and geometry-based exterior ordering without the prior."
        ),
    )
    direct_release_switch = parser.add_mutually_exclusive_group()
    direct_release_switch.add_argument(
        "--enable-direct-release",
        dest="enable_direct_release",
        action="store_true",
        help="Enable the straight swept-mesh extraction shortcut.",
    )
    direct_release_switch.add_argument(
        "--disable-direct-release",
        dest="enable_direct_release",
        action="store_false",
        help="Disable the shortcut and use force-driven BFS only.",
    )
    parser.set_defaults(enable_direct_release=None)
    parser.add_argument("--trans-dist-th", type=float, default=0.0075)
    parser.add_argument("--quat-dist-th", type=float, default=0.20)
    parser.add_argument("--max-action-chunks", type=int, default=200)
    parser.add_argument(
        "--max-progressive-depth",
        type=int,
        default=12,
        help="Maximum multi-action BFS depth for one remaining-part state",
    )
    parser.add_argument(
        "--max-sequence-trials",
        type=int,
        default=500,
        help="Hard cap preventing endless blocker cycles",
    )
    parser.add_argument(
        "--auto-rotation-depth",
        type=int,
        default=2,
        help="Automatically enable 6-DoF actions from this progressive depth",
    )
    parser.add_argument("--vis-plan", action="store_true")
    parser.add_argument(
        "--search-strategy",
        choices=[
            "adaptive-best-first", "bfs", "dfs", "rrt", "rrt-star",
            "bi-rrt", "bk-rrt",
        ],
        default="adaptive-best-first",
        help=(
            "Expert/debug selector. In benchmark runs --approach overrides this "
            "value so method definitions remain reproducible."
        ),
    )
    parser.add_argument(
        "--adaptive-beam-width",
        type=int,
        default=8,
        help="Number of highest-scoring successor states retained per expansion",
    )
    parser.add_argument(
        "--adaptive-action-top-k",
        type=int,
        default=24,
        help="Maximum ranked wrench candidates evaluated at one adaptive node",
    )
    parser.add_argument(
        "--disable-diagonal-actions",
        action="store_true",
        help="Disable XY/XZ/YZ and 3D diagonal force directions",
    )
    parser.add_argument(
        "--coupled-action-scale",
        type=float,
        default=0.45,
        help="Torque-to-force ratio in simultaneous translation/rotation actions",
    )
    parser.add_argument(
        "--rrt-max-nodes",
        type=int,
        default=400,
        help="Maximum tree nodes per part for RRT-family planners",
    )
    parser.add_argument(
        "--rrt-goal-bias",
        type=float,
        default=0.20,
        help="Probability of sampling an exterior disassembly-biased target",
    )
    parser.add_argument(
        "--rrt-neighbor-count",
        type=int,
        default=8,
        help="Nearest-neighbor count for RRT* parent selection/rewiring and Bi-RRT",
    )
    parser.add_argument(
        "--rrt-connect-distance-factor",
        type=float,
        default=0.35,
        help="Tree connection/rewiring radius normalized by the moving-part diagonal",
    )
    parser.add_argument(
        "--rrt-goal-root-count",
        type=int,
        default=8,
        help="Number of collision-free exterior roots used by Bi-RRT",
    )
    parser.add_argument(
        "--diagnostic-max-frames-per-trial",
        type=int,
        default=900,
        help="Maximum saved pose frames for each successful or failed trial",
    )

    # Kept as hidden compatibility flags. --approach is authoritative:
    # full_phyd2a always enables the prior and every other approach disables it.
    prior_switch = parser.add_mutually_exclusive_group()
    prior_switch.add_argument(
        "--use-llm-prior",
        dest="use_llm_prior",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    prior_switch.add_argument(
        "--disable-llm-prior",
        dest="use_llm_prior",
        action="store_false",
        help=argparse.SUPPRESS,
    )
    parser.set_defaults(use_llm_prior=None)
    parser.add_argument(
        "--llm-prior-file",
        default=None,
        help=(
            "JSON prior containing priority_order and per-part operations. "
            "Used only by --approach full_phyd2a; ignored by all baselines."
        ),
    )
    parser.add_argument("--prior-part-ids", default=None, help="Comma-separated priority ids; overrides the JSON order")
    parser.add_argument("--no-strict-prior", action="store_true")
    parser.add_argument("--prior-initial-depth", type=int, default=1)
    parser.add_argument(
        "--base-part-id",
        default=None,
        help=(
            "Hard stationary base constraint. Required for every non-Full "
            "approach. For full_phyd2a, this value overrides base_part_id in "
            "the LLM prior JSON."
        ),
    )
    parser.add_argument("--disable-llm-unscrew", action="store_true")
    parser.add_argument("--nut-unscrew-turns", type=float, default=3.0)
    parser.add_argument("--nut-unscrew-distance", type=float, default=None)
    parser.add_argument("--nut-unscrew-frames", type=int, default=120)
    parser.add_argument("--nut-axis-map", default=None)
    parser.add_argument("--axis-estimation-mode", choices=["auto", "min", "max"], default="auto")
    parser.add_argument(
        "--nut-unscrew-handedness",
        type=float,
        default=0.0,
        help="Preferred torque sign; 0 probes both directions automatically",
    )
    parser.add_argument("--unscrew-axial-force", type=float, default=0.8)
    parser.add_argument("--unscrew-torque", type=float, default=0.05)
    parser.add_argument("--unscrew-probe-chunks", type=int, default=3)
    parser.add_argument("--unscrew-max-chunks", type=int, default=500)
    parser.add_argument("--unscrew-stall-chunks", type=int, default=4)
    parser.add_argument("--unscrew-lateral-factor", type=float, default=0.35)
    parser.add_argument("--unscrew-force-growth", type=float, default=1.35)
    parser.add_argument("--unscrew-max-force-scale", type=float, default=3.0)
    parser.add_argument("--prior-unscrew-max-retries", type=int, default=2)

    # Parking and replay.
    parser.add_argument("--disable-parking", action="store_true")
    parser.add_argument("--parking-distance-scale", type=float, default=1.35)
    parser.add_argument("--parking-spacing-scale", type=float, default=0.30)
    parser.add_argument("--parking-frames", type=int, default=60)
    parser.add_argument("--no-vis", action="store_true", help="Disable final high-resolution replay")
    parser.add_argument("--replay-fps", type=float, default=30.0)
    parser.add_argument("--replay-repeat", action="store_true")
    parser.add_argument("--replay-intro-seconds", type=float, default=1.5)

    video_switch = parser.add_mutually_exclusive_group()
    video_switch.add_argument(
        "--save-video",
        dest="save_video",
        action="store_true",
        help="Save one complete successful disassembly demonstration video (default)",
    )
    video_switch.add_argument(
        "--no-save-video",
        dest="save_video",
        action="store_false",
        help="Disable automatic MP4 recording",
    )
    parser.set_defaults(save_video=True)
    parser.add_argument(
        "--video-filename",
        default="disassembly_demo.mp4",
        help="Absolute path or filename relative to --output-dir",
    )
    parser.add_argument("--video-width", type=int, default=1280)
    parser.add_argument("--video-height", type=int, default=720)
    parser.add_argument(
        "--video-fps",
        type=float,
        default=None,
        help="Encoded video FPS; defaults to --replay-fps",
    )
    parser.add_argument(
        "--video-step-hold-seconds",
        type=float,
        default=0.5,
        help="Pause after each removed part in the saved video",
    )
    parser.add_argument(
        "--video-final-hold-seconds",
        type=float,
        default=2.0,
        help="Hold the final fully disassembled state in the saved video",
    )
    parser.add_argument(
        "--video-frame-stride",
        type=int,
        default=1,
        help="Record every Nth replay frame; 1 preserves every frame",
    )
    diagnostic_video_switch = parser.add_mutually_exclusive_group()
    diagnostic_video_switch.add_argument(
        "--save-diagnostic-video",
        dest="save_diagnostic_video",
        action="store_true",
        help="Save a planning-trial video when the complete plan fails (default)",
    )
    diagnostic_video_switch.add_argument(
        "--no-save-diagnostic-video",
        dest="save_diagnostic_video",
        action="store_false",
        help="Disable automatic failed-run diagnostic video recording",
    )
    parser.set_defaults(save_diagnostic_video=True)
    parser.add_argument(
        "--diagnostic-video-filename",
        default="planning_diagnostics.mp4",
        help="Filename relative to --output-dir for incomplete-run trial replay",
    )
    parser.add_argument(
        "--diagnostic-reset-hold-seconds",
        type=float,
        default=0.25,
        help="Pause after resetting a failed attempted part",
    )
    parser.add_argument(
        "--replay-backend",
        choices=["cpu", "gpu", "same"],
        default="cpu",
        help="Backend used by the fresh replay process; CPU is the safest for visualization.",
    )
    parser.add_argument("--replay-only", action="store_true", help=argparse.SUPPRESS)
    return parser


def main():
    run_start_epoch = time.time()
    run_perf_start = time.perf_counter()
    run_started_at_utc = utc_now_iso()

    parser = build_argument_parser()
    args = parser.parse_args()

    try:
        args.approach = normalize_approach_name(args.approach)
    except ValueError as exc:
        parser.error(str(exc))
    approach_config = APPROACH_CONFIGS[args.approach]

    # Only Full PhyD2A is allowed to consume LLM-generated information.  Every
    # baseline is isolated from the prior file and receives the fixed base via
    # the independent --base-part-id argument.
    if args.approach != "full_phyd2a":
        if args.base_part_id in (None, ""):
            parser.error(
                f"--approach {args.approach} does not use an LLM prior; "
                "specify the stationary reference with --base-part-id."
            )
        ignored_prior_args = []
        if args.llm_prior_file:
            ignored_prior_args.append("--llm-prior-file")
        if args.prior_part_ids:
            ignored_prior_args.append("--prior-part-ids")
        if ignored_prior_args:
            print(
                "[Approach isolation] Ignoring "
                + ", ".join(ignored_prior_args)
                + f" because {args.approach} is a no-LLM baseline."
            )

    # --approach is the single authoritative method selector.  Common numerical
    # parameters remain user-configurable, while method-defining switches are
    # fixed here to prevent accidental hybrid baselines.
    args.use_llm_prior = approach_config.use_llm_prior
    args.search_strategy = approach_config.search_strategy
    args.candidate_order = approach_config.candidate_order
    args.enable_direct_release = approach_config.enable_direct_release
    args.rotation = approach_config.force_rotation
    args.auto_rotation_depth = approach_config.auto_rotation_depth
    args.no_strict_prior = not approach_config.strict_prior
    args.disable_llm_unscrew = not approach_config.use_llm_prior
    args.disable_diagonal_actions = not approach_config.enable_diagonal_actions

    if args.approach == "full_phyd2a":
        assert args.use_llm_prior, "Full PhyD2A must enable the LLM prior"
    else:
        assert not args.use_llm_prior, (
            f"Baseline {args.approach} must not enable the LLM prior"
        )
    args.coupled_action_scale = approach_config.coupled_action_scale

    if args.candidate_order is None:
        args.candidate_order = (
            "top-down-left-right" if args.use_llm_prior else "exterior"
        )
    if args.enable_direct_release is None:
        # Direct release is a physics-planner capability, not LLM information.
        # Keep it enabled in both prior and no-prior runs so the ablation changes
        # only semantic ordering/operations instead of silently weakening the
        # baseline planner. It can still be disabled explicitly from the CLI.
        args.enable_direct_release = True

    print("\n========== Unified Approach Configuration ==========")
    print("Approach:", approach_config.display_name, f"({args.approach})")
    print(
        "[Planner mode] LLM prior=", args.use_llm_prior,
        "| strict_prior=", approach_config.strict_prior,
        "| candidate_order=", args.candidate_order,
        "| direct_release=", args.enable_direct_release,
        "| search_strategy=", args.search_strategy,
        "| rotation=", args.rotation,
        "| diagonal_actions=", not args.disable_diagonal_actions,
        "| blocker_guidance=", approach_config.enable_blocker_guidance,
        "| structural_guidance=", approach_config.enable_structural_guidance,
    )

    if args.normalize_size <= 0:
        parser.error("--normalize-size must be positive.")
    if args.collision_face_num < 4:
        parser.error("--collision-face-num must be at least 4.")
    if args.sdf_min_res > args.sdf_max_res:
        parser.error("--sdf-min-res cannot exceed --sdf-max-res.")
    if args.friction < 0.01:
        parser.error("Genesis requires --friction >= 0.01 in current releases.")
    if args.penetration_tolerance < 0:
        parser.error("--penetration-tolerance must be non-negative.")
    if args.sweep_translation_step <= 0 or args.sweep_rotation_step <= 0:
        parser.error("Swept-collision sampling steps must be positive.")
    if args.min_removal_distance < 0 or args.min_removal_distance_factor < 0:
        parser.error("Removal-distance thresholds must be non-negative.")
    if args.path_max_time <= 0:
        parser.error("--path-max-time must be positive.")
    if args.max_progressive_depth < 1:
        parser.error("--max-progressive-depth must be at least 1.")
    if args.max_sequence_trials < 1:
        parser.error("--max-sequence-trials must be at least 1.")
    if args.auto_rotation_depth < 1:
        parser.error("--auto-rotation-depth must be at least 1.")
    if args.adaptive_beam_width < 1:
        parser.error("--adaptive-beam-width must be at least 1.")
    if args.adaptive_action_top_k < 6:
        parser.error("--adaptive-action-top-k must be at least 6.")
    if args.coupled_action_scale < 0:
        parser.error("--coupled-action-scale must be non-negative.")
    if args.rrt_max_nodes < 2:
        parser.error("--rrt-max-nodes must be at least 2.")
    if not 0.0 <= args.rrt_goal_bias <= 1.0:
        parser.error("--rrt-goal-bias must be in [0, 1].")
    if args.rrt_neighbor_count < 1:
        parser.error("--rrt-neighbor-count must be at least 1.")
    if args.rrt_connect_distance_factor <= 0:
        parser.error("--rrt-connect-distance-factor must be positive.")
    if args.rrt_goal_root_count < 1:
        parser.error("--rrt-goal-root-count must be at least 1.")
    if args.diagnostic_max_frames_per_trial < 2:
        parser.error("--diagnostic-max-frames-per-trial must be at least 2.")
    if args.diagnostic_reset_hold_seconds < 0:
        parser.error("--diagnostic-reset-hold-seconds must be non-negative.")
    if args.replay_fps <= 0:
        parser.error("--replay-fps must be positive.")
    if args.video_fps is None:
        args.video_fps = args.replay_fps
    if args.video_fps <= 0:
        parser.error("--video-fps must be positive.")
    if args.video_width < 16 or args.video_height < 16:
        parser.error("Video width and height must be at least 16 pixels.")
    if args.video_frame_stride < 1:
        parser.error("--video-frame-stride must be at least 1.")
    if args.video_step_hold_seconds < 0 or args.video_final_hold_seconds < 0:
        parser.error("Video hold durations must be non-negative.")

    assembly_dir = os.path.abspath(os.path.join(args.assets_root, args.dir, args.id))
    if not os.path.isdir(assembly_dir):
        raise FileNotFoundError(f"Assembly directory not found: {assembly_dir}")

    cache_dir = args.cache_dir or os.path.join(assembly_dir, ".genesis_cache_v14")
    output_dir = args.output_dir or os.path.join(
        assembly_dir, f"genesis_plan_result_{args.approach}"
    )

    if args.replay_only:
        if args.replay_backend == "cpu":
            selected_backend = gs.cpu
        elif args.replay_backend == "gpu":
            selected_backend = gs.gpu
        else:
            selected_backend = gs.cpu if args.cpu else gs.gpu
    else:
        selected_backend = gs.cpu if args.cpu else gs.gpu

    genesis_init_started = time.perf_counter()
    gs.init(
        backend=selected_backend,
        precision="32",
        seed=args.seed,
        logging_level="warning",
    )
    genesis_initialization_time = time.perf_counter() - genesis_init_started

    require_python_fcl()

    preprocessing_started = time.perf_counter()
    assets, normalization = preprocess_assembly_assets(
        assembly_dir=assembly_dir,
        cache_dir=cache_dir,
        normalize_size=args.normalize_size,
        collision_face_num=args.collision_face_num,
        simplify_aggressiveness=args.simplify_aggressiveness,
        color_seed=args.color_seed,
        rebuild_cache=args.rebuild_cache,
        force_hull_for_nonwatertight=args.force_hull_for_nonwatertight,
    )
    preprocessing_time = time.perf_counter() - preprocessing_started

    if args.replay_only:
        (
            summary,
            replay_sequence,
            replay_paths,
            diagnostic_events,
        ) = load_saved_replay(output_dir)
        print("[Replay] Base/reference part:", summary.get("base_part_id"))
        complete_success = bool(
            summary.get("status") == "Success"
            and summary.get("complete_sequence_available", False)
        )
        replay_mode = "final" if complete_success else "diagnostic"
        replay_save_video = (
            args.save_video if complete_success else args.save_diagnostic_video
        )
        replay_video_filename = (
            args.video_filename
            if complete_success
            else args.diagnostic_video_filename
        )
        try:
            visualize_disassembly_replay(
                assets=assets,
                sequence=replay_sequence,
                paths=replay_paths,
                dt=args.dt,
                replay_fps=args.replay_fps,
                repeat=args.replay_repeat,
                intro_seconds=args.replay_intro_seconds,
                output_dir=output_dir,
                save_video=replay_save_video,
                video_filename=replay_video_filename,
                video_width=args.video_width,
                video_height=args.video_height,
                video_fps=args.video_fps,
                video_step_hold_seconds=args.video_step_hold_seconds,
                video_final_hold_seconds=args.video_final_hold_seconds,
                video_frame_stride=args.video_frame_stride,
                show_viewer=not args.no_vis,
                replay_mode=replay_mode,
                diagnostic_events=diagnostic_events,
                diagnostic_reset_hold_seconds=args.diagnostic_reset_hold_seconds,
            )
        except KeyboardInterrupt:
            pass
        except Exception as exc:
            print(f"[Replay] Viewer stopped: {type(exc).__name__}: {exc}")
        finally:
            sys.stdout.flush()
            sys.stderr.flush()
            os._exit(0)

    explicit_prior_part_ids = None
    prior_file_for_run = None
    if args.approach == "full_phyd2a":
        prior_file_for_run = args.llm_prior_file
        if args.prior_part_ids:
            explicit_prior_part_ids = [
                x.strip() for x in args.prior_part_ids.split(",") if x.strip()
            ]

    llm_prior = load_llm_prior_plan(
        assets=assets,
        enabled=(args.approach == "full_phyd2a"),
        prior_file=prior_file_for_run,
        explicit_part_ids=explicit_prior_part_ids,
        default_part_ids=DEFAULT_LLM_PRIOR_NUT_IDS,
        axis_map=load_axis_map(args.nut_axis_map),
        enable_unscrew=not args.disable_llm_unscrew,
        default_turns=args.nut_unscrew_turns,
        default_distance=args.nut_unscrew_distance,
        default_frames=args.nut_unscrew_frames,
        default_handedness=args.nut_unscrew_handedness,
        default_axis_mode=args.axis_estimation_mode,
        explicit_base_part_id=args.base_part_id,
    )

    planner = GenesisProgressiveSequencePlanner(
        assets=assets,
        rotation=args.rotation,
        show_viewer=args.vis_plan,
        dt=args.dt,
        substeps=args.substeps,
        force_mag=args.force_mag,
        density=args.density,
        friction=args.friction,
        sdf_cell_size=args.sdf_cell_size,
        sdf_min_res=args.sdf_min_res,
        sdf_max_res=args.sdf_max_res,
        frame_skip=args.frame_skip,
        trans_dist_th=args.trans_dist_th,
        quat_dist_th=args.quat_dist_th,
        collision_face_num=args.collision_face_num,
        collision_mode=args.collision_mode,
        watertighten=args.watertighten,
        free_dof_order=args.free_dof_order,
        max_action_chunks=args.max_action_chunks,
        candidate_order=args.candidate_order,
        enable_direct_release=args.enable_direct_release,
        penetration_tolerance=args.penetration_tolerance,
        sweep_translation_step=args.sweep_translation_step,
        sweep_rotation_step=args.sweep_rotation_step,
        min_removal_distance=args.min_removal_distance,
        min_removal_distance_factor=args.min_removal_distance_factor,
        unscrew_axial_force=args.unscrew_axial_force,
        unscrew_torque=args.unscrew_torque,
        unscrew_probe_chunks=args.unscrew_probe_chunks,
        unscrew_max_chunks=args.unscrew_max_chunks,
        unscrew_stall_chunks=args.unscrew_stall_chunks,
        unscrew_lateral_factor=args.unscrew_lateral_factor,
        unscrew_force_growth=args.unscrew_force_growth,
        unscrew_max_force_scale=args.unscrew_max_force_scale,
        prior_unscrew_max_retries=args.prior_unscrew_max_retries,
        max_progressive_depth=args.max_progressive_depth,
        max_sequence_trials=args.max_sequence_trials,
        auto_rotation_depth=args.auto_rotation_depth,
        search_strategy=args.search_strategy,
        adaptive_beam_width=args.adaptive_beam_width,
        adaptive_action_top_k=args.adaptive_action_top_k,
        enable_diagonal_actions=not args.disable_diagonal_actions,
        coupled_action_scale=args.coupled_action_scale,
        diagnostic_max_frames_per_trial=args.diagnostic_max_frames_per_trial,
        rrt_max_nodes=args.rrt_max_nodes,
        rrt_goal_bias=args.rrt_goal_bias,
        rrt_neighbor_count=args.rrt_neighbor_count,
        rrt_connect_distance_factor=args.rrt_connect_distance_factor,
        rrt_goal_root_count=args.rrt_goal_root_count,
        enable_blocker_guidance=approach_config.enable_blocker_guidance,
        enable_structural_guidance=approach_config.enable_structural_guidance,
        approach_name=args.approach,
    )

    result = planner.plan_sequence(
        seq_max_time=args.seq_max_time,
        path_max_time=args.path_max_time,
        seed=args.seed,
        save_dir=output_dir,
        llm_prior=llm_prior,
        base_part_id=llm_prior.base_part_id,
        strict_prior=approach_config.strict_prior,
        prior_initial_depth=args.prior_initial_depth,
        parking_enabled=not args.disable_parking,
        parking_distance_scale=args.parking_distance_scale,
        parking_spacing_scale=args.parking_spacing_scale,
        parking_frames=args.parking_frames,
        verbose=True,
    )

    result["run_started_at_utc"] = run_started_at_utc
    result["run_start_epoch"] = run_start_epoch
    result["genesis_initialization_time"] = round(genesis_initialization_time, 6)
    result["preprocessing_time"] = round(preprocessing_time, 6)
    result["elapsed_before_save"] = round(time.perf_counter() - run_perf_start, 6)

    save_results(output_dir, result, assets, normalization, args)
    del planner
    gc.collect()
    time.sleep(0.3)

    complete_success = bool(
        result.get("status") == "Success"
        and result.get("complete_sequence_available", False)
    )
    replay_requested = (
        (complete_success and (args.save_video or not args.no_vis))
        or (
            not complete_success
            and bool(result.get("diagnostic_events"))
            and (args.save_diagnostic_video or not args.no_vis)
        )
    )
    if replay_requested:
        # A fresh process avoids native Genesis viewer state left by planning.
        # Incomplete runs are replayed in diagnostic mode instead of skipped.
        relaunch_replay_in_fresh_process()
    else:
        print("[Replay] Visualization and requested video recording are disabled.")


if __name__ == "__main__":
    main()
